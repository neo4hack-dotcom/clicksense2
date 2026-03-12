import os
import json
import uuid
import re
import math
import time
import requests as http_requests
import urllib3
from collections import Counter
from datetime import datetime, timezone, timedelta
from urllib.parse import urlparse
from dotenv import load_dotenv
from flask import Flask, request, jsonify, send_from_directory
from copy import deepcopy

try:
    from zoneinfo import ZoneInfo
except Exception:  # pragma: no cover - fallback for very old Python runtimes
    ZoneInfo = None

# Suppress InsecureRequestWarning for plain-HTTP endpoints
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


def _http_get(url, **kwargs):
    """GET with SSL verification disabled (supports self-signed certificates)."""
    kwargs.setdefault("verify", False)
    return http_requests.get(url, **kwargs)


def _http_post(url, **kwargs):
    """POST with SSL verification disabled (supports self-signed certificates)."""
    kwargs.setdefault("verify", False)
    return http_requests.post(url, **kwargs)


load_dotenv()

PORT = int(os.environ.get("PORT", 3000))
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DIST_DIR = os.path.join(BASE_DIR, "dist")

# ---------------------------------------------------------------------------
# In-memory console log buffer (real-time log streaming)
# ---------------------------------------------------------------------------
import threading as _threading
_console_buffer: list = []
_console_lock = _threading.Lock()
_CONSOLE_MAX = 500

def _log(msg: str, level: str = "info", source: str = "system") -> None:
    """Append a timestamped log entry to the in-memory console buffer and print it."""
    import datetime as _dt
    entry = {
        "ts": _dt.datetime.now().strftime("%H:%M:%S"),
        "level": level,
        "source": source,
        "msg": str(msg),
    }
    with _console_lock:
        _console_buffer.append(entry)
        if len(_console_buffer) > _CONSOLE_MAX:
            del _console_buffer[0]
    print(f"[{entry['ts']}] [{level.upper():<5}] [{source}] {msg}")

app = Flask(__name__, static_folder=DIST_DIR, static_url_path="")


# ---------------------------------------------------------------------------
# CORS
# ---------------------------------------------------------------------------
@app.after_request
def add_cors_headers(response):
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS"
    return response


@app.route("/api", methods=["OPTIONS"])
@app.route("/api/<path:path>", methods=["OPTIONS"])
def options_handler(path=""):
    return "", 204


# ---------------------------------------------------------------------------
# JSON file database
# ---------------------------------------------------------------------------
DB_DIR = os.path.join(BASE_DIR, ".data")
os.makedirs(DB_DIR, exist_ok=True)
DB_FILE = os.path.join(DB_DIR, "app.json")

OLD_DB_FILE = os.path.join(BASE_DIR, "app.json")
if os.path.exists(OLD_DB_FILE) and not os.path.exists(DB_FILE):
    os.rename(OLD_DB_FILE, DB_FILE)

DEFAULT_DB = {
    "users": [{"id": 1, "name": "Default User"}],
    "saved_queries": [],
    "query_history": [],
    "table_metadata": [],
    "knowledge_folders": [],
    "table_mappings": [],
    "fk_relations": [],
    "agent_manager_workflows": [],
    "agent_manager_runs": [],
}


def read_db():
    try:
        if os.path.exists(DB_FILE):
            with open(DB_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception as e:
        print(f"Error reading DB: {e}")
    return {**DEFAULT_DB}


def write_db(data):
    try:
        with open(DB_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
    except Exception as e:
        print(f"Error writing DB: {e}")


if not os.path.exists(DB_FILE):
    write_db(DEFAULT_DB)


# ---------------------------------------------------------------------------
# In-memory configuration
# ---------------------------------------------------------------------------
clickhouse_config = {
    "host": os.environ.get("CLICKHOUSE_HOST", "http://localhost:8123"),
    "username": os.environ.get("CLICKHOUSE_USER", "default"),
    "password": os.environ.get("CLICKHOUSE_PASSWORD", ""),
    "database": os.environ.get("CLICKHOUSE_DB", "default"),
    "databases": [os.environ.get("CLICKHOUSE_DB", "default")],
}

llm_config = {
    "provider": "ollama",
    "model": "llama3",
    "baseUrl": "http://localhost:11434",
    "apiKey": "",
    # Optional runtime overrides (mainly useful for local_http / n8n transports)
    "contextWindow": "",
    "maxOutputTokens": "",
}

knowledge_base = ""

rag_config = {
    "esHost": os.environ.get("ES_HOST", "http://localhost:9200"),
    "esIndex": os.environ.get("ES_INDEX", "clicksense_rag"),
    "esUsername": os.environ.get("ES_USER", ""),
    "esPassword": os.environ.get("ES_PASSWORD", ""),
    "embeddingModel": "",
    "topK": 5,
    "chunkSize": 500,
}


# ---------------------------------------------------------------------------
# ClickHouse helper
# ---------------------------------------------------------------------------
def _parse_clickhouse_url(url: str) -> dict:
    parsed = urlparse(url)
    return {
        "host": parsed.hostname or "localhost",
        "port": parsed.port or 8123,
        "secure": parsed.scheme == "https",
    }


def get_clickhouse_client(cfg=None):
    if cfg is None:
        cfg = clickhouse_config
    try:
        import clickhouse_connect
    except ImportError as exc:
        raise RuntimeError(
            "clickhouse-connect is not installed. Run: pip install clickhouse-connect"
        ) from exc
    parts = _parse_clickhouse_url(cfg["host"])
    return clickhouse_connect.get_client(
        host=parts["host"],
        port=parts["port"],
        username=cfg["username"],
        password=cfg["password"],
        database=cfg["database"],
        secure=parts["secure"],
    )


def _rows_to_dicts(result) -> list:
    cols = result.column_names
    rows = []
    for row in result.result_rows:
        record = {}
        for col, val in zip(cols, row):
            if isinstance(val, (int, float, str, bool)) or val is None:
                record[col] = val
            else:
                record[col] = str(val)
        rows.append(record)
    return rows


# ---------------------------------------------------------------------------
# Embedding helper — uses the configured local LLM connection
# ---------------------------------------------------------------------------
def _resolve_local_http_api_base(raw_base_url: str | None) -> str:
    """Return the API base root from a local_http URL.

    Examples:
    - http://localhost:8000/v1/chat/completions -> http://localhost:8000
    - http://host:1234/proxy/v1/chat/completions -> http://host:1234/proxy
    - localhost:8000/v1/chat/completions -> http://localhost:8000
    """
    from urllib.parse import urlparse as _urlparse

    raw = (raw_base_url or "").strip() or "http://localhost:8000"
    if "://" not in raw:
        raw = "http://" + raw
    parsed = _urlparse(raw)
    if not parsed.netloc:
        parsed = _urlparse("http://localhost:8000")

    base = f"{parsed.scheme}://{parsed.netloc}"
    path = parsed.path or ""
    lower_path = path.lower()

    v1_idx = lower_path.find("/v1")
    if v1_idx >= 0:
        prefix = path[:v1_idx]
    elif lower_path.endswith("/chat/completions"):
        prefix = path[:-len("/chat/completions")]
    elif lower_path.endswith("/completions"):
        prefix = path[:-len("/completions")]
    elif lower_path.endswith("/models"):
        prefix = path[:-len("/models")]
    elif lower_path.endswith("/embeddings"):
        prefix = path[:-len("/embeddings")]
    else:
        prefix = path.rstrip("/")

    return (base + prefix).rstrip("/")


def _get_embedding(text: str, rag_cfg: dict | None = None, llm_cfg: dict | None = None) -> list:
    """Compute an embedding using the current connection with optional runtime overrides.

    - Reuses the same LLM connection (provider/baseUrl/apiKey).
    - Allows a dedicated embedding model via rag_cfg['embeddingModel'].
    - Supports temporary unsaved overrides from UI calls.
    """
    eff_llm = dict(llm_config)
    if isinstance(llm_cfg, dict):
        eff_llm.update(llm_cfg)

    eff_rag = dict(rag_config)
    if isinstance(rag_cfg, dict):
        eff_rag.update(rag_cfg)

    provider = (eff_llm.get("provider") or "ollama").strip()
    base_url = (eff_llm.get("baseUrl") or "").rstrip("/")
    model = (eff_rag.get("embeddingModel") or "").strip() or eff_llm.get("model") or "llama3"
    headers = {"Content-Type": "application/json"}

    if provider == "ollama":
        ollama_url = base_url or "http://localhost:11434"
        # Ollama /api/embed (v0.3+) — supports batch; use single input as list
        resp = _http_post(
            f"{ollama_url}/api/embed",
            json={"model": model, "input": text},
            headers=headers,
            timeout=120,
        )
        if resp.ok:
            data = resp.json()
            # /api/embed returns {"embeddings": [[...], ...]}
            if "embeddings" in data and data["embeddings"]:
                return data["embeddings"][0]
        elif resp.status_code == 404 or (not resp.ok and "not found" in resp.text.lower()):
            # Model not found — no point trying the legacy endpoint with the same model
            is_fallback_model = not (eff_rag.get("embeddingModel") or "").strip()
            if is_fallback_model:
                raise Exception(
                    f"The LLM model '{model}' does not support embeddings via Ollama. "
                    f"Please configure a dedicated embedding model (e.g. nomic-embed-text, bge-m3, all-minilm) "
                    f"in Settings > Embedding Model, then click 'Save RAG Config'."
                )
            raise Exception(
                f"Embedding model '{model}' not found in Ollama. "
                f"Pull it first with: ollama pull {model}"
            )
        # Fallback: legacy /api/embeddings endpoint
        resp2 = _http_post(
            f"{ollama_url}/api/embeddings",
            json={"model": model, "prompt": text},
            headers=headers,
            timeout=120,
        )
        if not resp2.ok:
            err_body = resp2.text
            # Detect model-not-found or embedding-unsupported errors and give a
            # clear, actionable message instead of the raw Ollama error.
            is_model_error = resp2.status_code == 404 or "not found" in err_body.lower()
            is_fallback_model = not (eff_rag.get("embeddingModel") or "").strip()
            if is_model_error:
                if is_fallback_model:
                    raise Exception(
                        f"The LLM model '{model}' does not support embeddings via Ollama. "
                        f"Please configure a dedicated embedding model (e.g. nomic-embed-text, bge-m3, all-minilm) "
                        f"in Settings > Embedding Model, then click 'Save RAG Config'."
                    )
                else:
                    raise Exception(
                        f"Embedding model '{model}' not found in Ollama. "
                        f"Pull it first with: ollama pull {model}"
                    )
            raise Exception(f"Ollama embedding error {resp2.status_code}: {err_body}")
        data2 = resp2.json()
        if "embedding" in data2:
            return data2["embedding"]
        raise Exception(f"Unexpected Ollama embedding response: {list(data2.keys())}")

    elif provider == "local_http":
        api_base = _resolve_local_http_api_base(base_url)
        candidates = [f"{api_base}/v1/embeddings", f"{api_base}/embeddings"]
        if eff_llm.get("apiKey"):
            headers["Authorization"] = f"Bearer {eff_llm['apiKey']}"

        last_error = ""
        for endpoint in candidates:
            resp = _http_post(
                endpoint,
                json={"model": model, "input": text},
                headers=headers,
                timeout=120,
            )
            if not resp.ok:
                last_error = f"{resp.status_code}: {resp.text}"
                continue

            data = resp.json()
            if "data" in data and data["data"]:
                emb = data["data"][0].get("embedding")
                if isinstance(emb, list):
                    return emb
            if isinstance(data.get("embedding"), list):
                return data["embedding"]
            if isinstance(data.get("embeddings"), list) and data["embeddings"]:
                first = data["embeddings"][0]
                if isinstance(first, list):
                    return first
            raise Exception(f"Unexpected embedding response: {list(data.keys())}")

        raise Exception(
            f"Embedding error on local_http for model '{model}' "
            f"(tried {', '.join(candidates)}): {last_error}"
        )

    else:
        raise Exception(
            f"Embedding not supported for provider '{provider}'. Use 'ollama' or 'local_http'."
        )


# ---------------------------------------------------------------------------
# Elasticsearch helper
# ---------------------------------------------------------------------------
def _es_request(method: str, path: str, cfg: dict, **kwargs):
    """Make a request to Elasticsearch."""
    host = cfg.get("esHost", "http://localhost:9200").rstrip("/")
    url = f"{host}{path}"
    auth = None
    if cfg.get("esUsername"):
        auth = (cfg["esUsername"], cfg.get("esPassword", ""))
    fn = _http_post if method.upper() == "POST" else _http_get
    if method.upper() == "PUT":
        fn = lambda u, **kw: http_requests.put(u, verify=False, **kw)
    elif method.upper() == "DELETE":
        fn = lambda u, **kw: http_requests.delete(u, verify=False, **kw)
    elif method.upper() == "HEAD":
        fn = lambda u, **kw: http_requests.head(u, verify=False, **kw)

    return fn(url, auth=auth, **kwargs)


def _chunk_text(text: str, chunk_size: int) -> list:
    """Split text into chunks of roughly chunk_size chars, trying to split on newlines."""
    chunks = []
    while len(text) > chunk_size:
        split_at = text.rfind("\n", 0, chunk_size)
        if split_at == -1:
            split_at = chunk_size
        chunks.append(text[:split_at].strip())
        text = text[split_at:].strip()
    if text:
        chunks.append(text)
    return chunks


def _get_knowledge_context_by_similarity(query: str, top_k: int = 5) -> str | None:
    """Embed *query* and return the top-K most relevant knowledge chunks from ES.

    Returns a formatted context string on success, or None when ES is
    unavailable / not indexed so the caller can fall back gracefully.
    """
    try:
        query_embedding = _get_embedding(query)
    except Exception:
        return None  # embedding not configured — skip similarity search

    try:
        host = rag_config.get("esHost", "http://localhost:9200").rstrip("/")
        index = rag_config.get("esIndex", "clicksense_rag")
        auth = None
        if rag_config.get("esUsername"):
            auth = (rag_config["esUsername"], rag_config.get("esPassword", ""))

        knn_query = {
            "knn": {
                "field": "embedding",
                "query_vector": query_embedding,
                "k": top_k,
                "num_candidates": top_k * 10,
            },
            "_source": ["title", "content"],
        }

        resp = http_requests.post(
            f"{host}/{index}/_search",
            auth=auth,
            json=knn_query,
            verify=False,
            timeout=10,
        )

        if not resp.ok:
            return None

        hits = resp.json().get("hits", {}).get("hits", [])
        if not hits:
            return None

        parts = []
        for hit in hits:
            src = hit.get("_source", {})
            parts.append(f"[{src.get('title', 'Knowledge')}]\n{src.get('content', '')}")

        return "\n\n---\n\n".join(parts)

    except Exception:
        return None


def _ensure_es_index(cfg: dict, dims: int):
    """Create ES index with dense_vector mapping if it doesn't exist."""
    index = cfg.get("esIndex", "clicksense_rag")
    host = cfg.get("esHost", "http://localhost:9200").rstrip("/")
    auth = None
    if cfg.get("esUsername"):
        auth = (cfg["esUsername"], cfg.get("esPassword", ""))

    check = http_requests.head(f"{host}/{index}", auth=auth, verify=False, timeout=10)
    if check.status_code == 404:
        mapping = {
            "mappings": {
                "properties": {
                    "folder_id": {"type": "integer"},
                    "title": {"type": "text"},
                    "content": {"type": "text"},
                    "chunk_index": {"type": "integer"},
                    "embedding": {
                        "type": "dense_vector",
                        "dims": dims,
                        "index": True,
                        "similarity": "cosine",
                    },
                }
            }
        }
        r = http_requests.put(
            f"{host}/{index}",
            auth=auth,
            json=mapping,
            verify=False,
            timeout=10,
        )
        if not r.ok and r.status_code != 400:
            raise Exception(f"Failed to create ES index: {r.text}")


# ---------------------------------------------------------------------------
# LLM call helper (shared)
# ---------------------------------------------------------------------------
def _strip_llm_markdown(text: str) -> str:
    """Remove markdown code fences that LLMs sometimes wrap around JSON."""
    text = text.strip()
    # Remove ```json ... ``` or ``` ... ``` blocks
    for fence in ("```json", "```JSON", "```"):
        if text.startswith(fence):
            text = text[len(fence):]
            break
    if text.endswith("```"):
        text = text[:-3]
    return text.strip()


def _clean_llm_output(text: str) -> str:
    """Remove <think>...</think> blocks (DeepSeek etc.) and residual HTML tags."""
    import re
    text = re.sub(r'<think>.*?</think>', '', text, flags=re.DOTALL)
    text = re.sub(r'<[^>]+>', '', text)
    return text.strip()


def _describe_payload_shape(payload, depth: int = 0) -> str:
    """Return a compact structural description for debugging unknown LLM payloads."""
    if depth > 2:
        return "..."
    if isinstance(payload, dict):
        keys = list(payload.keys())
        preview = keys[:8]
        suffix = "..." if len(keys) > 8 else ""
        if preview:
            first_key = preview[0]
            child = _describe_payload_shape(payload.get(first_key), depth + 1)
            return f"dict(keys={preview}{suffix}, first[{first_key}]={child})"
        return "dict(keys=[])"
    if isinstance(payload, list):
        if not payload:
            return "list(len=0)"
        return f"list(len={len(payload)}, first={_describe_payload_shape(payload[0], depth + 1)})"
    return type(payload).__name__


def _extract_llm_content(payload) -> str:
    """Extract assistant text from heterogeneous provider payload schemas."""

    def _extract(value, depth: int = 0) -> str:
        if depth > 7 or value is None:
            return ""

        if isinstance(value, str):
            return value.strip()

        if isinstance(value, list):
            if not value:
                return ""
            parts = []
            for item in value:
                txt = _extract(item, depth + 1)
                if txt:
                    parts.append(txt)
            return "\n".join(parts).strip() if parts else ""

        if isinstance(value, dict):
            # Common structured text chunk shape: {"type":"text","text":"..."}
            if str(value.get("type", "")).lower() == "text":
                txt = _extract(value.get("text"), depth + 1)
                if txt:
                    return txt

            # Most likely text-carrying keys first.
            for key in (
                "content",
                "text",
                "response",
                "output",
                "answer",
                "completion",
                "generated_text",
                "reasoning_content",
                "reasoning",
                "assistant_response",
            ):
                if key in value:
                    txt = _extract(value.get(key), depth + 1)
                    if txt:
                        return txt

            # OpenAI-compatible / streamed chunk variants.
            for key in ("message", "delta"):
                if key in value:
                    txt = _extract(value.get(key), depth + 1)
                    if txt:
                        return txt

            # Common container keys used by gateways/proxies/webhooks.
            for key in ("choices", "data", "results", "messages", "items"):
                if key in value:
                    txt = _extract(value.get(key), depth + 1)
                    if txt:
                        return txt

            return ""

        return ""

    return _extract(payload, 0)


def _parse_response_json(resp, label: str = "LLM") -> dict:
    """Safely parse JSON from an HTTP response with descriptive errors.

    Also handles SSE (Server-Sent Events) streaming responses by extracting
    the last complete data chunk that contains the assistant message.
    """
    if not resp.text or not resp.text.strip():
        raise Exception(
            f"{label} returned an empty response body (HTTP {resp.status_code}). "
            "Check that the server is running and the model is loaded."
        )
    try:
        return resp.json()
    except Exception:
        pass

    # Fallback: try to parse as an SSE stream (lines starting with "data: ")
    text = resp.text.strip()
    if "data:" in text:
        import re
        chunks = re.findall(r"^data:\s*(.+)$", text, re.MULTILINE)
        # Accumulate all delta content from streaming chunks
        accumulated_content = ""
        has_valid_chunk = False
        for chunk in chunks:
            chunk = chunk.strip()
            if chunk == "[DONE]":
                continue
            try:
                chunk_data = json.loads(chunk)
                delta_content = (
                    chunk_data.get("choices", [{}])[0]
                    .get("delta", {})
                    .get("content") or ""
                )
                accumulated_content += delta_content
                has_valid_chunk = True
            except json.JSONDecodeError:
                continue
        if has_valid_chunk and accumulated_content:
            return {"choices": [{"message": {"content": accumulated_content}}]}
        # If no delta content found, try returning the last parseable chunk as-is
        for chunk in reversed(chunks):
            chunk = chunk.strip()
            if chunk == "[DONE]":
                continue
            try:
                return json.loads(chunk)
            except json.JSONDecodeError:
                continue

    raise Exception(
        f"{label} response is not valid JSON (HTTP {resp.status_code}): "
        f"{resp.text[:300]!r}"
    )


def _parse_llm_json(content: str) -> dict:
    """Parse JSON from LLM output, handling markdown fences and extra surrounding text."""
    if not content or not content.strip():
        raise ValueError("LLM returned an empty response")

    def _normalize_candidate(txt: str) -> str:
        out = str(txt or "").strip().replace("\ufeff", "")
        out = out.replace("\r\n", "\n").replace("\r", "\n")
        out = out.replace("“", '"').replace("”", '"').replace("’", "'").replace("‘", "'")
        # Common malformed prefix from local outputs: "{n ..." instead of "{\n ..."
        out = re.sub(r"^(\{|\[)\s*n(?=\s*[\"{\[])", r"\1\n", out)
        out = re.sub(r"^(json|JSON)\s*", "", out)
        return out.strip()

    def _extract_first_json_block(txt: str) -> str:
        if not txt:
            return ""
        starts = [i for i, ch in enumerate(txt) if ch in "{["]
        closers = {"{": "}", "[": "]"}
        for start in starts[:12]:
            open_ch = txt[start]
            stack = [closers[open_ch]]
            in_str = False
            escaped = False
            for idx in range(start + 1, len(txt)):
                ch = txt[idx]
                if in_str:
                    if escaped:
                        escaped = False
                    elif ch == "\\":
                        escaped = True
                    elif ch == '"':
                        in_str = False
                    continue

                if ch == '"':
                    in_str = True
                    continue
                if ch in "{[":
                    stack.append(closers[ch])
                    continue
                if ch in "}]":
                    if not stack or ch != stack[-1]:
                        break
                    stack.pop()
                    if not stack:
                        return txt[start: idx + 1]
        return ""

    def _escape_newlines_in_strings(txt: str) -> str:
        if not txt:
            return txt
        out = []
        in_str = False
        escaped = False
        for ch in txt:
            if in_str:
                if escaped:
                    out.append(ch)
                    escaped = False
                    continue
                if ch == "\\":
                    out.append(ch)
                    escaped = True
                    continue
                if ch == '"':
                    out.append(ch)
                    in_str = False
                    continue
                if ch == "\n":
                    out.append("\\n")
                    continue
                if ch == "\t":
                    out.append("\\t")
                    continue
                out.append(ch)
                continue

            if ch == '"':
                in_str = True
            out.append(ch)
        return "".join(out)

    def _iter_candidates(txt: str):
        base = _normalize_candidate(_strip_llm_markdown(txt))
        candidates = [base]

        extracted = _extract_first_json_block(base)
        if extracted:
            candidates.append(_normalize_candidate(extracted))

        expanded = []
        for cand in candidates:
            if not cand:
                continue
            expanded.append(cand)
            # Remove trailing commas before closing braces/brackets.
            expanded.append(re.sub(r",(\s*[}\]])", r"\1", cand))
            # Quote unquoted keys: {action: "x"} -> {"action":"x"}.
            expanded.append(re.sub(r'([{,]\s*)([A-Za-z_][A-Za-z0-9_]*)(\s*:)', r'\1"\2"\3', cand))
            # Normalize Python literals.
            py_norm = re.sub(r"\bTrue\b", "true", cand)
            py_norm = re.sub(r"\bFalse\b", "false", py_norm)
            py_norm = re.sub(r"\bNone\b", "null", py_norm)
            expanded.append(py_norm)
            # Escape raw newlines/tabs inside JSON strings.
            expanded.append(_escape_newlines_in_strings(cand))

        seen = set()
        for cand in expanded:
            norm = _normalize_candidate(cand)
            if not norm or norm in seen:
                continue
            seen.add(norm)
            yield norm

    def _normalize_parsed_root(parsed):
        if isinstance(parsed, dict):
            return parsed
        # Some transports wrap a single JSON object into an array.
        if isinstance(parsed, list) and len(parsed) == 1 and isinstance(parsed[0], dict):
            return parsed[0]
        raise ValueError(
            "LLM JSON root must be an object. "
            f"Got {type(parsed).__name__} instead."
        )

    last_json_error = None
    for candidate in _iter_candidates(content):
        try:
            parsed = json.loads(candidate)
            return _normalize_parsed_root(parsed)
        except (json.JSONDecodeError, ValueError) as exc:
            last_json_error = exc
            continue

    raise ValueError(
        "LLM response is not valid JSON. "
        f"Raw content (first 500 chars): {content[:500]!r}"
        + (f" | last parser error: {last_json_error}" if last_json_error else "")
    )


# ---------------------------------------------------------------------------
# Token budget utilities
# ---------------------------------------------------------------------------
import re as _re

# Cache: "<provider>::<model>" -> max context window (tokens)
_model_context_cache: dict = {}

# Known context windows for common models
_KNOWN_CONTEXT_LIMITS: dict = {
    "llama3": 8192,
    "llama3.1": 131072,
    "llama3.2": 131072,
    "llama3.3": 131072,
    "mistral": 32768,
    "mixtral": 32768,
    "gemma2": 8192,
    "gemma3": 131072,
    "qwen2.5": 131072,
    "qwen2": 32768,
    "phi3": 131072,
    "phi4": 131072,
    "deepseek": 65536,
    "codellama": 16384,
    "vicuna": 4096,
    "gpt-4": 128000,
    "gpt-3.5": 16385,
    "claude": 200000,
}

# Budget: reserve this fraction of the effective context for system prompt context
_SYSTEM_PROMPT_BUDGET_FRACTION = 0.45

# Keep a conservative margin: local model tokenization/runtime context can differ
# from nominal architecture limits.
_CONTEXT_SAFETY_RATIO = 0.72
_CONTEXT_OUTPUT_RESERVE = 256

# For generic HTTP LLM providers, avoid optimistic context assumptions unless
# the user explicitly sets a context window. Many local OpenAI-compatible
# runtimes are configured around 2k-4k even when model architecture supports more.
_HTTP_CONTEXT_FALLBACK = 3072
_HTTP_CONTEXT_AUTO_CAP = 4096


def _estimate_tokens(text: str) -> int:
    """Conservative token estimate (biased high to avoid overflow)."""
    if not text:
        return 0
    # JSON-heavy prompts tokenize denser than plain prose.
    base = len(text) / 3.2
    structure_penalty = (
        text.count("{")
        + text.count("}")
        + text.count("[")
        + text.count("]")
        + text.count(":")
        + text.count(",")
        + text.count("\n")
    ) * 0.04
    return max(1, int(base + structure_penalty))


def _parse_int_like(value, default: int = 0) -> int:
    """Parse int from config values that may be int/float/string."""
    if value is None:
        return default
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str):
        m = _re.search(r"\d+", value)
        if m:
            try:
                return int(m.group(0))
            except Exception:
                return default
    return default


def _is_context_overflow_error(error_text: str) -> bool:
    """Detect provider error messages caused by context/token overflow."""
    text = (error_text or "").lower()
    if not text:
        return False
    markers = (
        "context window",
        "context length",
        "maximum context",
        "prompt is too long",
        "input is too long",
        "too many tokens",
        "token limit",
        "input length exceeds",
        "requested tokens",
        "num_ctx",
    )
    return any(m in text for m in markers)


def _get_ollama_runtime_context_limit(model: str, base_url: str) -> int | None:
    """Ask Ollama /api/show for the effective model/runtime context."""
    try:
        ollama_url = (base_url or "http://localhost:11434").rstrip("/")
        resp = _http_post(
            f"{ollama_url}/api/show",
            json={"name": model},
            timeout=20,
        )
        if not resp.ok:
            return None
        data = _parse_response_json(resp, "Ollama show")
    except Exception:
        return None

    candidates: list[int] = []

    details = data.get("details", {})
    if isinstance(details, dict):
        for key in ("context_length", "num_ctx"):
            v = _parse_int_like(details.get(key), 0)
            if v > 0:
                candidates.append(v)

    model_info = data.get("model_info", {})
    if isinstance(model_info, dict):
        for key, value in model_info.items():
            if "context_length" in str(key):
                v = _parse_int_like(value, 0)
                if v > 0:
                    candidates.append(v)

    # "parameters" is often a raw multiline string (e.g. "num_ctx 8192").
    params_raw = data.get("parameters")
    if isinstance(params_raw, str):
        for pattern in (r"\bnum_ctx\s+(\d+)\b", r"\bcontext_length\s+(\d+)\b"):
            m = _re.search(pattern, params_raw, flags=_re.IGNORECASE)
            if m:
                v = _parse_int_like(m.group(1), 0)
                if v > 0:
                    candidates.append(v)

    if not candidates:
        return None

    # Use the smallest discovered limit: this is safer than architecture max.
    return max(512, min(candidates))


def _get_model_context_limit() -> int:
    """Return the context window (tokens) for the current LLM model.

    Order of precedence:
    1. In-memory cache (from a previous call).
    2. Provider-specific runtime detection (Ollama) or conservative known cap (HTTP).
    3. Known limits dict (non-HTTP providers).
    4. Ask the LLM itself, cache the result (last resort).
    5. Conservative default (4 096 tokens).
    """
    provider = (llm_config.get("provider") or "ollama").lower()
    model = (llm_config.get("model") or "unknown").lower()
    cache_key = f"{provider}::{model}"

    if cache_key in _model_context_cache:
        return _model_context_cache[cache_key]

    # Explicit override from config/UI (if provided) has highest priority.
    explicit = _parse_int_like(
        llm_config.get("contextWindow")
        or llm_config.get("maxContextTokens")
        or llm_config.get("numCtx"),
        0,
    )
    if explicit > 0:
        _model_context_cache[cache_key] = explicit
        print(f"[Token budget] Using explicit contextWindow={explicit} for '{cache_key}'")
        return explicit

    # For generic HTTP providers, keep context assumptions conservative.
    # Many OpenAI-compatible servers expose architecture model names but run with
    # smaller runtime context windows.
    if provider in ("local_http", "n8n"):
        for known_name, limit in _KNOWN_CONTEXT_LIMITS.items():
            if known_name in model:
                capped = max(512, min(limit, _HTTP_CONTEXT_AUTO_CAP))
                _model_context_cache[cache_key] = capped
                print(
                    f"[Token budget] Conservative HTTP context for '{cache_key}': "
                    f"{capped} tokens (from known={limit})"
                )
                return capped
        _model_context_cache[cache_key] = _HTTP_CONTEXT_FALLBACK
        print(
            f"[Token budget] Unknown HTTP model '{cache_key}', using fallback: "
            f"{_HTTP_CONTEXT_FALLBACK} tokens"
        )
        return _HTTP_CONTEXT_FALLBACK

    # For Ollama, prefer runtime context from /api/show over theoretical model limits.
    if provider == "ollama":
        runtime_limit = _get_ollama_runtime_context_limit(
            llm_config.get("model") or "llama3",
            llm_config.get("baseUrl") or "http://localhost:11434",
        )
        if runtime_limit:
            _model_context_cache[cache_key] = runtime_limit
            print(f"[Token budget] Ollama runtime context for '{model}': {runtime_limit} tokens")
            return runtime_limit

        # Safe fallback for local Ollama runs when runtime discovery is unavailable.
        fallback = 8192
        _model_context_cache[cache_key] = fallback
        print(
            f"[Token budget] Could not detect Ollama runtime context for '{model}', "
            f"using conservative fallback: {fallback} tokens"
        )
        return fallback

    for known_name, limit in _KNOWN_CONTEXT_LIMITS.items():
        if known_name in model:
            _model_context_cache[cache_key] = limit
            print(f"[Token budget] Known context limit for '{cache_key}': {limit} tokens")
            return limit

    # Unknown model – ask the LLM
    try:
        answer = _call_llm(
            "You are a helpful assistant. Answer with a single integer only.",
            [{"role": "user", "content": "What is your maximum context window size in tokens? Reply with just the number, no other text."}],
            temperature=0.0,
        )
        # Extract first integer-like sequence from the answer
        match = _re.search(r"\d[\d,. ]*", answer)
        if match:
            limit = int(_re.sub(r"[^0-9]", "", match.group()))
            _model_context_cache[cache_key] = limit
            print(f"[Token budget] Model '{cache_key}' self-reported context limit: {limit} tokens")
            return limit
    except Exception as exc:
        print(f"[Token budget] Could not query model context limit: {exc}")

    default = 4096
    _model_context_cache[cache_key] = default
    print(f"[Token budget] Unknown model '{cache_key}', using conservative default: {default} tokens")
    return default


def _get_effective_context_limit() -> int:
    """Return a conservative usable context limit with safety margin."""
    raw = _get_model_context_limit()
    effective = min(int(raw * _CONTEXT_SAFETY_RATIO), raw - _CONTEXT_OUTPUT_RESERVE)
    # Never go below 512 tokens to keep minimal functionality.
    return max(512, effective)


def _truncate_prompt_context(
    schema: dict,
    table_metadata: dict,
    knowledge_context: str,
    base_tokens: int,
) -> tuple:
    """Progressively shrink schema/metadata/knowledge to fit within the token budget.

    Returns (schema, table_metadata, knowledge_context) – possibly truncated.
    """
    context_limit = _get_effective_context_limit()
    budget = int(context_limit * _SYSTEM_PROMPT_BUDGET_FRACTION)
    remaining = budget - base_tokens

    if remaining <= 0:
        # Base prompt alone is already too large – send nothing extra
        print("[Token budget] Base prompt exceeds budget, sending empty schema/metadata/knowledge")
        return {}, {}, ""

    def _fits(s_str: str, m_str: str, k_str: str) -> bool:
        return _estimate_tokens(s_str + m_str + k_str) <= remaining

    schema_full = json.dumps(schema, indent=2)
    meta_full = json.dumps(table_metadata, indent=2)

    # Step 1 – full schema + metadata + knowledge
    if _fits(schema_full, meta_full, knowledge_context):
        return schema, table_metadata, knowledge_context

    # Step 2 – compress schema: keep only column names (drop types)
    schema_compact = {tbl: [c["name"] if isinstance(c, dict) else c for c in cols]
                      for tbl, cols in schema.items()}
    schema_compact_str = json.dumps(schema_compact, indent=2)
    if _fits(schema_compact_str, meta_full, knowledge_context):
        print("[Token budget] Schema compressed (column names only) to fit budget")
        return schema_compact, table_metadata, knowledge_context

    # Step 3 – compact schema + no metadata
    if _fits(schema_compact_str, "{}", knowledge_context):
        print("[Token budget] Table metadata dropped to fit budget")
        return schema_compact, {}, knowledge_context

    # Step 4 – compact schema + no metadata + truncated knowledge
    available_for_knowledge = remaining - _estimate_tokens(schema_compact_str) - 20
    if available_for_knowledge > 50:
        truncated_knowledge = knowledge_context[:available_for_knowledge * 4]
        print(f"[Token budget] Knowledge context truncated to ~{available_for_knowledge} tokens")
        return schema_compact, {}, truncated_knowledge

    # Step 5 – only table names (no columns), capped to remaining budget
    max_schema_tokens = max(40, remaining - 10)
    schema_names_only = {}
    consumed = 0
    for tbl in schema:
        cost = _estimate_tokens(tbl) + 3
        if schema_names_only and consumed + cost > max_schema_tokens:
            break
        schema_names_only[tbl] = []
        consumed += cost
    if not schema_names_only and schema:
        first_tbl = next(iter(schema.keys()))
        schema_names_only[first_tbl] = []
    print(
        "[Token budget] Sending capped table names only "
        f"({len(schema_names_only)}/{len(schema)}) – context extremely tight"
    )
    return schema_names_only, {}, ""


def _trim_messages_to_budget(
    messages: list,
    token_budget: int,
    keep_last: int = 10,
) -> list:
    """Keep the latest messages that fit into a token budget."""
    if not messages or token_budget <= 0:
        return []

    tail = messages[-keep_last:] if keep_last > 0 else messages
    trimmed_reversed = []
    consumed = 0
    for msg in reversed(tail):
        msg_tokens = _estimate_tokens(str(msg.get("content", ""))) + 8
        if trimmed_reversed and consumed + msg_tokens > token_budget:
            break
        if not trimmed_reversed and msg_tokens > token_budget:
            # Always keep at least the latest user message.
            trimmed_reversed.append(msg)
            break
        trimmed_reversed.append(msg)
        consumed += msg_tokens
    return list(reversed(trimmed_reversed))


def _truncate_text_to_budget(text: str, token_budget: int) -> str:
    """Trim free text to a token budget using the local estimator."""
    if not text:
        return ""
    if token_budget <= 0:
        return ""
    if _estimate_tokens(text) <= token_budget:
        return text
    # Convert budget back to chars with conservative multiplier.
    max_chars = max(80, int(token_budget * 3))
    return text[:max_chars]


def _compact_llm_inputs(
    system_prompt: str,
    messages: list,
    *,
    system_token_budget: int = 220,
    message_token_budget: int = 520,
    keep_last: int = 5,
) -> tuple[str, list]:
    """Build a compact prompt/messages pair for transport-level retries."""
    compact_system = _truncate_text_to_budget(system_prompt or "", system_token_budget)
    safe_messages = [
        {"role": m.get("role", "user"), "content": m.get("content", "")}
        for m in (messages or [])
    ]
    compact_messages = _trim_messages_to_budget(
        safe_messages,
        token_budget=message_token_budget,
        keep_last=keep_last,
    )
    if not compact_messages and safe_messages:
        compact_messages = [safe_messages[-1]]
    return compact_system, compact_messages


def _get_http_completion_budget() -> int:
    """Return a conservative completion budget for HTTP LLM transports."""
    explicit = _parse_int_like(
        llm_config.get("maxOutputTokens") or llm_config.get("maxTokens"),
        0,
    )
    if explicit > 0:
        return max(64, min(4096, explicit))

    provider = (llm_config.get("provider") or "ollama").lower()
    model = (llm_config.get("model") or "unknown").lower()
    cache_key = f"{provider}::{model}"

    context_hint = _parse_int_like(
        llm_config.get("contextWindow")
        or llm_config.get("maxContextTokens")
        or llm_config.get("numCtx"),
        0,
    )
    if context_hint <= 0:
        context_hint = int(_model_context_cache.get(cache_key, _HTTP_CONTEXT_FALLBACK))

    budget = int(context_hint * 0.22)
    return max(128, min(2048, budget))


def _summarize_agent_steps_for_prompt(
    steps: list,
    max_steps: int = 8,
    max_result_chars: int = 240,
) -> str:
    """Compact step history to reduce context pressure in agent prompts."""
    if not steps:
        return "None yet — this is the first step."

    selected = steps[-max_steps:]
    lines = []
    for s in selected:
        action_label = "SQL" if s.get("type") == "query" else "Knowledge search"
        status = "OK" if s.get("ok", True) else "FAILED"
        payload = (s.get("sql") or s.get("search_query") or "").strip()
        result_preview = str(s.get("result_summary", "")).replace("\n", " ")
        result_preview = result_preview[:max_result_chars]
        lines.append(
            f"Step {s.get('step', '?')} [{action_label} | {status}] "
            f"Reasoning: {s.get('reasoning', '')[:180]} | "
            f"Action: {payload[:220]} | "
            f"Result: {result_preview}"
        )
    return "\n".join(lines)


_READONLY_ALLOWED_PREFIXES = ("SELECT", "SHOW", "DESCRIBE", "EXPLAIN", "WITH")
_READONLY_FORBIDDEN_KEYWORDS = (
    "INSERT", "UPDATE", "DELETE", "DROP", "ALTER", "TRUNCATE", "CREATE",
    "RENAME", "ATTACH", "DETACH", "OPTIMIZE", "SYSTEM", "GRANT", "REVOKE",
    "KILL", "SET ROLE", "USE",
)
_SIMPLE_COMPAT_FORBIDDEN_FUNCTION_PATTERNS = (
    r"\bwindowfunnel\s*\(",
    r"\bretention\s*\(",
    r"\bsequencematch\s*\(",
    r"\bsequencecount\s*\(",
    r"\bsequencenextnode\s*\(",
    r"\bargmax\s*\(",
    r"\bargmin\s*\(",
    r"\btopk\s*\(",
    r"\buniqhll12\s*\(",
    r"\buniqtheta\s*\(",
    r"\bquantiles?[a-z0-9_]*\s*\(",
    r"\bjsonextract[a-z0-9_]*\s*\(",
    r"\bsumif\s*\(",
    r"\bcountif\s*\(",
)
_SIMPLE_COMPAT_FORBIDDEN_DATE_FUNCTION_PATTERNS = (
    r"\btoyear\s*\(",
    r"\btomonth\s*\(",
    r"\btoquarter\s*\(",
    r"\byear\s*\(",
    r"\bmonth\s*\(",
    r"\bdate_trunc\s*\(",
    r"\btostartof[a-z0-9_]*\s*\(",
    r"\bextract\s*\(",
)
_DATE_LITERAL_PATTERN = (
    r"(?:'\d{4}-\d{2}-\d{2}(?:[ T]\d{2}:\d{2}:\d{2})?'"
    r"|toDate\s*\(\s*'\d{4}-\d{2}-\d{2}'\s*\)"
    r"|toDateTime\s*\(\s*'\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}:\d{2}'\s*\))"
)


def _classify_clickhouse_error(error_text: str) -> str:
    """Map raw ClickHouse errors to stable categories for retry strategies."""
    text = (error_text or "").lower()
    if not text:
        return "unknown"
    if "syntax" in text or "parse" in text:
        return "syntax_error"
    if "unknown table" in text or "doesn't exist" in text:
        return "unknown_table"
    if "unknown column" in text or "missing columns" in text:
        return "unknown_column"
    if "type mismatch" in text or "cannot parse" in text or "illegal type" in text:
        return "type_mismatch"
    if "memory limit" in text:
        return "memory_limit"
    if "timeout" in text or "max_execution_time" in text:
        return "timeout"
    if "not enough privileges" in text or "readonly" in text:
        return "permission"
    if "aggregate function" in text:
        return "aggregation_error"
    if "join" in text and ("not found" in text or "cannot" in text):
        return "join_error"
    if "simple compatibility" in text or "between" in text or "date filter" in text:
        return "simple_compat"
    return "execution_error"


def _validate_simple_clickhouse_sql(statement: str) -> None:
    """Validate SQL against simple compatibility + strict date filtering rules."""
    sql = (statement or "").strip()
    low = sql.lower()

    for pattern in _SIMPLE_COMPAT_FORBIDDEN_FUNCTION_PATTERNS:
        if re.search(pattern, low, flags=re.IGNORECASE):
            raise ValueError(
                "Simple compatibility mode: advanced ClickHouse functions are not allowed. "
                "Use COUNT, SUM, AVG, MIN, MAX, COUNT DISTINCT, ROUND, COALESCE."
            )

    for pattern in _SIMPLE_COMPAT_FORBIDDEN_DATE_FUNCTION_PATTERNS:
        if re.search(pattern, low, flags=re.IGNORECASE):
            raise ValueError(
                "Simple compatibility mode: date extraction functions are not allowed "
                "(year/month/toYear/toMonth/date_trunc/toStartOf*)."
            )

    # Date filtering policy:
    # - If a date literal is used, filter must use BETWEEN ... AND ... or '='
    # - Comparators >, <, >=, <= with date literals are forbidden.
    has_date_literal = bool(re.search(_DATE_LITERAL_PATTERN, sql, flags=re.IGNORECASE))
    if not has_date_literal:
        return

    left_cmp = rf"\b[\w`.]+\b\s*(?:>=|<=|>|<)\s*{_DATE_LITERAL_PATTERN}"
    right_cmp = rf"{_DATE_LITERAL_PATTERN}\s*(?:>=|<=|>|<)\s*\b[\w`.]+\b"
    if re.search(left_cmp, sql, flags=re.IGNORECASE) or re.search(right_cmp, sql, flags=re.IGNORECASE):
        raise ValueError(
            "Simple compatibility mode: date filters must use BETWEEN ... AND ... or '=' "
            "(no >, <, >=, <= with date values)."
        )

    has_between = bool(
        re.search(
            rf"\bbetween\s+{_DATE_LITERAL_PATTERN}\s+and\s+{_DATE_LITERAL_PATTERN}",
            sql,
            flags=re.IGNORECASE,
        )
    )
    has_equal_date = bool(
        re.search(rf"\b[\w`.]+\b\s*=\s*{_DATE_LITERAL_PATTERN}", sql, flags=re.IGNORECASE)
        or re.search(rf"{_DATE_LITERAL_PATTERN}\s*=\s*\b[\w`.]+\b", sql, flags=re.IGNORECASE)
    )
    if not (has_between or has_equal_date):
        raise ValueError(
            "Simple compatibility mode: when filtering by date, use either "
            "`column BETWEEN 'YYYY-MM-DD' AND 'YYYY-MM-DD'` or `column = 'YYYY-MM-DD'`."
        )


def _normalize_sql_fingerprint(sql: str) -> str:
    """Build a semantic-ish fingerprint to detect near-duplicate SQL attempts."""
    raw = (sql or "").strip().lower()
    if not raw:
        return ""
    # Remove comments
    raw = re.sub(r"--.*?$", " ", raw, flags=re.MULTILINE)
    raw = re.sub(r"/\*[\s\S]*?\*/", " ", raw)
    # Normalize literals to reduce false negatives on trivial value changes
    raw = re.sub(r"'(?:''|[^'])*'", "?", raw)
    raw = re.sub(r"\b\d+(?:\.\d+)?\b", "?", raw)
    # Normalize spacing/punctuation
    raw = re.sub(r"\s+", " ", raw).strip().rstrip(";")
    return raw


_MEMORY_SQL_PLACEHOLDER_RE = re.compile(r"\{\{\s*([a-zA-Z0-9_]+)\.([a-zA-Z0-9_]+)\s*\}\}")


def _sql_literal(value) -> str:
    """Serialize a primitive Python value into a safe SQL literal."""
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, (int, float)):
        return str(value)
    txt = str(value).replace("\\", "\\\\").replace("'", "''")
    return f"'{txt}'"


def _coerce_float(value):
    """Best-effort numeric coercion for descriptive summaries."""
    if value is None:
        return None
    if isinstance(value, bool):
        return float(int(value))
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        txt = value.strip()
        if not txt:
            return None
        if re.fullmatch(r"-?\d+(?:\.\d+)?", txt):
            try:
                return float(txt)
            except Exception:
                return None
    return None


def _build_query_result_summary(
    rows: list,
    columns: list,
    preview_rows: list,
    *,
    large_result_threshold: int = 250,
) -> str:
    """Return a compact query summary; auto-condense large datasets."""
    total_rows = len(rows)
    safe_columns = list(columns or [])
    if total_rows <= large_result_threshold:
        sample_lines = [
            ", ".join(str(r.get(c, "")) for c in safe_columns)
            for r in preview_rows[:10]
        ]
        return (
            f"{total_rows} row(s) returned. Columns: {safe_columns}."
            + (f"\nSample:\n{chr(10).join(sample_lines)}" if sample_lines else "")
        )

    inspected = rows[: min(total_rows, 300)]
    lines = [
        f"{total_rows} row(s) returned. Condensed summary applied for token safety.",
        f"Columns: {safe_columns}.",
        f"Descriptive stats computed on a {len(inspected)}-row analysis window.",
    ]

    for col in safe_columns[:8]:
        vals = []
        for row in inspected:
            if not isinstance(row, dict):
                continue
            val = row.get(col)
            if val is None:
                continue
            if isinstance(val, str) and not val.strip():
                continue
            vals.append(val)
        if not vals:
            continue

        numeric_vals = [v for v in (_coerce_float(vv) for vv in vals) if v is not None]
        numeric_coverage = len(numeric_vals) / max(1, len(vals))
        if numeric_coverage >= 0.8 and len(numeric_vals) >= 2:
            n_min = min(numeric_vals)
            n_max = max(numeric_vals)
            n_avg = sum(numeric_vals) / len(numeric_vals)
            lines.append(
                f"- {col}: min={n_min:.4g}, max={n_max:.4g}, avg={n_avg:.4g} "
                f"(sample n={len(numeric_vals)})."
            )
            continue

        uniq_all = set()
        examples = []
        for val in vals:
            txt = str(val).strip()
            if not txt:
                continue
            uniq_all.add(txt)
            if len(examples) < 4 and txt not in examples:
                examples.append(txt[:36])
        lines.append(
            f"- {col}: non-null sample={len(vals)}, approx unique(sample)={len(uniq_all)}, "
            f"examples={'; '.join(examples) if examples else 'n/a'}."
        )

    return "\n".join(lines[:14])


def _extract_intermediate_id_sets(
    rows: list,
    columns: list,
    *,
    max_scan_rows: int = 800,
    max_values_per_set: int = 200,
) -> dict:
    """Extract reusable ID lists from a query result for multi-hop retrieval."""
    if not rows or not columns:
        return {}
    out = {}
    for col in columns:
        col_name = str(col or "").strip()
        if not col_name:
            continue
        lower = col_name.lower()
        if lower != "id" and not lower.endswith("_id"):
            continue

        vals = []
        seen = set()
        for row in rows[:max_scan_rows]:
            if not isinstance(row, dict):
                continue
            value = row.get(col_name)
            if value is None:
                continue
            if isinstance(value, str):
                value = value.strip()
                if not value:
                    continue
            if isinstance(value, (list, tuple, dict, set)):
                continue
            fp = f"{type(value).__name__}:{value}"
            if fp in seen:
                continue
            seen.add(fp)
            vals.append(value)
            if len(vals) >= max_values_per_set:
                break
        if vals:
            out[col_name] = vals
    return out


def _resolve_sql_memory_placeholders(sql: str, working_memory: dict | None) -> str:
    """Resolve placeholders like {{step1.client_id}} or {{last.order_id}}."""
    raw = (sql or "").strip()
    if not raw or "{{" not in raw:
        return raw

    memory = working_memory or {}
    artifacts = memory.get("artifacts") or {}
    order = memory.get("order") or []

    def _replace(match: re.Match) -> str:
        artifact_ref = str(match.group(1) or "").strip()
        field_ref = str(match.group(2) or "").strip()
        if not artifact_ref or not field_ref:
            raise ValueError("Invalid working-memory placeholder.")

        target_key = artifact_ref
        if artifact_ref.lower() == "last":
            if not order:
                raise ValueError("Placeholder {{last.*}} used but working memory is empty.")
            target_key = str(order[-1])

        artifact = artifacts.get(target_key)
        if not isinstance(artifact, dict):
            raise ValueError(f"Unknown working-memory artifact '{artifact_ref}'.")

        id_sets = artifact.get("id_sets") or {}
        values = id_sets.get(field_ref)
        if values is None:
            for k, v in id_sets.items():
                if str(k).lower() == field_ref.lower():
                    values = v
                    break
        if not isinstance(values, list) or not values:
            raise ValueError(f"No stored values for placeholder '{artifact_ref}.{field_ref}'.")

        literals = ", ".join(_sql_literal(v) for v in values[:200])
        return literals or "NULL"

    return _MEMORY_SQL_PLACEHOLDER_RE.sub(_replace, raw)


def _force_limit_for_retry(sql: str, preferred_limit: int = 200, hard_cap: int = 5000) -> str:
    """Force a bounded LIMIT in retries to reduce timeout/memory failures."""
    statement = (sql or "").strip().rstrip(";")
    if not statement:
        return statement
    limit_match = re.search(r"\bLIMIT\s+(\d+)\b", statement, flags=re.IGNORECASE)
    if limit_match:
        current = int(limit_match.group(1))
        # In retry mode, prefer smaller result sets.
        target = max(1, min(preferred_limit, hard_cap, current))
        return re.sub(
            r"\bLIMIT\s+\d+\b",
            f"LIMIT {target}",
            statement,
            count=1,
            flags=re.IGNORECASE,
        )
    return f"{statement}\nLIMIT {max(1, min(preferred_limit, hard_cap))}"


def _apply_sql_retry_playbook(
    sql: str,
    *,
    error_class: str,
    error_text: str = "",
) -> str:
    """Apply deterministic SQL simplifications before spending an LLM retry."""
    statement = (sql or "").strip().rstrip(";")
    if not statement:
        return ""

    err = (error_class or "").strip().lower()
    candidate = ""

    if err in {"timeout", "memory_limit"}:
        # Remove ORDER BY to reduce sorting overhead and force a smaller LIMIT.
        no_order = re.sub(
            r"\bORDER\s+BY\b[\s\S]*?(?=\bLIMIT\b|$)",
            " ",
            statement,
            flags=re.IGNORECASE,
        )
        candidate = _force_limit_for_retry(no_order, preferred_limit=200, hard_cap=5000)
    elif err == "simple_compat":
        simplified = statement
        # Rewrites for common non-compatible aggregates.
        simplified = re.sub(
            r"\bcountIf\s*\(([^()]+)\)",
            r"sum(if(\1, 1, 0))",
            simplified,
            flags=re.IGNORECASE,
        )
        simplified = re.sub(
            r"\bsumIf\s*\(\s*([^,()]+)\s*,\s*([^()]+)\)",
            r"sum(if(\2, \1, 0))",
            simplified,
            flags=re.IGNORECASE,
        )
        # If date comparators are used with literals, degrade to BETWEEN same date.
        simplified = re.sub(
            rf"(\b[\w`.]+\b)\s*(?:>=|<=|>|<)\s*({_DATE_LITERAL_PATTERN})",
            r"\1 BETWEEN \2 AND \2",
            simplified,
            flags=re.IGNORECASE,
        )
        candidate = _force_limit_for_retry(simplified, preferred_limit=500, hard_cap=5000)
    elif err == "syntax_error":
        # Minimal cleanup only.
        candidate = statement
    elif err == "aggregation_error":
        # Keep query shape but reduce volume to improve stability.
        candidate = _force_limit_for_retry(statement, preferred_limit=300, hard_cap=5000)

    if not candidate:
        return ""
    if _normalize_sql_fingerprint(candidate) == _normalize_sql_fingerprint(statement):
        return ""

    # Keep only syntactically acceptable read-only single statement here.
    try:
        checked = _normalize_sql_for_execution(
            candidate,
            read_only=True,
            default_limit=500,
            hard_limit_cap=5000,
        )
        if err == "simple_compat":
            _validate_simple_clickhouse_sql(checked)
        return checked
    except Exception:
        return ""


def _normalize_sql_for_execution(
    sql: str,
    *,
    read_only: bool,
    default_limit: int = 1000,
    hard_limit_cap: int = 5000,
) -> str:
    """Validate and normalize SQL before execution."""
    raw = (sql or "").strip()
    if not raw:
        raise ValueError("Empty SQL query.")

    # Disallow multi-statement payloads.
    statement = raw.rstrip(";").strip()
    if ";" in statement:
        raise ValueError("Multiple SQL statements are not allowed.")
    upper = statement.upper()

    if read_only:
        if not upper.startswith(_READONLY_ALLOWED_PREFIXES):
            raise ValueError(
                "Read-only agent accepts only SELECT/SHOW/DESCRIBE/EXPLAIN/WITH queries."
            )
        for keyword in _READONLY_FORBIDDEN_KEYWORDS:
            if re.search(rf"\b{re.escape(keyword)}\b", upper):
                raise ValueError(f"Forbidden SQL keyword for read-only mode: {keyword}")

        # Ensure bounded result sets for better reliability with local LLM loops.
        limit_match = re.search(r"\bLIMIT\s+(\d+)\b", statement, flags=re.IGNORECASE)
        if limit_match:
            current = int(limit_match.group(1))
            if current > hard_limit_cap:
                statement = re.sub(
                    r"\bLIMIT\s+\d+\b",
                    f"LIMIT {hard_limit_cap}",
                    statement,
                    count=1,
                    flags=re.IGNORECASE,
                )
        elif upper.startswith(("SELECT", "WITH")):
            statement = f"{statement}\nLIMIT {default_limit}"

    return statement


def _execute_sql_guarded(
    sql: str,
    *,
    read_only: bool,
    enforce_simple_compat: bool = False,
    max_preview_rows: int = 20,
    max_execution_time: int = 15,
    default_limit: int = 1000,
    hard_limit_cap: int = 5000,
    client=None,
) -> dict:
    """Execute SQL with safety rules and stable structured output."""
    try:
        normalized_sql = _normalize_sql_for_execution(
            sql,
            read_only=read_only,
            default_limit=default_limit,
            hard_limit_cap=hard_limit_cap,
        )
        if enforce_simple_compat:
            _validate_simple_clickhouse_sql(normalized_sql)
    except Exception as exc:
        err = str(exc)
        return {
            "ok": False,
            "sql": (sql or "").strip(),
            "normalized_sql": (sql or "").strip(),
            "rows": [],
            "preview_rows": [],
            "columns": [],
            "total_rows": 0,
            "summary": f"Query blocked before execution: {err}",
            "error": err,
            "error_class": _classify_clickhouse_error(err),
        }

    try:
        db_client = client or get_clickhouse_client()
        settings = {"max_execution_time": max(1, int(max_execution_time))}
        if read_only:
            settings["readonly"] = 1

        sql_upper = normalized_sql.lstrip().upper()
        is_query = any(sql_upper.startswith(p) for p in _READONLY_ALLOWED_PREFIXES)
        if is_query:
            result = db_client.query(normalized_sql, settings=settings)
            rows = _rows_to_dicts(result)
            preview = rows[:max_preview_rows]
            columns = list(preview[0].keys()) if preview else list(result.column_names or [])
            summary = _build_query_result_summary(rows, columns, preview)
            return {
                "ok": True,
                "sql": sql,
                "normalized_sql": normalized_sql,
                "rows": rows,
                "preview_rows": preview,
                "columns": columns,
                "total_rows": len(rows),
                "summary": summary,
                "error": "",
                "error_class": "",
            }

        # Non-read query (used by writer agent only)
        db_client.command(normalized_sql)
        return {
            "ok": True,
            "sql": sql,
            "normalized_sql": normalized_sql,
            "rows": [],
            "preview_rows": [],
            "columns": [],
            "total_rows": 0,
            "summary": "Command executed successfully.",
            "error": "",
            "error_class": "",
        }
    except Exception as exc:
        err = str(exc)
        return {
            "ok": False,
            "sql": sql,
            "normalized_sql": normalized_sql,
            "rows": [],
            "preview_rows": [],
            "columns": [],
            "total_rows": 0,
            "summary": f"Query failed: {err}",
            "error": err,
            "error_class": _classify_clickhouse_error(err),
        }


def _is_list_tables_request(messages: list) -> bool:
    """Return True when the user's last message is asking to list all tables."""
    for msg in reversed(messages):
        if msg.get("role") == "user":
            text = msg.get("content", "").lower().strip()
            patterns = [
                "list of all tables", "show tables", "list tables",
                "all tables", "what tables", "which tables", "show all tables",
                "liste des tables", "lister les tables", "toutes les tables",
                "liste toutes les tables",
            ]
            return any(p in text for p in patterns)
    return False


def _messages_to_prompt(system_prompt: str, messages: list) -> str:
    """Flatten system prompt + conversation history into a single prompt string."""
    parts = []
    if system_prompt:
        parts.append(f"System: {system_prompt}")
    for msg in messages:
        role = msg.get("role", "user").capitalize()
        content = msg.get("content", "")
        parts.append(f"{role}: {content}")
    parts.append("Assistant:")
    return "\n\n".join(parts)


def _call_llm(system_prompt: str, messages: list, temperature: float = 0.7,
              language: str = None) -> str:
    """Call the configured LLM and return the content string.

    Supports providers: ollama, local_http, n8n.
    """
    # Append language instruction to the last user message if requested
    if language:
        lang_instruction = (
            "Vous DEVEZ écrire votre réponse ENTIÈRE en FRANÇAIS"
            if language == "fr"
            else "You MUST write your entire response in ENGLISH"
        )
        if messages:
            last = messages[-1]
            messages = messages[:-1] + [{
                **last,
                "content": last.get("content", "") + f"\n\n{lang_instruction}",
            }]
        else:
            system_prompt = system_prompt + f"\n\n{lang_instruction}"

    provider = llm_config.get("provider", "ollama")
    base_url = (llm_config.get("baseUrl") or "").rstrip("/")

    if provider == "local_http":
        endpoint = base_url or "http://localhost:8000"
        if "/v1/chat" not in endpoint:
            endpoint = endpoint.rstrip("/") + "/v1/chat/completions"
        headers = {"Content-Type": "application/json"}
        if llm_config.get("apiKey"):
            headers["Authorization"] = f"Bearer {llm_config['apiKey']}"

        def _post_local_http(sys_prompt: str, msg_list: list, max_tokens: int):
            return _http_post(
                endpoint,
                json={
                    "model": llm_config.get("model") or "local-model",
                    "messages": [{"role": "system", "content": sys_prompt}] + msg_list,
                    "temperature": temperature,
                    "max_tokens": max_tokens,
                    "stream": False,
                },
                headers=headers,
                timeout=120,
            )

        completion_budget = _get_http_completion_budget()
        resp = _post_local_http(system_prompt, messages, completion_budget)
        if not resp.ok:
            if _is_context_overflow_error(resp.text):
                compact_system, compact_messages = _compact_llm_inputs(
                    system_prompt,
                    messages,
                    system_token_budget=200,
                    message_token_budget=420,
                    keep_last=4,
                )
                compact_budget = max(96, min(512, completion_budget // 2))
                resp = _post_local_http(compact_system, compact_messages, compact_budget)
            if not resp.ok:
                raise Exception(f"local_http LLM Error: {resp.status_code} - {resp.text}")

        try:
            resp_data = _parse_response_json(resp, "local_http LLM")
        except Exception as exc:
            if _is_context_overflow_error(str(exc)):
                compact_system, compact_messages = _compact_llm_inputs(
                    system_prompt,
                    messages,
                    system_token_budget=180,
                    message_token_budget=360,
                    keep_last=3,
                )
                compact_budget = max(96, min(384, completion_budget // 2))
                retry_resp = _post_local_http(compact_system, compact_messages, compact_budget)
                if not retry_resp.ok:
                    raise Exception(f"local_http LLM Error: {retry_resp.status_code} - {retry_resp.text}")
                resp_data = _parse_response_json(retry_resp, "local_http LLM")
            else:
                raise

        content = _extract_llm_content(resp_data)
        if not content:
            raw_txt = str(getattr(resp, "text", "") or "").strip()
            if raw_txt and not raw_txt.startswith("{") and not raw_txt.startswith("["):
                content = raw_txt
        if not content:
            raise Exception(
                "local_http LLM returned empty content. "
                f"Payload shape: {_describe_payload_shape(resp_data)}"
            )
        return _clean_llm_output(_strip_llm_markdown(content))

    elif provider == "ollama":
        ollama_url = base_url or "http://localhost:11434"
        prompt = _messages_to_prompt(system_prompt, messages)
        body = {
            "model": llm_config.get("model", "llama3"),
            "prompt": prompt,
            "stream": False,
        }
        resp = _http_post(
            f"{ollama_url}/api/generate",
            json=body,
            timeout=120,
        )
        if not resp.ok:
            raise Exception(f"Ollama LLM Error: {resp.status_code} - {resp.text}")
        resp_data = _parse_response_json(resp, "Ollama LLM")
        content = _extract_llm_content(resp_data)
        if not content:
            raise Exception(
                "Ollama LLM returned empty content. "
                f"Payload shape: {_describe_payload_shape(resp_data)}"
            )
        return _clean_llm_output(_strip_llm_markdown(content))

    elif provider == "n8n":
        endpoint = base_url or ""
        if not endpoint:
            raise Exception("n8n provider requires a baseUrl (webhook URL)")
        headers = {"Content-Type": "application/json"}
        if llm_config.get("apiKey"):
            headers["Authorization"] = llm_config["apiKey"]  # raw value, no Bearer

        def _post_n8n(sys_prompt: str, msg_list: list, max_tokens: int):
            prompt = _messages_to_prompt(sys_prompt, msg_list)
            return _http_post(
                endpoint,
                json={
                    "prompt": prompt,
                    "model": llm_config.get("model", ""),
                    "max_tokens": max_tokens,
                    "timestamp": datetime.now(timezone.utc).isoformat(),
                },
                headers=headers,
                timeout=120,
            )

        completion_budget = _get_http_completion_budget()
        resp = _post_n8n(system_prompt, messages, completion_budget)
        if not resp.ok:
            if _is_context_overflow_error(resp.text):
                compact_system, compact_messages = _compact_llm_inputs(
                    system_prompt,
                    messages,
                    system_token_budget=180,
                    message_token_budget=380,
                    keep_last=4,
                )
                compact_budget = max(96, min(512, completion_budget // 2))
                resp = _post_n8n(compact_system, compact_messages, compact_budget)
            if not resp.ok:
                raise Exception(f"n8n LLM Error: {resp.status_code} - {resp.text}")

        try:
            resp_data = _parse_response_json(resp, "n8n LLM")
        except Exception as exc:
            if _is_context_overflow_error(str(exc)):
                compact_system, compact_messages = _compact_llm_inputs(
                    system_prompt,
                    messages,
                    system_token_budget=160,
                    message_token_budget=320,
                    keep_last=3,
                )
                compact_budget = max(96, min(384, completion_budget // 2))
                retry_resp = _post_n8n(compact_system, compact_messages, compact_budget)
                if not retry_resp.ok:
                    raise Exception(f"n8n LLM Error: {retry_resp.status_code} - {retry_resp.text}")
                resp_data = _parse_response_json(retry_resp, "n8n LLM")
            else:
                raise

        content = _extract_llm_content(resp_data)
        if not content:
            raw_txt = str(getattr(resp, "text", "") or "").strip()
            if raw_txt and not raw_txt.startswith("{") and not raw_txt.startswith("["):
                content = raw_txt
        if not content:
            raise Exception(
                "n8n LLM returned empty content. "
                f"Payload shape: {_describe_payload_shape(resp_data)}"
            )
        return _clean_llm_output(_strip_llm_markdown(content))

    else:
        raise Exception(f"Invalid LLM provider: {provider!r}")


# ---------------------------------------------------------------------------
# Configuration endpoints
# ---------------------------------------------------------------------------
@app.route("/api/config", methods=["GET"])
def get_config():
    cfg = dict(clickhouse_config)
    # Always expose databases as a list (back-compat: derive from database if missing)
    if not cfg.get("databases"):
        cfg["databases"] = [cfg.get("database", "default")]
    return jsonify({
        "clickhouseConfig": cfg,
        "llmConfig": llm_config,
        "knowledgeBase": knowledge_base,
    })


@app.route("/api/config", methods=["POST"])
def update_config():
    global clickhouse_config, llm_config, knowledge_base
    data = request.get_json()
    if data.get("clickhouse"):
        ch = data["clickhouse"]
        clickhouse_config.update(ch)
        # Sync: if databases list provided, keep database = databases[0]
        if ch.get("databases") and isinstance(ch["databases"], list):
            dbs = [d.strip() for d in ch["databases"] if d.strip()]
            if dbs:
                clickhouse_config["databases"] = dbs
                clickhouse_config["database"] = dbs[0]
        # If only database string provided (old clients), sync databases list
        elif ch.get("database") and not ch.get("databases"):
            clickhouse_config["databases"] = [ch["database"]]
    if data.get("llm"):
        llm_config.update(data["llm"])
        _model_context_cache.clear()
    if "knowledge" in data:
        knowledge_base = data["knowledge"]
    return jsonify({"success": True})


# ---------------------------------------------------------------------------
# RAG configuration endpoints
# ---------------------------------------------------------------------------
@app.route("/api/rag/config", methods=["GET"])
def get_rag_config():
    return jsonify(rag_config)


@app.route("/api/rag/config", methods=["POST"])
def update_rag_config():
    global rag_config
    data = request.get_json()
    # Drop legacy fields that were replaced by the LLM connection
    for legacy in ("embeddingUrl", "embeddingApiKey"):
        data.pop(legacy, None)
    rag_config.update(data)
    return jsonify({"success": True})


@app.route("/api/rag/test", methods=["POST"])
def test_elasticsearch():
    data = request.get_json()
    try:
        host = data.get("esHost", "http://localhost:9200").rstrip("/")
        auth = None
        if data.get("esUsername"):
            auth = (data["esUsername"], data.get("esPassword", ""))
        resp = http_requests.get(f"{host}/_cluster/health", auth=auth, verify=False, timeout=10)
        if not resp.ok:
            raise Exception(f"ES returned {resp.status_code}: {resp.text}")
        return jsonify({"success": True, "status": resp.json().get("status")})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/rag/embedding-models", methods=["POST"])
def get_embedding_models():
    """List models available at the LLM local endpoint for use as embedding models."""
    data = request.get_json(silent=True) or {}
    llm_override = data.get("llm") if isinstance(data.get("llm"), dict) else {}
    eff_llm = dict(llm_config)
    if isinstance(llm_override, dict):
        eff_llm.update(llm_override)

    provider = eff_llm.get("provider", "ollama")
    base_url = (eff_llm.get("baseUrl") or "").rstrip("/")
    headers = {"Content-Type": "application/json"}
    if eff_llm.get("apiKey"):
        headers["Authorization"] = f"Bearer {eff_llm['apiKey']}"
    try:
        if provider == "ollama":
            ollama_url = base_url or "http://localhost:11434"
            resp = _http_get(f"{ollama_url}/api/tags", headers=headers, timeout=10)
            if not resp.ok:
                return jsonify({"models": []})
            models = [m.get("name", "") for m in resp.json().get("models", [])]
            return jsonify({"models": [m for m in models if m]})
        elif provider == "local_http":
            api_base = _resolve_local_http_api_base(base_url)
            candidates = [f"{api_base}/v1/models", f"{api_base}/models"]
            for endpoint in candidates:
                resp = _http_get(endpoint, headers=headers, timeout=10)
                if not resp.ok:
                    continue
                payload = resp.json()
                models = [
                    m.get("id") or m.get("name", "")
                    for m in (payload.get("data", []) if isinstance(payload.get("data"), list) else [])
                    if isinstance(m, dict)
                ]
                # Fallback shape: {"models":[...]}
                if not models and isinstance(payload.get("models"), list):
                    models = [str(m) for m in payload.get("models", []) if str(m).strip()]
                return jsonify({"models": [m for m in models if m]})
            return jsonify({"models": []})
        return jsonify({"models": []})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/rag/test-embedding", methods=["POST"])
def test_embedding():
    data = request.get_json(silent=True) or {}
    rag_override = {}
    llm_override = {}

    if isinstance(data.get("ragConfig"), dict):
        rag_override.update(data["ragConfig"])
    else:
        for key in ("embeddingModel", "topK", "chunkSize", "esHost", "esIndex", "esUsername", "esPassword"):
            if key in data:
                rag_override[key] = data.get(key)

    if isinstance(data.get("llm"), dict):
        llm_override.update(data["llm"])
    elif isinstance(data.get("llmConfig"), dict):
        llm_override.update(data["llmConfig"])
    else:
        for key in ("provider", "model", "baseUrl", "apiKey", "contextWindow", "maxOutputTokens"):
            if key in data:
                llm_override[key] = data.get(key)

    try:
        embedding = _get_embedding(
            "test",
            rag_cfg=rag_override or None,
            llm_cfg=llm_override or None,
        )
        return jsonify({
            "success": True,
            "dims": len(embedding),
            "model": (rag_override.get("embeddingModel") or "").strip() or (llm_override.get("model") or llm_config.get("model")),
            "provider": (llm_override.get("provider") or llm_config.get("provider", "ollama")),
        })
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/rag/index", methods=["POST"])
def index_knowledge_to_es():
    """Embed all knowledge folders and index them in Elasticsearch."""
    data = request.get_json(silent=True) or {}
    cfg = data.get("ragConfig", rag_config)
    llm_override = data.get("llm") or data.get("llmConfig") or None
    db = read_db()
    folders = db.get("knowledge_folders", [])

    if not folders:
        return jsonify({"error": "No knowledge folders to index"}), 400

    try:
        # Determine embedding dimension by running one test embedding
        test_vec = _get_embedding("test", rag_cfg=cfg, llm_cfg=llm_override)
        dims = len(test_vec)
        _ensure_es_index(cfg, dims)

        host = cfg.get("esHost", "http://localhost:9200").rstrip("/")
        index = cfg.get("esIndex", "clicksense_rag")
        chunk_size = int(cfg.get("chunkSize", 500))
        auth = None
        if cfg.get("esUsername"):
            auth = (cfg["esUsername"], cfg.get("esPassword", ""))

        # Delete existing documents for a fresh index
        http_requests.post(
            f"{host}/{index}/_delete_by_query",
            auth=auth,
            json={"query": {"match_all": {}}},
            verify=False,
            timeout=30,
        )

        total_chunks = 0
        for folder in folders:
            content = folder.get("content", "")
            title = folder.get("title", "")
            folder_id = folder.get("id")
            if not content.strip():
                continue

            chunks = _chunk_text(content, chunk_size)
            for chunk_idx, chunk in enumerate(chunks):
                embedding = _get_embedding(
                    f"{title}\n\n{chunk}",
                    rag_cfg=cfg,
                    llm_cfg=llm_override,
                )
                doc = {
                    "folder_id": folder_id,
                    "title": title,
                    "content": chunk,
                    "chunk_index": chunk_idx,
                    "embedding": embedding,
                }
                http_requests.post(
                    f"{host}/{index}/_doc",
                    auth=auth,
                    json=doc,
                    verify=False,
                    timeout=30,
                )
                total_chunks += 1

        # Refresh index
        http_requests.post(f"{host}/{index}/_refresh", auth=auth, verify=False, timeout=10)

        return jsonify({"success": True, "indexed": total_chunks, "folders": len(folders)})

    except Exception as exc:
        print(f"Index error: {exc}")
        return jsonify({"error": str(exc)}), 500


@app.route("/api/rag/chat", methods=["POST"])
def rag_chat():
    """RAG chat: embed query → ES kNN search → LLM augmented generation."""
    data = request.get_json(silent=True) or {}
    query = data.get("query", "")
    history = data.get("history", [])
    cfg = data.get("ragConfig", rag_config)
    llm_override = data.get("llm") or data.get("llmConfig") or None
    top_k = int(cfg.get("topK", 5))

    if not query:
        return jsonify({"error": "Empty query"}), 400

    try:
        # 1. Embed the user query
        query_embedding = _get_embedding(query, rag_cfg=cfg, llm_cfg=llm_override)

        # 2. kNN search in Elasticsearch
        host = cfg.get("esHost", "http://localhost:9200").rstrip("/")
        index = cfg.get("esIndex", "clicksense_rag")
        auth = None
        if cfg.get("esUsername"):
            auth = (cfg["esUsername"], cfg.get("esPassword", ""))

        knn_query = {
            "knn": {
                "field": "embedding",
                "query_vector": query_embedding,
                "k": top_k,
                "num_candidates": top_k * 10,
            },
            "_source": ["title", "content", "folder_id", "chunk_index"],
        }

        es_resp = http_requests.post(
            f"{host}/{index}/_search",
            auth=auth,
            json=knn_query,
            verify=False,
            timeout=30,
        )

        sources = []
        context_parts = []

        if es_resp.ok:
            hits = es_resp.json().get("hits", {}).get("hits", [])
            for hit in hits:
                src = hit.get("_source", {})
                score = hit.get("_score", 0)
                # Normalize cosine score (ES returns 0-1 for cosine similarity mapping)
                sources.append({
                    "title": src.get("title", ""),
                    "score": round(float(score), 4),
                    "excerpt": src.get("content", "")[:300],
                    "folder_id": src.get("folder_id"),
                })
                context_parts.append(
                    f"[Source: {src.get('title', 'Unknown')}]\n{src.get('content', '')}"
                )

        # 3. Build augmented prompt
        context_text = "\n\n---\n\n".join(context_parts) if context_parts else "No relevant knowledge found."

        system_prompt = f"""You are a knowledgeable assistant with access to a specialized knowledge base.
Use the following retrieved context to answer the user's question accurately and concisely.
If the context doesn't contain enough information, say so clearly rather than guessing.

RETRIEVED CONTEXT:
{context_text}

Instructions:
- Base your answer primarily on the retrieved context
- Be specific and cite which source you are referencing when relevant
- If multiple sources provide complementary information, synthesize them
- If the context is insufficient, indicate what information is missing
"""

        # 4. Call LLM
        formatted_messages = [{"role": m["role"], "content": m["content"]} for m in history]
        formatted_messages.append({"role": "user", "content": query})

        answer = _call_llm(system_prompt, formatted_messages, temperature=0.3)
        return jsonify({"answer": answer, "sources": sources})

    except Exception as exc:
        print(f"RAG chat error: {exc}")
        return jsonify({"error": str(exc)}), 500


# ---------------------------------------------------------------------------
# ClickHouse connection test
# ---------------------------------------------------------------------------
@app.route("/api/clickhouse/test", methods=["POST"])
def test_clickhouse():
    data = request.get_json()
    try:
        # Support both databases[] and legacy database field
        databases = data.get("databases") or []
        db = databases[0].strip() if databases else data.get("database", "default")
        client = get_clickhouse_client({
            "host": data["host"],
            "username": data["username"],
            "password": data["password"],
            "database": db,
        })
        client.query("SELECT 1")
        return jsonify({"success": True})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


# ---------------------------------------------------------------------------
# LLM endpoints
# ---------------------------------------------------------------------------
@app.route("/api/llm/test", methods=["POST"])
def test_llm():
    """Test LLM connection by sending a simple hello prompt."""
    data = request.get_json()
    provider = data.get("provider")
    base_url = (data.get("baseUrl") or "").rstrip("/")
    test_prompt = "Hello, are you online? Respond with 'Yes'."
    try:
        if provider == "ollama":
            ollama_url = base_url or "http://localhost:11434"
            resp = _http_post(
                f"{ollama_url}/api/generate",
                json={"model": data.get("model", "llama3"), "prompt": test_prompt, "stream": False},
                timeout=30,
            )
            if not resp.ok:
                raise Exception(f"Ollama error: {resp.status_code} - {resp.text}")
        elif provider == "local_http":
            endpoint = base_url or "http://localhost:8000"
            if "/v1/chat" not in endpoint:
                endpoint = endpoint.rstrip("/") + "/v1/chat/completions"
            headers = {"Content-Type": "application/json"}
            if data.get("apiKey"):
                headers["Authorization"] = f"Bearer {data['apiKey']}"
            resp = _http_post(
                endpoint,
                json={
                    "model": data.get("model") or "local-model",
                    "messages": [{"role": "user", "content": test_prompt}],
                    "temperature": 0.7,
                },
                headers=headers,
                timeout=30,
            )
            if not resp.ok:
                raise Exception(f"local_http LLM error: {resp.status_code} - {resp.text}")
        elif provider == "n8n":
            endpoint = base_url
            if not endpoint:
                raise Exception("n8n provider requires a baseUrl (webhook URL)")
            headers = {"Content-Type": "application/json"}
            if data.get("apiKey"):
                headers["Authorization"] = data["apiKey"]
            resp = _http_post(
                endpoint,
                json={
                    "prompt": test_prompt,
                    "model": data.get("model", ""),
                    "timestamp": datetime.now(timezone.utc).isoformat(),
                },
                headers=headers,
                timeout=30,
            )
            if not resp.ok:
                raise Exception(f"n8n LLM error: {resp.status_code} - {resp.text}")
        return jsonify({"success": True})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/llm/models", methods=["GET"])
def get_llm_models():
    """Fetch available model list (Ollama and local_http only)."""
    try:
        provider = llm_config.get("provider", "ollama")
        base_url = (llm_config.get("baseUrl") or "").rstrip("/")
        if provider == "ollama":
            ollama_url = base_url or "http://localhost:11434"
            resp = _http_get(f"{ollama_url}/api/tags", timeout=10)
            if not resp.ok:
                raise Exception(f"Ollama error: {resp.status_code}")
            models = [m["name"] for m in resp.json().get("models", [])]
            return jsonify({"models": models})
        elif provider == "local_http":
            # Derive base from endpoint URL (strip /v1/chat/completions suffix if present)
            endpoint = base_url or "http://localhost:8000/v1/chat/completions"
            api_base = endpoint.split("/v1/chat")[0].split("/v1/models")[0]
            headers = {}
            if llm_config.get("apiKey"):
                headers["Authorization"] = f"Bearer {llm_config['apiKey']}"
            resp = _http_get(f"{api_base}/v1/models", headers=headers, timeout=10)
            if not resp.ok:
                if resp.status_code == 404:
                    return jsonify({"models": []})
                raise Exception(f"HTTP error: {resp.status_code}")
            models = [m.get("id") or m.get("name", "") for m in resp.json().get("data", []) if m.get("id") or m.get("name")]
            return jsonify({"models": models})
        return jsonify({"models": []})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


# ---------------------------------------------------------------------------
# Schema
# ---------------------------------------------------------------------------
@app.route("/api/schema", methods=["GET"])
def get_schema():
    try:
        client = get_clickhouse_client()
        # Resolve the list of databases to expose
        databases = clickhouse_config.get("databases") or []
        if not databases:
            databases = [clickhouse_config.get("database", "default")]
        databases = [d.strip() for d in databases if d.strip()]
        if not databases:
            databases = ["default"]

        multi_db = len(databases) > 1
        dbs_in = ", ".join(f"'{d}'" for d in databases)
        result = client.query(
            f"SELECT database, table, name, type FROM system.columns"
            f" WHERE database IN ({dbs_in})"
            f" ORDER BY database, table, name"
        )
        rows = _rows_to_dicts(result)
        schema: dict = {}
        for row in rows:
            db_name = row["database"]
            tbl = row["table"]
            key = f"{db_name}.{tbl}" if multi_db else tbl
            if key not in schema:
                schema[key] = []
            schema[key].append({"name": row["name"], "type": row["type"]})
        return jsonify({"schema": schema})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


# ---------------------------------------------------------------------------
# Users
# ---------------------------------------------------------------------------
@app.route("/api/users", methods=["GET"])
def get_users():
    db = read_db()
    return jsonify(db["users"])


# ---------------------------------------------------------------------------
# Table metadata
# ---------------------------------------------------------------------------
@app.route("/api/tables/metadata", methods=["GET"])
def get_table_metadata():
    db = read_db()
    result = {}
    for row in db["table_metadata"]:
        result[row["table_name"]] = {
            "description": row["description"],
            "is_favorite": bool(row["is_favorite"]),
        }
    return jsonify(result)


@app.route("/api/tables/metadata", methods=["POST"])
def update_table_metadata():
    data = request.get_json()
    table_name = data["table_name"]
    description = data.get("description", "")
    is_favorite = 1 if data.get("is_favorite") else 0

    db = read_db()
    idx = next((i for i, m in enumerate(db["table_metadata"]) if m["table_name"] == table_name), None)
    if idx is not None:
        db["table_metadata"][idx]["description"] = description
        db["table_metadata"][idx]["is_favorite"] = is_favorite
    else:
        db["table_metadata"].append({
            "table_name": table_name,
            "description": description,
            "is_favorite": is_favorite,
        })
    write_db(db)
    return jsonify({"success": True})


# ---------------------------------------------------------------------------
# Knowledge Base Folders
# ---------------------------------------------------------------------------
@app.route("/api/knowledge/folders", methods=["GET"])
def get_knowledge_folders():
    db = read_db()
    folders = db.get("knowledge_folders", [])
    return jsonify(folders)


@app.route("/api/knowledge/folders", methods=["POST"])
def create_knowledge_folder():
    data = request.get_json()
    db = read_db()
    if "knowledge_folders" not in db:
        db["knowledge_folders"] = []

    new_id = max((f["id"] for f in db["knowledge_folders"]), default=0) + 1
    now = datetime.now(timezone.utc).isoformat()
    folder = {
        "id": new_id,
        "title": data.get("title", ""),
        "content": data.get("content", ""),
        "created_at": now,
        "updated_at": now,
    }
    db["knowledge_folders"].append(folder)
    write_db(db)
    return jsonify(folder)


@app.route("/api/knowledge/folders/<int:folder_id>", methods=["PUT"])
def update_knowledge_folder(folder_id):
    data = request.get_json()
    db = read_db()
    folders = db.get("knowledge_folders", [])
    idx = next((i for i, f in enumerate(folders) if f["id"] == folder_id), None)
    if idx is None:
        return jsonify({"error": "Folder not found"}), 404

    now = datetime.now(timezone.utc).isoformat()
    if "title" in data:
        folders[idx]["title"] = data["title"]
    if "content" in data:
        folders[idx]["content"] = data["content"]
    folders[idx]["updated_at"] = now

    write_db(db)
    return jsonify(folders[idx])


@app.route("/api/knowledge/folders/<int:folder_id>", methods=["DELETE"])
def delete_knowledge_folder(folder_id):
    db = read_db()
    db["knowledge_folders"] = [f for f in db.get("knowledge_folders", []) if f["id"] != folder_id]
    write_db(db)
    return jsonify({"success": True})


# ---------------------------------------------------------------------------
# Table Mappings (friendly names for ClickHouse tables)
# ---------------------------------------------------------------------------
@app.route("/api/table-mappings", methods=["GET"])
def get_table_mappings():
    db = read_db()
    return jsonify(db.get("table_mappings", []))


@app.route("/api/table-mappings", methods=["POST"])
def upsert_table_mapping():
    data = request.get_json()
    table_name = data.get("table_name", "").strip()
    mapping_name = data.get("mapping_name", "").strip()
    if not table_name:
        return jsonify({"error": "table_name is required"}), 400

    db = read_db()
    mappings = db.get("table_mappings", [])
    idx = next((i for i, m in enumerate(mappings) if m["table_name"] == table_name), None)
    if mapping_name:
        if idx is not None:
            mappings[idx]["mapping_name"] = mapping_name
        else:
            mappings.append({"table_name": table_name, "mapping_name": mapping_name})
    else:
        # Empty mapping_name removes the entry
        if idx is not None:
            mappings.pop(idx)

    db["table_mappings"] = mappings
    write_db(db)
    return jsonify({"success": True})


@app.route("/api/table-mappings/<path:table_name>", methods=["DELETE"])
def delete_table_mapping(table_name):
    db = read_db()
    db["table_mappings"] = [m for m in db.get("table_mappings", []) if m["table_name"] != table_name]
    write_db(db)
    return jsonify({"success": True})


# ---------------------------------------------------------------------------
# FK Relations (foreign-key relationships discovered by the Key Identifier agent)
# ---------------------------------------------------------------------------
@app.route("/api/fk-relations", methods=["GET"])
def get_fk_relations():
    db = read_db()
    return jsonify(db.get("fk_relations", []))


@app.route("/api/fk-relations", methods=["POST"])
def create_fk_relation():
    data = request.get_json(silent=True) or {}
    required = ["table_a", "field_a", "table_b", "field_b"]
    for field in required:
        if not data.get(field, "").strip():
            return jsonify({"error": f"'{field}' is required"}), 400
    db = read_db()
    relations = db.setdefault("fk_relations", [])
    new_id = max((r["id"] for r in relations), default=0) + 1
    record = {
        "id": new_id,
        "table_a": data["table_a"].strip(),
        "field_a": data["field_a"].strip(),
        "table_b": data["table_b"].strip(),
        "field_b": data["field_b"].strip(),
        "direction": data.get("direction", "").strip(),
        "llm_reason": data.get("llm_reason", "").strip(),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    relations.append(record)
    write_db(db)
    return jsonify(record), 201


@app.route("/api/fk-relations/<int:relation_id>", methods=["DELETE"])
def delete_fk_relation(relation_id):
    db = read_db()
    db["fk_relations"] = [r for r in db.get("fk_relations", []) if r["id"] != relation_id]
    write_db(db)
    return jsonify({"success": True})


# ---------------------------------------------------------------------------
# Query history
# ---------------------------------------------------------------------------
@app.route("/api/history", methods=["POST"])
def add_history():
    data = request.get_json()
    db = read_db()
    new_id = max((h["id"] for h in db["query_history"]), default=0) + 1
    db["query_history"].append({
        "id": new_id,
        "user_id": data["user_id"],
        "query_text": data["query_text"],
        "sql": data["sql"],
        "created_at": datetime.now(timezone.utc).isoformat(),
    })
    write_db(db)
    return jsonify({"success": True})


@app.route("/api/history/<int:user_id>", methods=["GET"])
def get_history(user_id):
    db = read_db()
    history = [h for h in db["query_history"] if h["user_id"] == user_id]
    history.sort(key=lambda h: h["created_at"], reverse=True)
    return jsonify(history[:20])


# ---------------------------------------------------------------------------
# Saved queries (dashboard)
# ---------------------------------------------------------------------------
@app.route("/api/saved_queries", methods=["POST"])
def add_saved_query():
    data = request.get_json()
    db = read_db()
    new_id = max((q["id"] for q in db["saved_queries"]), default=0) + 1
    db["saved_queries"].append({
        "id": new_id,
        "user_id": data["user_id"],
        "name": data["name"],
        "sql": data["sql"],
        "config": json.dumps(data.get("config", {})),
        "visual_type": data.get("visual_type", "table"),
    })
    write_db(db)
    return jsonify({"success": True})


@app.route("/api/saved_queries/<int:user_id>", methods=["GET"])
def get_saved_queries(user_id):
    db = read_db()
    queries = [q for q in db["saved_queries"] if q["user_id"] == user_id]
    result = []
    for q in queries:
        q_copy = {**q}
        try:
            q_copy["config"] = json.loads(q_copy["config"])
        except Exception:
            q_copy["config"] = {}
        result.append(q_copy)
    return jsonify(result)


@app.route("/api/saved_queries/<int:query_id>", methods=["DELETE"])
def delete_saved_query(query_id):
    db = read_db()
    db["saved_queries"] = [q for q in db["saved_queries"] if q["id"] != query_id]
    write_db(db)
    return jsonify({"success": True})


# ---------------------------------------------------------------------------
# Query execution
# ---------------------------------------------------------------------------
@app.route("/api/query", methods=["POST"])
def execute_query():
    data = request.get_json()
    query = data["query"]
    enforce_simple_compat = bool(data.get("enforce_simple_compat", False))
    exec_result = _execute_sql_guarded(
        query,
        read_only=True,
        enforce_simple_compat=enforce_simple_compat,
        max_preview_rows=200,
        max_execution_time=15,
        default_limit=2000,
        hard_limit_cap=10000,
    )
    if not exec_result["ok"]:
        return jsonify({
            "error": exec_result["error"] or exec_result["summary"],
            "error_class": exec_result.get("error_class", ""),
            "normalized_sql": exec_result.get("normalized_sql", query),
        }), 500
    return jsonify({
        "data": exec_result["rows"],
        "normalized_sql": exec_result["normalized_sql"],
    })


# ---------------------------------------------------------------------------
# AI chat (SQL generation) with ambiguity detection
# ---------------------------------------------------------------------------
@app.route("/api/chat", methods=["POST"])
def chat():
    data = request.get_json()
    messages = data["messages"]
    schema = data.get("schema", {})
    table_metadata = data.get("tableMetadata", {})
    # tableMappingFilter: list of technical table names to restrict the schema to
    table_mapping_filter = data.get("tableMappingFilter", [])
    # use_knowledge_base: when False, skip knowledge context entirely
    use_knowledge_base = data.get("use_knowledge_base", True)

    # Build knowledge context from folders using similarity search when possible.
    # Only the most relevant chunks (matched via embedding kNN) are sent to the
    # LLM instead of the full knowledge base, reducing token usage and noise.
    db = read_db()
    folders = db.get("knowledge_folders", [])
    user_messages = [m for m in messages if m.get("role") == "user"]
    last_user_query = user_messages[-1]["content"] if user_messages else ""

    if not use_knowledge_base:
        knowledge_context = ""
    else:
        # Attempt similarity search first; fall back to full content if unavailable
        knowledge_context = None
        if last_user_query and folders:
            knowledge_context = _get_knowledge_context_by_similarity(
                last_user_query, top_k=int(rag_config.get("topK", 5))
            )

        if not knowledge_context:
            # Fallback: concatenate all folder content (old behaviour)
            knowledge_context = "\n\n".join(
                f"[{f['title']}]\n{f['content']}" for f in folders if f.get("content")
            ) or knowledge_base
            knowledge_context = knowledge_context[:20000]

    # Build a map of technical name -> friendly mapping name
    all_mappings = {m["table_name"]: m["mapping_name"] for m in db.get("table_mappings", [])}

    # Build FK relations context (confirmed by user via Key Identifier agent)
    fk_relations = db.get("fk_relations", [])
    if fk_relations:
        fk_lines = []
        for r in fk_relations:
            direction = r.get("direction") or f"{r['table_a']}.{r['field_a']} → {r['table_b']}.{r['field_b']}"
            fk_lines.append(f"  - {direction}")
        fk_context = (
            "KNOWN FOREIGN KEY RELATIONS — use these to build JOIN clauses when relevant:\n"
            + "\n".join(fk_lines)
        )
    else:
        fk_context = ""

    # If a filter is active, restrict the schema to selected tables only
    if table_mapping_filter:
        schema = {t: cols for t, cols in schema.items() if t in table_mapping_filter}
        table_metadata = {t: v for t, v in table_metadata.items() if t in table_mapping_filter}

    formatted_messages = [{"role": m["role"], "content": m["content"]} for m in messages]
    message_budget = int(_get_effective_context_limit() * 0.14)
    formatted_messages = _trim_messages_to_budget(
        formatted_messages,
        token_budget=message_budget,
        keep_last=10,
    )

    # ------------------------------------------------------------------
    # Special case: "list all tables" → generate SQL directly, no need
    # to send the full schema to the LLM (would be very token-expensive).
    # ------------------------------------------------------------------
    if _is_list_tables_request(messages):
        return jsonify({
            "sql": "SHOW TABLES",
            "explanation": "Lists all tables available in the current ClickHouse database.",
            "suggestedVisual": "table",
        })

    # Build a mapping note for the system prompt
    mapping_lines = []
    for tbl in schema:
        if tbl in all_mappings:
            mapping_lines.append(f"  - {tbl}  →  \"{all_mappings[tbl]}\"")
    mapping_note = (
        "The following tables have friendly business names. When communicating with the user use the friendly name, but always use the technical name in SQL:\n"
        + "\n".join(mapping_lines)
    ) if mapping_lines else ""
    mapping_note = _truncate_text_to_budget(mapping_note, token_budget=280)
    fk_context = _truncate_text_to_budget(fk_context, token_budget=220)

    # ------------------------------------------------------------------
    # Token budget: base prompt without dynamic content
    # ------------------------------------------------------------------
    base_prompt_template = """You are an expert ClickHouse data analyst and a proactive guide.
Your goal is to help the user query their database by asking smart clarifying questions BEFORE generating SQL.

KNOWLEDGE BASE — CRITICAL:
Before generating SQL or asking for clarification, you MUST consult the functional knowledge base provided.
If the knowledge base contains a definition or mapping for a concept mentioned by the user
(e.g., "a trade corresponds to a row in table toto"), use that information directly to build the SQL.
Do NOT ask for clarification on a concept that is already explained in the knowledge base.

PROACTIVE CLARIFICATION — CRITICAL:
You must ask for clarification in ALL of these situations. Prefer asking over guessing.

1. TABLE AMBIGUITY: If the user mentions a concept (e.g., "orders") but there are multiple tables that could match,
   AND the knowledge base does not resolve which table to use, return:
   {"needs_clarification": true, "question": "...", "options": ["table1 — description", "table2 — description"], "type": "table_selection"}

2. FIELD AMBIGUITY: If the user asks to display or use a field type (e.g., "date", "id", "name")
   and there are MULTIPLE fields of that type in the same table, return:
   {"needs_clarification": true, "question": "...", "options": ["field1 (type)", "field2 (type)"], "type": "field_selection"}

3. VALUE AMBIGUITY: If the user wants to filter on a categorical/enum field but does NOT specify the exact value
   (e.g., "orders from region X", "sales for department Y", "employees in position Z"), return:
   {"needs_clarification": true, "question": "Which value for [field]?", "options": [], "type": "value_selection",
    "context": {"table": "exact_table_name", "field": "exact_field_name"}}
   Leave options empty — the backend will populate them with real values from the database.

4. METRIC AMBIGUITY: If the user wants to measure/analyze something but has NOT specified how to calculate it
   (e.g., "show me sales" without saying count or sum), return:
   {"needs_clarification": true, "question": "How do you want to measure this?",
    "options": ["COUNT — number of records", "SUM — total amount", "AVG — average value", "COUNT DISTINCT — unique count", "MIN / MAX — extremes"],
    "type": "metric_selection"}

5. PERIOD AMBIGUITY: If the user uses vague time expressions like "recently", "last period", "this year" without a clear date range, return:
   {"needs_clarification": true, "question": "Which time period?",
    "options": ["Last 7 days", "Last 30 days", "Current month", "Last 3 months", "Last 6 months", "Current year"],
    "type": "period_selection"}

6. DIMENSION AMBIGUITY: If the user asks to group or break down data but does NOT specify the grouping dimension
   (e.g., "show sales by category" but multiple grouping options exist), return:
   {"needs_clarification": true, "question": "How do you want to group the results?",
    "options": ["field1 — description", "field2 — description", "field3 — description"],
    "type": "dimension_selection",
    "context": {"table": "exact_table_name"}}

CLICKHOUSE INSTRUCTIONS:
- Use SIMPLE and highly compatible ClickHouse SQL only.
- Prefer: COUNT, SUM, AVG, MIN, MAX, COUNT DISTINCT, ROUND, COALESCE, IF, simple GROUP BY.
- Avoid advanced/specialized functions: windowFunnel, retention, sequenceMatch, argMax, topK, uniqHLL12, quantiles*, JSONExtract*, sumIf, countIf.
- Keep SQL straightforward and robust.

DATE HANDLING — CRITICAL:
- When filtering by date, ALWAYS use either:
  1) `date_col BETWEEN 'YYYY-MM-DD' AND 'YYYY-MM-DD'`
  2) `date_col = 'YYYY-MM-DD'`
- NEVER use date extraction/transformation functions for filtering: no year(), month(), toYear(), toMonth(), date_trunc(), toStartOf*().
- NEVER use >, <, >=, <= with date literals.
- If the user asks a relative period, translate it to explicit literal dates and use BETWEEN.

When NOT ambiguous, return ONLY this JSON:
{
  "sql": "SELECT ...",
  "explanation": "Brief explanation of what the query does.",
  "suggestedVisual": "table" | "bar" | "line" | "pie"
}

Do not include markdown formatting. Just the raw JSON.
"""
    messages_tokens = sum(_estimate_tokens(m.get("content", "")) for m in formatted_messages)
    base_tokens = _estimate_tokens(base_prompt_template) + _estimate_tokens(mapping_note) + _estimate_tokens(fk_context) + messages_tokens

    # Truncate dynamic context (schema / metadata / knowledge) to fit the budget
    schema, table_metadata, knowledge_context = _truncate_prompt_context(
        schema, table_metadata, knowledge_context, base_tokens
    )

    # Build optional knowledge base block for the system prompt
    if knowledge_context:
        _kb_block = f"""Here is the functional knowledge base:
{knowledge_context}

KNOWLEDGE BASE — CRITICAL:
Before generating SQL or asking for clarification, you MUST consult the functional knowledge base above.
If the knowledge base contains a definition or mapping for a concept mentioned by the user
(e.g., "a trade corresponds to a row in table toto"), use that information directly to build the SQL.
Do NOT ask for clarification on a concept that is already explained in the knowledge base."""
    else:
        _kb_block = ""

    system_prompt = f"""You are an expert ClickHouse data analyst and a proactive guide.
Your goal is to help the user query their database by asking smart clarifying questions BEFORE generating SQL.

Here is the database schema:
{json.dumps(schema, indent=2)}

Here is the table metadata (functional descriptions):
{json.dumps(table_metadata, indent=2)}

{mapping_note}

{fk_context}

{_kb_block}

PROACTIVE CLARIFICATION — CRITICAL:
You must ask for clarification in ALL of these situations. Prefer asking over guessing.

1. TABLE AMBIGUITY: If the user mentions a concept (e.g., "orders") but there are multiple tables that could match,
   AND the knowledge base does not resolve which table to use, return:
   {{"needs_clarification": true, "question": "...", "options": ["table1 — description", "table2 — description"], "type": "table_selection"}}

2. FIELD AMBIGUITY: If the user asks to display or use a field type (e.g., "date", "id", "name")
   and there are MULTIPLE fields of that type in the same table, return:
   {{"needs_clarification": true, "question": "...", "options": ["field1 (type)", "field2 (type)"], "type": "field_selection"}}

3. VALUE AMBIGUITY: If the user wants to filter on a categorical/enum field but does NOT specify the exact value
   (e.g., "orders from region X", "sales for department Y", "employees in position Z"), return:
   {{"needs_clarification": true, "question": "Which value for [field]?", "options": [], "type": "value_selection",
    "context": {{"table": "exact_table_name", "field": "exact_field_name"}}}}
   Leave options empty — the backend will populate them with real values from the database.

4. METRIC AMBIGUITY: If the user wants to measure/analyze something but has NOT specified how to calculate it
   (e.g., "show me sales" without saying count or sum), return:
   {{"needs_clarification": true, "question": "How do you want to measure this?",
    "options": ["COUNT — number of records", "SUM — total amount", "AVG — average value", "COUNT DISTINCT — unique count", "MIN / MAX — extremes"],
    "type": "metric_selection"}}

5. PERIOD AMBIGUITY: If the user uses vague time expressions like "recently", "last period", "this year" without a clear date range, return:
   {{"needs_clarification": true, "question": "Which time period?",
    "options": ["Last 7 days", "Last 30 days", "Current month", "Last 3 months", "Last 6 months", "Current year"],
    "type": "period_selection"}}

6. DIMENSION AMBIGUITY: If the user asks to group or break down data but does NOT specify the grouping dimension
   (e.g., "show sales by category" but multiple grouping options exist), return:
   {{"needs_clarification": true, "question": "How do you want to group the results?",
    "options": ["field1 — description", "field2 — description", "field3 — description"],
    "type": "dimension_selection",
    "context": {{"table": "exact_table_name"}}}}

CLICKHOUSE INSTRUCTIONS:
- Use SIMPLE and highly compatible ClickHouse SQL only.
- Prefer: COUNT, SUM, AVG, MIN, MAX, COUNT DISTINCT, ROUND, COALESCE, IF, simple GROUP BY.
- Avoid advanced/specialized functions: windowFunnel, retention, sequenceMatch, argMax, topK, uniqHLL12, quantiles*, JSONExtract*, sumIf, countIf.
- Keep SQL straightforward and robust.

DATE HANDLING — CRITICAL:
- When filtering by date, ALWAYS use either:
  1) `date_col BETWEEN 'YYYY-MM-DD' AND 'YYYY-MM-DD'`
  2) `date_col = 'YYYY-MM-DD'`
- NEVER use date extraction/transformation functions for filtering: no year(), month(), toYear(), toMonth(), date_trunc(), toStartOf*().
- NEVER use >, <, >=, <= with date literals.
- If the user asks a relative period, translate it to explicit literal dates and use BETWEEN.

SQL QUERY — PRIORITY RULE:
The SQL query is ALWAYS the primary deliverable. Even for descriptive or analytical questions,
generate a SQL query that retrieves the relevant data. Never respond with text only when a SQL query
can answer the question. The "sql" field is mandatory in every non-clarification response.

When NOT ambiguous, return ONLY this JSON:
{{
  "sql": "SELECT ...",
  "explanation": "Brief explanation of what the query does.",
  "suggestedVisual": "table" | "bar" | "line" | "pie"
}}

Do not include markdown formatting. Just the raw JSON.
"""

    try:
        content = _call_llm(system_prompt, formatted_messages, temperature=0.3)
    except Exception as exc:
        if _is_context_overflow_error(str(exc)):
            print(f"[Chat] Context overflow detected, retrying with compact prompt: {exc}")
            compact_tables = list(schema.items())[:24]
            compact_schema = {}
            for tbl, cols in compact_tables:
                names = [c["name"] if isinstance(c, dict) else str(c) for c in cols][:8]
                compact_schema[tbl] = names
            compact_prompt = f"""You are an expert ClickHouse data analyst.
Return ONLY valid JSON:
{{
  "sql": "SELECT ...",
  "explanation": "Brief explanation.",
  "suggestedVisual": "table" | "bar" | "line" | "pie"
}}
If critical ambiguity remains, return:
{{"needs_clarification": true, "question": "...", "options": [], "type": "table_selection"}}

Database schema hint (truncated):
{json.dumps(compact_schema, indent=2)}
"""
            tiny_messages = _trim_messages_to_budget(
                formatted_messages,
                token_budget=int(_get_effective_context_limit() * 0.08),
                keep_last=6,
            )
            try:
                content = _call_llm(compact_prompt, tiny_messages, temperature=0.2)
            except Exception as retry_exc:
                print(f"Chat compact retry error: {retry_exc}")
                return jsonify({"error": str(retry_exc)}), 500
        else:
            print(f"Chat error: {exc}")
            return jsonify({"error": str(exc)}), 500

    conversation_id = (
        data.get("conversationId")
        or data.get("conversation_id")
        or str(uuid.uuid4())
    )
    _log(f"Chat SQL generated for conversation {conversation_id[:8]}", source="chat")
    try:
        parsed = _parse_llm_json(content)
    except ValueError:
        return jsonify({"explanation": content})

    # Post-process: if value_selection, populate options from real DB data
    if parsed.get("needs_clarification") and parsed.get("type") == "value_selection":
        ctx = parsed.get("context", {})
        tbl = ctx.get("table", "")
        fld = ctx.get("field", "")
        if tbl and fld:
            try:
                client = get_clickhouse_client()
                safe_tbl = tbl.replace("`", "")
                safe_fld = fld.replace("`", "")
                result = client.query(
                    f"SELECT DISTINCT `{safe_fld}` FROM `{safe_tbl}` "
                    f"WHERE `{safe_fld}` IS NOT NULL AND toString(`{safe_fld}`) != '' "
                    f"ORDER BY `{safe_fld}` LIMIT 25"
                )
                values = [str(row[0]) for row in result.result_rows if row[0] is not None]
                if values:
                    parsed["options"] = values
            except Exception as exc:
                print(f"value_selection DB lookup failed: {exc}")

    # Post-validation: enforce read-only + simple compatibility SQL policy.
    # This keeps /api/chat output aligned with the agent execution constraints.
    if not parsed.get("needs_clarification") and parsed.get("sql"):
        candidate_sql = str(parsed.get("sql", "")).strip()
        try:
            normalized_sql = _normalize_sql_for_execution(
                candidate_sql,
                read_only=True,
                default_limit=1000,
                hard_limit_cap=5000,
            )
            _validate_simple_clickhouse_sql(normalized_sql)
            parsed["sql"] = normalized_sql
        except Exception as compat_exc:
            rewrite_prompt = f"""Rewrite the SQL so it is ClickHouse simple-compatible.

USER QUESTION:
{last_user_query}

INVALID SQL:
{candidate_sql}

VALIDATION ERROR:
{compat_exc}

RULES:
- Read-only SQL only (SELECT/SHOW/DESCRIBE/EXPLAIN/WITH).
- Use only simple functions: COUNT, SUM, AVG, MIN, MAX, COUNT DISTINCT, ROUND, COALESCE, IF.
- No advanced functions (windowFunnel, retention, sequenceMatch, argMax, topK, uniqHLL12, quantiles*, JSONExtract*, sumIf, countIf).
- Date filters must use BETWEEN 'YYYY-MM-DD' AND 'YYYY-MM-DD' or = 'YYYY-MM-DD'.
- No year()/month()/toYear()/toMonth()/date_trunc()/toStartOf*().
- No >, <, >=, <= with date literals.
- Keep LIMIT <= 5000.

Return JSON only:
{{
  "sql": "SELECT ...",
  "explanation": "Brief explanation.",
  "suggestedVisual": "table" | "bar" | "line" | "pie"
}}"""
            try:
                fixed_content = _call_llm(
                    rewrite_prompt,
                    [{"role": "user", "content": "Rewrite now."}],
                    temperature=0.1,
                )
                fixed = _parse_llm_json(fixed_content)
                fixed_sql = str(fixed.get("sql", "")).strip()
                if not fixed_sql:
                    raise ValueError("No rewritten SQL returned.")
                normalized_fixed = _normalize_sql_for_execution(
                    fixed_sql,
                    read_only=True,
                    default_limit=1000,
                    hard_limit_cap=5000,
                )
                _validate_simple_clickhouse_sql(normalized_fixed)
                parsed["sql"] = normalized_fixed
                parsed["explanation"] = (
                    fixed.get("explanation")
                    or parsed.get("explanation")
                    or "SQL rewritten to satisfy compatibility constraints."
                )
                if fixed.get("suggestedVisual"):
                    parsed["suggestedVisual"] = fixed.get("suggestedVisual")
                parsed["compatibility_rewrite"] = True
            except Exception as rewrite_exc:
                return jsonify({
                    "error": (
                        "Generated SQL is not compatible with the enforced ClickHouse policy "
                        f"and auto-rewrite failed: {rewrite_exc}"
                    ),
                    "error_class": "simple_compat",
                    "invalid_sql": candidate_sql,
                }), 400

    return jsonify(parsed)


# ---------------------------------------------------------------------------
# Executive Summary — bullet-point synthesis
# ---------------------------------------------------------------------------
@app.route("/api/summarize_executive", methods=["POST"])
def summarize_executive():
    """Condense an executive summary into 5 key bullet points with risk flags.

    Request body (JSON):
        text  – the full executive summary / agent final answer
        lang  – optional language code ("fr" | "en", default "fr")

    Response (JSON):
        bullets – list of dicts {point: str, risk: bool, severity: "high"|"medium"|"info"}
        preamble – short introductory sentence
    """
    data = request.get_json(silent=True) or {}
    text = (data.get("text") or "").strip()
    lang = data.get("lang", "fr")
    count = max(5, min(10, _parse_int_like(data.get("count", 5), 5)))
    functional_focus = _coerce_bool_any(
        data.get("functional_focus", data.get("functionalFocus", False)),
        False,
    )

    if not text:
        return jsonify({"error": "No text provided."}), 400

    system_prompt = (
        "You are an expert executive analyst. Your task is to distill a detailed "
        f"analysis report into exactly {count} concise bullet points for a C-level audience. "
        "For each point, identify if it contains a risk or warning flag. "
        "Reply ONLY with valid JSON — no markdown, no extra text."
    )

    if lang == "fr":
        focus_note = (
            "Priorise fortement la lecture FONCTIONNELLE: explique les implications métier, "
            "les déductions opérationnelles et les conclusions fonctionnelles."
            if functional_focus
            else ""
        )
        format_instruction = (
            f"Réponds en FRANÇAIS. Rédige exactement {count} points synthétiques en langage "
            "de comité exécutif. Pour chaque point indique s'il constitue un risque/point "
            "d'attention (risk: true/false) et sa sévérité (severity: 'high', 'medium' ou 'info').\n"
            f"{focus_note + chr(10) if focus_note else ''}"
            "Format JSON attendu:\n"
            '{"preamble": "Phrase introductive courte.", "bullets": ['
            '{"point": "Texte du bullet point", "risk": false, "severity": "info"}, ...]}'
        )
    else:
        focus_note = (
            "Strongly prioritize FUNCTIONAL/business interpretation: implications, operational deductions, "
            "and business conclusions."
            if functional_focus
            else ""
        )
        format_instruction = (
            f"Write in ENGLISH. Draft exactly {count} concise bullet points in executive committee "
            "language. For each point indicate if it is a risk/watch point (risk: true/false) "
            "and its severity (severity: 'high', 'medium' or 'info').\n"
            f"{focus_note + chr(10) if focus_note else ''}"
            "Expected JSON format:\n"
            '{"preamble": "Short introductory sentence.", "bullets": ['
            '{"point": "Bullet point text", "risk": false, "severity": "info"}, ...]}'
        )

    messages = [{"role": "user", "content": f"Analysis to summarize:\n\n{text}\n\n{format_instruction}"}]

    try:
        raw = _call_llm(system_prompt, messages, temperature=0.3)
        try:
            result = json.loads(raw)
        except json.JSONDecodeError:
            import re as _re
            m = _re.search(r"\{.*\}", raw, _re.DOTALL)
            result = json.loads(m.group(0)) if m else {}

        bullets = result.get("bullets", [])
        # Ensure requested bullet count and normalize fields
        normalized = []
        for b in bullets[:count]:
            if isinstance(b, dict):
                normalized.append({
                    "point": str(b.get("point", "")),
                    "risk": bool(b.get("risk", False)),
                    "severity": b.get("severity", "info") if b.get("severity") in ("high", "medium", "info") else "info",
                })
        if len(normalized) < count:
            chunks = re.split(r"(?<=[\.\!\?])\s+", text)
            for chunk in chunks:
                c = str(chunk or "").strip()
                if not c:
                    continue
                normalized.append({
                    "point": c[:240],
                    "risk": False,
                    "severity": "info",
                })
                if len(normalized) >= count:
                    break
        while len(normalized) < count:
            idx = len(normalized) + 1
            normalized.append({
                "point": (
                    f"Point complémentaire {idx}: confirmer l'impact fonctionnel avec une vérification ciblée."
                    if lang == "fr"
                    else f"Additional point {idx}: validate functional impact with a targeted follow-up check."
                ),
                "risk": False,
                "severity": "info",
            })
        normalized = normalized[:count]
        return jsonify({
            "preamble": str(result.get("preamble", "")),
            "bullets": normalized,
            "requested_count": count,
            "actual_count": len(normalized),
        })
    except Exception as exc:
        return jsonify({"error": f"LLM error: {exc}"}), 500


# ---------------------------------------------------------------------------
# Agent Analysis — agentic loop (up to 10 distinct ClickHouse queries)
# ---------------------------------------------------------------------------

def _identify_relevant_tables(
    question: str,
    all_tables_schema: dict,
    all_table_metadata: dict,
    table_mappings: dict,
) -> dict:
    """Lightweight LLM call to identify which tables are relevant to the user's question.

    Returns a dict:
      {
        "tables": [list of table names to use],
        "needs_selection": bool,   # True when the LLM is unsure → ask the user
        "candidates": [all table names, sorted by likely relevance],
      }
    """
    all_table_names = list(all_tables_schema.keys())

    if not all_tables_schema:
        return {"tables": [], "needs_selection": False, "candidates": []}

    # Skip identification for tiny schemas
    if len(all_tables_schema) <= 3:
        return {"tables": all_table_names, "needs_selection": False, "candidates": all_table_names}

    # Build a compact one-line summary per table (names + columns, no full schema)
    table_summaries = []
    for tbl, cols in all_tables_schema.items():
        col_names = [c["name"] if isinstance(c, dict) else str(c) for c in cols]
        friendly = table_mappings.get(tbl, "")
        meta = all_table_metadata.get(tbl, {})
        description = (meta.get("description", "") if isinstance(meta, dict) else str(meta))[:150]
        line = f"- {tbl}"
        if friendly:
            line += f' (alias: "{friendly}")'
        if description:
            line += f" — {description}"
        shown_cols = col_names[:25]
        line += f"\n  Colonnes: {', '.join(shown_cols)}"
        if len(col_names) > 25:
            line += f" … (+{len(col_names) - 25} autres)"
        table_summaries.append(line)

    tables_overview = "\n".join(table_summaries)

    identification_prompt = f"""Tu es un assistant data analyst. Ta seule tâche est d'identifier quelles tables de base de données sont nécessaires pour répondre à la question de l'utilisateur.

TABLES DISPONIBLES (noms exacts à réutiliser tels quels) :
{tables_overview}

QUESTION UTILISATEUR :
{question}

RÈGLES :
1. Retourne UNIQUEMENT les noms de tables tels qu'ils apparaissent dans la liste ci-dessus (respect de la casse, aucune modification).
2. Inclus toutes les tables nécessaires pour les JOINs.
3. Si tu n'es pas sûr de quelle table utiliser, indique confidence = "low".
4. Si la question ne correspond à aucune table évidente, indique confidence = "low".

Réponds UNIQUEMENT avec du JSON valide (sans markdown) :
{{
  "relevant_tables": ["nom_exact_table_1", "nom_exact_table_2"],
  "confidence": "high" | "medium" | "low",
  "reasoning": "Explication courte"
}}"""

    # Build a normalized lookup for fuzzy matching: lowercase → real name
    normalized_lookup = {t.lower().strip(): t for t in all_table_names}

    try:
        content = _call_llm(
            identification_prompt,
            [{"role": "user", "content": "Identifie les tables pertinentes."}],
            temperature=0.0,
        )
        result = _parse_llm_json(content)
        if result and isinstance(result.get("relevant_tables"), list):
            confidence = result.get("confidence", "medium")

            # First pass: exact match
            exact_valid = [t for t in result["relevant_tables"] if t in all_tables_schema]

            # Second pass: case-insensitive fuzzy match for names the LLM may have mangled
            fuzzy_valid = list(exact_valid)
            for raw in result["relevant_tables"]:
                norm = raw.lower().strip()
                if norm in normalized_lookup and normalized_lookup[norm] not in fuzzy_valid:
                    fuzzy_valid.append(normalized_lookup[norm])
                    print(f"[Table ID] Fuzzy-matched '{raw}' → '{normalized_lookup[norm]}'")

            if fuzzy_valid:
                if confidence == "low":
                    print(
                        f"[Table ID] Low confidence — will ask user to confirm. "
                        f"Candidates: {fuzzy_valid}"
                    )
                    return {
                        "tables": fuzzy_valid,
                        "needs_selection": True,
                        "candidates": all_table_names,
                    }

                print(
                    f"[Table ID] {len(fuzzy_valid)}/{len(all_tables_schema)} tables — "
                    f"conf={confidence}: {fuzzy_valid} — {result.get('reasoning', '')}"
                )
                return {
                    "tables": fuzzy_valid,
                    "needs_selection": False,
                    "candidates": all_table_names,
                }

            # LLM returned table names that don't match anything → ask user
            print(
                f"[Table ID] No valid tables found in LLM response "
                f"({result.get('relevant_tables')}) — asking user to select"
            )
            return {"tables": [], "needs_selection": True, "candidates": all_table_names}

    except Exception as exc:
        print(f"[Table ID] Identification failed: {exc}")

    # LLM call failed entirely → ask user rather than blindly dumping everything
    print("[Table ID] LLM call failed — asking user to select tables")
    return {"tables": [], "needs_selection": True, "candidates": all_table_names}


_KNOWLEDGE_MODE_CONTEXT_ONCE = "kb_context_once"
_KNOWLEDGE_MODE_AGENTIC = "kb_agentic"
_KNOWLEDGE_MODE_SCHEMA_ONLY = "schema_only"
_KNOWLEDGE_MODE_MINIMAL = "minimal"


def _coerce_bool_any(value, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "on", "oui"}
    return bool(value)


def _normalize_knowledge_mode(raw_mode) -> str:
    txt = str(raw_mode or "").strip().lower()
    aliases = {
        _KNOWLEDGE_MODE_CONTEXT_ONCE: _KNOWLEDGE_MODE_CONTEXT_ONCE,
        "context_once": _KNOWLEDGE_MODE_CONTEXT_ONCE,
        "default": _KNOWLEDGE_MODE_CONTEXT_ONCE,
        "standard": _KNOWLEDGE_MODE_CONTEXT_ONCE,
        _KNOWLEDGE_MODE_AGENTIC: _KNOWLEDGE_MODE_AGENTIC,
        "agentic": _KNOWLEDGE_MODE_AGENTIC,
        "knowledge_agent": _KNOWLEDGE_MODE_AGENTIC,
        _KNOWLEDGE_MODE_SCHEMA_ONLY: _KNOWLEDGE_MODE_SCHEMA_ONLY,
        "schema_only_no_kb": _KNOWLEDGE_MODE_SCHEMA_ONLY,
        "no_kb_schema": _KNOWLEDGE_MODE_SCHEMA_ONLY,
        _KNOWLEDGE_MODE_MINIMAL: _KNOWLEDGE_MODE_MINIMAL,
        "minimal_no_context": _KNOWLEDGE_MODE_MINIMAL,
        "no_kb_no_context": _KNOWLEDGE_MODE_MINIMAL,
    }
    return aliases.get(txt, "")


def _resolve_knowledge_mode_flags(
    *,
    knowledge_mode_raw,
    use_knowledge_base_raw,
    use_knowledge_agent_raw,
) -> tuple[str, bool, bool]:
    mode = _normalize_knowledge_mode(knowledge_mode_raw)
    if not mode:
        use_knowledge_base = _coerce_bool_any(use_knowledge_base_raw, True)
        use_knowledge_agent = _coerce_bool_any(use_knowledge_agent_raw, False)
        if use_knowledge_base and not use_knowledge_agent:
            mode = _KNOWLEDGE_MODE_CONTEXT_ONCE
        elif use_knowledge_base and use_knowledge_agent:
            mode = _KNOWLEDGE_MODE_AGENTIC
        elif (not use_knowledge_base) and (not use_knowledge_agent):
            mode = _KNOWLEDGE_MODE_SCHEMA_ONLY
        else:
            mode = _KNOWLEDGE_MODE_MINIMAL

    mapping = {
        _KNOWLEDGE_MODE_CONTEXT_ONCE: (True, False),
        _KNOWLEDGE_MODE_AGENTIC: (True, True),
        _KNOWLEDGE_MODE_SCHEMA_ONLY: (False, False),
        _KNOWLEDGE_MODE_MINIMAL: (False, True),
    }
    use_knowledge_base, use_knowledge_agent = mapping.get(
        mode,
        mapping[_KNOWLEDGE_MODE_CONTEXT_ONCE],
    )
    return mode, use_knowledge_base, use_knowledge_agent


@app.route("/api/agent", methods=["POST"])
def agent_analysis():
    """Orchestrated multi-step analysis: the LLM autonomously decides which
    ClickHouse queries to run (up to MAX_AGENT_STEPS), analyses each result,
    and finally synthesises a detailed answer for the user.

    This endpoint is intentionally READ-ONLY.
    """

    data = request.get_json()
    MAX_AGENT_STEPS = max(1, min(50, int(data.get("maxSteps", 10))))
    # Global cap: at most 3 technical alternatives across the whole run.
    # Worst-case total SQL attempts = MAX_AGENT_STEPS + 3.
    MAX_TECHNICAL_RETRIES = 3

    user_question = data.get("question", "")
    schema = data.get("schema", {})
    table_metadata = data.get("tableMetadata", {})
    table_mapping_filter = data.get("tableMappingFilter", [])
    # Tables explicitly confirmed by the user via the selection UI (skips auto-identification)
    confirmed_tables = data.get("confirmedTables", [])
    knowledge_mode, use_knowledge_base, use_knowledge_agent = _resolve_knowledge_mode_flags(
        knowledge_mode_raw=data.get("knowledge_mode", data.get("knowledgeMode", "")),
        use_knowledge_base_raw=data.get("use_knowledge_base", True),
        use_knowledge_agent_raw=data.get("use_knowledge_agent", data.get("useKnowledgeAgent", False)),
    )
    # "knowledge agent" mode: do not inject static context in prompts.
    no_prompt_context_injection = use_knowledge_agent
    control_session_id = str(data.get("control_session_id", "")).strip()

    def _check_external_interrupt() -> str:
        """Allow session agent controls (pause/stop) to interrupt long runs."""
        if not control_session_id:
            return ""
        try:
            with _data_analyst_sessions_lock:
                sess = _data_analyst_sessions.get(control_session_id)
                if not sess:
                    return ""
                if sess.get("stop_requested"):
                    return "stopped"
                if sess.get("pause_requested"):
                    return "paused"
        except Exception:
            return ""
        return ""

    if not user_question.strip():
        return jsonify({"error": "No question provided"}), 400

    # Apply table filter
    if table_mapping_filter:
        schema = {t: cols for t, cols in schema.items() if t in table_mapping_filter}
        table_metadata = {t: v for t, v in table_metadata.items() if t in table_mapping_filter}

    # Knowledge base context (prefer semantic retrieval, keep fallback compact)
    db = read_db()
    folders = db.get("knowledge_folders", [])
    knowledge_context = ""
    if use_knowledge_base:
        kb_hit = None
        if user_question.strip() and folders:
            kb_hit = _get_knowledge_context_by_similarity(
                user_question.strip(),
                top_k=int(rag_config.get("topK", 5)),
            )
        if kb_hit:
            knowledge_context = kb_hit
        else:
            knowledge_context = "\n\n".join(
                f"[{f['title']}]\n{f['content']}" for f in folders if f.get("content")
            ) or knowledge_base
        knowledge_context = (knowledge_context or "")[:25000]

    # Build full mapping dict (used by table identification and mapping note)
    all_mappings = {m["table_name"]: m["mapping_name"] for m in db.get("table_mappings", [])}

    # ── Step 0: Table identification ──────────────────────────────────────────
    if confirmed_tables:
        # User already selected tables through the UI — trust their choice, skip LLM call
        schema = {t: cols for t, cols in schema.items() if t in confirmed_tables}
        table_metadata = {t: v for t, v in table_metadata.items() if t in confirmed_tables}
        print(f"[Table ID] Using user-confirmed tables: {list(schema.keys())}")
    else:
        # Ask the LLM to identify which tables are relevant BEFORE building the full
        # schema string, so we never send unnecessary context to the model.
        id_result = _identify_relevant_tables(
            user_question, schema, table_metadata, all_mappings
        )

        if id_result["needs_selection"]:
            # LLM is not confident enough — ask the user to pick the tables
            candidates = id_result["candidates"] or list(schema.keys())
            print(f"[Table ID] Asking user to confirm tables. Candidates: {candidates}")
            return jsonify({
                "needs_table_selection": True,
                "candidate_tables": candidates,
                "question": (
                    "Je n'ai pas pu identifier avec certitude quelles tables concernent "
                    "votre demande. Merci de sélectionner la ou les tables à analyser :"
                ),
                "steps": [],
                "final_answer": "",
            })

        relevant_tables = id_result["tables"]
        schema = {t: cols for t, cols in schema.items() if t in relevant_tables}
        table_metadata = {t: v for t, v in table_metadata.items() if t in relevant_tables}

    try:
        agent_client = get_clickhouse_client()
    except Exception as exc:
        return jsonify({"error": f"Connexion ClickHouse impossible: {exc}"}), 500

    # Accumulate steps
    steps: list = []
    seen_query_fingerprints: set[str] = set()
    context_injected_once = False
    halfway_step = max(1, (MAX_AGENT_STEPS + 1) // 2)

    def _sanitize_plan_steps(raw_steps, max_items: int) -> list[str]:
        """Normalize plan steps into concise, unique text items."""
        if not isinstance(raw_steps, list):
            return []
        clean = []
        seen = set()
        for item in raw_steps:
            text = str(item or "").strip()
            if not text:
                continue
            text = " ".join(text.split())
            if len(text) > 180:
                text = text[:180].rstrip() + "..."
            key = text.lower()
            if key in seen:
                continue
            seen.add(key)
            clean.append(text)
            if len(clean) >= max_items:
                break
        return clean

    def _build_fallback_plan() -> list[str]:
        fallback = [
            "Valider les tables/colonnes pertinentes et la métrique principale.",
            "Exécuter une requête de base avec filtre date explicite pour établir un baseline.",
            "Comparer le résultat par une dimension métier clé pour expliquer les écarts.",
            "Confirmer la robustesse (métrique alternative simple ou contrôle croisé).",
            "Conclure avec limites et niveau de confiance.",
        ]
        return fallback[:max(1, min(MAX_AGENT_STEPS, 5))]

    def _build_initial_agent_plan() -> dict:
        """Create an initial compact SQL action plan."""
        schema_brief = {}
        for tbl, cols in list(schema.items())[:12]:
            col_names = [c["name"] if isinstance(c, dict) else str(c) for c in cols][:8]
            schema_brief[tbl] = col_names

        plan_prompt = f"""You are planning an efficient ClickHouse analysis strategy.
Create a concise action plan for SQL exploration.

USER QUESTION:
{user_question}

TABLE HINTS (truncated):
{json.dumps(schema_brief, ensure_ascii=False, indent=2)}

CONSTRAINTS:
- Max action credits: {MAX_AGENT_STEPS}
- Prefer 3 to 5 focused SQL actions.
- Use only simple SQL functions (COUNT, SUM, AVG, MIN, MAX, COUNT DISTINCT, ROUND, COALESCE).
- Date filters must use BETWEEN 'YYYY-MM-DD' AND 'YYYY-MM-DD' or = 'YYYY-MM-DD'.
- Avoid advanced ClickHouse functions.

Return ONLY JSON:
{{
  "plan_steps": ["step 1", "step 2", "step 3"],
  "reasoning": "short strategy rationale"
}}"""
        try:
            max_plan_tokens = int(_get_effective_context_limit() * 0.55)
            plan_prompt = _truncate_text_to_budget(plan_prompt, max_plan_tokens)
            content = _call_llm(
                plan_prompt,
                [{"role": "user", "content": "Build the plan."}],
                temperature=0.1,
            )
            parsed = _parse_llm_json(content)
            plan_steps = _sanitize_plan_steps(
                parsed.get("plan_steps"),
                max_items=max(1, min(MAX_AGENT_STEPS, 6)),
            )
            if not plan_steps:
                plan_steps = _build_fallback_plan()
            return {
                "plan_steps": plan_steps,
                "reasoning": str(parsed.get("reasoning", ""))[:220],
                "source": "llm" if parsed else "fallback",
            }
        except Exception as exc:
            print(f"[Agent Plan] Initial planning fallback: {exc}")
            return {
                "plan_steps": _build_fallback_plan(),
                "reasoning": "Fallback heuristic plan",
                "source": "fallback",
            }

    def _run_midcourse_review(steps_so_far: list, used_steps: int, current_plan_steps: list[str]) -> dict:
        """Reassess direction at midpoint and optionally revise remaining plan."""
        steps_context = _summarize_agent_steps_for_prompt(
            steps_so_far,
            max_steps=8,
            max_result_chars=180,
        )
        remaining = max(0, MAX_AGENT_STEPS - used_steps)
        plan_block = "\n".join(f"- {s}" for s in current_plan_steps[:6]) or "- (none)"
        review_prompt = f"""You are reviewing the current SQL investigation strategy.
Decide if the agent is on track or drifting, then revise the remaining plan.

USER QUESTION:
{user_question}

CURRENT PLAN:
{plan_block}

USED CREDITS: {used_steps}/{MAX_AGENT_STEPS}
REMAINING CREDITS: {remaining}

EVIDENCE SUMMARY:
{steps_context}

Return ONLY JSON:
{{
  "judgement": "on_track" | "off_track" | "blocked",
  "guidance": "short directive for next actions",
  "should_finish_early": false,
  "updated_plan_steps": ["next step 1", "next step 2"]
}}"""
        try:
            max_review_tokens = int(_get_effective_context_limit() * 0.5)
            review_prompt = _truncate_text_to_budget(review_prompt, max_review_tokens)
            content = _call_llm(
                review_prompt,
                [{"role": "user", "content": "Review and revise now."}],
                temperature=0.1,
            )
            parsed = _parse_llm_json(content)
            updated_steps = _sanitize_plan_steps(
                parsed.get("updated_plan_steps"),
                max_items=max(1, min(remaining, 5)),
            )
            return {
                "judgement": str(parsed.get("judgement", "on_track")),
                "guidance": str(parsed.get("guidance", ""))[:260],
                "should_finish_early": bool(parsed.get("should_finish_early", False)),
                "updated_plan_steps": updated_steps,
                "source": "llm",
            }
        except Exception as exc:
            print(f"[Agent Plan] Midcourse review fallback: {exc}")
            failed = sum(1 for s in steps_so_far if s.get("type") == "query" and not s.get("ok"))
            judgement = "off_track" if failed >= 2 else "on_track"
            fallback_guidance = (
                "Recentrer la stratégie: 1) baseline agrégé simple, 2) une seule dimension de découpage, "
                "3) conclure rapidement."
            )
            fallback_steps = _sanitize_plan_steps(
                [
                    "Calculer un baseline agrégé strictement aligné à la question.",
                    "Comparer une dimension métier principale avec la même plage de dates.",
                    "Finaliser la réponse avec niveau de confiance.",
                ],
                max_items=max(1, min(remaining, 3)),
            )
            return {
                "judgement": judgement,
                "guidance": fallback_guidance,
                "should_finish_early": False,
                "updated_plan_steps": fallback_steps,
                "source": "fallback",
            }

    def _should_trigger_early_midcourse_review(steps_so_far: list, used_steps: int) -> bool:
        """Trigger a review early when the agent appears to drift."""
        if used_steps < 2:
            return False
        recent_queries = [s for s in steps_so_far if s.get("type") == "query"][-3:]
        if len(recent_queries) < 2:
            return False
        failed_recent = sum(1 for s in recent_queries if not s.get("ok"))
        weak_recent = 0
        for step in recent_queries:
            eval_obj = step.get("self_evaluation") or {}
            score = int(eval_obj.get("score", 0)) if isinstance(eval_obj, dict) else 0
            if score < 60:
                weak_recent += 1
        zero_rows_streak = (
            len(recent_queries) >= 2
            and all(int(s.get("row_count", 0) or 0) == 0 for s in recent_queries[-2:])
        )
        duplicate_recent = any(
            str(s.get("error_class", "")).strip().lower() == "duplicate_query"
            for s in recent_queries[-2:]
        )
        return failed_recent >= 2 or weak_recent >= 2 or zero_rows_streak or duplicate_recent

    def _last_query_score(steps_so_far: list) -> int:
        for step in reversed(steps_so_far):
            if step.get("type") != "query":
                continue
            eval_obj = step.get("self_evaluation") or {}
            if isinstance(eval_obj, dict):
                try:
                    return int(eval_obj.get("score", 60))
                except Exception:
                    return 60
        return 60

    def _compute_confidence_delta(steps_so_far: list, current_score: int) -> int:
        return int(current_score) - int(_last_query_score(steps_so_far))

    def _build_synthesis_evidence_lines(
        steps_so_far: list,
        *,
        max_steps: int,
        max_result_chars: int,
    ) -> list[str]:
        lines = []
        for s in steps_so_far[-max_steps:]:
            eval_note = ""
            if isinstance(s.get("self_evaluation"), dict):
                eval_note = (
                    f" | SelfEval(status={s['self_evaluation'].get('status')}, "
                    f"score={s['self_evaluation'].get('score')}, "
                    f"delta={s.get('confidence_delta', 0)})"
                )
            if s.get("type") == "search_knowledge":
                result_preview = str(s.get("result_summary", "")).replace("\n", " ")[:max_result_chars]
                lines.append(
                    f"Step {s.get('step')} [KB] reason={s.get('reasoning', '')[:140]} "
                    f"search={str(s.get('search_query', ''))[:120]} result={result_preview}{eval_note}"
                )
            elif s.get("type") == "export_csv":
                lines.append(
                    f"Step {s.get('step')} [EXPORT] sql={str(s.get('sql', ''))[:180]} "
                    f"path={str(s.get('suggested_path', ''))[:120]}{eval_note}"
                )
            else:
                result_preview = str(s.get("result_summary", "")).replace("\n", " ")[:max_result_chars]
                lines.append(
                    f"Step {s.get('step')} [SQL {'OK' if s.get('ok') else 'FAIL'}] "
                    f"hyp={str(s.get('hypothesis', ''))[:120]} reason={str(s.get('reasoning', ''))[:140]} "
                    f"sql={str(s.get('sql', ''))[:220]} result={result_preview}{eval_note}"
                )
        return lines

    def _synthesize_final_answer_from_steps(steps_so_far: list) -> str:
        """Build final answer with token-budgeted fallbacks."""
        detailed_lines = _build_synthesis_evidence_lines(
            steps_so_far,
            max_steps=12,
            max_result_chars=240,
        )
        synth_prompt = f"""You are a senior business data analyst.
Write a clear, highly detailed and professional final answer with strong FUNCTIONAL focus.
Use the same language as the USER QUESTION.

USER QUESTION: {user_question}

STEPS AND RESULTS:
{chr(10).join(detailed_lines)}

Output format: MARKDOWN with explicit section titles (## headings).
Required sections in this exact order:
## Résumé exécutif
Short executive summary for decision-makers.

## Explications fonctionnelles détaillées
Explain business meaning of observed patterns (process, product, revenue, risk, operations), with explicit business impacts.

## Deep Analysis et insights issus des résultats
Provide deeper analytical insights from the evidence (patterns, ruptures, asymmetries, anomalies, operational implications).
Each insight must reference concrete observed signals.

## Déductions et conclusions fonctionnelles
List key deductions and functional conclusions, each tied to evidence.

## Faits chiffrés et preuves
List quantified facts and references to step evidence.

## Conclusion finale détaillée
Provide a detailed closing synthesis with implications, what can/cannot be concluded, and a decision-oriented narrative.

## Niveau de confiance et limites
State confidence level and important caveats.

Requirements:
- Prioritize functional interpretation, not only numbers.
- Connect conclusions to explicit evidence from the steps.
- Make the conclusion substantially detailed (not a short paragraph).
- Do NOT include a section named "Recommandations actionnables" (or equivalent).
- If data is missing or inconclusive, say it clearly."""
        context_limit = _get_effective_context_limit()
        if _estimate_tokens(synth_prompt) > context_limit:
            compact_lines = _build_synthesis_evidence_lines(
                steps_so_far,
                max_steps=8,
                max_result_chars=120,
            )
            synth_prompt = f"""Write a concise final answer from this compact evidence.
Use the same language as the USER QUESTION.

USER QUESTION: {user_question}
EVIDENCE:
{chr(10).join(compact_lines)}

Requirements:
- Keep a clear functional/business focus.
- Mention confidence level.
- Mention gaps/limitations.
- Include a short "Deep Analysis et insights issus des résultats" section.
- Include a detailed closing conclusion.
- Do NOT include "Recommandations actionnables"."""

        try:
            return _call_llm(
                synth_prompt,
                [{"role": "user", "content": "Synthesise the final answer."}],
                temperature=0.3,
            )
        except Exception as exc:
            if not _is_context_overflow_error(str(exc)):
                raise
            compact_lines = _build_synthesis_evidence_lines(
                steps_so_far,
                max_steps=6,
                max_result_chars=80,
            )
            emergency_prompt = (
                "Write final answer in 5 short bullets with functional focus from minimal evidence only.\n\n"
                f"USER QUESTION: {user_question}\n"
                "EVIDENCE:\n"
                + "\n".join(compact_lines)
            )
            try:
                return _call_llm(
                    emergency_prompt,
                    [{"role": "user", "content": "Final answer now."}],
                    temperature=0.2,
                )
            except Exception:
                return (
                    "Synthèse partielle: la génération finale a été limitée par la fenêtre contextuelle du modèle. "
                    "Réduisez le périmètre (tables/période) pour une conclusion plus fiable."
                )

    def _sanitize_final_answer_text(answer_text: str, steps_so_far: list) -> str:
        """Normalize final answer formatting and remove disallowed recommendation section."""
        txt = str(answer_text or "").strip()
        if not txt:
            return txt

        # Remove an explicit actionable recommendations section if generated anyway.
        section_pattern = re.compile(
            r'(?ims)^\s{0,3}(?:#{1,6}\s*)?(?:recommandations?\s+actionnables?|actionable\s+recommendations?)\s*:?\s*$'
            r'.*?(?=^\s{0,3}(?:#{1,6}\s+\S|[A-ZÀ-ÖØ-Þ][^:\n]{1,80}:\s*$)|\Z)'
        )
        txt = re.sub(section_pattern, "", txt)
        txt = re.sub(r"\n{3,}", "\n\n", txt).strip()

        # If deep-analysis section is missing, append a compact evidence-backed fallback.
        has_deep_analysis = re.search(r"(?i)(deep\s+analysis|analyse\s+approfondie|insights\s+issus\s+des\s+résultats)", txt)
        if not has_deep_analysis:
            fallback_insights = []
            for step in steps_so_far:
                if step.get("type") != "query" or not step.get("ok"):
                    continue
                summary = " ".join(str(step.get("result_summary", "")).split())[:180]
                if not summary:
                    continue
                fallback_insights.append(
                    f"- Step {step.get('step')}: {summary}"
                )
                if len(fallback_insights) >= 3:
                    break
            if fallback_insights:
                txt += (
                    "\n\n## Deep Analysis et insights issus des résultats\n"
                    + "\n".join(fallback_insights)
                )

        return txt

    initial_plan = _build_initial_agent_plan()
    current_plan_steps = list(initial_plan.get("plan_steps", []))
    midcourse_review = None
    midcourse_review_done = False
    midcourse_guidance = ""
    force_finish_now = False
    # Transient per-run working memory for multi-hop retrieval.
    working_memory = {"artifacts": {}, "order": []}

    def _working_memory_snapshot(max_items: int = 4) -> str:
        order = working_memory.get("order") or []
        artifacts = working_memory.get("artifacts") or {}
        if not order:
            return "No reusable intermediate memory yet."
        lines = []
        for key in order[-max_items:]:
            artifact = artifacts.get(key) or {}
            id_sets = artifact.get("id_sets") or {}
            if id_sets:
                id_desc = ", ".join(
                    f"{name}({len(vals)})"
                    for name, vals in list(id_sets.items())[:5]
                )
            else:
                id_desc = "none"
            lines.append(
                f"- {key}: rows={artifact.get('row_count', 0)}, "
                f"columns={len(artifact.get('columns', []))}, id_sets={id_desc}"
            )
        return "\n".join(lines)

    def _update_working_memory(step_number: int, sql_text: str, exec_result: dict) -> None:
        if not exec_result.get("ok"):
            return
        rows = exec_result.get("rows") or []
        cols = exec_result.get("columns") or []
        if not rows or not cols:
            return
        id_sets = _extract_intermediate_id_sets(rows, cols, max_scan_rows=800, max_values_per_set=200)
        artifact_key = f"step{step_number}"
        artifacts = working_memory.setdefault("artifacts", {})
        order = working_memory.setdefault("order", [])
        artifacts[artifact_key] = {
            "step": step_number,
            "sql": str(exec_result.get("normalized_sql") or sql_text or "")[:380],
            "row_count": int(exec_result.get("total_rows", 0) or 0),
            "columns": list(cols)[:14],
            "id_sets": id_sets,
        }
        if artifact_key in order:
            order.remove(artifact_key)
        order.append(artifact_key)
        while len(order) > 12:
            old = order.pop(0)
            artifacts.pop(old, None)

    def _search_knowledge_for_agent(query: str) -> str:
        """Search the knowledge base (ES RAG) for context relevant to the query.

        Falls back to static knowledge text when semantic retrieval is unavailable.
        """
        if not use_knowledge_base:
            return "Knowledge base search is disabled for this run."
        try:
            if query.strip():
                hit_context = _get_knowledge_context_by_similarity(
                    query.strip(),
                    top_k=int(rag_config.get("topK", 5)),
                )
                if hit_context:
                    return hit_context[:3000]
        except Exception as exc:
            print(f"[Agent] Knowledge search failed: {exc}")
        return (knowledge_context or "No knowledge base content available.")[:3000]

    def _agent_self_evaluate(step_entry: dict, total_used_steps: int) -> dict:
        """Score result quality (technical retries are handled separately)."""
        if step_entry.get("type") != "query":
            return {
                "status": "good",
                "score": 80,
                "reason": "Non-SQL action completed.",
                "should_retry": False,
                "should_finish": False,
            }
        if not step_entry.get("ok"):
            return {
                "status": "bad",
                "score": 0,
                "reason": "SQL execution failed.",
                "should_retry": True,
                "should_finish": False,
            }

        row_count = int(step_entry.get("row_count", 0) or 0)
        if row_count == 0:
            return {
                "status": "weak",
                "score": 35,
                "reason": "Query returned zero rows; broaden filters or verify dimensions.",
                "should_retry": False,
                "should_finish": False,
            }
        if row_count < 3:
            return {
                "status": "weak",
                "score": 55,
                "reason": "Very few rows; confidence may be limited.",
                "should_retry": False,
                "should_finish": False,
            }
        if row_count > 4000:
            return {
                "status": "medium",
                "score": 70,
                "reason": "Large result set; could benefit from tighter aggregation.",
                "should_retry": False,
                "should_finish": False,
            }
        return {
            "status": "good",
            "score": 85,
            "reason": "Result set is usable and supports analysis.",
            "should_retry": False,
            "should_finish": total_used_steps >= max(2, MAX_AGENT_STEPS - 1),
        }

    def _run_agent_step(steps_so_far: list, used_steps: int = None) -> dict:
        """Ask the LLM for its next action given accumulated context."""
        nonlocal context_injected_once
        steps_context = _summarize_agent_steps_for_prompt(
            steps_so_far,
            max_steps=8,
            max_result_chars=220,
        )

        effective_used = used_steps if used_steps is not None else len(steps_so_far)
        plan_block = "\n".join(
            f"{idx + 1}. {step}"
            for idx, step in enumerate(current_plan_steps[:8])
        ) or "1. Build a direct baseline SQL query.\n2. Validate with one comparison query.\n3. Conclude."
        midcourse_block = midcourse_guidance or "No mid-course revision yet."
        mandatory_finish_note = ""
        if effective_used >= MAX_AGENT_STEPS or force_finish_now:
            mandatory_finish_note = (
                "IMPORTANT: based on credit constraints/review, you MUST respond with action 'finish' now."
            )

        if use_knowledge_base:
            actions_block = """You have FOUR possible actions:
1. Run a ClickHouse SQL query to fetch data -> action "query"
2. Search the knowledge base for business rules, definitions or context -> action "search_knowledge"
3. Export data to a CSV file (pipe-separated) - ONLY if the user explicitly asks to export or save results -> action "export_csv"
4. Produce the final answer when you have enough information -> action "finish"""
            kb_json_schema = """
For a knowledge base search:
{
  "action": "search_knowledge",
  "reasoning": "Why you need to search the knowledge base and what concept you're looking for",
  "search_query": "the search terms or question to look up"
}
"""
        else:
            actions_block = """You have THREE possible actions:
1. Run a ClickHouse SQL query to fetch data -> action "query"
2. Export data to a CSV file (pipe-separated) - ONLY if the user explicitly asks to export or save results -> action "export_csv"
3. Produce the final answer when you have enough information -> action "finish"
Knowledge base search is DISABLED for this run."""
            kb_json_schema = ""

        # ── Static parts of the prompt (no schema/metadata/knowledge yet) ────
        static_header = f"""You are an autonomous ClickHouse data analyst agent.
Your goal is to answer the user's question by executing a sequence of targeted actions.
You may run up to {MAX_AGENT_STEPS} distinct actions in total. Each action must bring NEW information.
After gathering enough evidence you MUST produce a final answer.
"""
        working_memory_block = _working_memory_snapshot(max_items=5)
        static_footer = f"""

CLICKHOUSE INSTRUCTIONS:
- Use SIMPLE and highly compatible ClickHouse SQL only.
- Prefer: COUNT, SUM, AVG, MIN, MAX, COUNT DISTINCT, ROUND, COALESCE, IF, simple GROUP BY.
- Avoid advanced/specialized functions: windowFunnel, retention, sequenceMatch, argMax, topK, uniqHLL12, quantiles*, JSONExtract*, sumIf, countIf.
- SQL MUST remain read-only in this agent: SELECT/SHOW/DESCRIBE/EXPLAIN/WITH only.
- Always write highly optimised SQL with LIMIT <= 5000.
- For exploratory steps, default to LIMIT 200 unless a broader scan is strictly required.
- Each query must explore a genuinely distinct angle (different aggregation, filter, dimension, or sub-question).
- Do not repeat the same SQL logic twice.
- Emit exactly one SQL statement (no semicolon-separated batch).

DATE HANDLING — CRITICAL:
- When filtering by date, ALWAYS use either:
  1) `date_col BETWEEN 'YYYY-MM-DD' AND 'YYYY-MM-DD'`
  2) `date_col = 'YYYY-MM-DD'`
- NEVER use date extraction/transformation functions for filtering: no year(), month(), toYear(), toMonth(), date_trunc(), toStartOf*().
- NEVER use >, <, >=, <= with date literals.
- If the user asks a relative period, translate it to explicit literal dates and use BETWEEN.

REACT DISCIPLINE:
- Think before acting: each step MUST contain a clear reasoning thought before the action.
- For action "query", explicitly include a hypothesis and expected signal before SQL execution.
- If the previous query failed or was weak, first explain how the next action corrects trajectory.

FINAL ANSWER QUALITY:
- The final answer must remain professional and detailed with a strong functional/business focus.
- Include functional explanations, deep analytical insights, deductions, and business conclusions in addition to factual metrics.
- The final answer must be formatted in MARKDOWN using clear section headings (## ...).
- Never include a section called "Recommandations actionnables".

WORKING MEMORY (TRANSIENT, FOR MULTI-HOP RETRIEVAL):
{working_memory_block}
- You may reuse stored ID sets with SQL placeholders (best used inside IN (...)):
  `{{{{step1.client_id}}}}`
  `{{{{last.order_id}}}}`  (last = most recent successful SQL step)
- Placeholders expand to a bounded list (max 200 values).
- If no suitable memory set exists, run a query first to create it.

CURRENT ANALYSIS PROGRESS ({effective_used}/{MAX_AGENT_STEPS} steps used):
{steps_context if steps_context else "None yet — this is the first step."}

CURRENT EXECUTION PLAN:
{plan_block}

MID-COURSE GUIDANCE:
{midcourse_block}

USER QUESTION:
{user_question}

INSTRUCTIONS:
{actions_block}

{mandatory_finish_note}

Respond ONLY with valid JSON in one of these forms:

For a SQL query:
{{
  "action": "query",
  "reasoning": "Why this specific query is needed and what new insight it will provide",
  "hypothesis": "What you expect to confirm or refute",
  "expected_signal": "What pattern in the result would validate the hypothesis",
  "sql": "SELECT ..."
}}
{kb_json_schema}

For an export request (only when the user asks to save/export data):
{{
  "action": "export_csv",
  "reasoning": "Why this export is needed",
  "sql": "SELECT ... (the query whose results will be exported, include LIMIT 1000000)",
  "suggested_path": "/home/user/export.csv"
}}

For the final answer:
{{
  "action": "finish",
  "reasoning": "Why you have enough information to conclude",
  "final_answer": "A detailed markdown answer with functional explanations, deep analysis insights from results, deductions, quantified evidence, a detailed final conclusion, and confidence limits"
}}

No markdown fences. Only raw JSON."""

        inject_context_now = (not no_prompt_context_injection) and (not context_injected_once)
        if inject_context_now:
            # Inject heavy static context only once per run to avoid prompt bloat.
            base_tokens = _estimate_tokens(static_header + static_footer + "Proceed with the next step.")
            pruned_schema, pruned_metadata, pruned_knowledge = _truncate_prompt_context(
                schema, table_metadata, knowledge_context, base_tokens
            )

            pruned_schema_str = json.dumps(pruned_schema, indent=2)
            pruned_metadata_str = json.dumps(pruned_metadata, indent=2)
            pruned_mapping_note = (
                "Friendly business names for tables (use technical name in SQL, friendly name when talking to the user):\n"
                + "\n".join(
                    f"  - {tbl}  ->  \"{all_mappings[tbl]}\""
                    for tbl in pruned_schema if tbl in all_mappings
                )
            ) if any(tbl in all_mappings for tbl in pruned_schema) else ""

            context_section = (
                "DATABASE SCHEMA:\n"
                + pruned_schema_str
                + f"\n\nTABLE METADATA (functional descriptions):\n{pruned_metadata_str}"
                + (f"\n\n{pruned_mapping_note}" if pruned_mapping_note else "")
                + (f"\n\nSTATIC KNOWLEDGE BASE (available context):\n{pruned_knowledge}" if pruned_knowledge else "")
            )
            context_injected_once = True
        else:
            if no_prompt_context_injection:
                context_section = """CONTEXT MODE:
- Static schema/metadata/knowledge injection is disabled for this run.
- Rely on the user question plus previous step evidence only."""
            else:
                context_section = """CONTEXT MODE:
- Static schema/metadata/knowledge were already injected earlier in this run.
- Do not request reinjection; rely on previous step evidence."""

        system_prompt = static_header + "\n\n" + context_section + static_footer

        # ── Final safety check: if still too large, fallback to ultra-compact prompt ──
        context_limit = _get_effective_context_limit()
        total_tokens = _estimate_tokens(system_prompt + "Proceed with the next step.")
        if total_tokens > context_limit:
            print(
                f"[Token safety] Prompt too large even after truncation: "
                f"~{total_tokens} tokens > {context_limit} limit. Falling back to compact step prompt."
            )
            compact_steps_context = _summarize_agent_steps_for_prompt(
                steps_so_far,
                max_steps=4,
                max_result_chars=120,
            )
            system_prompt = f"""You are an autonomous ClickHouse data analyst agent.
Your goal is to decide the next best action with minimal context.

CURRENT ANALYSIS PROGRESS:
{compact_steps_context if compact_steps_context else "None yet — first step."}

USER QUESTION:
{user_question}

INSTRUCTIONS:
{actions_block}
Return raw JSON only.

For SQL:
{{
  "action": "query",
  "reasoning": "Short reason",
  "sql": "SELECT ..."
}}
{kb_json_schema}
For final answer:
{{
  "action": "finish",
  "reasoning": "Why enough information is available",
  "final_answer": "Structured markdown answer with functional focus, deep insights, detailed conclusion, confidence and limitations"
}}"""
            compact_total = _estimate_tokens(system_prompt + "Proceed with the next step.")
            if compact_total > context_limit:
                return {
                    "action": "finish",
                    "reasoning": "Context window exceeded",
                    "final_answer": (
                        "Désolé, la demande dépasse la fenêtre contextuelle du modèle, même en mode compact. "
                        "Veuillez préciser un périmètre plus réduit (ex: moins de tables ou une période plus courte)."
                    ),
                }

        try:
            content = _call_llm(
                system_prompt,
                [{"role": "user", "content": "Proceed with the next step."}],
                temperature=0.2,
            )
        except Exception as exc:
            if _is_context_overflow_error(str(exc)):
                compact_steps_context = _summarize_agent_steps_for_prompt(
                    steps_so_far,
                    max_steps=3,
                    max_result_chars=90,
                )
                emergency_prompt = f"""You are an autonomous ClickHouse data analyst agent.
Pick ONLY the next action in raw JSON.
{actions_block}
Current progress:
{compact_steps_context}
User question:
{user_question}
Allowed JSON:
{{"action":"query","reasoning":"...","sql":"SELECT ..."}}
{{"action":"finish","reasoning":"...","final_answer":"..."}}"""
                try:
                    content = _call_llm(
                        emergency_prompt,
                        [{"role": "user", "content": "Proceed with the next step."}],
                        temperature=0.1,
                    )
                except Exception:
                    return {
                        "action": "finish",
                        "reasoning": "Context window exceeded",
                        "final_answer": (
                            "Désolé, l'agent est bloqué par la limite de contexte du modèle local. "
                            "Essayez de filtrer les tables (Table Mapping Filter) ou de désactiver la knowledge base."
                        ),
                    }
                return _parse_llm_json(content)
            raise
        return _parse_llm_json(content)

    def _run_agent_retry(steps_so_far: list, retry_info: dict) -> dict:
        """Generate a corrective SQL after a technical SQL failure.

        This retry does NOT consume a step credit.
        """
        retry_mode = "technical"
        failed_sql = retry_info.get("failed_sql", "")
        error_msg = retry_info.get("error", "")
        error_class = retry_info.get("error_class", "")

        steps_context = _summarize_agent_steps_for_prompt(
            steps_so_far,
            max_steps=8,
            max_result_chars=220,
        )
        working_memory_block = _working_memory_snapshot(max_items=4)
        if no_prompt_context_injection:
            retry_context_note = (
                "No static schema/metadata/knowledge context is available in this run. "
                "Reuse only what can be inferred from previous steps."
            )
        else:
            retry_context_note = (
                "Static context is injected only once per run. Do not request reinjection; "
                "reuse table/column hints from previous steps."
            )

        issue_block = (
            "A previous SQL query was rejected by ClickHouse with a technical error. "
            "Generate a SIMPLER alternative that achieves the same analytical goal."
        )

        retry_prompt = f"""You are an autonomous ClickHouse data analyst agent.
{issue_block}

RETRY MODE: {retry_mode}
ERROR CLASS: {error_class or "n/a"}

CONTEXT POLICY:
{retry_context_note}

PREVIOUS SQL:
{failed_sql}

ISSUE DETAIL:
{error_msg}

PREVIOUS STEPS FOR CONTEXT:
{steps_context if steps_context else "None yet."}

TRANSIENT WORKING MEMORY:
{working_memory_block}
You can reference stored IDs with placeholders like `{{{{step1.client_id}}}}` or `{{{{last.order_id}}}}`.

RULES:
- Read-only SQL only: SELECT/SHOW/DESCRIBE/EXPLAIN/WITH.
- Use FEWER and SIMPLER functions when fixing technical errors.
- Use only simple compatible functions: COUNT, SUM, AVG, MIN, MAX, COUNT DISTINCT, ROUND, COALESCE, IF.
- Do NOT use: windowFunnel, retention, sequenceMatch, argMax, topK, uniqHLL12, quantiles*, JSONExtract*, sumIf, countIf.
- Avoid nested complexity unless strictly required.
- For date filters use only: BETWEEN 'YYYY-MM-DD' AND 'YYYY-MM-DD' or = 'YYYY-MM-DD'.
- No year()/month()/toYear()/toMonth()/date_trunc()/toStartOf*() in filters.
- No >, <, >=, <= with date literals.
- Add LIMIT 200 if not present; never exceed LIMIT 5000.
- Return exactly one SQL statement.

Respond ONLY with valid JSON (no markdown):
{{
  "action": "query",
  "reasoning": "What was fixed and why this should work better",
  "sql": "SELECT ..."
}}"""
        retry_prompt = _truncate_text_to_budget(
            retry_prompt,
            int(_get_effective_context_limit() * 0.62),
        )

        content = _call_llm(
            retry_prompt,
            [{"role": "user", "content": "Generate the corrective query."}],
            temperature=0.2,
        )
        result = _parse_llm_json(content)
        if not result or not result.get("sql"):
            return {"action": "finish", "reasoning": "Could not generate alternative", "final_answer": ""}
        return result

    def _execute_and_summarise(sql: str, max_rows: int = 20) -> dict:
        """Execute a read-only ClickHouse query and return structured result."""
        return _execute_sql_guarded(
            sql,
            read_only=True,
            enforce_simple_compat=True,
            max_preview_rows=max_rows,
            max_execution_time=15,
            default_limit=500,
            hard_limit_cap=5000,
            client=agent_client,
        )

    # ---- Agentic loop ----
    try:
        final_answer = None
        used_steps = 0        # steps counting against the credit budget
        technical_retries_used = 0  # retries not counting against credits
        query_attempts_used = 0
        retry_pending = None  # {"failed_sql": "...", "error": "...", "error_class": "...", "playbook_attempted": bool}
        safety_limit = MAX_AGENT_STEPS * 4  # absolute ceiling to prevent infinite loops
        safety_counter = 0

        while safety_counter < safety_limit:
            if used_steps >= MAX_AGENT_STEPS and not retry_pending:
                break
            safety_counter += 1

            interrupt_reason = _check_external_interrupt()
            if interrupt_reason:
                final_answer = (
                    _synthesize_final_answer_from_steps(steps)
                    if steps
                    else (
                        "Analyse interrompue avant exécution complète. "
                        "Ajoutez du contexte puis relancez la session."
                    )
                )
                final_answer = _sanitize_final_answer_text(final_answer, steps)
                return jsonify({
                    "steps": steps,
                    "final_answer": final_answer,
                    "total_steps": len(steps),
                    "technical_retries_used": technical_retries_used,
                    "free_retries_used": technical_retries_used,
                    "query_attempts_used": query_attempts_used,
                    "max_technical_retries": MAX_TECHNICAL_RETRIES,
                    "max_total_query_attempts": MAX_AGENT_STEPS + MAX_TECHNICAL_RETRIES,
                    "initial_plan": initial_plan.get("plan_steps", []),
                    "current_plan": current_plan_steps,
                    "midcourse_review": midcourse_review or {},
                    "no_prompt_context_injection": bool(no_prompt_context_injection),
                    "knowledge_mode": knowledge_mode,
                    "interrupted": True,
                    "interrupt_reason": interrupt_reason,
                })

            # ── Mid-course review (halfway OR early stagnation trigger) ─────────
            if (
                not midcourse_review_done
                and steps
                and not retry_pending
                and (
                    used_steps >= halfway_step
                    or _should_trigger_early_midcourse_review(steps, used_steps)
                )
            ):
                midcourse_review = _run_midcourse_review(
                    steps,
                    used_steps,
                    current_plan_steps,
                )
                midcourse_review_done = True
                midcourse_guidance = str(midcourse_review.get("guidance", "")).strip()
                revised_steps = midcourse_review.get("updated_plan_steps") or []
                if revised_steps:
                    current_plan_steps = revised_steps
                if midcourse_review.get("should_finish_early"):
                    force_finish_now = True

            # ── Handle a pending corrective retry (free, no credit consumed) ────
            if retry_pending:
                if technical_retries_used >= MAX_TECHNICAL_RETRIES:
                    if steps and steps[-1].get("type") == "query" and not steps[-1].get("ok"):
                        steps[-1]["result_summary"] += (
                            f" | Technical retry budget exhausted ({MAX_TECHNICAL_RETRIES} max)."
                        )
                    retry_pending = None
                else:
                    retry_mode = "technical"
                    alt_sql = ""
                    playbook_attempted = bool(retry_pending.get("playbook_attempted", False))
                    if not playbook_attempted:
                        alt_sql = _apply_sql_retry_playbook(
                            retry_pending.get("failed_sql", ""),
                            error_class=str(retry_pending.get("error_class", "")),
                            error_text=str(retry_pending.get("error", "")),
                        )
                        if alt_sql:
                            retry_mode = "playbook"
                        retry_pending["playbook_attempted"] = True

                    if not alt_sql:
                        alt_decision = _run_agent_retry(steps, retry_pending)
                        alt_sql = (alt_decision.get("sql") or "").strip()
                        retry_mode = "technical"
                        retry_pending = None
                    else:
                        retry_pending = None

                    technical_retries_used += 1

                    if alt_sql and steps:
                        resolved_alt_sql = alt_sql
                        alt_resolution_error = ""
                        try:
                            resolved_alt_sql = _resolve_sql_memory_placeholders(alt_sql, working_memory)
                        except Exception as mem_exc:
                            alt_resolution_error = str(mem_exc)

                        if alt_resolution_error:
                            alt_result = {
                                "ok": False,
                                "normalized_sql": alt_sql.strip(),
                                "total_rows": 0,
                                "summary": f"Query blocked before execution: {alt_resolution_error}",
                                "error_class": "memory_reference",
                            }
                        else:
                            query_attempts_used += 1
                            alt_result = _execute_and_summarise(resolved_alt_sql)
                        alt_norm = (alt_result.get("normalized_sql") or alt_sql).strip()
                        alt_fp = _normalize_sql_fingerprint(alt_norm)
                        is_dup = bool(alt_fp and alt_fp in seen_query_fingerprints)
                        if is_dup:
                            alt_result = {
                                "ok": False,
                                "normalized_sql": alt_norm,
                                "total_rows": 0,
                                "summary": "Query failed: duplicate query pattern detected.",
                                "error_class": "duplicate_query",
                            }
                        if alt_result["ok"]:
                            if alt_fp:
                                seen_query_fingerprints.add(alt_fp)
                            steps[-1].update({
                                "sql": alt_norm,
                                "result_summary": alt_result["summary"],
                                "row_count": alt_result["total_rows"],
                                "ok": True,
                                "retried": True,
                                "retry_mode": retry_mode,
                                "technical_retry_used": True,
                                "error_class": "",
                            })
                            steps[-1]["self_evaluation"] = _agent_self_evaluate(
                                steps[-1],
                                used_steps,
                            )
                            if isinstance(steps[-1].get("self_evaluation"), dict):
                                score = int(steps[-1]["self_evaluation"].get("score", 0))
                                steps[-1]["confidence_delta"] = _compute_confidence_delta(
                                    steps[:-1],
                                    score,
                                )
                            _update_working_memory(
                                int(steps[-1].get("step", used_steps) or used_steps),
                                alt_norm,
                                alt_result,
                            )
                        else:
                            steps[-1]["result_summary"] += (
                                f" | Corrective retry failed ({retry_mode}): {alt_result['summary']}"
                            )
                            steps[-1]["error_class"] = alt_result.get("error_class", "")
                            steps[-1]["technical_retry_used"] = True
                            # Keep room for another alternative path until global +3 retry cap is hit.
                            if technical_retries_used < MAX_TECHNICAL_RETRIES:
                                retry_pending = {
                                    "failed_sql": alt_norm or alt_sql,
                                    "error": alt_result.get("summary", ""),
                                    "error_class": alt_result.get("error_class", ""),
                                    "playbook_attempted": retry_mode == "playbook",
                                }
                    continue

            if used_steps >= MAX_AGENT_STEPS:
                break

            # ── Normal agent step ────────────────────────────────────────────
            decision = _run_agent_step(steps, used_steps)

            action = decision.get("action", "finish")
            reasoning = decision.get("reasoning", "")

            if force_finish_now and action != "finish":
                # Mid-course review requested early stop to avoid inefficient chains.
                break

            if action == "query":
                used_steps += 1
                sql_template = decision.get("sql", "").strip()
                hypothesis = str(decision.get("hypothesis", "")).strip()[:220]
                expected_signal = str(decision.get("expected_signal", "")).strip()[:220]
                resolved_sql = sql_template
                resolution_error = ""
                try:
                    resolved_sql = _resolve_sql_memory_placeholders(sql_template, working_memory)
                except Exception as mem_exc:
                    resolution_error = str(mem_exc)

                if resolution_error:
                    exec_result = {
                        "ok": False,
                        "normalized_sql": sql_template,
                        "total_rows": 0,
                        "summary": f"Query blocked before execution: {resolution_error}",
                        "error_class": "memory_reference",
                    }
                else:
                    query_attempts_used += 1
                    exec_result = _execute_and_summarise(resolved_sql)

                normalized_sql = (exec_result.get("normalized_sql") or resolved_sql or sql_template).strip()
                normalized_fp = _normalize_sql_fingerprint(normalized_sql)

                # Prevent pointless loops on repeated queries.
                if normalized_fp and normalized_fp in seen_query_fingerprints:
                    exec_result = {
                        "ok": False,
                        "normalized_sql": normalized_sql,
                        "total_rows": 0,
                        "summary": "Query failed: duplicate query pattern detected.",
                        "error_class": "duplicate_query",
                    }
                elif exec_result["ok"] and normalized_fp:
                    seen_query_fingerprints.add(normalized_fp)

                step_entry = {
                    "step": used_steps,
                    "type": "query",
                    "reasoning": reasoning,
                    "hypothesis": hypothesis,
                    "expected_signal": expected_signal,
                    "sql": normalized_sql,
                    "result_summary": exec_result["summary"],
                    "row_count": exec_result.get("total_rows", 0),
                    "ok": exec_result["ok"],
                    "error_class": exec_result.get("error_class", ""),
                    "technical_retry_used": False,
                }
                if sql_template and sql_template != normalized_sql:
                    step_entry["sql_template"] = sql_template
                step_entry["self_evaluation"] = _agent_self_evaluate(step_entry, used_steps)
                if isinstance(step_entry.get("self_evaluation"), dict):
                    step_entry["confidence_delta"] = _compute_confidence_delta(
                        steps,
                        int(step_entry["self_evaluation"].get("score", 0)),
                    )
                if step_entry["self_evaluation"]["reason"]:
                    step_entry["result_summary"] += (
                        f"\nSelf-evaluation: {step_entry['self_evaluation']['reason']}"
                    )
                steps.append(step_entry)
                if exec_result["ok"]:
                    _update_working_memory(used_steps, normalized_sql, exec_result)

                if not exec_result["ok"]:
                    retry_pending = {
                        "failed_sql": normalized_sql or sql_template,
                        "error": exec_result["summary"],
                        "error_class": exec_result.get("error_class", ""),
                        "playbook_attempted": False,
                    }

            elif action == "search_knowledge":
                used_steps += 1
                search_query = decision.get("search_query", "").strip()
                kb_result = _search_knowledge_for_agent(search_query)
                kb_summary = kb_result[:1500] if len(kb_result) > 1500 else kb_result
                steps.append({
                    "step": used_steps,
                    "type": "search_knowledge",
                    "reasoning": reasoning,
                    "sql": None,
                    "search_query": search_query,
                    "result_summary": f"Knowledge base results:\n{kb_summary}",
                    "row_count": 0,
                    "ok": True,
                    "self_evaluation": {
                        "status": "good",
                        "score": 80,
                        "reason": "Knowledge lookup completed.",
                        "should_retry": False,
                        "should_finish": False,
                    },
                    "confidence_delta": 0,
                })

            elif action == "export_csv":
                used_steps += 1
                export_sql = decision.get("sql", "").strip()
                suggested_path = decision.get("suggested_path", "/tmp/export.csv").strip()
                steps.append({
                    "step": used_steps,
                    "type": "export_csv",
                    "reasoning": reasoning,
                    "sql": export_sql,
                    "suggested_path": suggested_path,
                    "result_summary": f"Export CSV requested → {suggested_path}",
                    "row_count": 0,
                    "ok": True,
                    "self_evaluation": {
                        "status": "good",
                        "score": 75,
                        "reason": "Export instruction prepared.",
                        "should_retry": False,
                        "should_finish": False,
                    },
                    "confidence_delta": 0,
                })
            else:
                # action == "finish"
                decision_final = str(decision.get("final_answer", "")).strip()
                final_answer = (
                    _sanitize_final_answer_text(decision_final, steps)
                    if decision_final else None
                )
                break

        # If the loop exhausted all steps without finish, synthesize from evidence.
        if final_answer is None:
            final_answer = _synthesize_final_answer_from_steps(steps)
        final_answer = _sanitize_final_answer_text(final_answer, steps)

        return jsonify({
            "steps": steps,
            "final_answer": final_answer,
            "total_steps": len(steps),
            "technical_retries_used": technical_retries_used,
            "free_retries_used": technical_retries_used,
            "query_attempts_used": query_attempts_used,
            "max_technical_retries": MAX_TECHNICAL_RETRIES,
            "max_total_query_attempts": MAX_AGENT_STEPS + MAX_TECHNICAL_RETRIES,
            "initial_plan": initial_plan.get("plan_steps", []),
            "current_plan": current_plan_steps,
            "midcourse_review": midcourse_review or {},
            "no_prompt_context_injection": bool(no_prompt_context_injection),
            "knowledge_mode": knowledge_mode,
        })

    except Exception as exc:
        print(f"Agent error: {exc}")
        return jsonify({"error": str(exc)}), 500

# ---------------------------------------------------------------------------
# AI Query Analyzer
# ---------------------------------------------------------------------------
@app.route("/api/analyze", methods=["POST"])
def analyze_query():
    data = request.get_json()
    sql = data.get("sql", "")
    schema = data.get("schema", {})

    # Token budget: base prompt without the schema blob
    base_analyze_prompt = """You are an expert ClickHouse SQL performance analyst.
Analyze the following SQL query and provide:
1. Performance alerts (full table scans, missing LIMIT, unoptimized aggregations)
2. Correctness concerns (wrong results, type mismatches, NULL handling)
3. Optimization suggestions (better functions, indexes, partitioning hints)
4. Data projections (estimated result size, cardinality warnings)

Return ONLY a JSON object (no markdown) with:
{
  "alerts": ["alert message 1"],
  "suggestions": ["suggestion 1"],
  "projections": ["projection 1"],
  "optimized_sql": "improved SQL or empty string",
  "risk_level": "low" | "medium" | "high"
}
"""
    user_message = f"Analyze this ClickHouse SQL query:\n\n{sql}"
    base_tokens = _estimate_tokens(base_analyze_prompt) + _estimate_tokens(user_message)

    # Truncate schema context to fit within budget
    schema, _, _ = _truncate_prompt_context(schema, {}, "", base_tokens)

    system_prompt = f"""You are an expert ClickHouse SQL performance analyst.
Analyze the following SQL query and provide:
1. Performance alerts (full table scans, missing LIMIT, unoptimized aggregations)
2. Correctness concerns (wrong results, type mismatches, NULL handling)
3. Optimization suggestions (better functions, indexes, partitioning hints)
4. Data projections (estimated result size, cardinality warnings)

Database schema context:
{json.dumps(schema, indent=2)}

Return ONLY a JSON object (no markdown) with:
{{
  "alerts": ["alert message 1"],
  "suggestions": ["suggestion 1"],
  "projections": ["projection 1"],
  "optimized_sql": "improved SQL or empty string",
  "risk_level": "low" | "medium" | "high"
}}
"""

    try:
        content = _call_llm(system_prompt, [{"role": "user", "content": user_message}], temperature=0.3)
        return jsonify(_parse_llm_json(content))
    except Exception as exc:
        print(f"Analyze error: {exc}")
        return jsonify({"error": str(exc)}), 500


# ---------------------------------------------------------------------------
# Table Profiling
# ---------------------------------------------------------------------------
@app.route("/api/profile/<table_name>", methods=["GET"])
def profile_table(table_name):
    try:
        client = get_clickhouse_client()

        # 1. Column types
        desc_result = client.query(f"DESCRIBE TABLE `{table_name}`")
        columns = [{"name": r[0], "type": r[1]} for r in desc_result.result_rows]

        if not columns:
            return jsonify({"error": "Table not found or empty schema"}), 404

        # 2. Total rows
        count_result = client.query(f"SELECT count() FROM `{table_name}`")
        total_rows = int(count_result.result_rows[0][0]) if count_result.result_rows else 0

        # 3. Single stats query against a row-limited subquery.
        #
        # KEY FIX: The previous approach used "FROM table SAMPLE 0.1" or a
        # trailing LIMIT on the outer query.  Neither correctly limits the
        # rows fed into aggregate functions:
        #   • SAMPLE 0.1 requires a sampling key and fails silently on tables
        #     that don't have one — stats_row stays {}, every column gets
        #     notnull=0, which produces null_pct=100% for every column.
        #   • A trailing LIMIT on a single-row aggregate result is a no-op for
        #     the underlying scan.
        #
        # The correct pattern for large tables is to wrap the table in a
        # subquery with LIMIT so ClickHouse stops reading after N rows:
        #   SELECT agg(col) FROM (SELECT col FROM table LIMIT N)
        #
        # We also use uniq() (HyperLogLog, ~2% error) instead of uniqExact()
        # to avoid materializing huge hash sets in memory.
        profile_sample = 500_000 if total_rows > 500_000 else total_rows
        col_names_quoted = ", ".join(f"`{c['name'].replace('`','``')}`" for c in columns)
        sub = f"(SELECT {col_names_quoted} FROM `{table_name}` LIMIT {profile_sample})"

        select_parts = []
        for col in columns:
            cn = col["name"].replace("`", "``")
            select_parts.append(f"countIf(`{cn}` IS NOT NULL) AS `{cn}__notnull`")
            select_parts.append(f"uniq(`{cn}`) AS `{cn}__distinct`")

        stats_row = {}
        if select_parts:
            stats_sql = f"SELECT {', '.join(select_parts)} FROM {sub}"
            try:
                sr = client.query(stats_sql)
                if sr.result_rows:
                    stats_row = dict(zip(sr.column_names, sr.result_rows[0]))
            except Exception as e:
                print(f"Stats query warning: {e}")

        # 4. Build per-column stats
        # Scale sampled counts back to total_rows estimate when a partial sample
        # was used.  When stats_row is empty (query failed), mark as unknown
        # rather than reporting a misleading 100% null rate.
        scale = (total_rows / profile_sample) if profile_sample < total_rows and profile_sample > 0 else 1.0
        col_stats = []
        for col in columns:
            cn = col["name"]
            notnull_raw = stats_row.get(f"{cn}__notnull")
            distinct_raw = stats_row.get(f"{cn}__distinct")

            if notnull_raw is None:
                # Stats query failed for this column — report unknown
                col_stats.append({
                    "name": cn,
                    "type": col["type"],
                    "notnull": None,
                    "null_count": None,
                    "null_pct": None,
                    "distinct": None,
                    "distinct_pct": None,
                    "top_values": [],
                })
                continue

            notnull = int(notnull_raw)
            distinct = int(distinct_raw) if distinct_raw is not None else 0
            estimated_notnull = round(notnull * scale)
            null_count = max(0, total_rows - estimated_notnull)
            null_pct = round(null_count / total_rows * 100, 1) if total_rows > 0 else 0.0
            distinct_pct = round(distinct / profile_sample * 100, 1) if profile_sample > 0 else 0.0
            col_stats.append({
                "name": cn,
                "type": col["type"],
                "notnull": estimated_notnull,
                "null_count": null_count,
                "null_pct": null_pct,
                "distinct": distinct,
                "distinct_pct": distinct_pct,
                "top_values": [],
            })

        # 5. Top values for low-cardinality columns (distinct <= 30)
        # Use the same LIMIT-bounded subquery to avoid full-table scans.
        for cs in col_stats:
            if cs["distinct"] is not None and 0 < cs["distinct"] <= 30:
                try:
                    cn_q = cs['name'].replace('`', '``')
                    tv = client.query(
                        f"SELECT toString(`{cn_q}`) AS val, count() AS cnt"
                        f" FROM (SELECT `{cn_q}` FROM `{table_name}` LIMIT {profile_sample})"
                        f" GROUP BY val ORDER BY cnt DESC LIMIT 8"
                    )
                    cs["top_values"] = [
                        {"value": str(r[0]), "count": int(r[1])}
                        for r in tv.result_rows
                    ]
                except Exception:
                    pass

        return jsonify({
            "table": table_name,
            "total_rows": total_rows,
            "total_columns": len(columns),
            "columns": col_stats,
        })
    except Exception as exc:
        print(f"Profile error: {exc}")
        return jsonify({"error": str(exc)}), 500


@app.route("/api/profile/<table_name>/insights", methods=["POST"])
def profile_insights(table_name):
    data = request.get_json() or {}
    stats = data.get("stats", {})

    # Compact stats for LLM — metadata only, no raw rows
    compact = {
        "table": stats.get("table"),
        "total_rows": stats.get("total_rows"),
        "total_columns": stats.get("total_columns"),
        "columns": [
            {
                "name": c["name"],
                "type": c["type"],
                "null_pct": c["null_pct"],
                "distinct": c["distinct"],
                "top_values": c.get("top_values", [])[:3],
            }
            for c in stats.get("columns", [])
        ],
    }

    system_prompt = """You are a senior data analyst. Given a table profile (metadata only, no raw data),
identify data quality issues, patterns, and recommendations.

Return ONLY a JSON object:
{
  "summary": "2-3 sentence overview of the table and its likely purpose",
  "quality_issues": ["issue 1", "issue 2"],
  "key_insights": ["insight 1", "insight 2", "insight 3"],
  "recommendations": ["recommendation 1", "recommendation 2"]
}"""

    user_msg = f"Profile data:\n{json.dumps(compact)}"
    try:
        content = _call_llm(system_prompt, [{"role": "user", "content": user_msg}], temperature=0.4)
        return jsonify(_parse_llm_json(content))
    except Exception as exc:
        print(f"Profile insights error: {exc}")
        return jsonify({"error": str(exc)}), 500


# ---------------------------------------------------------------------------
# Data Quality Analysis
# ---------------------------------------------------------------------------

def _dq_column_stats(
    client,
    table: str,
    column: str,
    col_type: str,
    sample_size: int,
    filter_col: str | None = None,
    filter_op: str | None = None,
    filter_val: str | None = None,
    filter_val2: str | None = None,
) -> dict:
    """Collect a statistical profile for one column using ClickHouse queries.

    Uses a subquery with LIMIT to bound the number of rows processed — essential
    for tables with billions of rows.  A trailing LIMIT on an aggregate query only
    limits the *result* rows (always 1 for a single aggregate), NOT the rows read
    from the table, so we must wrap the table scan in a subquery.

    If filter_col/filter_op/filter_val are provided, only rows matching that
    condition are included in the analysis (applied inside the bounding subquery).
    BETWEEN operator uses filter_val (start) and filter_val2 (end).
    """
    import re as _re
    stats: dict = {"column": column, "type": col_type}
    safe_table = _re.sub(r"[^\w.]", "", table)
    safe_col = _re.sub(r"[^\w]", "", column)
    n = int(sample_size) if sample_size is not None else None

    # Build optional WHERE filter clause (injected inside the bounding subquery)
    where_clause = ""
    if filter_col and filter_op and filter_val is not None:
        safe_fcol = _re.sub(r"[^\w]", "", filter_col)
        # Escape single quotes in the filter value
        escaped_val = str(filter_val).replace("'", "''")
        if filter_op == "BETWEEN" and filter_val2 is not None:
            escaped_val2 = str(filter_val2).replace("'", "''")
            where_clause = f" WHERE `{safe_fcol}` BETWEEN '{escaped_val}' AND '{escaped_val2}'"
            stats["filter_applied"] = f"`{safe_fcol}` BETWEEN '{escaped_val}' AND '{escaped_val2}'"
        else:
            op_map = {"=": "=", "!=": "!=", "<": "<", ">": ">", "<=": "<=", ">=": ">=", "LIKE": "LIKE"}
            op = op_map.get(filter_op, "=")
            where_clause = f" WHERE `{safe_fcol}` {op} '{escaped_val}'"
            stats["filter_applied"] = f"`{safe_fcol}` {op} '{escaped_val}'"

    # Subquery that limits input rows — this is the correct way to bound scans
    # on large tables in ClickHouse.  Every aggregate below queries this sub.
    limit_clause = f" LIMIT {n}" if n is not None else ""
    sub = f"(SELECT `{safe_col}` FROM {safe_table}{where_clause}{limit_clause})"

    try:
        # ── Basic counts ──────────────────────────────────────────────────────
        # uniq() is approximate (~2% error) but orders-of-magnitude faster than
        # uniqExact() on large cardinality columns.
        r = client.query(
            f"SELECT count() AS total,"
            f" countIf(`{safe_col}` IS NULL) AS null_count,"
            f" uniq(`{safe_col}`) AS approx_distinct"
            f" FROM {sub}"
        )
        row = r.result_rows[0]
        total = int(row[0])
        null_count = int(row[1])
        distinct_count = int(row[2])
        stats.update({
            "total": total,
            "null_count": null_count,
            "null_pct": round(100 * null_count / max(total, 1), 2),
            "distinct_count": distinct_count,
            "distinct_pct": round(100 * distinct_count / max(total, 1), 2),
        })

        # ── Type-specific stats ───────────────────────────────────────────────
        ct_upper = col_type.upper()
        is_numeric = any(t in ct_upper for t in ("INT", "FLOAT", "DECIMAL", "DOUBLE", "NUMBER"))
        is_string = any(t in ct_upper for t in ("STRING", "VARCHAR", "FIXEDSTRING", "TEXT"))
        is_date = any(t in ct_upper for t in ("DATE", "DATETIME"))

        if is_numeric:
            r2 = client.query(
                f"SELECT min(`{safe_col}`), max(`{safe_col}`),"
                f" avg(`{safe_col}`), stddevPop(`{safe_col}`),"
                f" quantile(0.25)(`{safe_col}`), quantile(0.5)(`{safe_col}`),"
                f" quantile(0.75)(`{safe_col}`),"
                f" countIf(`{safe_col}` < 0),"
                f" countIf(`{safe_col}` = 0)"
                f" FROM {sub}"
                f" WHERE `{safe_col}` IS NOT NULL"
            )
            rr = r2.result_rows[0]
            def _f(v):
                return round(float(v), 6) if v is not None else None
            avg_val = _f(rr[2])
            stddev_val = _f(rr[3])
            p50_val = _f(rr[5])
            stats.update({
                "min": _f(rr[0]), "max": _f(rr[1]),
                "avg": avg_val, "stddev": stddev_val,
                "p25": _f(rr[4]), "p50": p50_val, "p75": _f(rr[6]),
                "negative_count": int(rr[7]) if rr[7] is not None else 0,
                "zero_count": int(rr[8]) if rr[8] is not None else 0,
            })
            # Coefficient of variation (stddev / avg) — indicates relative spread
            if avg_val and avg_val != 0 and stddev_val is not None:
                stats["coeff_variation"] = round(abs(stddev_val / avg_val), 4)
            # Pearson skewness approximation: 3*(mean - median) / stddev
            if avg_val is not None and p50_val is not None and stddev_val and stddev_val != 0:
                stats["skewness_approx"] = round(3 * (avg_val - p50_val) / stddev_val, 4)
            # Outlier detection via IQR
            if stats["p25"] is not None and stats["p75"] is not None:
                iqr = stats["p75"] - stats["p25"]
                lower = stats["p25"] - 1.5 * iqr
                upper = stats["p75"] + 1.5 * iqr
                r3 = client.query(
                    f"SELECT countIf(`{safe_col}` < {lower} OR `{safe_col}` > {upper})"
                    f" FROM {sub}"
                    f" WHERE `{safe_col}` IS NOT NULL"
                )
                oc = int(r3.result_rows[0][0])
                stats["outlier_count"] = oc
                stats["outlier_pct"] = round(100 * oc / max(total - null_count, 1), 2)
            # Outlier detection via Z-Score (|value - mean| > 3 * stddev)
            if avg_val is not None and stddev_val and stddev_val > 0:
                r4 = client.query(
                    f"SELECT countIf(abs(`{safe_col}` - {avg_val}) > 3 * {stddev_val})"
                    f" FROM {sub}"
                    f" WHERE `{safe_col}` IS NOT NULL"
                )
                zoc = int(r4.result_rows[0][0])
                stats["zscore_outlier_count"] = zoc
                stats["zscore_outlier_pct"] = round(100 * zoc / max(total - null_count, 1), 2)

        elif is_string:
            r2 = client.query(
                f"SELECT countIf(`{safe_col}` = ''),"
                f" min(length(`{safe_col}`)), max(length(`{safe_col}`)),"
                f" avg(length(`{safe_col}`)),"
                f" countIf(length(`{safe_col}`) > 1000),"
                f" countIf(length(trimBoth(`{safe_col}`)) < length(`{safe_col}`)),"
                f" countIf(upperUTF8(`{safe_col}`) = `{safe_col}` AND lowerUTF8(`{safe_col}`) != `{safe_col}`),"
                f" countIf(match(`{safe_col}`, '^[0-9]+$')),"
                f" countIf(match(`{safe_col}`, '^[^@\\\\s]+@[^@\\\\s]+\\\\.[^@\\\\s]+$'))"
                f" FROM {sub}"
                f" WHERE `{safe_col}` IS NOT NULL"
            )
            rr = r2.result_rows[0]
            empty_count = int(rr[0]) if rr[0] is not None else 0
            stats.update({
                "empty_count": empty_count,
                "empty_pct": round(100 * empty_count / max(total, 1), 2),
                "min_length": int(rr[1]) if rr[1] is not None else 0,
                "max_length": int(rr[2]) if rr[2] is not None else 0,
                "avg_length": round(float(rr[3]), 2) if rr[3] is not None else 0,
                "very_long_count": int(rr[4]) if rr[4] is not None else 0,
                "whitespace_padded_count": int(rr[5]) if rr[5] is not None else 0,
                "all_caps_count": int(rr[6]) if rr[6] is not None else 0,
                "numeric_string_count": int(rr[7]) if rr[7] is not None else 0,
                "email_like_count": int(rr[8]) if rr[8] is not None else 0,
            })
            # Detect common sentinel/garbage values
            r3 = client.query(
                f"SELECT countIf(lower(toString(`{safe_col}`)) IN"
                f" ('null','none','n/a','na','#na','#n/a','unknown','undefined','nan','nil','-','0'))"
                f" FROM {sub}"
                f" WHERE `{safe_col}` IS NOT NULL"
            )
            stats["sentinel_count"] = int(r3.result_rows[0][0])

        elif is_date:
            r2 = client.query(
                f"SELECT min(`{safe_col}`), max(`{safe_col}`),"
                f" countIf(`{safe_col}` > now()),"
                f" countIf(toDate(`{safe_col}`) = '1970-01-01'),"
                f" countIf(toDayOfWeek(`{safe_col}`) >= 6),"
                f" countIf(toYear(`{safe_col}`) < 1900)"
                f" FROM {sub}"
                f" WHERE `{safe_col}` IS NOT NULL"
            )
            rr = r2.result_rows[0]
            stats.update({
                "min_date": str(rr[0]) if rr[0] is not None else None,
                "max_date": str(rr[1]) if rr[1] is not None else None,
                "future_count": int(rr[2]) if rr[2] is not None else 0,
                "epoch_sentinel_count": int(rr[3]) if rr[3] is not None else 0,
                "weekend_count": int(rr[4]) if rr[4] is not None else 0,
                "pre_1900_count": int(rr[5]) if rr[5] is not None else 0,
            })

        # ── Top 10 most frequent values (from the same limited sample) ────────
        r_top = client.query(
            f"SELECT toString(`{safe_col}`) AS val, count() AS cnt"
            f" FROM {sub}"
            f" WHERE `{safe_col}` IS NOT NULL"
            f" GROUP BY `{safe_col}` ORDER BY cnt DESC LIMIT 10"
        )
        stats["top_values"] = [
            {"value": str(row[0]), "count": int(row[1])} for row in r_top.result_rows
        ]

    except Exception as e:
        stats["query_error"] = str(e)

    return stats


_DQ_SYSTEM_PROMPT = """You are an expert data quality analyst specializing in database analytics and anomaly detection.
Analyze the statistical profiles of the provided columns and identify all data quality issues.

Evaluate each column for:
1. **Null/Empty anomalies**: High null rates, sentinel values used as nulls (0, -1, 'N/A', 'null', 'none'…)
2. **Format anomalies**: Inconsistent patterns, unexpected length distributions, mixed formats
3. **Content anomalies**: Impossible values, values outside expected business domain
4. **Cardinality anomalies**: Suspiciously low distinct counts (near-constant column), or unexpectedly high cardinality
5. **Distribution anomalies**: Extreme skew (use skewness_approx), statistical outliers (IQR method), Z-Score outliers (zscore_outlier_count/zscore_outlier_pct), zero inflation (zero_count), highly unbalanced categories
6. **Temporal anomalies**: Future dates, epoch sentinel dates (1970-01-01, epoch_sentinel_count), implausible date ranges (pre_1900_count), weekend concentration (weekend_count)
7. **String quality**: Whitespace padding (whitespace_padded_count), ALL-CAPS inconsistency (all_caps_count), numeric strings in text fields (numeric_string_count), unexpected email-like values (email_like_count)
8. **Spread anomalies**: High coefficient of variation (coeff_variation > 1 indicates extreme spread), high skewness (|skewness_approx| > 2 indicates strong skew)
9. **Volume/Temporal consistency** (if provided): Periods with abnormally low record counts may indicate pipeline issues or data loss

For outlier severity guidance:
- IQR outliers > 10% of rows = warning, > 25% = critical
- Z-Score outliers (3σ) > 5% = warning, > 15% = critical
- If both IQR and Z-Score detect outliers, the column likely has heavy tails

Return ONLY valid JSON with this exact structure:
{
  "summary": "<2-3 sentence overall data quality assessment>",
  "quality_score": <integer 0-100>,
  "columns": [
    {
      "column": "<column_name>",
      "quality_score": <integer 0-100>,
      "issues": [
        {
          "severity": "critical|warning|info",
          "category": "nulls|format|content|cardinality|distribution|temporal|string_quality|spread|volume",
          "title": "<short issue title>",
          "description": "<detailed description with specific numbers>",
          "affected_rows": <number or null>,
          "recommendation": "<concrete actionable fix>"
        }
      ],
      "insights": "<what looks good or any useful context about this column>"
    }
  ],
  "recommendations": ["<global recommendation 1>", "<global recommendation 2>"]
}"""

_dq_prepared_runs_lock = _threading.Lock()
_dq_prepared_runs: dict[str, dict] = {}
_DQ_PREPARED_TTL_SECONDS = 30 * 60
_DQ_PREPARED_MAX_ITEMS = 64


def _dq_prune_prepared_runs() -> None:
    import time as _time

    now = _time.time()
    with _dq_prepared_runs_lock:
        stale_ids = [
            run_id
            for run_id, payload in _dq_prepared_runs.items()
            if now - float(payload.get("created_at", 0)) > _DQ_PREPARED_TTL_SECONDS
        ]
        for run_id in stale_ids:
            _dq_prepared_runs.pop(run_id, None)
        while len(_dq_prepared_runs) > _DQ_PREPARED_MAX_ITEMS:
            oldest = min(
                _dq_prepared_runs.items(),
                key=lambda item: float(item[1].get("created_at", 0)),
            )[0]
            _dq_prepared_runs.pop(oldest, None)


def _dq_store_prepared_run(payload: dict) -> str:
    import time as _time

    _dq_prune_prepared_runs()
    run_id = f"dq_{uuid.uuid4().hex}"
    body = dict(payload or {})
    body["created_at"] = _time.time()
    with _dq_prepared_runs_lock:
        _dq_prepared_runs[run_id] = body
    return run_id


def _dq_pop_prepared_run(run_id: str) -> dict | None:
    _dq_prune_prepared_runs()
    with _dq_prepared_runs_lock:
        return _dq_prepared_runs.pop(run_id, None)


def _dq_filter_note(filter_col, filter_op, filter_val, filter_val2) -> str:
    if not filter_col or filter_val is None:
        return ""
    if filter_op == "BETWEEN" and filter_val2 is not None:
        return (
            f" (filtered to rows where `{filter_col}` BETWEEN '{filter_val}' "
            f"AND '{filter_val2}')"
        )
    return f" (filtered to rows where `{filter_col}` {filter_op} '{filter_val}')"


def _dq_compact_column_stat_for_llm(stat: dict, aggressive: bool = False) -> dict:
    keep = (
        "column", "type", "total", "null_count", "null_pct", "distinct_count", "distinct_pct",
        "empty_count", "empty_pct", "min", "max", "avg", "stddev", "p25", "p50", "p75",
        "outlier_count", "outlier_pct", "zscore_outlier_count", "zscore_outlier_pct",
        "negative_count", "zero_count", "coeff_variation", "skewness_approx", "min_length",
        "max_length", "avg_length", "sentinel_count", "whitespace_padded_count",
        "all_caps_count", "numeric_string_count", "email_like_count", "min_date",
        "max_date", "future_count", "epoch_sentinel_count", "weekend_count", "pre_1900_count",
        "filter_applied", "query_error",
    )
    compact = {}
    for key in keep:
        if key in stat and stat.get(key) is not None:
            compact[key] = stat.get(key)

    top_values = stat.get("top_values") or []
    if isinstance(top_values, list) and top_values:
        max_values = 3 if aggressive else 6
        max_len = 42 if aggressive else 80
        compact["top_values"] = [
            {
                "value": str(v.get("value", ""))[:max_len],
                "count": int(v.get("count", 0)),
            }
            for v in top_values[:max_values]
            if isinstance(v, dict)
        ]
    if compact.get("query_error"):
        compact["query_error"] = str(compact["query_error"])[:200]
    return compact


def _dq_compact_volume_for_llm(volume_analysis: dict | None) -> dict | None:
    if not isinstance(volume_analysis, dict):
        return None
    if volume_analysis.get("error"):
        return {
            "time_column": volume_analysis.get("time_column"),
            "error": str(volume_analysis.get("error"))[:200],
        }
    return {
        "time_column": volume_analysis.get("time_column"),
        "granularity": volume_analysis.get("granularity"),
        "periods": volume_analysis.get("periods"),
        "avg_volume": volume_analysis.get("avg_volume"),
        "stddev_volume": volume_analysis.get("stddev_volume"),
        "min_volume": volume_analysis.get("min_volume"),
        "max_volume": volume_analysis.get("max_volume"),
        "p25_volume": volume_analysis.get("p25_volume"),
        "p75_volume": volume_analysis.get("p75_volume"),
        "low_volume_threshold": volume_analysis.get("low_volume_threshold"),
        "anomaly_count": volume_analysis.get("anomaly_count"),
        "anomaly_periods": (volume_analysis.get("anomaly_periods") or [])[:8],
        "recent_periods": (volume_analysis.get("recent_periods") or [])[:6],
    }


def _dq_build_llm_plan(
    table: str,
    scan_desc: str,
    filter_note: str,
    column_stats: list[dict],
    volume_analysis: dict | None,
) -> dict:
    compact_stats = [_dq_compact_column_stat_for_llm(s) for s in column_stats]
    compact_volume = _dq_compact_volume_for_llm(volume_analysis)

    raw_context_window = _get_model_context_limit()
    effective_context = _get_effective_context_limit()

    prompt_scaffold = (
        f"Analyze data quality for table `{table}` ({scan_desc}{filter_note}).\n\n"
        "Column Statistics:\n[]\n\n"
        "Provide a thorough analysis identifying all data quality issues with specific numbers."
    )
    scaffold_tokens = _estimate_tokens(prompt_scaffold)
    system_tokens = _estimate_tokens(_DQ_SYSTEM_PROMPT)
    output_reserve = max(180, min(900, int(effective_context * 0.20)))
    token_budget_per_call = max(
        120,
        effective_context - system_tokens - scaffold_tokens - output_reserve,
    )

    volume_tokens = 0
    if compact_volume:
        volume_tokens = _estimate_tokens(
            "VOLUME CONSISTENCY:\n"
            + json.dumps(compact_volume, separators=(",", ":"), ensure_ascii=False, default=str)
        )

    serialized_stats = [
        json.dumps(s, separators=(",", ":"), ensure_ascii=False, default=str)
        for s in compact_stats
    ]
    stat_tokens = [max(8, _estimate_tokens(s)) for s in serialized_stats]

    batches = []
    i = 0
    while i < len(compact_stats):
        batch_budget = token_budget_per_call - (volume_tokens if len(batches) == 0 else 0)
        batch_budget = max(80, batch_budget)
        current = []
        used = 0
        while i < len(compact_stats):
            entry = compact_stats[i]
            entry_tokens = stat_tokens[i]
            if entry_tokens > batch_budget:
                compact_entry = _dq_compact_column_stat_for_llm(column_stats[i], aggressive=True)
                compact_tokens = _estimate_tokens(
                    json.dumps(compact_entry, separators=(",", ":"), ensure_ascii=False, default=str)
                )
                if compact_tokens < entry_tokens:
                    compact_stats[i] = compact_entry
                    stat_tokens[i] = compact_tokens
                    entry = compact_entry
                    entry_tokens = compact_tokens

            if current and used + entry_tokens > batch_budget:
                break

            current.append(entry)
            used += entry_tokens
            i += 1

            # Always advance even when a single field exceeds the budget.
            if entry_tokens > batch_budget:
                break

        if not current:
            current.append(compact_stats[i])
            i += 1
        batches.append(current)

    prompt_tokens_per_call = []
    for idx, batch in enumerate(batches):
        batch_tokens = _estimate_tokens(
            json.dumps(batch, separators=(",", ":"), ensure_ascii=False, default=str)
        )
        prompt_tokens_per_call.append(
            scaffold_tokens + batch_tokens + (volume_tokens if idx == 0 else 0)
        )

    total_synthesis_tokens = scaffold_tokens + volume_tokens + sum(stat_tokens)
    calls_by_ratio = max(1, math.ceil(total_synthesis_tokens / max(1, token_budget_per_call)))

    return {
        "raw_context_window_tokens": raw_context_window,
        "effective_context_window_tokens": effective_context,
        "token_budget_per_call": token_budget_per_call,
        "estimated_total_synthesis_tokens": total_synthesis_tokens,
        "estimated_total_prompt_tokens": sum(prompt_tokens_per_call),
        "estimated_calls_ratio": calls_by_ratio,
        "estimated_calls": len(batches),
        "estimated_prompt_tokens_per_call": prompt_tokens_per_call,
        "columns_per_call": [len(b) for b in batches],
        "batches": batches,
        "compact_volume": compact_volume,
    }


def _dq_public_plan(plan: dict) -> dict:
    if not isinstance(plan, dict):
        return {}
    return {
        "raw_context_window_tokens": plan.get("raw_context_window_tokens"),
        "effective_context_window_tokens": plan.get("effective_context_window_tokens"),
        "token_budget_per_call": plan.get("token_budget_per_call"),
        "estimated_total_synthesis_tokens": plan.get("estimated_total_synthesis_tokens"),
        "estimated_total_prompt_tokens": plan.get("estimated_total_prompt_tokens"),
        "estimated_calls_ratio": plan.get("estimated_calls_ratio"),
        "estimated_calls": plan.get("estimated_calls"),
        "estimated_prompt_tokens_per_call": plan.get("estimated_prompt_tokens_per_call") or [],
        "columns_per_call": plan.get("columns_per_call") or [],
    }


def _dq_call_batch_llm(
    table: str,
    scan_desc: str,
    filter_note: str,
    batch_stats: list[dict],
    batch_idx: int,
    total_batches: int,
    compact_volume: dict | None = None,
) -> dict:
    batch_json = json.dumps(batch_stats, indent=2, ensure_ascii=False, default=str)
    volume_section = ""
    if compact_volume and batch_idx == 1:
        volume_section = (
            "\n\nVOLUME CONSISTENCY:\n"
            + json.dumps(compact_volume, indent=2, ensure_ascii=False, default=str)
        )
    batch_note = f" [batch {batch_idx}/{total_batches}]" if total_batches > 1 else ""
    user_msg = (
        f"Analyze data quality for table `{table}` ({scan_desc}{filter_note}){batch_note}.\n\n"
        f"Column Statistics:\n{batch_json}{volume_section}\n\n"
        "Provide a thorough analysis identifying all data quality issues with specific numbers."
    )
    safe_user_msg = _truncate_text_to_budget(user_msg, int(_get_effective_context_limit() * 0.68))
    raw = _call_llm(_DQ_SYSTEM_PROMPT, [{"role": "user", "content": safe_user_msg}], temperature=0.1)

    try:
        parsed = _parse_llm_json(raw)
    except Exception:
        parsed = None

    if not isinstance(parsed, dict):
        return {
            "summary": str(raw)[:320],
            "quality_score": None,
            "columns": [
                {"column": s.get("column"), "quality_score": None, "issues": [], "insights": ""}
                for s in batch_stats
            ],
            "recommendations": [],
        }
    return parsed


def _dq_merge_batch_results(column_stats: list[dict], batch_results: list[dict]) -> dict:
    scores = []
    summaries = []
    recs = []
    by_column: dict[str, dict] = {}

    for br in batch_results:
        if isinstance(br.get("quality_score"), (int, float)):
            scores.append(int(br.get("quality_score")))
        if br.get("summary"):
            summaries.append(str(br.get("summary")))
        for rec in (br.get("recommendations") or []):
            if isinstance(rec, str) and rec.strip():
                recs.append(rec.strip())
        for col in (br.get("columns") or []):
            if not isinstance(col, dict):
                continue
            cname = str(col.get("column") or "").strip()
            if not cname:
                continue
            clean = {
                "column": cname,
                "quality_score": col.get("quality_score"),
                "issues": col.get("issues") if isinstance(col.get("issues"), list) else [],
                "insights": str(col.get("insights", ""))[:400],
            }
            by_column[cname] = clean

    ordered_columns = []
    for stat in column_stats:
        cname = str(stat.get("column", "")).strip()
        fallback = {
            "column": cname,
            "quality_score": None,
            "issues": [],
            "insights": "",
        }
        ordered_columns.append(by_column.get(cname, fallback))

    merged_score = round(sum(scores) / len(scores)) if scores else None
    merged_summary = (
        f"Analysis of {len(column_stats)} columns in {len(batch_results)} LLM call(s). "
        + " ".join(summaries[:2])
    ).strip()
    return {
        "summary": merged_summary[:700],
        "quality_score": merged_score,
        "columns": ordered_columns,
        "recommendations": list(dict.fromkeys(recs))[:10],
    }


def _dq_run_llm_analysis(
    table: str,
    scan_desc: str,
    filter_note: str,
    column_stats: list[dict],
    volume_analysis: dict | None,
    llm_plan: dict,
) -> dict:
    batches = llm_plan.get("batches") or [[_dq_compact_column_stat_for_llm(s) for s in column_stats]]
    compact_volume = llm_plan.get("compact_volume")
    if compact_volume is None:
        compact_volume = _dq_compact_volume_for_llm(volume_analysis)

    batch_results = []
    for idx, batch in enumerate(batches):
        br = _dq_call_batch_llm(
            table=table,
            scan_desc=scan_desc,
            filter_note=filter_note,
            batch_stats=batch,
            batch_idx=idx + 1,
            total_batches=len(batches),
            compact_volume=compact_volume,
        )
        batch_results.append(br)

    return _dq_merge_batch_results(column_stats, batch_results)


def _dq_parse_request_payload(data: dict) -> dict:
    data = data or {}
    table = str(data.get("table") or "").strip()

    columns_raw = data.get("columns", [])
    if isinstance(columns_raw, str):
        columns = [c.strip() for c in columns_raw.split(",") if c.strip()]
    elif isinstance(columns_raw, list):
        columns = [str(c).strip() for c in columns_raw if str(c).strip()]
    else:
        columns = []
    columns = list(dict.fromkeys(columns))

    sample_raw = data.get("sample_size")
    sample_size = None
    if sample_raw is not None and str(sample_raw).strip() != "":
        try:
            sample_size = int(sample_raw)
        except Exception as exc:
            raise ValueError("sample_size must be an integer") from exc
        sample_size = min(max(sample_size, 1), 500000)

    filter_col = str(data.get("filter_column") or "").strip() or None
    filter_op = str(data.get("filter_operator") or "=").strip().upper()
    allowed_ops = {"=", "!=", "<", ">", "<=", ">=", "LIKE", "BETWEEN"}
    if filter_op not in allowed_ops:
        filter_op = "="
    filter_val = data.get("filter_value")
    if filter_val is not None and str(filter_val) != "":
        filter_val = str(filter_val)
    else:
        filter_val = None
    filter_val2 = data.get("filter_value2")
    if filter_val2 is not None and str(filter_val2) != "":
        filter_val2 = str(filter_val2)
    else:
        filter_val2 = None

    time_col = str(data.get("time_column") or "").strip() or None

    if not table:
        raise ValueError("Table name is required")
    if not columns:
        raise ValueError("At least one column is required")
    if not _re.match(r"^[\w.]+$", table):
        raise ValueError("Invalid table name")
    for col in columns:
        if not _re.match(r"^\w+$", col):
            raise ValueError(f"Invalid column name: {col}")
    if filter_col and not _re.match(r"^\w+$", filter_col):
        raise ValueError("Invalid filter column name")
    if time_col and not _re.match(r"^\w+$", time_col):
        raise ValueError("Invalid time column name")
    if filter_op == "BETWEEN" and filter_col and filter_val is not None and filter_val2 is None:
        raise ValueError("filter_value2 is required when filter_operator is BETWEEN")

    return {
        "table": table,
        "columns": columns,
        "sample_size": sample_size,
        "filter_col": filter_col,
        "filter_op": filter_op,
        "filter_val": filter_val,
        "filter_val2": filter_val2,
        "time_col": time_col,
    }


def _dq_collect_profiles(params: dict) -> dict:
    table = params["table"]
    columns = params["columns"]
    sample_size = params.get("sample_size")
    filter_col = params.get("filter_col")
    filter_op = params.get("filter_op")
    filter_val = params.get("filter_val")
    filter_val2 = params.get("filter_val2")
    time_col = params.get("time_col")

    safe_table = _re.sub(r"[^\w.]", "", table)

    where_clause = ""
    if filter_col and filter_op and filter_val is not None:
        safe_fcol = _re.sub(r"[^\w]", "", filter_col)
        escaped_val = str(filter_val).replace("'", "''")
        if filter_op == "BETWEEN" and filter_val2 is not None:
            escaped_val2 = str(filter_val2).replace("'", "''")
            where_clause = f" WHERE `{safe_fcol}` BETWEEN '{escaped_val}' AND '{escaped_val2}'"
        else:
            where_clause = f" WHERE `{safe_fcol}` {filter_op} '{escaped_val}'"

    client = get_clickhouse_client()

    desc = client.query(f"DESCRIBE TABLE {table}")
    col_types = {row[0]: row[1] for row in desc.result_rows}

    column_stats = []
    for col in columns:
        col_type = col_types.get(col, "String")
        cs = _dq_column_stats(
            client,
            table,
            col,
            col_type,
            sample_size,
            filter_col=filter_col,
            filter_op=filter_op,
            filter_val=filter_val,
            filter_val2=filter_val2,
        )
        column_stats.append(cs)

    volume_analysis = None
    if time_col:
        safe_time_col = _re.sub(r"[^\w]", "", time_col)
        try:
            range_r = client.query(
                f"SELECT min(`{safe_time_col}`), max(`{safe_time_col}`) FROM {safe_table}"
            )
            rr = range_r.result_rows[0]
            min_ts, max_ts = rr[0], rr[1]
            try:
                delta_days = (max_ts - min_ts).days if hasattr(max_ts, "days") else 999
            except Exception:
                delta_days = 999
            granularity = "hour" if delta_days <= 7 else "day"
            trunc_fn = "toStartOfHour" if granularity == "hour" else "toStartOfDay"

            if where_clause:
                where_vol = where_clause + f" AND `{safe_time_col}` IS NOT NULL"
            else:
                where_vol = f" WHERE `{safe_time_col}` IS NOT NULL"
            vol_r = client.query(
                f"SELECT {trunc_fn}(`{safe_time_col}`) AS period, count() AS cnt"
                f" FROM {safe_table}{where_vol}"
                f" GROUP BY period ORDER BY period"
            )
            vol_rows = [{"period": str(row[0]), "count": int(row[1])} for row in vol_r.result_rows]
            if vol_rows:
                counts = [r["count"] for r in vol_rows]
                vol_avg = sum(counts) / len(counts)
                vol_stddev = (sum((c - vol_avg) ** 2 for c in counts) / max(len(counts), 1)) ** 0.5
                sorted_counts = sorted(counts)
                vol_p25 = sorted_counts[len(sorted_counts) // 4]
                vol_p75 = sorted_counts[3 * len(sorted_counts) // 4]
                iqr_vol = vol_p75 - vol_p25
                low_threshold = max(1, vol_p25 - 1.5 * iqr_vol)
                anomaly_periods = [r for r in vol_rows if r["count"] < low_threshold]
                volume_analysis = {
                    "time_column": time_col,
                    "granularity": granularity,
                    "periods": len(vol_rows),
                    "avg_volume": round(vol_avg, 1),
                    "stddev_volume": round(vol_stddev, 1),
                    "min_volume": min(counts),
                    "max_volume": max(counts),
                    "p25_volume": vol_p25,
                    "p75_volume": vol_p75,
                    "low_volume_threshold": round(low_threshold, 1),
                    "anomaly_count": len(anomaly_periods),
                    "anomaly_periods": anomaly_periods[:20],
                    "recent_periods": vol_rows[-10:],
                }
        except Exception as exc:
            volume_analysis = {"error": str(exc), "time_column": time_col}

    return {
        "table": table,
        "sample_size": sample_size,
        "column_stats": column_stats,
        "volume_analysis": volume_analysis,
        "scan_desc": "full scan" if sample_size is None else f"sample of {sample_size:,} rows",
        "filter_note": _dq_filter_note(filter_col, filter_op, filter_val, filter_val2),
    }


def _dq_prepare_run(params: dict) -> dict:
    profile = _dq_collect_profiles(params)
    llm_plan = _dq_build_llm_plan(
        table=profile["table"],
        scan_desc=profile["scan_desc"],
        filter_note=profile["filter_note"],
        column_stats=profile["column_stats"],
        volume_analysis=profile.get("volume_analysis"),
    )
    prepared_run_id = _dq_store_prepared_run(
        {
            "table": profile["table"],
            "sample_size": profile["sample_size"],
            "column_stats": profile["column_stats"],
            "volume_analysis": profile.get("volume_analysis"),
            "scan_desc": profile["scan_desc"],
            "filter_note": profile["filter_note"],
            "llm_plan": llm_plan,
        }
    )
    return {
        "status": "awaiting_llm_approval",
        "approval_required": True,
        "prepared_run_id": prepared_run_id,
        "table": profile["table"],
        "sample_size": profile["sample_size"],
        "columns_count": len(profile["column_stats"]),
        "llm_plan": _dq_public_plan(llm_plan),
        "message": (
            f"{llm_plan.get('estimated_calls', 1)} appel(s) LLM local estimé(s). "
            "Confirmez OUI pour lancer l'analyse ou NON pour annuler."
        ),
    }


@app.route("/api/data-quality/plan", methods=["POST"])
def data_quality_plan():
    """Collect stats + estimate token-aware LLM batches, without calling the LLM."""
    data = request.get_json(silent=True) or {}
    try:
        params = _dq_parse_request_payload(data)
        return jsonify(_dq_prepare_run(params))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        print(f"Data quality plan error: {exc}")
        return jsonify({"error": str(exc)}), 500


@app.route("/api/data-quality/analyze", methods=["POST"])
def analyze_data_quality():
    """Run AI-powered data quality analysis on selected table columns."""
    data = request.get_json(silent=True) or {}
    prepared_run_id = str(data.get("prepared_run_id") or "").strip()

    try:
        if prepared_run_id:
            approval_raw = data.get("llm_approval")
            approval_text = str(approval_raw).strip().lower() if approval_raw is not None else ""
            approved = approval_raw is True or approval_text in {"oui", "yes", "true", "1"}
            if not approved:
                return jsonify({
                    "error": "LLM approval is required (OUI/NON).",
                    "approval_required": True,
                    "prepared_run_id": prepared_run_id,
                }), 400

            prepared = _dq_pop_prepared_run(prepared_run_id)
            if not prepared:
                return jsonify({
                    "error": "Prepared run not found or expired. Please run planning again.",
                    "prepared_run_id": prepared_run_id,
                }), 404

            analysis = _dq_run_llm_analysis(
                table=prepared["table"],
                scan_desc=prepared["scan_desc"],
                filter_note=prepared["filter_note"],
                column_stats=prepared["column_stats"],
                volume_analysis=prepared.get("volume_analysis"),
                llm_plan=prepared["llm_plan"],
            )
            plan_public = _dq_public_plan(prepared["llm_plan"])
            plan_public["executed_calls"] = len(prepared["llm_plan"].get("batches") or [])
            return jsonify({
                "table": prepared["table"],
                "sample_size": prepared.get("sample_size"),
                "column_stats": prepared["column_stats"],
                "analysis": analysis,
                "volume_analysis": prepared.get("volume_analysis"),
                "llm_plan": plan_public,
            })

        if bool(data.get("prepare_only")):
            params = _dq_parse_request_payload(data)
            return jsonify(_dq_prepare_run(params))

        params = _dq_parse_request_payload(data)
        profile = _dq_collect_profiles(params)
        llm_plan = _dq_build_llm_plan(
            table=profile["table"],
            scan_desc=profile["scan_desc"],
            filter_note=profile["filter_note"],
            column_stats=profile["column_stats"],
            volume_analysis=profile.get("volume_analysis"),
        )
        analysis = _dq_run_llm_analysis(
            table=profile["table"],
            scan_desc=profile["scan_desc"],
            filter_note=profile["filter_note"],
            column_stats=profile["column_stats"],
            volume_analysis=profile.get("volume_analysis"),
            llm_plan=llm_plan,
        )
        plan_public = _dq_public_plan(llm_plan)
        plan_public["executed_calls"] = len(llm_plan.get("batches") or [])
        return jsonify({
            "table": profile["table"],
            "sample_size": profile["sample_size"],
            "column_stats": profile["column_stats"],
            "analysis": analysis,
            "volume_analysis": profile.get("volume_analysis"),
            "llm_plan": plan_public,
        })
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        print(f"Data quality error: {exc}")
        return jsonify({"error": str(exc)}), 500


# ---------------------------------------------------------------------------
# CSV Export
# ---------------------------------------------------------------------------
MAX_EXPORT_ROWS = 1_000_000


@app.route("/api/export_csv", methods=["POST"])
def export_csv():
    """Export a ClickHouse query result to a pipe-separated CSV file.

    Body parameters:
      - sql  (required): the SELECT query to execute
      - output_path (optional): absolute server-side path to write the file.
        When provided the file is written locally and the path + row count are
        returned. When absent the CSV content is sent as an HTTP download.

    Row count is capped at MAX_EXPORT_ROWS (1 000 000).
    """
    import csv
    import io as _io

    data = request.get_json(silent=True) or {}
    sql = (data.get("sql") or "").strip()
    output_path = (data.get("output_path") or "").strip()

    if not sql:
        return jsonify({"error": "Aucune requête SQL fournie."}), 400

    # Ensure the query has a LIMIT that does not exceed MAX_EXPORT_ROWS.
    # We inject one when absent; when the user already has LIMIT we leave it
    # so that their intentional cap is respected (it will still be capped
    # client-side at MAX_EXPORT_ROWS rows after fetch).
    if "LIMIT" not in sql.upper():
        sql = f"{sql} LIMIT {MAX_EXPORT_ROWS}"

    try:
        client = get_clickhouse_client()
        result = client.query(sql)
        rows = _rows_to_dicts(result)
        rows = rows[:MAX_EXPORT_ROWS]

        if not rows:
            return jsonify({"error": "La requête n'a retourné aucune donnée."}), 400

        # Build pipe-separated CSV
        output = _io.StringIO()
        writer = csv.writer(output, delimiter="|", quoting=csv.QUOTE_MINIMAL)
        writer.writerow(rows[0].keys())
        for row in rows:
            writer.writerow(row.values())
        csv_content = output.getvalue()

        if output_path:
            parent_dir = os.path.dirname(output_path)
            if parent_dir:
                os.makedirs(parent_dir, exist_ok=True)
            with open(output_path, "w", encoding="utf-8", newline="") as f:
                f.write(csv_content)
            return jsonify({"path": output_path, "row_count": len(rows)})
        else:
            from flask import Response
            return Response(
                csv_content,
                mimetype="text/csv; charset=utf-8",
                headers={"Content-Disposition": 'attachment; filename="export.csv"'},
            )

    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


# ---------------------------------------------------------------------------
# Config export / import
# ---------------------------------------------------------------------------
@app.route("/api/config/export", methods=["GET"])
def export_config():
    """Export all configuration and data as a single JSON file."""
    db = read_db()
    payload = {
        "version": 1,
        "clickhouseConfig": clickhouse_config,
        "llmConfig": llm_config,
        "ragConfig": rag_config,
        "knowledge_folders": db.get("knowledge_folders", []),
        "table_mappings": db.get("table_mappings", []),
        "table_metadata": db.get("table_metadata", []),
        "saved_queries": db.get("saved_queries", []),
    }
    return app.response_class(
        json.dumps(payload, indent=2, ensure_ascii=False),
        mimetype="application/json",
        headers={"Content-Disposition": "attachment; filename=clicksense-config.json"},
    )


@app.route("/api/config/import", methods=["POST"])
def import_config():
    """Import configuration and data from a previously exported JSON file."""
    global clickhouse_config, llm_config, rag_config
    data = request.get_json()
    if not data:
        return jsonify({"error": "Empty payload"}), 400

    if "clickhouseConfig" in data:
        clickhouse_config.update(data["clickhouseConfig"])
        # Sync database with databases[0] if databases list is present
        if data["clickhouseConfig"].get("databases"):
            dbs = [d.strip() for d in data["clickhouseConfig"]["databases"] if d.strip()]
            if dbs:
                clickhouse_config["databases"] = dbs
                clickhouse_config["database"] = dbs[0]
    if "llmConfig" in data:
        llm_config.update(data["llmConfig"])
        _model_context_cache.clear()
    if "ragConfig" in data:
        rag_config.update(data["ragConfig"])

    db = read_db()
    if "knowledge_folders" in data:
        db["knowledge_folders"] = data["knowledge_folders"]
    if "table_mappings" in data:
        db["table_mappings"] = data["table_mappings"]
    if "table_metadata" in data:
        db["table_metadata"] = data["table_metadata"]
    if "saved_queries" in data:
        db["saved_queries"] = data["saved_queries"]
    write_db(db)

    return jsonify({"success": True})


# ---------------------------------------------------------------------------
# Agents system
# ---------------------------------------------------------------------------

# In-memory session store for the ClickHouse Writer agent
_writer_sessions: dict = {}

# In-memory session store for the ETL agent
_etl_sessions: dict = {}

# In-memory session store for the AI Data Analyst session agent
_data_analyst_sessions: dict = {}
_data_analyst_sessions_lock = _threading.Lock()
_DATA_ANALYST_MAX_EVENTS = 500
_DATA_ANALYST_SCHEMA_CACHE_TTL_SEC = 180


def _da_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _da_coerce_bool(value, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "on", "oui"}
    return bool(value)


def _da_coerce_int(value, default: int, min_value: int, max_value: int) -> int:
    parsed = _parse_int_like(value, default)
    return max(min_value, min(max_value, parsed))


def _da_parse_table_filter(raw_value) -> list:
    if isinstance(raw_value, list):
        values = [str(v).strip() for v in raw_value if str(v).strip()]
    else:
        txt = str(raw_value or "").strip()
        values = [p.strip() for p in txt.split(",") if p.strip()] if txt else []
    return values[:200]


def _da_normalize_params(raw_params: dict | None) -> dict:
    params = raw_params or {}
    knowledge_mode, use_knowledge_base, use_knowledge_agent = _resolve_knowledge_mode_flags(
        knowledge_mode_raw=params.get("knowledge_mode", params.get("knowledgeMode", "")),
        use_knowledge_base_raw=params.get("use_knowledge_base", "yes"),
        use_knowledge_agent_raw=params.get("use_knowledge_agent", "no"),
    )
    return {
        "max_steps": _da_coerce_int(params.get("max_steps", 8), 8, 1, 50),
        "knowledge_mode": knowledge_mode,
        "use_knowledge_base": use_knowledge_base,
        "use_knowledge_agent": use_knowledge_agent,
        "memory_turn_limit": _da_coerce_int(params.get("memory_turn_limit", 8), 8, 2, 24),
        "memory_token_budget": _da_coerce_int(params.get("memory_token_budget", 700), 700, 120, 2400),
        "auto_run": _da_coerce_bool(params.get("auto_run", "yes"), True),
        "table_filter": _da_parse_table_filter(params.get("table_filter", "")),
    }


def _da_log_event(session: dict, message: str, level: str = "info", kind: str = "runtime") -> None:
    level_norm = str(level or "info").lower()
    if level_norm not in {"info", "warn", "error"}:
        level_norm = "info"
    entry = {
        "seq": int(session.get("event_seq", 0)) + 1,
        "ts": _da_now_iso(),
        "level": level_norm,
        "kind": str(kind or "runtime"),
        "message": str(message or "").strip(),
    }
    session["event_seq"] = entry["seq"]
    session.setdefault("event_log", []).append(entry)
    if len(session["event_log"]) > _DATA_ANALYST_MAX_EVENTS:
        session["event_log"] = session["event_log"][-_DATA_ANALYST_MAX_EVENTS:]
    session["updated_at"] = entry["ts"]
    try:
        _log(
            f"[{session.get('id', '?')[:8]}] {entry['message']}",
            level=level_norm,
            source="ai-data-analyst",
        )
    except Exception:
        pass


def _da_refresh_memory_summary(session: dict) -> None:
    params = session.get("params", {})
    turn_limit = _da_coerce_int(params.get("memory_turn_limit", 8), 8, 2, 24)
    token_budget = _da_coerce_int(params.get("memory_token_budget", 700), 700, 120, 2400)
    convo = session.get("conversation", [])
    lines = []
    for msg in convo[-(turn_limit * 2):]:
        role = "U" if msg.get("role") == "user" else "A"
        text = " ".join(str(msg.get("content", "")).split())
        if not text:
            continue
        lines.append(f"{role}: {text[:220]}")

    last_result = session.get("last_result") or {}
    if last_result:
        lines.append(
            "Last run: "
            f"steps={last_result.get('total_steps', 0)}, "
            f"retries={last_result.get('technical_retries_used', 0)}, "
            f"queries={last_result.get('query_attempts_used', 0)}."
        )

    raw_summary = "\n".join(lines)
    session["memory_summary"] = _truncate_text_to_budget(raw_summary, token_budget)
    session["updated_at"] = _da_now_iso()


def _da_compose_question(session: dict, question: str) -> str:
    user_question = (question or "").strip()
    if not user_question:
        return ""
    params = session.get("params", {})
    token_budget = _da_coerce_int(params.get("memory_token_budget", 700), 700, 120, 2400)
    memory_summary = (session.get("memory_summary") or "").strip()
    notes = session.get("paused_notes", [])[-3:]

    blocks = [user_question]
    if memory_summary:
        blocks.append(
            "CONVERSATION MEMORY (compact, may be partial):\n"
            + memory_summary
            + "\nUse this memory only when relevant to the new question."
        )
    if notes:
        note_text = "\n".join(f"- {str(n)[:220]}" for n in notes)
        blocks.append(
            "ADDITIONAL USER NOTES PROVIDED DURING PAUSE:\n"
            + note_text
        )

    combined = "\n\n".join(blocks)
    # Keep enough room for the current question while bounding historical memory.
    return _truncate_text_to_budget(combined, max(180, int(token_budget * 1.35)))


def _da_build_table_metadata_map() -> dict:
    db = read_db()
    out = {}
    for row in db.get("table_metadata", []):
        table_name = str(row.get("table_name", "")).strip()
        if not table_name:
            continue
        out[table_name] = {
            "description": row.get("description", ""),
            "is_favorite": bool(row.get("is_favorite", False)),
        }
    return out


def _da_fetch_schema_for_agent() -> dict:
    client = get_clickhouse_client()
    databases = clickhouse_config.get("databases") or []
    if not databases:
        databases = [clickhouse_config.get("database", "default")]
    databases = [d.strip() for d in databases if d.strip()]
    if not databases:
        databases = ["default"]
    multi_db = len(databases) > 1
    dbs_in = ", ".join(f"'{d}'" for d in databases)
    result = client.query(
        f"SELECT database, table, name, type FROM system.columns "
        f"WHERE database IN ({dbs_in}) "
        f"ORDER BY database, table, name"
    )
    rows = _rows_to_dicts(result)
    schema = {}
    for row in rows:
        db_name = row.get("database", "")
        tbl = row.get("table", "")
        key = f"{db_name}.{tbl}" if multi_db else tbl
        schema.setdefault(key, []).append({
            "name": row.get("name", ""),
            "type": row.get("type", ""),
        })
    return schema


def _da_get_schema_with_cache(session: dict) -> dict:
    now_ts = datetime.now(timezone.utc).timestamp()
    cached_schema = session.get("cached_schema")
    cached_ts = float(session.get("cached_schema_ts", 0) or 0)
    if cached_schema and (now_ts - cached_ts) <= _DATA_ANALYST_SCHEMA_CACHE_TTL_SEC:
        return cached_schema

    schema = _da_fetch_schema_for_agent()
    session["cached_schema"] = schema
    session["cached_schema_ts"] = now_ts
    return schema


def _da_session_payload(session: dict) -> dict:
    return {
        "session_id": session.get("id"),
        "status": session.get("status", "idle"),
        "running": bool(session.get("running", False)),
        "pause_requested": bool(session.get("pause_requested", False)),
        "stop_requested": bool(session.get("stop_requested", False)),
        "pending_user_inputs": len(session.get("pending_user_inputs", [])),
        "memory_summary": session.get("memory_summary", ""),
        "event_log": session.get("event_log", [])[-220:],
        "response_seq": int(session.get("response_seq", 0)),
        "latest_assistant": session.get("latest_assistant"),
        "last_result": session.get("last_result"),
        "params": session.get("params", {}),
        "created_at": session.get("created_at"),
        "updated_at": session.get("updated_at"),
    }


def _da_start_worker_if_needed(session: dict) -> bool:
    if session.get("running"):
        return False
    if not session.get("pending_user_inputs"):
        return False
    session["running"] = True
    session["status"] = "running"
    session["run_seq"] = int(session.get("run_seq", 0)) + 1
    run_seq = session["run_seq"]
    session["updated_at"] = _da_now_iso()
    t = _threading.Thread(
        target=_da_worker_loop,
        args=(session.get("id"), run_seq),
        daemon=True,
    )
    session["worker"] = t
    t.start()
    return True


def _da_extract_response_payload(resp_obj):
    status_code = 200
    response = resp_obj
    if isinstance(resp_obj, tuple):
        response = resp_obj[0]
        if len(resp_obj) > 1 and isinstance(resp_obj[1], int):
            status_code = resp_obj[1]
    if hasattr(response, "status_code"):
        status_code = int(getattr(response, "status_code", status_code) or status_code)
    if hasattr(response, "get_json"):
        data = response.get_json(silent=True) or {}
    else:
        data = {}
    return status_code, data


def _da_worker_loop(session_id: str, run_seq: int) -> None:
    while True:
        with _data_analyst_sessions_lock:
            session = _data_analyst_sessions.get(session_id)
            if not session:
                return
            if int(session.get("run_seq", 0)) != int(run_seq):
                # A newer run superseded this worker.
                return
            if session.get("stop_requested"):
                session["running"] = False
                session["status"] = "stopped"
                _da_log_event(session, "Run cancelled before start (stop requested).", "warn", "control")
                return
            if session.get("pause_requested"):
                session["running"] = False
                session["status"] = "paused"
                _da_log_event(session, "Run paused before start.", "warn", "control")
                return
            if not session.get("pending_user_inputs"):
                session["running"] = False
                if session.get("status") not in {"paused", "stopped", "error"}:
                    session["status"] = "idle"
                session["updated_at"] = _da_now_iso()
                return
            user_question = str(session["pending_user_inputs"].pop(0)).strip()
            params = dict(session.get("params", {}))
            _da_refresh_memory_summary(session)
            composed_question = _da_compose_question(session, user_question)
            _da_log_event(
                session,
                f"Run #{run_seq}: starting analysis for question '{user_question[:120]}'",
                "info",
                "run",
            )

            try:
                schema = _da_get_schema_with_cache(session)
                table_metadata = _da_build_table_metadata_map()
            except Exception as schema_exc:
                schema = {}
                table_metadata = {}
                _da_log_event(
                    session,
                    f"Schema/metadata loading failed ({schema_exc}). Running with empty schema context.",
                    "warn",
                    "schema",
                )

            table_filter = _da_parse_table_filter(params.get("table_filter", []))

        payload = {
            "question": composed_question or user_question,
            "schema": schema,
            "tableMetadata": table_metadata,
            "tableMappingFilter": table_filter,
            "maxSteps": _da_coerce_int(params.get("max_steps", 8), 8, 1, 50),
            "knowledge_mode": str(params.get("knowledge_mode", _KNOWLEDGE_MODE_CONTEXT_ONCE)),
            "use_knowledge_base": _da_coerce_bool(params.get("use_knowledge_base", True), True),
            "use_knowledge_agent": _da_coerce_bool(params.get("use_knowledge_agent", False), False),
            "control_session_id": session_id,
        }

        try:
            with app.app_context():
                with app.test_request_context("/api/agent", method="POST", json=payload):
                    resp = agent_analysis()
            status_code, result = _da_extract_response_payload(resp)
        except Exception as run_exc:
            status_code = 500
            result = {"error": str(run_exc)}

        with _data_analyst_sessions_lock:
            session = _data_analyst_sessions.get(session_id)
            if not session or int(session.get("run_seq", 0)) != int(run_seq):
                return

            assistant_payload = {
                "content": "",
                "analyst_result": None,
                "error": "",
            }

            if status_code >= 400 or result.get("error"):
                err_txt = str(result.get("error") or f"Agent returned HTTP {status_code}")
                session["status"] = "error"
                session["running"] = False
                assistant_payload["content"] = f"Agent error: {err_txt}"
                assistant_payload["error"] = err_txt
                _da_log_event(session, f"Run failed: {err_txt}", "error", "run")
            else:
                interrupted = bool(result.get("interrupted", False))
                interrupt_reason = str(result.get("interrupt_reason", "")).strip().lower()
                analyst_result = {
                    "final_answer": result.get("final_answer", ""),
                    "steps": result.get("steps", []),
                    "total_steps": result.get("total_steps", len(result.get("steps", []))),
                    "technical_retries_used": result.get("technical_retries_used", 0),
                    "query_attempts_used": result.get("query_attempts_used", 0),
                    "max_total_query_attempts": result.get("max_total_query_attempts", 0),
                    "initial_plan": result.get("initial_plan", []),
                    "current_plan": result.get("current_plan", []),
                    "midcourse_review": result.get("midcourse_review", {}),
                    "no_prompt_context_injection": bool(result.get("no_prompt_context_injection", False)),
                    "knowledge_mode": str(result.get("knowledge_mode", params.get("knowledge_mode", _KNOWLEDGE_MODE_CONTEXT_ONCE))),
                    "interrupted": interrupted,
                    "interrupt_reason": interrupt_reason,
                }
                session["last_result"] = analyst_result
                assistant_payload["analyst_result"] = analyst_result
                assistant_payload["content"] = (
                    str(result.get("final_answer", "")).strip()
                    or "Analysis complete."
                )
                if interrupted and interrupt_reason == "paused":
                    session["status"] = "paused"
                    _da_log_event(session, "Run paused after interruption request.", "warn", "control")
                elif interrupted and interrupt_reason == "stopped":
                    session["status"] = "stopped"
                    _da_log_event(session, "Run stopped after interruption request.", "warn", "control")
                else:
                    session["status"] = "completed"
                    _da_log_event(
                        session,
                        (
                            "Run completed: "
                            f"{analyst_result['total_steps']} step(s), "
                            f"{analyst_result['technical_retries_used']} free retry(ies), "
                            f"{analyst_result['query_attempts_used']} SQL attempt(s)."
                        ),
                        "info",
                        "run",
                    )

            if assistant_payload["content"]:
                session.setdefault("conversation", []).append({
                    "role": "assistant",
                    "content": assistant_payload["content"],
                    "ts": _da_now_iso(),
                })
            session["latest_assistant"] = assistant_payload
            session["response_seq"] = int(session.get("response_seq", 0)) + 1
            _da_refresh_memory_summary(session)

            if session.get("stop_requested"):
                session["running"] = False
                session["status"] = "stopped"
                session["updated_at"] = _da_now_iso()
                return
            if session.get("pause_requested"):
                session["running"] = False
                session["status"] = "paused"
                session["updated_at"] = _da_now_iso()
                return

            should_continue = bool(
                session.get("params", {}).get("auto_run", True)
                and session.get("pending_user_inputs")
            )
            if should_continue:
                session["running"] = True
                session["status"] = "running"
                session["updated_at"] = _da_now_iso()
                _da_log_event(session, "Continuing with queued follow-up question.", "info", "queue")
                continue

            session["running"] = False
            if session.get("status") not in {"paused", "stopped", "error"}:
                session["status"] = "idle"
            session["updated_at"] = _da_now_iso()
            return


def _run_ai_data_analyst_session_agent():
    data = request.get_json(silent=True) or {}
    params_raw = data.get("params", {}) or {}
    control = str(data.get("control", "")).strip().lower()
    session_id = str(data.get("session_id", "")).strip()
    messages = data.get("messages", []) or []

    last_user_text = ""
    if isinstance(messages, list):
        for m in reversed(messages):
            if str(m.get("role", "")) == "user":
                last_user_text = str(m.get("content", "")).strip()
                if last_user_text:
                    break

    with _data_analyst_sessions_lock:
        if session_id and session_id in _data_analyst_sessions:
            session = _data_analyst_sessions[session_id]
        else:
            session_id = str(uuid.uuid4())
            session = {
                "id": session_id,
                "created_at": _da_now_iso(),
                "updated_at": _da_now_iso(),
                "status": "idle",
                "running": False,
                "pause_requested": False,
                "stop_requested": False,
                "params": _da_normalize_params(params_raw),
                "conversation": [],
                "pending_user_inputs": [],
                "paused_notes": [],
                "memory_summary": "",
                "event_log": [],
                "event_seq": 0,
                "response_seq": 0,
                "latest_assistant": None,
                "last_result": None,
                "cached_schema": None,
                "cached_schema_ts": 0,
                "run_seq": 0,
                "worker": None,
            }
            _data_analyst_sessions[session_id] = session
            _da_log_event(session, "New AI data analyst session created.", "info", "session")

        # Merge latest params at each call (keeps runtime tunable in the UI)
        merged_params = dict(session.get("params", {}))
        merged_params.update(_da_normalize_params(params_raw))
        session["params"] = merged_params
        session["updated_at"] = _da_now_iso()

        # Control actions
        if control == "status":
            payload = _da_session_payload(session)
            payload["content"] = "Runtime status fetched."
            return jsonify(payload)

        if control == "pause":
            session["pause_requested"] = True
            if session.get("running"):
                session["status"] = "pausing"
                _da_log_event(session, "Pause requested (will apply between steps).", "warn", "control")
                content = "Pause requested. The agent will pause at the next safe checkpoint."
            else:
                session["status"] = "paused"
                _da_log_event(session, "Session paused.", "warn", "control")
                content = "Session paused. You can add more context, then click Resume."
            payload = _da_session_payload(session)
            payload["content"] = content
            return jsonify(payload)

        if control == "stop":
            session["stop_requested"] = True
            if session.get("running"):
                session["status"] = "stopping"
                _da_log_event(session, "Stop requested (will apply between steps).", "warn", "control")
                content = "Stop requested. Current run will stop at the next safe checkpoint."
            else:
                session["running"] = False
                session["status"] = "stopped"
                _da_log_event(session, "Session stopped.", "warn", "control")
                content = "Session stopped."
            payload = _da_session_payload(session)
            payload["content"] = content
            return jsonify(payload)

        if control == "resume":
            session["pause_requested"] = False
            session["stop_requested"] = False
            if last_user_text:
                session["conversation"].append({
                    "role": "user",
                    "content": last_user_text,
                    "ts": _da_now_iso(),
                })
                session["pending_user_inputs"].append(last_user_text)
                _da_log_event(session, "User added a new message while resuming.", "info", "input")
            started = _da_start_worker_if_needed(session)
            if started:
                content = "Session resumed. Agent run started."
                _da_log_event(session, "Session resumed and run started.", "info", "control")
            else:
                session["status"] = "idle" if not session.get("running") else session.get("status")
                content = "Session resumed. No pending message to process."
                _da_log_event(session, "Session resumed without pending run.", "info", "control")
            payload = _da_session_payload(session)
            payload["content"] = content
            return jsonify(payload)

        if control == "run":
            session["pause_requested"] = False
            session["stop_requested"] = False
            if last_user_text:
                session["conversation"].append({
                    "role": "user",
                    "content": last_user_text,
                    "ts": _da_now_iso(),
                })
                session["pending_user_inputs"].append(last_user_text)
                _da_log_event(session, "Queued new user message for manual run.", "info", "input")
            started = _da_start_worker_if_needed(session)
            payload = _da_session_payload(session)
            payload["content"] = "Run started." if started else "No pending message to run."
            return jsonify(payload)

        if control == "note":
            if last_user_text:
                session["conversation"].append({
                    "role": "user",
                    "content": last_user_text,
                    "ts": _da_now_iso(),
                })
                session.setdefault("paused_notes", []).append(last_user_text)
                session["paused_notes"] = session["paused_notes"][-20:]
                _da_refresh_memory_summary(session)
                _da_log_event(session, "User note captured while paused.", "info", "input")
            payload = _da_session_payload(session)
            payload["content"] = "Context note saved. Resume the agent when ready."
            return jsonify(payload)

        # Default message flow
        if not last_user_text:
            payload = _da_session_payload(session)
            payload["content"] = "No user message provided."
            return jsonify(payload), 400

        session["conversation"].append({
            "role": "user",
            "content": last_user_text,
            "ts": _da_now_iso(),
        })
        if session.get("status") == "paused" or session.get("pause_requested"):
            session.setdefault("paused_notes", []).append(last_user_text)
            session["paused_notes"] = session["paused_notes"][-20:]
            _da_refresh_memory_summary(session)
            _da_log_event(session, "User context stored while session is paused.", "info", "input")
            payload = _da_session_payload(session)
            payload["content"] = "Context saved (paused). Click Resume to continue."
            return jsonify(payload)

        session["pending_user_inputs"].append(last_user_text)
        _da_refresh_memory_summary(session)
        _da_log_event(session, f"Queued user request: '{last_user_text[:120]}'", "info", "input")
        started = False
        if session.get("params", {}).get("auto_run", True):
            session["pause_requested"] = False
            session["stop_requested"] = False
            started = _da_start_worker_if_needed(session)

        payload = _da_session_payload(session)
        if started:
            payload["content"] = "Agent started. Follow live logs while it is running."
        else:
            payload["content"] = (
                "Message queued. Click Run to start."
                if not session.get("running")
                else "Message queued while a run is already in progress."
            )
        return jsonify(payload)


# In-memory session store for the Data Wrangling session agent
_data_wrangling_sessions: dict = {}
_data_wrangling_sessions_lock = _threading.Lock()
_DATA_WRANGLING_MAX_EVENTS = 800
_DATA_WRANGLING_SCHEMA_CACHE_TTL_SEC = 180
_DW_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
_DW_UUID_RE = re.compile(
    r"^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[1-5][0-9a-fA-F]{3}-[89abAB][0-9a-fA-F]{3}-[0-9a-fA-F]{12}$"
)
_DW_DATE_LITERAL_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def _dw_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _dw_coerce_bool(value, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "on", "oui"}
    return bool(value)


def _dw_coerce_int(value, default: int, min_value: int, max_value: int) -> int:
    parsed = _parse_int_like(value, default)
    return max(min_value, min(max_value, parsed))


def _dw_safe_identifier(name: str) -> str:
    return f"`{str(name or '').replace('`', '``')}`"


def _dw_parse_date_literal(value) -> str:
    txt = str(value or "").strip()
    if not txt:
        return ""
    if not _DW_DATE_LITERAL_RE.match(txt):
        return ""
    return txt


def _dw_normalize_params(raw_params: dict | None) -> dict:
    params = raw_params or {}
    return {
        "table": str(params.get("table", "") or "").strip(),
        "date_column": str(params.get("date_column", "") or "").strip(),
        "date_start": _dw_parse_date_literal(params.get("date_start", "")),
        "date_end": _dw_parse_date_literal(params.get("date_end", "")),
        "max_rows": _dw_coerce_int(params.get("max_rows", 5000), 5000, 50, 500000),
        "batch_size": _dw_coerce_int(params.get("batch_size", 400), 400, 50, 5000),
        "max_steps": _dw_coerce_int(params.get("max_steps", 24), 24, 2, 200),
        "row_start": _dw_coerce_int(params.get("row_start", 0), 0, 0, 100000000),
        "export_excel": _dw_coerce_bool(params.get("export_excel", "no"), False),
        "export_path": str(params.get("export_path", "") or "").strip(),
        "reset_watermark": _dw_coerce_bool(params.get("reset_watermark", "no"), False),
        "use_knowledge_base": _dw_coerce_bool(params.get("use_knowledge_base", "yes"), True),
        "auto_run": _dw_coerce_bool(params.get("auto_run", "yes"), True),
    }


def _dw_log_event(session: dict, message: str, level: str = "info", kind: str = "runtime") -> None:
    level_norm = str(level or "info").lower()
    if level_norm not in {"info", "warn", "error"}:
        level_norm = "info"
    entry = {
        "seq": int(session.get("event_seq", 0)) + 1,
        "ts": _dw_now_iso(),
        "level": level_norm,
        "kind": str(kind or "runtime"),
        "message": str(message or "").strip(),
    }
    session["event_seq"] = entry["seq"]
    session.setdefault("event_log", []).append(entry)
    if len(session["event_log"]) > _DATA_WRANGLING_MAX_EVENTS:
        session["event_log"] = session["event_log"][-_DATA_WRANGLING_MAX_EVENTS:]
    session["updated_at"] = entry["ts"]
    try:
        _log(
            f"[{session.get('id', '?')[:8]}] {entry['message']}",
            level=level_norm,
            source="data-wrangling",
        )
    except Exception:
        pass


def _dw_session_payload(session: dict) -> dict:
    return {
        "session_id": session.get("id"),
        "status": session.get("status", "idle"),
        "running": bool(session.get("running", False)),
        "pause_requested": bool(session.get("pause_requested", False)),
        "stop_requested": bool(session.get("stop_requested", False)),
        "pending_user_inputs": len(session.get("pending_user_inputs", [])),
        "memory_summary": session.get("memory_summary", ""),
        "event_log": session.get("event_log", [])[-240:],
        "response_seq": int(session.get("response_seq", 0)),
        "latest_assistant": session.get("latest_assistant"),
        "last_result": session.get("last_result"),
        "params": session.get("params", {}),
        "created_at": session.get("created_at"),
        "updated_at": session.get("updated_at"),
    }


def _dw_publish_message(session: dict, content: str, wrangling_result: dict | None = None, error: str = "") -> None:
    txt = str(content or "").strip()
    if not txt:
        return
    payload = {
        "content": txt,
        "wrangling_result": wrangling_result or None,
        "error": str(error or "").strip(),
    }
    session["latest_assistant"] = payload
    session["response_seq"] = int(session.get("response_seq", 0)) + 1
    session.setdefault("conversation", []).append({
        "role": "assistant",
        "content": txt,
        "ts": _dw_now_iso(),
    })
    session["conversation"] = session["conversation"][-100:]
    session["updated_at"] = _dw_now_iso()


def _dw_extract_response_payload(resp_obj):
    status_code = 200
    response = resp_obj
    if isinstance(resp_obj, tuple):
        response = resp_obj[0]
        if len(resp_obj) > 1 and isinstance(resp_obj[1], int):
            status_code = resp_obj[1]
    if hasattr(response, "status_code"):
        status_code = int(getattr(response, "status_code", status_code) or status_code)
    if hasattr(response, "get_json"):
        data = response.get_json(silent=True) or {}
    else:
        data = {}
    return status_code, data


def _dw_get_schema_with_cache(session: dict) -> dict:
    now_ts = datetime.now(timezone.utc).timestamp()
    cached_schema = session.get("cached_schema")
    cached_ts = float(session.get("cached_schema_ts", 0) or 0)
    if cached_schema and (now_ts - cached_ts) <= _DATA_WRANGLING_SCHEMA_CACHE_TTL_SEC:
        return cached_schema
    schema = _da_fetch_schema_for_agent()
    session["cached_schema"] = schema
    session["cached_schema_ts"] = now_ts
    return schema


def _dw_match_table(schema_map: dict, requested_table: str) -> str:
    req = str(requested_table or "").strip()
    if not req:
        return ""
    if req in schema_map:
        return req
    lower_map = {str(k).lower(): k for k in schema_map.keys()}
    if req.lower() in lower_map:
        return str(lower_map[req.lower()])
    short_map = {}
    for key in schema_map.keys():
        short = str(key).split(".")[-1].lower()
        short_map.setdefault(short, key)
    return str(short_map.get(req.lower(), ""))


def _dw_pick_order_column(columns: list[dict], preferred_date_col: str = "") -> str:
    col_names = [str(c.get("name", "")).strip() for c in columns if str(c.get("name", "")).strip()]
    if not col_names:
        return ""
    if preferred_date_col and preferred_date_col in col_names:
        return preferred_date_col
    for candidate in col_names:
        lower = candidate.lower()
        if lower == "id":
            return candidate
    for candidate in col_names:
        lower = candidate.lower()
        if lower.endswith("_id"):
            return candidate
    return col_names[0]


def _dw_build_scope_where(params: dict, available_cols: set[str]) -> str:
    date_col = str(params.get("date_column", "")).strip()
    start = str(params.get("date_start", "")).strip()
    end = str(params.get("date_end", "")).strip()
    if not date_col or date_col not in available_cols:
        return ""
    if start and end:
        return f"WHERE {_dw_safe_identifier(date_col)} BETWEEN '{start}' AND '{end}'"
    if start:
        return f"WHERE {_dw_safe_identifier(date_col)} = '{start}'"
    if end:
        return f"WHERE {_dw_safe_identifier(date_col)} = '{end}'"
    return ""


def _dw_scope_key(table: str, params: dict) -> str:
    date_col = str(params.get("date_column", "")).strip() or "-"
    date_start = str(params.get("date_start", "")).strip() or "-"
    date_end = str(params.get("date_end", "")).strip() or "-"
    return f"{table}|{date_col}|{date_start}|{date_end}"


def _dw_read_watermarks() -> list:
    db = read_db()
    raw = db.get("wrangling_watermarks", [])
    return raw if isinstance(raw, list) else []


def _dw_get_watermark(scope_key: str) -> dict:
    for row in _dw_read_watermarks():
        if str(row.get("scope_key", "")) == scope_key:
            return row
    return {}


def _dw_upsert_watermark(scope_key: str, payload: dict) -> None:
    db = read_db()
    items = db.get("wrangling_watermarks", [])
    if not isinstance(items, list):
        items = []
    found = False
    for idx, row in enumerate(items):
        if str(row.get("scope_key", "")) == scope_key:
            merged = dict(row)
            merged.update(payload or {})
            merged["scope_key"] = scope_key
            items[idx] = merged
            found = True
            break
    if not found:
        row = {"scope_key": scope_key}
        row.update(payload or {})
        items.append(row)
    db["wrangling_watermarks"] = items[-4000:]
    write_db(db)


def _dw_get_table_metadata(table_name: str) -> dict:
    db = read_db()
    for row in db.get("table_metadata", []):
        if str(row.get("table_name", "")).strip() == table_name:
            return row
    return {}


def _dw_get_knowledge_hint(
    question: str,
    table_name: str,
    table_description: str,
    use_knowledge_base: bool,
) -> str:
    if not use_knowledge_base:
        return ""
    seed = " ".join(
        part for part in [
            question or "",
            table_name or "",
            table_description or "",
            "anomalie format valeur métier qualité",
        ] if part
    )
    try:
        hit = _get_knowledge_context_by_similarity(seed, top_k=int(rag_config.get("topK", 5)))
        if hit:
            return _truncate_text_to_budget(hit, 1300)
    except Exception as exc:
        print(f"[Wrangling] KB retrieval warning: {exc}")
    return ""


def _dw_compact_sql_check_summary(name: str, sql: str, result: dict) -> str:
    status = "OK" if result.get("ok") else "FAIL"
    return (
        f"[{status}] {name}: sql={str(sql)[:180]} "
        f"→ {str(result.get('summary', ''))[:280]}"
    )


def _dw_plan_with_llm(
    question: str,
    table_name: str,
    columns: list[dict],
    where_clause: str,
    knowledge_hint: str,
    max_steps: int,
) -> dict:
    column_hints = []
    for col in columns[:80]:
        column_hints.append({
            "name": str(col.get("name", "")),
            "type": str(col.get("type", "")),
        })
    prompt = f"""You are designing a ClickHouse data-wrangling inspection plan.
Minimize token usage and maximize anomaly detection quality.

TASK:
- Build a concise action plan for anomaly detection on a single table.
- Propose only simple read-only SQL checks.

SCOPE:
- table: {table_name}
- optional where clause: {where_clause or "(none)"}
- max actionable steps: {max_steps}

USER REQUEST:
{question}

TABLE COLUMNS:
{json.dumps(column_hints, ensure_ascii=False)}

KNOWLEDGE HINTS:
{knowledge_hint or "(none)"}

RULES:
- Keep SQL simple and compatible.
- Read-only SQL only.
- Add LIMIT <= 5000 where relevant.
- Date predicates must use BETWEEN 'YYYY-MM-DD' AND 'YYYY-MM-DD' or = 'YYYY-MM-DD'.
- No advanced ClickHouse functions.

Return ONLY JSON:
{{
  "plan_steps": ["step 1", "step 2", "step 3"],
  "focus_columns": ["col_a", "col_b"],
  "sql_checks": [
    {{"name": "check name", "sql": "SELECT ... FROM {table_name} ... LIMIT 200"}}
  ],
  "reasoning": "short rationale"
}}"""
    safe_prompt = _truncate_text_to_budget(prompt, int(_get_effective_context_limit() * 0.58))
    content = _call_llm(
        safe_prompt,
        [{"role": "user", "content": "Build wrangling plan."}],
        temperature=0.1,
    )
    parsed = _parse_llm_json(content)
    return parsed if isinstance(parsed, dict) else {}


def _dw_fallback_plan(columns: list[dict], max_steps: int) -> dict:
    top_cols = [str(c.get("name", "")) for c in columns[:8] if str(c.get("name", ""))]
    return {
        "plan_steps": [
            "Calculer la distribution de null/vides et cardinalité.",
            "Scanner les lignes en batch pour outliers de format/taille/valeur.",
            "Vérifier cohérences croisées (dates, identifiants) et conclure.",
        ][:max(1, min(3, max_steps))],
        "focus_columns": top_cols[:10],
        "sql_checks": [],
        "reasoning": "Fallback deterministic wrangling plan",
    }


def _dw_finalize_plan(raw_plan: dict, columns: list[dict], max_steps: int) -> dict:
    col_set = {str(c.get("name", "")) for c in columns}
    plan_steps = []
    for item in (raw_plan.get("plan_steps") if isinstance(raw_plan, dict) else []) or []:
        text = " ".join(str(item or "").split())
        if not text:
            continue
        if len(text) > 180:
            text = text[:180].rstrip() + "..."
        if text.lower() in {p.lower() for p in plan_steps}:
            continue
        plan_steps.append(text)
        if len(plan_steps) >= max(1, min(6, max_steps)):
            break
    focus_columns = []
    for col in (raw_plan.get("focus_columns") if isinstance(raw_plan, dict) else []) or []:
        c = str(col or "").strip()
        if c in col_set and c not in focus_columns:
            focus_columns.append(c)
        if len(focus_columns) >= 24:
            break
    sql_checks = []
    for row in (raw_plan.get("sql_checks") if isinstance(raw_plan, dict) else []) or []:
        if not isinstance(row, dict):
            continue
        name = " ".join(str(row.get("name", "")).split())[:120]
        sql = str(row.get("sql", "")).strip()
        if not sql:
            continue
        sql_checks.append({"name": name or "SQL check", "sql": sql})
        if len(sql_checks) >= 6:
            break
    return {
        "plan_steps": plan_steps,
        "focus_columns": focus_columns,
        "sql_checks": sql_checks,
        "reasoning": str((raw_plan or {}).get("reasoning", ""))[:220],
    }


def _dw_parse_any_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    try:
        import datetime as _dt
        if isinstance(value, _dt.date):
            return value
    except Exception:
        pass
    txt = str(value).strip()
    if not txt:
        return None
    txt = txt.replace("T", " ")
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d"):
        try:
            return datetime.strptime(txt[:19], fmt).date()
        except Exception:
            continue
    return None


def _dw_export_anomalies_excel(anomalies: list[dict], output_path: str) -> str:
    if not anomalies:
        return ""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    target_path = output_path.strip() or os.path.join(
        BASE_DIR, ".data", f"wrangling_anomalies_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    os.makedirs(os.path.dirname(target_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Anomalies"
    headers = [
        "table",
        "line_number",
        "column",
        "issue_type",
        "severity",
        "value_preview",
        "reference",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="0B4F6C")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in anomalies:
        ws.append([
            row.get("table", ""),
            row.get("line_number", ""),
            row.get("column", ""),
            row.get("issue_type", ""),
            row.get("severity", ""),
            row.get("value_preview", ""),
            row.get("reference", ""),
        ])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{ws.max_row}"
    widths = {"A": 30, "B": 13, "C": 24, "D": 20, "E": 12, "F": 34, "G": 70}
    for key, width in widths.items():
        ws.column_dimensions[key].width = width
    wb.save(target_path)
    return target_path


def _dw_detect_batch_anomalies(
    rows: list[dict],
    *,
    table_name: str,
    line_offset: int,
    column_types: dict,
    focus_columns: set[str],
    column_state: dict,
    primary_id_column: str,
    primary_id_seen: set,
    date_pairs: list[tuple[str, str]],
) -> list[dict]:
    anomalies = []

    def _severity_from_score(score: float) -> str:
        if score >= 0.9:
            return "high"
        if score >= 0.6:
            return "medium"
        return "low"

    for idx, row in enumerate(rows):
        line_number = line_offset + idx + 1
        if not isinstance(row, dict):
            continue

        if primary_id_column:
            pid = row.get(primary_id_column)
            if pid is not None and str(pid).strip() != "":
                pid_fp = str(pid).strip()
                if pid_fp in primary_id_seen:
                    anomalies.append({
                        "table": table_name,
                        "line_number": line_number,
                        "column": primary_id_column,
                        "issue_type": "duplicate_primary_id",
                        "severity": "high",
                        "value_preview": str(pid)[:120],
                        "reference": "Duplicate primary identifier detected in scan scope.",
                    })
                else:
                    primary_id_seen.add(pid_fp)

        for col_name, value in row.items():
            cname = str(col_name or "")
            if not cname:
                continue
            cstate = column_state.setdefault(cname, {
                "seen": 0,
                "nulls": 0,
                "empty": 0,
                "len_n": 0,
                "len_mean": 0.0,
                "len_m2": 0.0,
                "num_n": 0,
                "num_mean": 0.0,
                "num_m2": 0.0,
                "freq": Counter(),
            })
            prev_seen = int(cstate.get("seen", 0))
            prev_nulls = int(cstate.get("nulls", 0))
            prev_empty = int(cstate.get("empty", 0))
            lower = cname.lower()
            in_focus = cname in focus_columns or lower in {c.lower() for c in focus_columns}
            ctype = str(column_types.get(cname, "")).lower()

            if value is None:
                if prev_seen >= 60:
                    null_rate = prev_nulls / max(1, prev_seen)
                    if null_rate < 0.02:
                        anomalies.append({
                            "table": table_name,
                            "line_number": line_number,
                            "column": cname,
                            "issue_type": "unexpected_null",
                            "severity": "medium",
                            "value_preview": "NULL",
                            "reference": f"Observed NULL whereas historical null-rate was {null_rate:.2%}.",
                        })
                cstate["seen"] = prev_seen + 1
                cstate["nulls"] = prev_nulls + 1
                continue

            txt = str(value).strip()
            if txt == "":
                if prev_seen >= 60:
                    empty_rate = prev_empty / max(1, prev_seen)
                    if empty_rate < 0.02:
                        anomalies.append({
                            "table": table_name,
                            "line_number": line_number,
                            "column": cname,
                            "issue_type": "unexpected_empty",
                            "severity": "medium",
                            "value_preview": "(empty)",
                            "reference": f"Observed empty value whereas historical empty-rate was {empty_rate:.2%}.",
                        })
                cstate["seen"] = prev_seen + 1
                cstate["empty"] = prev_empty + 1
                continue

            # Format anomalies for known business-like fields
            if "email" in lower and txt and not _DW_EMAIL_RE.match(txt):
                anomalies.append({
                    "table": table_name,
                    "line_number": line_number,
                    "column": cname,
                    "issue_type": "invalid_email_format",
                    "severity": "high",
                    "value_preview": txt[:120],
                    "reference": "Value does not match expected email format.",
                })
            if ("phone" in lower or "tel" in lower) and txt:
                digits = re.sub(r"\D+", "", txt)
                if len(digits) < 8 or len(digits) > 15:
                    anomalies.append({
                        "table": table_name,
                        "line_number": line_number,
                        "column": cname,
                        "issue_type": "invalid_phone_length",
                        "severity": "medium",
                        "value_preview": txt[:120],
                        "reference": "Phone-like field has an unusual number of digits.",
                    })
            if "uuid" in lower and txt and not _DW_UUID_RE.match(txt):
                anomalies.append({
                    "table": table_name,
                    "line_number": line_number,
                    "column": cname,
                    "issue_type": "invalid_uuid_format",
                    "severity": "high",
                    "value_preview": txt[:120],
                    "reference": "UUID-like field does not match canonical UUID format.",
                })

            # Length outlier on string-like columns.
            if not any(k in ctype for k in ("int", "float", "decimal")):
                length_val = len(txt)
                len_n = int(cstate.get("len_n", 0))
                len_mean = float(cstate.get("len_mean", 0.0))
                len_m2 = float(cstate.get("len_m2", 0.0))
                if len_n >= 40:
                    len_var = (len_m2 / max(1, len_n - 1)) if len_n > 1 else 0.0
                    len_std = math.sqrt(max(0.0, len_var))
                    if len_std > 0:
                        z = abs((length_val - len_mean) / len_std)
                        if z >= 5.0 and in_focus:
                            anomalies.append({
                                "table": table_name,
                                "line_number": line_number,
                                "column": cname,
                                "issue_type": "length_outlier",
                                "severity": _severity_from_score(min(1.0, z / 8)),
                                "value_preview": txt[:120],
                                "reference": f"Length {length_val} is an outlier (z-score {z:.2f}).",
                            })
                # Update length stats
                len_n_new = len_n + 1
                delta = length_val - len_mean
                len_mean_new = len_mean + delta / len_n_new
                len_m2_new = len_m2 + delta * (length_val - len_mean_new)
                cstate["len_n"] = len_n_new
                cstate["len_mean"] = len_mean_new
                cstate["len_m2"] = len_m2_new

            # Numeric outlier
            num_val = _coerce_float(value)
            if num_val is not None:
                num_n = int(cstate.get("num_n", 0))
                num_mean = float(cstate.get("num_mean", 0.0))
                num_m2 = float(cstate.get("num_m2", 0.0))
                if num_n >= 40:
                    var = (num_m2 / max(1, num_n - 1)) if num_n > 1 else 0.0
                    std = math.sqrt(max(0.0, var))
                    if std > 0:
                        z = abs((num_val - num_mean) / std)
                        if z >= 5.0:
                            anomalies.append({
                                "table": table_name,
                                "line_number": line_number,
                                "column": cname,
                                "issue_type": "numeric_outlier",
                                "severity": _severity_from_score(min(1.0, z / 8)),
                                "value_preview": str(value)[:120],
                                "reference": f"Value {num_val:.6g} is a numeric outlier (z-score {z:.2f}).",
                            })
                num_n_new = num_n + 1
                delta_num = num_val - num_mean
                num_mean_new = num_mean + delta_num / num_n_new
                num_m2_new = num_m2 + delta_num * (num_val - num_mean_new)
                cstate["num_n"] = num_n_new
                cstate["num_mean"] = num_mean_new
                cstate["num_m2"] = num_m2_new

            # Rare category check on focused columns
            freq = cstate.get("freq", Counter())
            if not isinstance(freq, Counter):
                freq = Counter()
            if prev_seen >= 200 and in_focus and len(freq) >= 6 and freq.get(txt, 0) == 0:
                anomalies.append({
                    "table": table_name,
                    "line_number": line_number,
                    "column": cname,
                    "issue_type": "rare_category",
                    "severity": "low",
                    "value_preview": txt[:120],
                    "reference": "Previously unseen category after a large baseline.",
                })
            if len(freq) < 600 or txt in freq:
                freq[txt] += 1
            cstate["freq"] = freq
            cstate["seen"] = prev_seen + 1

        # Cross-field date coherence checks
        for start_col, end_col in date_pairs:
            start_val = row.get(start_col)
            end_val = row.get(end_col)
            d1 = _dw_parse_any_date(start_val)
            d2 = _dw_parse_any_date(end_val)
            if d1 and d2 and d1 > d2:
                anomalies.append({
                    "table": table_name,
                    "line_number": line_number,
                    "column": f"{start_col} -> {end_col}",
                    "issue_type": "date_order_inconsistency",
                    "severity": "high",
                    "value_preview": f"{start_val} > {end_val}",
                    "reference": "Start-like date is greater than end-like date.",
                })

    return anomalies


def _dw_build_date_pairs(column_names: list[str]) -> list[tuple[str, str]]:
    lowers = {c.lower(): c for c in column_names}
    pairs = []
    for col in column_names:
        lower = col.lower()
        candidate = ""
        if "start" in lower:
            candidate = lower.replace("start", "end")
        elif "from" in lower:
            candidate = lower.replace("from", "to")
        elif "debut" in lower:
            candidate = lower.replace("debut", "fin")
        if candidate and candidate in lowers:
            pairs.append((col, lowers[candidate]))
    uniq = []
    seen = set()
    for a, b in pairs:
        key = f"{a}|{b}"
        if key in seen:
            continue
        seen.add(key)
        uniq.append((a, b))
    return uniq[:20]


def _dw_summarize_with_llm(
    question: str,
    table_name: str,
    plan_steps: list[str],
    anomaly_type_counter: Counter,
    anomalies_preview: list[dict],
    knowledge_hint: str,
    scanned_range_start: int,
    scanned_range_end: int,
) -> str:
    anomalies_compact = []
    for row in anomalies_preview[:25]:
        anomalies_compact.append(
            f"L{row.get('line_number')} {row.get('column')}: {row.get('issue_type')} | {row.get('reference')}"
        )
    counter_lines = [f"- {k}: {v}" for k, v in anomaly_type_counter.most_common(12)]
    prompt = f"""You are a senior Data Quality analyst.
Write a concise but high-value French synthesis.

QUESTION:
{question}

TABLE:
{table_name}

SCANNED LINE RANGE:
{scanned_range_start} -> {scanned_range_end}

PLAN STEPS:
{chr(10).join(f"- {s}" for s in plan_steps[:8])}

ANOMALY COUNTS:
{chr(10).join(counter_lines) if counter_lines else "- none"}

REPRESENTATIVE ANOMALIES:
{chr(10).join(anomalies_compact) if anomalies_compact else "(none)"}

KNOWLEDGE HINTS:
{knowledge_hint or "(none)"}

Output requirements:
- 1 short executive paragraph.
- 3 to 6 bullet findings with precise references (line and column).
- explain why some anomalies are subtle or hard to detect manually.
- include a confidence level and next remediation actions.
No markdown title."""
    safe_prompt = _truncate_text_to_budget(prompt, int(_get_effective_context_limit() * 0.56))
    return _call_llm(
        safe_prompt,
        [{"role": "user", "content": "Draft final wrangling synthesis."}],
        temperature=0.2,
        language="fr",
    )


def _dw_should_interrupt(session: dict) -> str:
    if session.get("stop_requested"):
        return "stopped"
    if session.get("pause_requested"):
        return "paused"
    return ""


def _dw_start_worker_if_needed(session: dict) -> bool:
    if session.get("running"):
        return False
    if not session.get("pending_user_inputs"):
        return False
    session["running"] = True
    session["status"] = "running"
    session["run_seq"] = int(session.get("run_seq", 0)) + 1
    run_seq = int(session["run_seq"])
    session["updated_at"] = _dw_now_iso()
    t = _threading.Thread(
        target=_dw_worker_loop,
        args=(session.get("id"), run_seq),
        daemon=True,
    )
    session["worker"] = t
    t.start()
    return True


def _dw_rethink_sql_check(check_name: str, sql: str, error_text: str, scope_where: str) -> str:
    prompt = f"""A planned wrangling SQL check failed.
Fix it with a simpler read-only ClickHouse query.

Check name: {check_name}
Previous SQL:
{sql}
Error:
{error_text}
Scope where clause:
{scope_where or "(none)"}

Rules:
- Keep read-only SQL.
- Keep it simple and compatible.
- LIMIT <= 5000.
- If date filter is needed use BETWEEN or equality date literals.

Return JSON only:
{{
  "sql": "SELECT ...",
  "reasoning": "short fix rationale"
}}"""
    safe_prompt = _truncate_text_to_budget(prompt, int(_get_effective_context_limit() * 0.45))
    content = _call_llm(
        safe_prompt,
        [{"role": "user", "content": "Fix failed check SQL."}],
        temperature=0.1,
    )
    parsed = _parse_llm_json(content)
    return str((parsed or {}).get("sql", "")).strip()


def _dw_worker_loop(session_id: str, run_seq: int) -> None:
    while True:
        with _data_wrangling_sessions_lock:
            session = _data_wrangling_sessions.get(session_id)
            if not session:
                return
            if int(session.get("run_seq", 0)) != int(run_seq):
                return
            if session.get("stop_requested"):
                session["running"] = False
                session["status"] = "stopped"
                _dw_log_event(session, "Run cancelled before start (stop requested).", "warn", "control")
                return
            if session.get("pause_requested"):
                session["running"] = False
                session["status"] = "paused"
                _dw_log_event(session, "Run paused before start.", "warn", "control")
                return
            if not session.get("pending_user_inputs"):
                session["running"] = False
                if session.get("status") not in {"paused", "stopped", "error"}:
                    session["status"] = "idle"
                session["updated_at"] = _dw_now_iso()
                return

            question = str(session["pending_user_inputs"].pop(0)).strip()
            params = dict(session.get("params", {}))
            session.setdefault("conversation", []).append({
                "role": "user",
                "content": question,
                "ts": _dw_now_iso(),
            })
            _dw_log_event(
                session,
                f"Run #{run_seq}: wrangling started for '{question[:120]}'",
                "info",
                "run",
            )
            _dw_publish_message(
                session,
                "Run lancé: préparation du plan de nettoyage et définition du scope d'analyse.",
                None,
                "",
            )

        try:
            with _data_wrangling_sessions_lock:
                live = _data_wrangling_sessions.get(session_id)
                if not live or int(live.get("run_seq", 0)) != int(run_seq):
                    return
                schema_map = _dw_get_schema_with_cache(live)

            requested_table = str(params.get("table", "")).strip()
            matched_table = _dw_match_table(schema_map, requested_table)
            if not matched_table:
                raise ValueError(
                    f"Table introuvable: '{requested_table}'. Renseignez le paramètre table exactement."
                )

            columns = schema_map.get(matched_table) or []
            if not columns:
                raise ValueError(f"Aucune colonne détectée pour la table '{matched_table}'.")

            col_names = [str(c.get("name", "")).strip() for c in columns if str(c.get("name", "")).strip()]
            col_set = set(col_names)
            col_types = {str(c.get("name", "")).strip(): str(c.get("type", "")) for c in columns}
            order_col = _dw_pick_order_column(columns, str(params.get("date_column", "")).strip())
            if not order_col:
                raise ValueError("Impossible de déterminer une colonne d'ordre pour le scan.")
            where_clause = _dw_build_scope_where(params, col_set)

            table_meta = _dw_get_table_metadata(matched_table)
            table_desc = str(table_meta.get("description", "") or "")
            knowledge_hint = _dw_get_knowledge_hint(
                question,
                matched_table,
                table_desc,
                bool(params.get("use_knowledge_base", True)),
            )

            try:
                raw_plan = _dw_plan_with_llm(
                    question,
                    matched_table,
                    columns,
                    where_clause,
                    knowledge_hint,
                    int(params.get("max_steps", 24)),
                )
                plan = _dw_finalize_plan(raw_plan, columns, int(params.get("max_steps", 24)))
            except Exception as plan_exc:
                plan = _dw_fallback_plan(columns, int(params.get("max_steps", 24)))
                with _data_wrangling_sessions_lock:
                    live = _data_wrangling_sessions.get(session_id)
                    if live and int(live.get("run_seq", 0)) == int(run_seq):
                        _dw_log_event(
                            live,
                            f"LLM planning fallback activated ({plan_exc}).",
                            "warn",
                            "planning",
                        )

            if not plan.get("plan_steps"):
                plan = _dw_fallback_plan(columns, int(params.get("max_steps", 24)))
            focus_columns = set(plan.get("focus_columns") or [])
            if not focus_columns:
                focus_columns = set(col_names[: min(20, len(col_names))])

            with _data_wrangling_sessions_lock:
                live = _data_wrangling_sessions.get(session_id)
                if not live or int(live.get("run_seq", 0)) != int(run_seq):
                    return
                plan_text = "\n".join(f"{idx+1}. {step}" for idx, step in enumerate(plan.get("plan_steps", [])))
                _dw_publish_message(
                    live,
                    (
                        f"Plan initial prêt pour `{matched_table}`.\n"
                        f"{plan_text}\n\n"
                        f"Colonnes focus: {', '.join(sorted(list(focus_columns))[:16])}"
                    ),
                    None,
                    "",
                )
                _dw_log_event(live, f"Plan initialized with {len(plan.get('plan_steps', []))} steps.", "info", "planning")

            # Execute planned SQL checks first (LLM call only at start, then ClickHouse-first flow)
            check_outputs = []
            client = get_clickhouse_client()
            safe_table = _dw_safe_identifier(matched_table)
            planned_checks = plan.get("sql_checks") or []
            for check_idx, check in enumerate(planned_checks[:6]):
                check_name = str(check.get("name", f"check_{check_idx+1}"))
                check_sql = str(check.get("sql", "")).strip()
                if not check_sql:
                    continue
                interrupt_reason = ""
                with _data_wrangling_sessions_lock:
                    live = _data_wrangling_sessions.get(session_id)
                    if not live or int(live.get("run_seq", 0)) != int(run_seq):
                        return
                    interrupt_reason = _dw_should_interrupt(live)
                if interrupt_reason:
                    raise RuntimeError(f"Run interrupted: {interrupt_reason}")

                rendered_sql = (
                    check_sql
                    .replace("{table}", safe_table)
                    .replace("{table_name}", safe_table)
                    .replace("{where}", where_clause)
                    .replace("{where_clause}", where_clause)
                )
                res = _execute_sql_guarded(
                    rendered_sql,
                    read_only=True,
                    enforce_simple_compat=True,
                    max_preview_rows=10,
                    max_execution_time=15,
                    default_limit=300,
                    hard_limit_cap=5000,
                    client=client,
                )
                if not res.get("ok"):
                    fixed_sql = ""
                    try:
                        fixed_sql = _dw_rethink_sql_check(check_name, rendered_sql, res.get("summary", ""), where_clause)
                    except Exception:
                        fixed_sql = ""
                    if fixed_sql:
                        res = _execute_sql_guarded(
                            fixed_sql,
                            read_only=True,
                            enforce_simple_compat=True,
                            max_preview_rows=10,
                            max_execution_time=15,
                            default_limit=300,
                            hard_limit_cap=5000,
                            client=client,
                        )
                        rendered_sql = fixed_sql
                check_outputs.append({
                    "name": check_name,
                    "sql": rendered_sql,
                    "ok": bool(res.get("ok")),
                    "summary": str(res.get("summary", ""))[:700],
                })
                with _data_wrangling_sessions_lock:
                    live = _data_wrangling_sessions.get(session_id)
                    if not live or int(live.get("run_seq", 0)) != int(run_seq):
                        return
                    _dw_log_event(
                        live,
                        _dw_compact_sql_check_summary(check_name, rendered_sql, res),
                        "info" if res.get("ok") else "warn",
                        "sql-check",
                    )
                    _dw_publish_message(
                        live,
                        (
                            f"Check SQL `{check_name}` {'OK' if res.get('ok') else 'FAIL'}\n"
                            f"{str(res.get('summary', ''))[:320]}"
                        ),
                        None,
                        "",
                    )

            scope_key = _dw_scope_key(matched_table, params)
            row_start_param = int(params.get("row_start", 0) or 0)
            watermark = _dw_get_watermark(scope_key)
            if _dw_coerce_bool(params.get("reset_watermark", False), False):
                start_offset = row_start_param
            else:
                saved_offset = _parse_int_like(watermark.get("next_offset", 0), 0) if watermark else 0
                start_offset = max(row_start_param, saved_offset)

            max_rows = int(params.get("max_rows", 5000))
            batch_size = int(params.get("batch_size", 400))
            max_steps = int(params.get("max_steps", 24))
            scan_cap = max(1, max_rows)

            selected_cols = col_names[: min(140, len(col_names))]
            select_cols_sql = ", ".join(_dw_safe_identifier(c) for c in selected_cols)
            order_sql = _dw_safe_identifier(order_col)

            column_state = {}
            primary_id_column = ""
            for c in selected_cols:
                if c.lower() == "id":
                    primary_id_column = c
                    break
            date_pairs = _dw_build_date_pairs(selected_cols)
            primary_seen = set()
            anomalies = []
            anomaly_types = Counter()

            current_offset = start_offset
            total_scanned = 0
            step_count = 0
            range_start = start_offset + 1
            range_end = start_offset

            while total_scanned < scan_cap and step_count < max_steps:
                step_count += 1
                interrupt_reason = ""
                with _data_wrangling_sessions_lock:
                    live = _data_wrangling_sessions.get(session_id)
                    if not live or int(live.get("run_seq", 0)) != int(run_seq):
                        return
                    interrupt_reason = _dw_should_interrupt(live)
                if interrupt_reason:
                    break

                remaining = scan_cap - total_scanned
                fetch_n = max(1, min(batch_size, remaining))
                sql = (
                    f"SELECT {select_cols_sql} "
                    f"FROM {safe_table} "
                    f"{where_clause} "
                    f"ORDER BY {order_sql} "
                    f"LIMIT {fetch_n} OFFSET {current_offset}"
                )
                run = _execute_sql_guarded(
                    sql,
                    read_only=True,
                    enforce_simple_compat=True,
                    max_preview_rows=5,
                    max_execution_time=20,
                    default_limit=fetch_n,
                    hard_limit_cap=5000,
                    client=client,
                )
                if not run.get("ok"):
                    with _data_wrangling_sessions_lock:
                        live = _data_wrangling_sessions.get(session_id)
                        if live and int(live.get("run_seq", 0)) == int(run_seq):
                            _dw_log_event(
                                live,
                                f"Batch query failed at offset {current_offset}: {run.get('summary', '')[:220]}",
                                "warn",
                                "scan",
                            )
                            _dw_publish_message(
                                live,
                                (
                                    "Échec technique pendant le scan; tentative de poursuite avec batch suivant plus petit.\n"
                                    + str(run.get("summary", ""))[:260]
                                ),
                                None,
                                "",
                            )
                    batch_size = max(50, int(batch_size * 0.6))
                    if batch_size <= 60:
                        break
                    continue

                rows = run.get("rows") or []
                if not rows:
                    break
                batch_start = current_offset + 1
                batch_end = current_offset + len(rows)
                range_end = batch_end

                batch_anomalies = _dw_detect_batch_anomalies(
                    rows,
                    table_name=matched_table,
                    line_offset=current_offset,
                    column_types=col_types,
                    focus_columns=focus_columns,
                    column_state=column_state,
                    primary_id_column=primary_id_column,
                    primary_id_seen=primary_seen,
                    date_pairs=date_pairs,
                )
                anomalies.extend(batch_anomalies)
                for a in batch_anomalies:
                    anomaly_types[str(a.get("issue_type", "unknown"))] += 1
                anomalies = anomalies[:12000]

                current_offset += len(rows)
                total_scanned += len(rows)

                watermark_payload = {
                    "table": matched_table,
                    "date_column": str(params.get("date_column", "")),
                    "date_start": str(params.get("date_start", "")),
                    "date_end": str(params.get("date_end", "")),
                    "next_offset": int(current_offset),
                    "last_scanned_range_start": int(range_start),
                    "last_scanned_range_end": int(range_end),
                    "updated_at": _dw_now_iso(),
                }
                _dw_upsert_watermark(scope_key, watermark_payload)

                top_findings = batch_anomalies[:2]
                finding_text = "\n".join(
                    f"- L{f.get('line_number')} `{f.get('column')}`: {f.get('issue_type')} ({f.get('severity')})"
                    for f in top_findings
                )
                with _data_wrangling_sessions_lock:
                    live = _data_wrangling_sessions.get(session_id)
                    if not live or int(live.get("run_seq", 0)) != int(run_seq):
                        return
                    live["memory_summary"] = (
                        f"Table {matched_table} | scope {range_start}-{range_end} "
                        f"| scanned={total_scanned}/{scan_cap} rows | anomalies={len(anomalies)}"
                    )
                    _dw_log_event(
                        live,
                        (
                            f"Batch {step_count}: scanned lines {batch_start}-{batch_end} "
                            f"({len(rows)} rows), findings +{len(batch_anomalies)} "
                            f"(total {len(anomalies)})."
                        ),
                        "info",
                        "scan",
                    )
                    _dw_publish_message(
                        live,
                        (
                            f"Batch {step_count} terminé ({batch_start}-{batch_end}). "
                            f"Anomalies détectées: +{len(batch_anomalies)} (total {len(anomalies)})."
                            + (f"\nFindings clés:\n{finding_text}" if finding_text else "")
                        ),
                        None,
                        "",
                    )

                if step_count == max(1, max_steps // 2):
                    with _data_wrangling_sessions_lock:
                        live = _data_wrangling_sessions.get(session_id)
                        if live and int(live.get("run_seq", 0)) == int(run_seq):
                            judgement = (
                                "Trajectoire efficace: anomalies pertinentes détectées."
                                if len(anomalies) > 0
                                else "Peu d'anomalies détectées; poursuite avec focus distributionnel."
                            )
                            _dw_log_event(live, f"Mid-course review: {judgement}", "info", "review")
                            _dw_publish_message(live, f"Réévaluation mi-parcours: {judgement}", None, "")

            final_anomalies = anomalies[:4000]
            export_path = ""
            if _dw_coerce_bool(params.get("export_excel", False), False) and final_anomalies:
                try:
                    export_path = _dw_export_anomalies_excel(final_anomalies, str(params.get("export_path", "")))
                except Exception as exp_exc:
                    with _data_wrangling_sessions_lock:
                        live = _data_wrangling_sessions.get(session_id)
                        if live and int(live.get("run_seq", 0)) == int(run_seq):
                            _dw_log_event(live, f"Excel export failed: {exp_exc}", "warn", "export")

            final_text = ""
            try:
                final_text = _dw_summarize_with_llm(
                    question,
                    matched_table,
                    plan.get("plan_steps", []),
                    anomaly_types,
                    final_anomalies,
                    knowledge_hint,
                    range_start,
                    range_end,
                )
            except Exception:
                top_issue = ", ".join(f"{k}:{v}" for k, v in anomaly_types.most_common(5)) or "none"
                final_text = (
                    f"Scan terminé sur `{matched_table}` (lignes {range_start}-{range_end}). "
                    f"{len(final_anomalies)} anomalies détectées. Types principaux: {top_issue}. "
                    "Consultez les références ligne/colonne et corrigez en priorité les anomalies de sévérité élevée."
                )

            with _data_wrangling_sessions_lock:
                live = _data_wrangling_sessions.get(session_id)
                if not live or int(live.get("run_seq", 0)) != int(run_seq):
                    return
                interrupt_reason = _dw_should_interrupt(live)
                wr_result = {
                    "table": matched_table,
                    "scope": {
                        "date_column": str(params.get("date_column", "")),
                        "date_start": str(params.get("date_start", "")),
                        "date_end": str(params.get("date_end", "")),
                        "line_start": int(range_start),
                        "line_end": int(range_end),
                    },
                    "plan_steps": plan.get("plan_steps", []),
                    "focus_columns": sorted(list(focus_columns))[:30],
                    "sql_checks": check_outputs,
                    "scanned_rows": int(total_scanned),
                    "anomaly_count": len(final_anomalies),
                    "anomaly_type_counts": dict(anomaly_types),
                    "anomalies_preview": final_anomalies[:120],
                    "watermark": {
                        "scope_key": scope_key,
                        "next_offset": int(current_offset),
                        "last_scanned_range_start": int(range_start),
                        "last_scanned_range_end": int(range_end),
                    },
                    "export_excel_path": export_path,
                    "interrupted": bool(interrupt_reason),
                    "interrupt_reason": interrupt_reason or "",
                }
                live["last_result"] = wr_result
                live["memory_summary"] = (
                    f"Table {matched_table} | last range {range_start}-{range_end} | "
                    f"anomalies={len(final_anomalies)} | watermark next_offset={current_offset}"
                )
                _dw_publish_message(live, final_text, wr_result, "")
                if interrupt_reason == "paused":
                    live["status"] = "paused"
                    live["running"] = False
                    _dw_log_event(live, "Run paused after interruption request.", "warn", "control")
                elif interrupt_reason == "stopped":
                    live["status"] = "stopped"
                    live["running"] = False
                    _dw_log_event(live, "Run stopped after interruption request.", "warn", "control")
                else:
                    live["status"] = "completed"
                    live["running"] = False
                    _dw_log_event(
                        live,
                        f"Run completed: scanned_rows={total_scanned}, anomalies={len(final_anomalies)}.",
                        "info",
                        "run",
                    )
        except Exception as run_exc:
            with _data_wrangling_sessions_lock:
                live = _data_wrangling_sessions.get(session_id)
                if not live or int(live.get("run_seq", 0)) != int(run_seq):
                    return
                err_txt = str(run_exc)
                live["running"] = False
                live["status"] = "error"
                _dw_log_event(live, f"Run failed: {err_txt}", "error", "run")
                _dw_publish_message(live, f"Erreur Data Wrangling: {err_txt}", None, err_txt)

        with _data_wrangling_sessions_lock:
            session = _data_wrangling_sessions.get(session_id)
            if not session or int(session.get("run_seq", 0)) != int(run_seq):
                return
            if session.get("stop_requested"):
                session["running"] = False
                session["status"] = "stopped"
                session["updated_at"] = _dw_now_iso()
                return
            if session.get("pause_requested"):
                session["running"] = False
                session["status"] = "paused"
                session["updated_at"] = _dw_now_iso()
                return
            should_continue = bool(
                session.get("params", {}).get("auto_run", True)
                and session.get("pending_user_inputs")
            )
            if should_continue:
                session["running"] = True
                session["status"] = "running"
                session["updated_at"] = _dw_now_iso()
                _dw_log_event(session, "Continuing with queued follow-up question.", "info", "queue")
                continue
            session["running"] = False
            if session.get("status") not in {"paused", "stopped", "error"}:
                session["status"] = "idle"
            session["updated_at"] = _dw_now_iso()
            return


def _run_data_wrangling_agent():
    data = request.get_json(silent=True) or {}
    params_raw = data.get("params", {}) or {}
    control = str(data.get("control", "")).strip().lower()
    session_id = str(data.get("session_id", "")).strip()
    messages = data.get("messages", []) or []

    last_user_text = ""
    if isinstance(messages, list):
        for m in reversed(messages):
            if str(m.get("role", "")) == "user":
                last_user_text = str(m.get("content", "")).strip()
                if last_user_text:
                    break

    with _data_wrangling_sessions_lock:
        if session_id and session_id in _data_wrangling_sessions:
            session = _data_wrangling_sessions[session_id]
        else:
            session_id = str(uuid.uuid4())
            session = {
                "id": session_id,
                "created_at": _dw_now_iso(),
                "updated_at": _dw_now_iso(),
                "status": "idle",
                "running": False,
                "pause_requested": False,
                "stop_requested": False,
                "params": _dw_normalize_params(params_raw),
                "conversation": [],
                "pending_user_inputs": [],
                "memory_summary": "",
                "event_log": [],
                "event_seq": 0,
                "response_seq": 0,
                "latest_assistant": None,
                "last_result": None,
                "cached_schema": None,
                "cached_schema_ts": 0,
                "run_seq": 0,
                "worker": None,
            }
            _data_wrangling_sessions[session_id] = session
            _dw_log_event(session, "New data wrangling session created.", "info", "session")

        merged_params = dict(session.get("params", {}))
        merged_params.update(_dw_normalize_params(params_raw))
        session["params"] = merged_params
        session["updated_at"] = _dw_now_iso()

        if control == "status":
            payload = _dw_session_payload(session)
            payload["content"] = "Runtime status fetched."
            return jsonify(payload)

        if control == "pause":
            session["pause_requested"] = True
            if session.get("running"):
                session["status"] = "pausing"
                _dw_log_event(session, "Pause requested (will apply between scan steps).", "warn", "control")
                content = "Pause requested. The wrangling process will pause at the next checkpoint."
            else:
                session["status"] = "paused"
                _dw_log_event(session, "Session paused.", "warn", "control")
                content = "Session paused. You can add context then click Resume."
            payload = _dw_session_payload(session)
            payload["content"] = content
            return jsonify(payload)

        if control == "stop":
            session["stop_requested"] = True
            if session.get("running"):
                session["status"] = "stopping"
                _dw_log_event(session, "Stop requested (will apply between scan steps).", "warn", "control")
                content = "Stop requested. Current scan will stop at the next checkpoint."
            else:
                session["running"] = False
                session["status"] = "stopped"
                _dw_log_event(session, "Session stopped.", "warn", "control")
                content = "Session stopped."
            payload = _dw_session_payload(session)
            payload["content"] = content
            return jsonify(payload)

        if control == "resume":
            session["pause_requested"] = False
            session["stop_requested"] = False
            if last_user_text:
                session["pending_user_inputs"].append(last_user_text)
                session.setdefault("conversation", []).append({
                    "role": "user",
                    "content": last_user_text,
                    "ts": _dw_now_iso(),
                })
                _dw_log_event(session, "User added context while resuming.", "info", "input")
            started = _dw_start_worker_if_needed(session)
            payload = _dw_session_payload(session)
            payload["content"] = "Session resumed. Wrangling run started." if started else "Session resumed. No pending message to process."
            return jsonify(payload)

        if control == "run":
            session["pause_requested"] = False
            session["stop_requested"] = False
            if last_user_text:
                session["pending_user_inputs"].append(last_user_text)
                session.setdefault("conversation", []).append({
                    "role": "user",
                    "content": last_user_text,
                    "ts": _dw_now_iso(),
                })
                _dw_log_event(session, "Queued new user request for wrangling run.", "info", "input")
            started = _dw_start_worker_if_needed(session)
            payload = _dw_session_payload(session)
            payload["content"] = "Run started." if started else "No pending message to run."
            return jsonify(payload)

        if control == "note":
            if last_user_text:
                session.setdefault("conversation", []).append({
                    "role": "user",
                    "content": last_user_text,
                    "ts": _dw_now_iso(),
                })
                session["conversation"] = session["conversation"][-100:]
                _dw_log_event(session, "Context note captured while paused.", "info", "input")
            payload = _dw_session_payload(session)
            payload["content"] = "Context note saved. Resume when ready."
            return jsonify(payload)

        if not last_user_text:
            payload = _dw_session_payload(session)
            payload["content"] = "No user message provided."
            return jsonify(payload), 400

        session.setdefault("conversation", []).append({
            "role": "user",
            "content": last_user_text,
            "ts": _dw_now_iso(),
        })
        session["conversation"] = session["conversation"][-100:]
        if session.get("status") == "paused" or session.get("pause_requested"):
            payload = _dw_session_payload(session)
            payload["content"] = "Context saved (paused). Click Resume to continue."
            return jsonify(payload)

        session.setdefault("pending_user_inputs", []).append(last_user_text)
        _dw_log_event(session, f"Queued wrangling request: '{last_user_text[:120]}'", "info", "input")
        started = False
        if session.get("params", {}).get("auto_run", True):
            session["pause_requested"] = False
            session["stop_requested"] = False
            started = _dw_start_worker_if_needed(session)
        payload = _dw_session_payload(session)
        payload["content"] = (
            "Wrangling started. Follow live findings in chat and runtime logs."
            if started
            else (
                "Message queued. Click Run to start."
                if not session.get("running")
                else "Message queued while another wrangling run is in progress."
            )
        )
        return jsonify(payload)

AGENTS_CATALOG = [
    {
        "id": "ai-data-analyst",
        "name": "AI Data Analyst (Session)",
        "description": (
            "Agent analyste data en lecture seule, conçu pour des discussions multi-tours. "
            "Il garde une mémoire compacte de la conversation, planifie des requêtes ClickHouse "
            "simples, se réévalue et permet pause/reprise/arrêt avec logs temps réel."
        ),
        "parameters": [
            {
                "name": "max_steps",
                "label": "Maximum steps",
                "type": "number",
                "default": 8,
                "description": "Crédit de steps par run (1-50). Les retries techniques gratuits restent plafonnés globalement.",
            },
            {
                "name": "knowledge_mode",
                "label": "Knowledge strategy",
                "type": "select",
                "options": ["kb_context_once", "kb_agentic", "schema_only", "minimal"],
                "default": "kb_context_once",
                "description": "Choix unique (4 modes) pour contrôler l'usage KB et l'injection de contexte.",
            },
            {
                "name": "memory_turn_limit",
                "label": "Memory turns kept",
                "type": "number",
                "default": 8,
                "description": "Nombre de tours récents conservés dans la mémoire compacte.",
            },
            {
                "name": "memory_token_budget",
                "label": "Memory token budget",
                "type": "number",
                "default": 700,
                "description": "Budget tokens maximal pour la mémoire injectée à chaque nouveau run.",
            },
            {
                "name": "table_filter",
                "label": "Table filter (optional)",
                "type": "string",
                "default": "",
                "description": "Liste de tables séparées par des virgules pour restreindre l'analyse.",
            },
            {
                "name": "auto_run",
                "label": "Auto-run after message",
                "type": "select",
                "options": ["yes", "no"],
                "default": "yes",
                "description": "Lance automatiquement une exécution après chaque nouveau message utilisateur.",
            },
        ],
    },
    {
        "id": "data-wrangling",
        "name": "Nettoyage et Préparation (Data Wrangling)",
        "description": (
            "Agent de data wrangling orienté qualité: planifie son audit, scanne ligne par ligne et champ par champ "
            "sur un scope ciblé, détecte des anomalies subtiles (format, outliers, cohérences croisées), "
            "publie ses findings en direct dans le chat et maintient un watermark par table/scope."
        ),
        "parameters": [
            {
                "name": "table",
                "label": "Table cible",
                "type": "string",
                "default": "",
                "description": "Nom technique de la table à scanner (obligatoire).",
            },
            {
                "name": "date_column",
                "label": "Colonne date (optionnelle)",
                "type": "string",
                "default": "",
                "description": "Colonne date utilisée pour restreindre le scope (BETWEEN ou =).",
            },
            {
                "name": "date_start",
                "label": "Date début (YYYY-MM-DD)",
                "type": "string",
                "default": "",
                "description": "Date de début de scope (optionnel).",
            },
            {
                "name": "date_end",
                "label": "Date fin (YYYY-MM-DD)",
                "type": "string",
                "default": "",
                "description": "Date de fin de scope (optionnel).",
            },
            {
                "name": "row_start",
                "label": "Ligne de départ",
                "type": "number",
                "default": 0,
                "description": "Offset de départ dans le scan (utilisé avec watermark).",
            },
            {
                "name": "max_rows",
                "label": "Nombre de lignes à scanner",
                "type": "number",
                "default": 5000,
                "description": "Volume max analysé par run.",
            },
            {
                "name": "batch_size",
                "label": "Taille de batch",
                "type": "number",
                "default": 400,
                "description": "Nombre de lignes lues par batch.",
            },
            {
                "name": "max_steps",
                "label": "Maximum steps",
                "type": "number",
                "default": 24,
                "description": "Nombre max d'itérations de scan/review.",
            },
            {
                "name": "use_knowledge_base",
                "label": "Use knowledge base",
                "type": "select",
                "options": ["yes", "no"],
                "default": "yes",
                "description": "Utilise la knowledge base pour guider les checks fonctionnels.",
            },
            {
                "name": "reset_watermark",
                "label": "Reset watermark",
                "type": "select",
                "options": ["no", "yes"],
                "default": "no",
                "description": "Si yes: ignore le watermark sauvegardé et rescane depuis row_start.",
            },
            {
                "name": "export_excel",
                "label": "Export anomalies to Excel",
                "type": "select",
                "options": ["no", "yes"],
                "default": "no",
                "description": "Export optionnel des anomalies détectées dans un fichier .xlsx formaté.",
            },
            {
                "name": "export_path",
                "label": "Excel export path",
                "type": "string",
                "default": "",
                "description": "Chemin serveur du fichier .xlsx (vide = chemin auto).",
            },
            {
                "name": "auto_run",
                "label": "Auto-run after message",
                "type": "select",
                "options": ["yes", "no"],
                "default": "yes",
                "description": "Lance automatiquement un run à chaque nouveau message.",
            },
        ],
    },
    {
        "id": "data-dictionary",
        "name": "Générateur de Data Dictionary dynamique",
        "description": (
            "Se connecte au Data Warehouse, analyse le contenu des tables, "
            "comprend le contexte métier et génère (puis maintient à jour) une "
            "documentation complète et compréhensible pour les utilisateurs métiers."
        ),
        "parameters": [
            {
                "name": "database",
                "label": "Base de données",
                "type": "string",
                "default": "",
                "description": "Nom de la base à documenter (vide = base par défaut configurée)",
            },
            {
                "name": "tables",
                "label": "Tables à analyser",
                "type": "string",
                "default": "",
                "description": "Liste de tables séparées par virgule (vide = toutes les tables)",
            },
            {
                "name": "language",
                "label": "Langue de la documentation",
                "type": "select",
                "options": ["fr", "en"],
                "default": "fr",
                "description": "Langue utilisée pour la génération des descriptions",
            },
            {
                "name": "sample_rows",
                "label": "Lignes d'échantillon",
                "type": "number",
                "default": 5,
                "description": "Nombre de lignes d'exemple par table pour aider le LLM",
            },
        ],
    },
    {
        "id": "key-identifier",
        "name": "Agent Key Identifier",
        "description": (
            "Scanne toutes les tables et identifie par similarité de nom de champ et de type de valeur "
            "les relations de clé étrangère potentielles entre les tables. Propose d'enregistrer les "
            "relations confirmées dans la Knowledge Base pour enrichir la génération de requêtes SQL."
        ),
        "parameters": [
            {
                "name": "database",
                "label": "Base de données",
                "type": "string",
                "default": "",
                "description": "Nom de la base à analyser (vide = base par défaut configurée)",
            },
            {
                "name": "sample_size",
                "label": "Valeurs échantillon par champ",
                "type": "number",
                "default": 5,
                "description": "Nombre de valeurs distinctes à échantillonner par champ (1–5) pour identifier les correspondances",
            },
            {
                "name": "confidence",
                "label": "Seuil de confiance",
                "type": "select",
                "options": ["low", "medium", "high"],
                "default": "medium",
                "description": "Niveau de confiance minimum pour qu'une relation soit proposée",
            },
        ],
    },
    {
        "id": "clickhouse-writer",
        "name": "ClickHouse Writer Agent",
        "description": (
            "Agent autonome capable d'enchaîner jusqu'à 12 opérations complexes en lecture ET "
            "écriture sur ClickHouse. Planifie ses actions, crée des tables intermédiaires (BOT_*), "
            "se réévalue toutes les 3 étapes, pose des questions si nécessaire et produit une "
            "synthèse complète avec réflexion."
        ),
        "parameters": [
            {
                "name": "database",
                "label": "Base de données",
                "type": "string",
                "default": "",
                "description": "Nom de la base ClickHouse (vide = base par défaut)",
            },
            {
                "name": "max_actions",
                "label": "Nombre max d'actions",
                "type": "number",
                "default": 12,
                "description": "Budget maximum d'actions successives (1-12)",
            },
            {
                "name": "sample_preview",
                "label": "Lignes de prévisualisation",
                "type": "number",
                "default": 5,
                "description": "Nombre de lignes à afficher par résultat intermédiaire",
            },
            {
                "name": "max_technical_retries",
                "label": "Retries techniques gratuits (global)",
                "type": "number",
                "default": 3,
                "description": "Nombre max de tentatives SQL simplifiées gratuites (0-8) sur tout le run.",
            },
        ],
    },
    {
        "id": "etl-agent",
        "name": "ETL Agent — Import fichiers",
        "description": (
            "Agent ETL autonome : parcourt un dossier (csv, excel, parquet, txt…), "
            "analyse les fichiers, crée des tables BOT_ETL_* dans ClickHouse, insère les données, "
            "ajoute des champs calculés (C_*) et peut enrichir depuis des tables existantes. "
            "Jusqu'à 15 actions paramétrables, demande confirmation à l'utilisateur en cas de doute."
        ),
        "parameters": [
            {
                "name": "folder_path",
                "label": "Dossier source",
                "type": "string",
                "default": "",
                "description": "Chemin absolu du dossier contenant les fichiers à importer",
            },
            {
                "name": "recursive",
                "label": "Sous-dossiers",
                "type": "select",
                "options": ["non", "oui"],
                "default": "non",
                "description": "Explorer aussi les sous-dossiers",
            },
            {
                "name": "database",
                "label": "Base de données cible",
                "type": "string",
                "default": "",
                "description": "Base ClickHouse cible (vide = base par défaut)",
            },
            {
                "name": "max_actions",
                "label": "Nombre max d'actions",
                "type": "number",
                "default": 15,
                "description": "Budget maximum d'actions successives (1-30)",
            },
        ],
    },
]


# ---------------------------------------------------------------------------
# Agent Manager — workflow orchestration + scheduling
# ---------------------------------------------------------------------------
_AGENT_MANAGER_WORKFLOWS_KEY = "agent_manager_workflows"
_AGENT_MANAGER_RUNS_KEY = "agent_manager_runs"
_AGENT_MANAGER_RUNTIME_AGENT_IDS = {"ai-data-analyst", "data-wrangling"}
_AGENT_MANAGER_INTERACTIVE_STATUSES = {
    "awaiting_user",
    "plan_ready",
    "awaiting_request",
    "awaiting_files",
}
_AGENT_MANAGER_ACTIVE_RUN_STATUSES = {"queued", "running", "stopping", "pausing"}
_AGENT_MANAGER_MAX_RUNS = 250
_AGENT_MANAGER_MAX_LOGS = 420
_AGENT_MANAGER_MAX_STEP_RESULTS = 120
_AGENT_MANAGER_SCHEDULER_INTERVAL_SECONDS = 20.0

_agent_manager_lock = _threading.Lock()
_agent_manager_runtime_lock = _threading.Lock()
_agent_manager_runtime: dict = {}
_agent_manager_scheduler_lock = _threading.Lock()
_agent_manager_scheduler_thread = None
_agent_manager_scheduler_started = False
_agent_manager_scheduler_stop = _threading.Event()


def _am_now_dt() -> datetime:
    return datetime.now(timezone.utc)


def _am_now_iso() -> str:
    return _am_now_dt().isoformat()


def _am_parse_iso(value: str | None):
    txt = str(value or "").strip()
    if not txt:
        return None
    try:
        if txt.endswith("Z"):
            txt = txt[:-1] + "+00:00"
        dt = datetime.fromisoformat(txt)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc)
    except Exception:
        return None


def _am_is_truthy(value, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    if isinstance(value, str):
        txt = value.strip().lower()
        if txt in {"1", "true", "yes", "y", "on", "oui"}:
            return True
        if txt in {"0", "false", "no", "n", "off", "non"}:
            return False
    return default


def _am_clamp_int(value, default: int, minimum: int, maximum: int) -> int:
    try:
        out = int(value)
    except Exception:
        out = int(default)
    if out < minimum:
        return minimum
    if out > maximum:
        return maximum
    return out


def _am_safe_tz_name(tz_name: str | None) -> str:
    raw = str(tz_name or "").strip() or "UTC"
    if ZoneInfo is None:
        return "UTC"
    try:
        ZoneInfo(raw)
        return raw
    except Exception:
        return "UTC"


def _am_sanitize_schedule(raw_schedule) -> dict:
    sched = raw_schedule if isinstance(raw_schedule, dict) else {}
    mode_raw = str(sched.get("mode", "disabled")).strip().lower()
    mode = mode_raw if mode_raw in {"disabled", "interval", "daily"} else "disabled"
    interval_minutes = _am_clamp_int(sched.get("interval_minutes", 60), 60, 5, 10080)

    raw_time = str(sched.get("daily_time", "09:00")).strip()
    m = re.match(r"^(\d{1,2}):(\d{2})$", raw_time)
    if m:
        hour = min(23, max(0, int(m.group(1))))
        minute = min(59, max(0, int(m.group(2))))
    else:
        hour, minute = 9, 0
    daily_time = f"{hour:02d}:{minute:02d}"

    timezone_name = _am_safe_tz_name(sched.get("timezone", "UTC"))

    return {
        "mode": mode,
        "interval_minutes": interval_minutes,
        "daily_time": daily_time,
        "timezone": timezone_name,
    }


def _am_compute_next_run(schedule: dict, now_utc: datetime | None = None):
    sched = _am_sanitize_schedule(schedule)
    mode = sched.get("mode", "disabled")
    now = (now_utc or _am_now_dt()).astimezone(timezone.utc)
    if mode == "disabled":
        return None
    if mode == "interval":
        minutes = _am_clamp_int(sched.get("interval_minutes", 60), 60, 5, 10080)
        return now + timedelta(minutes=minutes)

    tz_name = _am_safe_tz_name(sched.get("timezone", "UTC"))
    try:
        tzinfo = ZoneInfo(tz_name) if ZoneInfo else timezone.utc
    except Exception:
        tzinfo = timezone.utc
    local_now = now.astimezone(tzinfo)

    hh, mm = 9, 0
    try:
        ttxt = str(sched.get("daily_time", "09:00"))
        hh = min(23, max(0, int(ttxt.split(":")[0])))
        mm = min(59, max(0, int(ttxt.split(":")[1])))
    except Exception:
        hh, mm = 9, 0
    candidate = local_now.replace(hour=hh, minute=mm, second=0, microsecond=0)
    if candidate <= local_now:
        candidate += timedelta(days=1)
    return candidate.astimezone(timezone.utc)


def _am_truncate(text: str, max_chars: int) -> str:
    raw = str(text or "")
    if len(raw) <= max_chars:
        return raw
    return raw[:max_chars] + " …[truncated]"


def _am_load_state_unlocked():
    db = read_db()
    if not isinstance(db, dict):
        db = {}
    workflows = db.get(_AGENT_MANAGER_WORKFLOWS_KEY, [])
    if not isinstance(workflows, list):
        workflows = []
    runs = db.get(_AGENT_MANAGER_RUNS_KEY, [])
    if not isinstance(runs, list):
        runs = []
    db[_AGENT_MANAGER_WORKFLOWS_KEY] = workflows
    db[_AGENT_MANAGER_RUNS_KEY] = runs
    return db, workflows, runs


def _am_find_workflow_index(workflows: list, workflow_id: str) -> int:
    for i, wf in enumerate(workflows):
        if str(wf.get("id", "")).strip() == str(workflow_id).strip():
            return i
    return -1


def _am_find_run_index(runs: list, run_id: str) -> int:
    for i, run in enumerate(runs):
        if str(run.get("id", "")).strip() == str(run_id).strip():
            return i
    return -1


def _am_trim_runs(runs: list) -> list:
    if len(runs) <= _AGENT_MANAGER_MAX_RUNS:
        return runs

    active = [r for r in runs if str(r.get("status", "")).lower() in _AGENT_MANAGER_ACTIVE_RUN_STATUSES]
    archived = [r for r in runs if r not in active]
    archived.sort(key=lambda r: _am_parse_iso(r.get("created_at")) or _am_now_dt(), reverse=True)

    keep_slots = max(0, _AGENT_MANAGER_MAX_RUNS - len(active))
    kept = active + archived[:keep_slots]
    kept.sort(key=lambda r: _am_parse_iso(r.get("created_at")) or _am_now_dt())
    return kept


def _am_append_run_log_inplace(run: dict, level: str, kind: str, message: str, data=None) -> None:
    seq = int(run.get("log_seq", 0)) + 1
    entry = {
        "seq": seq,
        "ts": _am_now_iso(),
        "level": str(level or "info"),
        "kind": str(kind or "run"),
        "message": _am_truncate(str(message or ""), 800),
    }
    if data is not None:
        entry["data"] = data
    run["log_seq"] = seq
    run.setdefault("logs", []).append(entry)
    if len(run["logs"]) > _AGENT_MANAGER_MAX_LOGS:
        run["logs"] = run["logs"][-_AGENT_MANAGER_MAX_LOGS:]
    run["updated_at"] = entry["ts"]


def _am_get_agent_ids() -> set:
    return {str(a.get("id", "")).strip() for a in AGENTS_CATALOG}


def _am_parse_step_params(raw_params):
    if isinstance(raw_params, dict):
        return raw_params
    if isinstance(raw_params, str):
        txt = raw_params.strip()
        if txt:
            try:
                parsed = json.loads(txt)
                if isinstance(parsed, dict):
                    return parsed
            except Exception:
                return {}
    return {}


def _am_sanitize_step(raw_step: dict, index: int) -> dict:
    if not isinstance(raw_step, dict):
        raise ValueError(f"Step #{index + 1} must be an object.")
    agent_id = str(raw_step.get("agent_id", "")).strip()
    if not agent_id:
        raise ValueError(f"Step #{index + 1}: agent_id is required.")
    if agent_id not in _am_get_agent_ids():
        raise ValueError(f"Step #{index + 1}: unknown agent_id '{agent_id}'.")

    title = str(raw_step.get("title", "")).strip() or f"Step {index + 1}"
    prompt = str(
        raw_step.get("prompt")
        or raw_step.get("objective")
        or raw_step.get("prompt_template")
        or ""
    ).strip()
    if not prompt:
        prompt = "Execute this step and provide clear findings."

    timeout_seconds = _am_clamp_int(raw_step.get("timeout_seconds", 420), 420, 30, 3600)
    halt_on_error = _am_is_truthy(raw_step.get("halt_on_error"), True)
    auto_approve_questions = _am_is_truthy(raw_step.get("auto_approve_questions"), False)
    auto_reply_text = str(raw_step.get("auto_reply_text", "oui")).strip() or "oui"
    max_followups = _am_clamp_int(raw_step.get("max_followups", 2), 2, 0, 5)

    return {
        "id": str(raw_step.get("id", "")).strip() or str(uuid.uuid4()),
        "order": index + 1,
        "agent_id": agent_id,
        "title": _am_truncate(title, 160),
        "prompt": _am_truncate(prompt, 8000),
        "params": _am_parse_step_params(raw_step.get("params", {})),
        "halt_on_error": halt_on_error,
        "timeout_seconds": timeout_seconds,
        "auto_approve_questions": auto_approve_questions,
        "auto_reply_text": _am_truncate(auto_reply_text, 300),
        "max_followups": max_followups,
    }


def _am_sanitize_workflow_payload(raw_payload, existing: dict | None = None) -> dict:
    payload = raw_payload if isinstance(raw_payload, dict) else {}
    now_iso = _am_now_iso()
    current = existing or {}

    name = str(payload.get("name", current.get("name", ""))).strip()
    if not name:
        raise ValueError("Workflow name is required.")

    description = str(payload.get("description", current.get("description", ""))).strip()
    objective = str(payload.get("objective", current.get("objective", ""))).strip()
    default_input = str(payload.get("default_input", current.get("default_input", ""))).strip()

    if "enabled" in payload:
        enabled = _am_is_truthy(payload.get("enabled"), False)
    else:
        enabled = _am_is_truthy(current.get("enabled"), False)

    if "schedule" in payload:
        schedule = _am_sanitize_schedule(payload.get("schedule"))
    else:
        schedule = _am_sanitize_schedule(current.get("schedule", {}))

    raw_steps = payload.get("steps", current.get("steps", []))
    if not isinstance(raw_steps, list):
        raise ValueError("steps must be a list.")
    steps = [_am_sanitize_step(step, i) for i, step in enumerate(raw_steps)]
    if not steps:
        raise ValueError("Workflow must contain at least one step.")

    next_run_at = None
    if enabled and schedule.get("mode") != "disabled":
        next_dt = _am_compute_next_run(schedule, _am_now_dt())
        next_run_at = next_dt.isoformat() if next_dt else None

    return {
        "id": str(current.get("id", "")).strip() or str(uuid.uuid4()),
        "name": _am_truncate(name, 140),
        "description": _am_truncate(description, 1200),
        "objective": _am_truncate(objective, 4000),
        "default_input": _am_truncate(default_input, 4000),
        "enabled": enabled,
        "schedule": schedule,
        "steps": steps,
        "created_at": current.get("created_at") or now_iso,
        "updated_at": now_iso,
        "last_run_at": current.get("last_run_at"),
        "next_run_at": next_run_at,
    }


def _am_run_summary(run: dict) -> dict:
    return {
        "id": run.get("id"),
        "workflow_id": run.get("workflow_id"),
        "workflow_name": run.get("workflow_name"),
        "trigger": run.get("trigger"),
        "status": run.get("status"),
        "summary": run.get("summary", ""),
        "error": run.get("error", ""),
        "stop_requested": bool(run.get("stop_requested", False)),
        "created_at": run.get("created_at"),
        "started_at": run.get("started_at"),
        "completed_at": run.get("completed_at"),
        "updated_at": run.get("updated_at"),
        "step_results_count": len(run.get("step_results", [])),
    }


def _am_mutate_run(run_id: str, mutator):
    with _agent_manager_lock:
        db, workflows, runs = _am_load_state_unlocked()
        idx = _am_find_run_index(runs, run_id)
        if idx < 0:
            return None
        run = deepcopy(runs[idx])
        mutator(run)
        run["updated_at"] = _am_now_iso()
        runs[idx] = run
        db[_AGENT_MANAGER_WORKFLOWS_KEY] = workflows
        db[_AGENT_MANAGER_RUNS_KEY] = _am_trim_runs(runs)
        write_db(db)
        return deepcopy(run)


def _am_get_run(run_id: str):
    with _agent_manager_lock:
        _, _, runs = _am_load_state_unlocked()
        idx = _am_find_run_index(runs, run_id)
        if idx < 0:
            return None
        return deepcopy(runs[idx])


def _am_get_workflow(workflow_id: str):
    with _agent_manager_lock:
        _, workflows, _ = _am_load_state_unlocked()
        idx = _am_find_workflow_index(workflows, workflow_id)
        if idx < 0:
            return None
        return deepcopy(workflows[idx])


def _am_register_runtime_context(run_id: str, agent_id: str = "", session_id: str = "") -> None:
    with _agent_manager_runtime_lock:
        ctx = _agent_manager_runtime.get(run_id, {})
        if agent_id:
            ctx["agent_id"] = agent_id
        if session_id:
            ctx["session_id"] = session_id
        _agent_manager_runtime[run_id] = ctx


def _am_clear_runtime_context(run_id: str) -> None:
    with _agent_manager_runtime_lock:
        if run_id in _agent_manager_runtime:
            del _agent_manager_runtime[run_id]


def _am_get_runtime_context(run_id: str) -> dict:
    with _agent_manager_runtime_lock:
        ctx = _agent_manager_runtime.get(run_id, {})
        return {
            "agent_id": str(ctx.get("agent_id", "")).strip(),
            "session_id": str(ctx.get("session_id", "")).strip(),
        }


def _am_dispatch_agent_call(agent_id: str, body: dict) -> tuple[dict, int]:
    try:
        with app.test_request_context(
            f"/api/agents/{agent_id}/chat",
            method="POST",
            json=body,
        ):
            raw_response = agent_chat(agent_id)
    except Exception as exc:
        return {"error": f"Internal dispatch failed: {exc}"}, 500

    status_code = 200
    response_obj = raw_response
    if isinstance(raw_response, tuple):
        if len(raw_response) >= 1:
            response_obj = raw_response[0]
        if len(raw_response) >= 2:
            try:
                status_code = int(raw_response[1])
            except Exception:
                status_code = 500

    if hasattr(response_obj, "status_code"):
        try:
            status_code = int(response_obj.status_code)
        except Exception:
            pass

    if hasattr(response_obj, "get_json"):
        payload = response_obj.get_json(silent=True) or {}
    elif isinstance(response_obj, dict):
        payload = response_obj
    else:
        payload = {"content": str(response_obj)}

    if not isinstance(payload, dict):
        payload = {"content": str(payload)}
    return payload, status_code


def _am_is_interactive_payload(payload: dict) -> bool:
    status = str(payload.get("status", "")).strip().lower()
    if status in _AGENT_MANAGER_INTERACTIVE_STATUSES:
        return True
    return bool(payload.get("question"))


def _am_extract_step_text(payload: dict) -> str:
    if not isinstance(payload, dict):
        return ""
    latest = payload.get("latest_assistant")
    if isinstance(latest, dict):
        txt = str(latest.get("content", "")).strip()
        if txt:
            return txt
    content = str(payload.get("content", "")).strip()
    if content:
        return content
    msg = str(payload.get("message", "")).strip()
    if msg:
        return msg
    analyst_result = payload.get("analyst_result")
    if isinstance(analyst_result, dict):
        txt = str(analyst_result.get("final_answer", "")).strip()
        if txt:
            return txt
    synthesis = payload.get("synthesis")
    if isinstance(synthesis, dict):
        txt = str(synthesis.get("conclusion", "")).strip()
        if txt:
            return txt
    return ""


def _am_is_stop_requested(run_id: str) -> bool:
    run = _am_get_run(run_id)
    return bool(run and run.get("stop_requested"))


def _am_poll_runtime_agent(
    run_id: str,
    agent_id: str,
    initial_payload: dict,
    step_params: dict,
    timeout_seconds: int,
) -> tuple[dict, str, str]:
    payload = initial_payload if isinstance(initial_payload, dict) else {}
    session_id = str(payload.get("session_id", "")).strip()
    if session_id:
        _am_register_runtime_context(run_id, agent_id=agent_id, session_id=session_id)

    deadline = time.time() + float(timeout_seconds)
    stop_sent = False

    while True:
        if _am_is_stop_requested(run_id):
            if session_id and not stop_sent:
                _am_dispatch_agent_call(
                    agent_id,
                    {"control": "stop", "session_id": session_id, "params": step_params},
                )
                stop_sent = True
            return payload, "stopped", "Run stopped by user request."

        status_txt = str(payload.get("status", "")).strip().lower()
        running = bool(payload.get("running")) or status_txt in {"running", "pausing", "stopping"}
        if not running:
            return payload, "ok", ""

        if time.time() > deadline:
            return payload, "timeout", f"Runtime step timed out after {timeout_seconds}s."

        time.sleep(1.0)
        poll_payload, poll_status = _am_dispatch_agent_call(
            agent_id,
            {"control": "status", "session_id": session_id, "params": step_params},
        )
        payload = poll_payload if isinstance(poll_payload, dict) else {}
        if poll_status >= 400:
            err = payload.get("error") or f"Status polling failed ({poll_status})."
            return payload, "error", str(err)
        session_id = str(payload.get("session_id", "")).strip() or session_id
        if session_id:
            _am_register_runtime_context(run_id, agent_id=agent_id, session_id=session_id)


def _am_auto_followup(
    agent_id: str,
    base_payload: dict,
    step: dict,
) -> tuple[dict, str, str]:
    payload = base_payload if isinstance(base_payload, dict) else {}
    if not _am_is_interactive_payload(payload):
        return payload, "ok", ""

    if not _am_is_truthy(step.get("auto_approve_questions"), False):
        status_txt = str(payload.get("status", "")).strip().lower() or "awaiting_user"
        return payload, "needs_input", f"Agent requires user input (status={status_txt})."

    max_followups = _am_clamp_int(step.get("max_followups", 2), 2, 0, 5)
    auto_reply = str(step.get("auto_reply_text", "oui")).strip() or "oui"
    session_id = str(payload.get("session_id", "")).strip()
    step_params = step.get("params", {}) if isinstance(step.get("params"), dict) else {}

    for _ in range(max_followups):
        if not _am_is_interactive_payload(payload):
            return payload, "ok", ""
        body = {
            "messages": [{"role": "user", "content": auto_reply}],
            "params": step_params,
        }
        if session_id:
            body["session_id"] = session_id
        payload, status_code = _am_dispatch_agent_call(agent_id, body)
        if status_code >= 400:
            err = payload.get("error") or f"Agent follow-up failed ({status_code})."
            return payload, "error", str(err)
        session_id = str(payload.get("session_id", "")).strip() or session_id

    if _am_is_interactive_payload(payload):
        return payload, "needs_input", "Agent still requires input after auto follow-ups."
    return payload, "ok", ""


def _am_build_step_prompt(
    workflow: dict,
    step: dict,
    run_input: str,
    prior_outputs: list[str],
    step_idx: int,
    total_steps: int,
) -> str:
    objective = str(workflow.get("objective", "")).strip()
    context_snippets = [s for s in prior_outputs[-3:] if str(s or "").strip()]
    context_block = "\n\n".join(f"- { _am_truncate(s, 900) }" for s in context_snippets) if context_snippets else "- none"
    step_prompt = str(step.get("prompt", "")).strip()
    run_goal = str(run_input or workflow.get("default_input", "")).strip()
    return (
        f"You are executed by Agent Manager in workflow '{workflow.get('name','')}'.\n"
        f"Workflow objective: {objective or 'n/a'}\n"
        f"Global user need: {run_goal or 'n/a'}\n"
        f"Current step: {step_idx}/{total_steps} — {step.get('title','')}\n\n"
        f"Step instruction:\n{step_prompt}\n\n"
        "Prior step outputs (condensed):\n"
        f"{context_block}\n\n"
        "Return a concise, decision-useful answer. If data is missing, explain what is blocked."
    )


def _am_execute_step(
    run_id: str,
    workflow: dict,
    step: dict,
    step_idx: int,
    total_steps: int,
    run_input: str,
    prior_outputs: list[str],
) -> dict:
    started_at = _am_now_iso()
    agent_id = str(step.get("agent_id", "")).strip()
    step_params = step.get("params", {}) if isinstance(step.get("params"), dict) else {}
    timeout_seconds = _am_clamp_int(step.get("timeout_seconds", 420), 420, 30, 3600)

    prompt = _am_build_step_prompt(workflow, step, run_input, prior_outputs, step_idx, total_steps)
    payload, status_code = _am_dispatch_agent_call(
        agent_id,
        {
            "messages": [{"role": "user", "content": prompt}],
            "params": step_params,
        },
    )

    if status_code >= 400:
        err = payload.get("error") or f"Agent call failed ({status_code})."
        return {
            "step_id": step.get("id"),
            "step_index": step_idx,
            "title": step.get("title"),
            "agent_id": agent_id,
            "status": "failed",
            "started_at": started_at,
            "ended_at": _am_now_iso(),
            "error": str(err),
            "output_preview": "",
            "raw_payload": _am_truncate(json.dumps(payload, ensure_ascii=False), 10000),
        }

    session_id = str(payload.get("session_id", "")).strip()
    if session_id:
        _am_register_runtime_context(run_id, agent_id=agent_id, session_id=session_id)

    if agent_id in _AGENT_MANAGER_RUNTIME_AGENT_IDS:
        payload, state, state_msg = _am_poll_runtime_agent(
            run_id=run_id,
            agent_id=agent_id,
            initial_payload=payload,
            step_params=step_params,
            timeout_seconds=timeout_seconds,
        )
        if state == "stopped":
            return {
                "step_id": step.get("id"),
                "step_index": step_idx,
                "title": step.get("title"),
                "agent_id": agent_id,
                "status": "stopped",
                "started_at": started_at,
                "ended_at": _am_now_iso(),
                "error": state_msg,
                "output_preview": _am_truncate(_am_extract_step_text(payload), 3000),
                "raw_payload": _am_truncate(json.dumps(payload, ensure_ascii=False), 10000),
            }
        if state in {"error", "timeout"}:
            return {
                "step_id": step.get("id"),
                "step_index": step_idx,
                "title": step.get("title"),
                "agent_id": agent_id,
                "status": "failed",
                "started_at": started_at,
                "ended_at": _am_now_iso(),
                "error": state_msg,
                "output_preview": _am_truncate(_am_extract_step_text(payload), 3000),
                "raw_payload": _am_truncate(json.dumps(payload, ensure_ascii=False), 10000),
            }
    else:
        payload, state, state_msg = _am_auto_followup(agent_id, payload, step)
        if state == "needs_input":
            return {
                "step_id": step.get("id"),
                "step_index": step_idx,
                "title": step.get("title"),
                "agent_id": agent_id,
                "status": "failed",
                "started_at": started_at,
                "ended_at": _am_now_iso(),
                "error": state_msg,
                "output_preview": _am_truncate(_am_extract_step_text(payload), 3000),
                "raw_payload": _am_truncate(json.dumps(payload, ensure_ascii=False), 10000),
            }
        if state == "error":
            return {
                "step_id": step.get("id"),
                "step_index": step_idx,
                "title": step.get("title"),
                "agent_id": agent_id,
                "status": "failed",
                "started_at": started_at,
                "ended_at": _am_now_iso(),
                "error": state_msg,
                "output_preview": _am_truncate(_am_extract_step_text(payload), 3000),
                "raw_payload": _am_truncate(json.dumps(payload, ensure_ascii=False), 10000),
            }

    output_preview = _am_truncate(_am_extract_step_text(payload), 3000)
    if not output_preview:
        output_preview = _am_truncate(json.dumps(payload, ensure_ascii=False), 1200)
    return {
        "step_id": step.get("id"),
        "step_index": step_idx,
        "title": step.get("title"),
        "agent_id": agent_id,
        "status": "completed",
        "started_at": started_at,
        "ended_at": _am_now_iso(),
        "error": "",
        "output_preview": output_preview,
        "raw_payload": _am_truncate(json.dumps(payload, ensure_ascii=False), 10000),
    }


def _am_run_worker(run_id: str) -> None:
    try:
        def _mark_running(run: dict):
            if run.get("stop_requested"):
                run["status"] = "stopped"
                run["completed_at"] = run.get("completed_at") or _am_now_iso()
                run["summary"] = run.get("summary") or "Run cancelled before start."
                _am_append_run_log_inplace(run, "warn", "control", "Run cancelled before start.")
                return
            run["status"] = "running"
            run["started_at"] = run.get("started_at") or _am_now_iso()
            _am_append_run_log_inplace(run, "info", "run", "Run started.")

        run = _am_mutate_run(run_id, _mark_running)
        if not run:
            return
        if str(run.get("status", "")).lower() == "stopped":
            return

        workflow = _am_get_workflow(str(run.get("workflow_id", "")))
        if not workflow:
            def _mark_missing_wf(r: dict):
                r["status"] = "failed"
                r["completed_at"] = _am_now_iso()
                r["error"] = "Workflow introuvable."
                r["summary"] = "Run failed: workflow not found."
                _am_append_run_log_inplace(r, "error", "run", "Workflow not found.")

            _am_mutate_run(run_id, _mark_missing_wf)
            return

        steps = workflow.get("steps", []) if isinstance(workflow.get("steps"), list) else []
        total_steps = len(steps)
        run_input = str(run.get("input", "")).strip()
        prior_outputs: list[str] = []

        for idx, step in enumerate(steps, start=1):
            if _am_is_stop_requested(run_id):
                def _mark_stopped(r: dict):
                    r["status"] = "stopped"
                    r["completed_at"] = _am_now_iso()
                    r["summary"] = "Run stopped by user."
                    _am_append_run_log_inplace(r, "warn", "control", "Run stopped by user.")

                _am_mutate_run(run_id, _mark_stopped)
                return

            def _log_step_start(r: dict):
                _am_append_run_log_inplace(
                    r,
                    "info",
                    "step",
                    f"Step {idx}/{total_steps} started: {step.get('title','')} ({step.get('agent_id','')})",
                )

            _am_mutate_run(run_id, _log_step_start)

            step_result = _am_execute_step(
                run_id=run_id,
                workflow=workflow,
                step=step,
                step_idx=idx,
                total_steps=total_steps,
                run_input=run_input,
                prior_outputs=prior_outputs,
            )

            status_txt = str(step_result.get("status", "")).lower()
            if status_txt == "completed":
                prior_outputs.append(str(step_result.get("output_preview", "")))

            def _store_step_result(r: dict):
                r.setdefault("step_results", []).append(step_result)
                if len(r["step_results"]) > _AGENT_MANAGER_MAX_STEP_RESULTS:
                    r["step_results"] = r["step_results"][-_AGENT_MANAGER_MAX_STEP_RESULTS:]
                if status_txt == "completed":
                    _am_append_run_log_inplace(
                        r,
                        "info",
                        "step",
                        f"Step {idx}/{total_steps} completed.",
                        {"agent_id": step_result.get("agent_id")},
                    )
                elif status_txt == "stopped":
                    _am_append_run_log_inplace(
                        r,
                        "warn",
                        "step",
                        f"Step {idx}/{total_steps} stopped: {step_result.get('error','')}",
                    )
                else:
                    _am_append_run_log_inplace(
                        r,
                        "error",
                        "step",
                        f"Step {idx}/{total_steps} failed: {step_result.get('error','')}",
                    )

            _am_mutate_run(run_id, _store_step_result)

            if status_txt == "stopped":
                def _mark_stopped_after_step(r: dict):
                    r["status"] = "stopped"
                    r["completed_at"] = _am_now_iso()
                    r["summary"] = "Run stopped."

                _am_mutate_run(run_id, _mark_stopped_after_step)
                return

            if status_txt == "failed" and _am_is_truthy(step.get("halt_on_error"), True):
                def _mark_failed(r: dict):
                    r["status"] = "failed"
                    r["completed_at"] = _am_now_iso()
                    r["error"] = str(step_result.get("error", "Step failed"))
                    r["summary"] = (
                        f"Run failed at step {idx}/{total_steps}: {step_result.get('title','')}"
                    )
                    _am_append_run_log_inplace(
                        r,
                        "error",
                        "run",
                        f"Run halted on error at step {idx}.",
                    )

                _am_mutate_run(run_id, _mark_failed)
                return

        def _mark_completed(r: dict):
            r["status"] = "completed"
            r["completed_at"] = _am_now_iso()
            completed_steps = sum(
                1
                for s in r.get("step_results", [])
                if str(s.get("status", "")).lower() == "completed"
            )
            r["summary"] = f"Workflow completed. {completed_steps}/{total_steps} steps succeeded."
            _am_append_run_log_inplace(r, "info", "run", r["summary"])

        _am_mutate_run(run_id, _mark_completed)

    except Exception as exc:
        def _mark_crash(r: dict):
            r["status"] = "failed"
            r["completed_at"] = _am_now_iso()
            r["error"] = str(exc)
            r["summary"] = "Run failed due to internal manager error."
            _am_append_run_log_inplace(r, "error", "run", f"Unhandled manager error: {exc}")

        _am_mutate_run(run_id, _mark_crash)
    finally:
        _am_clear_runtime_context(run_id)


def _am_start_run_thread(run_id: str) -> bool:
    with _agent_manager_runtime_lock:
        existing = _agent_manager_runtime.get(run_id, {})
        t = existing.get("thread")
        if t and getattr(t, "is_alive", lambda: False)():
            return False
        worker = _threading.Thread(
            target=_am_run_worker,
            args=(run_id,),
            daemon=True,
            name=f"agent-manager-{run_id[:8]}",
        )
        _agent_manager_runtime[run_id] = {
            "thread": worker,
            "agent_id": "",
            "session_id": "",
        }
    worker.start()
    return True


def _am_enqueue_run(workflow_id: str, trigger: str, input_text: str = "") -> tuple[dict | None, str, int]:
    now_iso = _am_now_iso()
    with _agent_manager_lock:
        db, workflows, runs = _am_load_state_unlocked()
        wf_idx = _am_find_workflow_index(workflows, workflow_id)
        if wf_idx < 0:
            return None, "Workflow not found.", 404

        active_exists = any(
            str(r.get("workflow_id", "")) == str(workflow_id)
            and str(r.get("status", "")).lower() in _AGENT_MANAGER_ACTIVE_RUN_STATUSES
            for r in runs
        )
        if active_exists:
            return None, "A run is already active for this workflow.", 409

        workflow = workflows[wf_idx]
        run_input = str(input_text or workflow.get("default_input", "")).strip()
        run = {
            "id": str(uuid.uuid4()),
            "workflow_id": workflow.get("id"),
            "workflow_name": workflow.get("name"),
            "trigger": str(trigger or "manual"),
            "status": "queued",
            "summary": "Run queued.",
            "error": "",
            "input": _am_truncate(run_input, 4000),
            "stop_requested": False,
            "created_at": now_iso,
            "updated_at": now_iso,
            "started_at": None,
            "completed_at": None,
            "step_results": [],
            "logs": [],
            "log_seq": 0,
        }
        _am_append_run_log_inplace(run, "info", "run", f"Run queued (trigger={trigger}).")
        runs.append(run)
        db[_AGENT_MANAGER_RUNS_KEY] = _am_trim_runs(runs)

        workflows[wf_idx]["last_run_at"] = now_iso
        sched = workflows[wf_idx].get("schedule", {})
        if workflows[wf_idx].get("enabled") and _am_sanitize_schedule(sched).get("mode") != "disabled":
            next_dt = _am_compute_next_run(sched, _am_now_dt())
            workflows[wf_idx]["next_run_at"] = next_dt.isoformat() if next_dt else None
        else:
            workflows[wf_idx]["next_run_at"] = None
        workflows[wf_idx]["updated_at"] = now_iso

        db[_AGENT_MANAGER_WORKFLOWS_KEY] = workflows
        write_db(db)
        run_copy = deepcopy(run)

    _am_start_run_thread(run_copy["id"])
    return run_copy, "", 200


def _am_stop_run(run_id: str):
    def _mark_stop_requested(run: dict):
        run["stop_requested"] = True
        status = str(run.get("status", "")).lower()
        if status in {"queued"}:
            run["status"] = "stopped"
            run["completed_at"] = _am_now_iso()
            run["summary"] = "Run stopped before execution."
        elif status in {"running", "pausing"}:
            run["status"] = "stopping"
            run["summary"] = "Stop requested."
        _am_append_run_log_inplace(run, "warn", "control", "Stop requested by user.")

    run = _am_mutate_run(run_id, _mark_stop_requested)
    if not run:
        return None

    ctx = _am_get_runtime_context(run_id)
    agent_id = str(ctx.get("agent_id", "")).strip()
    session_id = str(ctx.get("session_id", "")).strip()
    if agent_id and session_id:
        _am_dispatch_agent_call(
            agent_id,
            {"control": "stop", "session_id": session_id, "params": {}},
        )
    return run


def _am_scheduler_tick() -> None:
    due_workflow_ids = []
    now = _am_now_dt()
    with _agent_manager_lock:
        db, workflows, runs = _am_load_state_unlocked()
        active_by_workflow = {
            str(r.get("workflow_id", ""))
            for r in runs
            if str(r.get("status", "")).lower() in _AGENT_MANAGER_ACTIVE_RUN_STATUSES
        }
        changed = False

        for wf in workflows:
            wf_id = str(wf.get("id", "")).strip()
            if not wf_id:
                continue
            enabled = _am_is_truthy(wf.get("enabled"), False)
            sched = _am_sanitize_schedule(wf.get("schedule", {}))
            wf["schedule"] = sched

            if not enabled or sched.get("mode") == "disabled":
                if wf.get("next_run_at"):
                    wf["next_run_at"] = None
                    changed = True
                continue

            next_dt = _am_parse_iso(wf.get("next_run_at"))
            if next_dt is None:
                computed = _am_compute_next_run(sched, now)
                wf["next_run_at"] = computed.isoformat() if computed else None
                changed = True
                next_dt = computed

            if next_dt and next_dt <= now:
                if wf_id in active_by_workflow:
                    wf["next_run_at"] = (now + timedelta(minutes=1)).isoformat()
                    changed = True
                    continue
                due_workflow_ids.append(wf_id)

        if changed:
            db[_AGENT_MANAGER_WORKFLOWS_KEY] = workflows
            db[_AGENT_MANAGER_RUNS_KEY] = _am_trim_runs(runs)
            write_db(db)

    for wf_id in due_workflow_ids:
        run, err, status = _am_enqueue_run(wf_id, "scheduled", "")
        if run:
            _log(
                f"Scheduled workflow '{run.get('workflow_name','')}' started (run {run.get('id','')[:8]}).",
                source="agent-manager",
            )
        elif status != 409:
            _log(f"Scheduler failed to enqueue workflow {wf_id}: {err}", "warn", "agent-manager")


def _am_scheduler_loop() -> None:
    _log("Agent Manager scheduler started.", source="agent-manager")
    while not _agent_manager_scheduler_stop.is_set():
        try:
            _am_scheduler_tick()
        except Exception as exc:
            _log(f"Scheduler tick error: {exc}", "error", "agent-manager")
        _agent_manager_scheduler_stop.wait(_AGENT_MANAGER_SCHEDULER_INTERVAL_SECONDS)


def _am_start_scheduler_if_needed() -> None:
    global _agent_manager_scheduler_started, _agent_manager_scheduler_thread
    if _agent_manager_scheduler_started:
        return
    if app.testing:
        return
    if _am_is_truthy(os.environ.get("CLICKSENSE_DISABLE_AGENT_MANAGER_SCHEDULER"), False):
        return
    with _agent_manager_scheduler_lock:
        if _agent_manager_scheduler_started:
            return
        _agent_manager_scheduler_stop.clear()
        _agent_manager_scheduler_thread = _threading.Thread(
            target=_am_scheduler_loop,
            daemon=True,
            name="agent-manager-scheduler",
        )
        _agent_manager_scheduler_thread.start()
        _agent_manager_scheduler_started = True


@app.before_request
def _ensure_agent_manager_scheduler():
    if request.path.startswith("/api/"):
        _am_start_scheduler_if_needed()


@app.route("/api/agent-manager/agents", methods=["GET"])
def agent_manager_agents():
    catalog = []
    for agent in AGENTS_CATALOG:
        aid = str(agent.get("id", "")).strip()
        catalog.append({
            "id": aid,
            "name": agent.get("name", aid),
            "description": agent.get("description", ""),
            "parameters": agent.get("parameters", []),
            "runtime": aid in _AGENT_MANAGER_RUNTIME_AGENT_IDS,
            "interactive_possible": aid in {"clickhouse-writer", "etl-agent"},
        })
    return jsonify({"agents": catalog})


@app.route("/api/agent-manager/workflows", methods=["GET"])
def agent_manager_list_workflows():
    with _agent_manager_lock:
        _, workflows, _ = _am_load_state_unlocked()
        result = sorted(
            [deepcopy(wf) for wf in workflows if isinstance(wf, dict)],
            key=lambda wf: _am_parse_iso(wf.get("updated_at")) or _am_now_dt(),
            reverse=True,
        )
    return jsonify({"workflows": result})


@app.route("/api/agent-manager/workflows", methods=["POST"])
def agent_manager_create_workflow():
    payload = request.get_json(silent=True) or {}
    try:
        workflow = _am_sanitize_workflow_payload(payload, existing=None)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    with _agent_manager_lock:
        db, workflows, runs = _am_load_state_unlocked()
        workflows.append(workflow)
        db[_AGENT_MANAGER_WORKFLOWS_KEY] = workflows
        db[_AGENT_MANAGER_RUNS_KEY] = _am_trim_runs(runs)
        write_db(db)
    return jsonify({"workflow": workflow}), 201


@app.route("/api/agent-manager/workflows/<workflow_id>", methods=["PUT"])
def agent_manager_update_workflow(workflow_id):
    payload = request.get_json(silent=True) or {}
    with _agent_manager_lock:
        db, workflows, runs = _am_load_state_unlocked()
        idx = _am_find_workflow_index(workflows, workflow_id)
        if idx < 0:
            return jsonify({"error": "Workflow not found."}), 404
        existing = workflows[idx]
        try:
            workflow = _am_sanitize_workflow_payload(payload, existing=existing)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        workflows[idx] = workflow
        db[_AGENT_MANAGER_WORKFLOWS_KEY] = workflows
        db[_AGENT_MANAGER_RUNS_KEY] = _am_trim_runs(runs)
        write_db(db)
    return jsonify({"workflow": workflow})


@app.route("/api/agent-manager/workflows/<workflow_id>", methods=["DELETE"])
def agent_manager_delete_workflow(workflow_id):
    with _agent_manager_lock:
        db, workflows, runs = _am_load_state_unlocked()
        idx = _am_find_workflow_index(workflows, workflow_id)
        if idx < 0:
            return jsonify({"error": "Workflow not found."}), 404
        active = any(
            str(r.get("workflow_id", "")) == str(workflow_id)
            and str(r.get("status", "")).lower() in _AGENT_MANAGER_ACTIVE_RUN_STATUSES
            for r in runs
        )
        if active:
            return jsonify({"error": "Cannot delete while a run is active."}), 409
        deleted = workflows.pop(idx)
        db[_AGENT_MANAGER_WORKFLOWS_KEY] = workflows
        db[_AGENT_MANAGER_RUNS_KEY] = _am_trim_runs(runs)
        write_db(db)
    return jsonify({"success": True, "deleted_id": deleted.get("id")})


@app.route("/api/agent-manager/workflows/<workflow_id>/run", methods=["POST"])
def agent_manager_run_workflow(workflow_id):
    data = request.get_json(silent=True) or {}
    run_input = str(data.get("input", "")).strip()
    trigger = str(data.get("trigger", "manual")).strip() or "manual"
    run, err, status = _am_enqueue_run(workflow_id, trigger, run_input)
    if not run:
        return jsonify({"error": err}), status
    return jsonify({"run": _am_run_summary(run), "run_id": run.get("id")})


@app.route("/api/agent-manager/runs", methods=["GET"])
def agent_manager_list_runs():
    workflow_id = str(request.args.get("workflow_id", "")).strip()
    limit = _am_clamp_int(request.args.get("limit", 50), 50, 1, 200)
    with _agent_manager_lock:
        _, _, runs = _am_load_state_unlocked()
        items = [deepcopy(r) for r in runs if isinstance(r, dict)]
    if workflow_id:
        items = [r for r in items if str(r.get("workflow_id", "")) == workflow_id]
    items.sort(key=lambda r: _am_parse_iso(r.get("created_at")) or _am_now_dt(), reverse=True)
    summaries = [_am_run_summary(r) for r in items[:limit]]
    return jsonify({"runs": summaries})


@app.route("/api/agent-manager/runs/<run_id>", methods=["GET"])
def agent_manager_get_run(run_id):
    run = _am_get_run(run_id)
    if not run:
        return jsonify({"error": "Run not found."}), 404
    return jsonify({"run": run})


@app.route("/api/agent-manager/runs/<run_id>/stop", methods=["POST"])
def agent_manager_stop_run(run_id):
    run = _am_stop_run(run_id)
    if not run:
        return jsonify({"error": "Run not found."}), 404
    return jsonify({"run": _am_run_summary(run), "success": True})


@app.route("/api/agent-manager/scheduler", methods=["GET"])
def agent_manager_scheduler_status():
    with _agent_manager_scheduler_lock:
        started = bool(_agent_manager_scheduler_started)
        alive = bool(_agent_manager_scheduler_thread and _agent_manager_scheduler_thread.is_alive())
    return jsonify({
        "started": started,
        "alive": alive,
        "interval_seconds": _AGENT_MANAGER_SCHEDULER_INTERVAL_SECONDS,
    })


@app.route("/api/console-logs", methods=["GET"])
def get_console_logs():
    """Return buffered console logs for the frontend Console panel."""
    since = int(request.args.get("since", -1))
    with _console_lock:
        total = len(_console_buffer)
        if since < 0:
            logs = list(_console_buffer[-100:])
            start_idx = max(0, total - 100)
        else:
            logs = list(_console_buffer[since + 1:])
            start_idx = since + 1
    return jsonify({"logs": logs, "next_idx": start_idx + len(logs) - 1, "total": total})


@app.route("/api/agents", methods=["GET"])
def list_agents():
    """Return the catalog of available agents."""
    return jsonify(AGENTS_CATALOG)


@app.route("/api/agents/<agent_id>/chat", methods=["POST"])
def agent_chat(agent_id):
    """Dispatch a chat message to the requested agent."""
    if agent_id == "ai-data-analyst":
        return _run_ai_data_analyst_session_agent()
    if agent_id == "data-wrangling":
        return _run_data_wrangling_agent()
    if agent_id == "data-dictionary":
        return _run_data_dictionary_agent()
    if agent_id == "clickhouse-writer":
        return _run_clickhouse_writer_agent()
    if agent_id == "key-identifier":
        return _run_key_identifier_agent()
    if agent_id == "etl-agent":
        return _run_etl_agent()
    return jsonify({"error": f"Agent '{agent_id}' introuvable."}), 404


@app.route("/api/agents/etl-agent/cleanup", methods=["POST"])
def etl_agent_cleanup():
    """Drop all BOT_ETL_ tables for a given ETL session."""
    data = request.get_json(silent=True) or {}
    session_id = data.get("session_id", "")
    session = _etl_sessions.get(session_id, {})
    created_tables = session.get("created_tables", [])
    database = session.get("database", clickhouse_config.get("database", "default"))
    try:
        client = get_clickhouse_client()
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500
    dropped, errors = [], []
    for table in created_tables:
        try:
            client.command(f"DROP TABLE IF EXISTS `{database}`.`{table}`")
            dropped.append(table)
        except Exception as exc:
            errors.append(f"{table}: {str(exc)}")
    if session_id in _etl_sessions:
        _etl_sessions[session_id]["created_tables"] = [
            t for t in created_tables if t not in dropped
        ]
    return jsonify({"dropped": dropped, "errors": errors})


@app.route("/api/etl/browse", methods=["POST"])
def etl_browse_files():
    """List supported data files in a folder."""
    data = request.get_json(silent=True) or {}
    folder = data.get("folder_path", "").strip()
    recursive = data.get("recursive", False)
    if not folder:
        return jsonify({"error": "folder_path est requis."}), 400
    if not os.path.isdir(folder):
        return jsonify({"error": f"Dossier introuvable: {folder}"}), 404
    files = _etl_list_files(folder, recursive)
    return jsonify({"files": files, "count": len(files)})


@app.route("/api/agents/clickhouse-writer/cleanup", methods=["POST"])
def clickhouse_writer_cleanup():
    """Drop BOT_* tables for a given writer session (including pre-existing BOT_*)."""
    data = request.get_json(silent=True) or {}
    session_id = data.get("session_id", "")
    session = _writer_sessions.get(session_id, {})
    created_tables = session.get("created_tables", [])
    database = (
        (data.get("database") or "").strip()
        or session.get("database", clickhouse_config.get("database", "default"))
    )

    raw_requested = data.get("tables")
    requested_tables = []
    if isinstance(raw_requested, list):
        requested_tables = [str(x) for x in raw_requested if str(x).strip()]
    elif isinstance(raw_requested, str):
        if "," in raw_requested:
            requested_tables = [s.strip() for s in raw_requested.split(",") if s.strip()]
        else:
            requested_tables = _cw_extract_bot_table_mentions(raw_requested) or [raw_requested.strip()]

    try:
        client = get_clickhouse_client()
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500

    existing_bot_tables = _cw_list_existing_bot_tables(client, database)
    resolved = _cw_resolve_cleanup_targets(
        existing_bot_tables=existing_bot_tables,
        session_tables=created_tables,
        requested_tables=requested_tables,
    )
    targets = resolved["targets"]
    dropped, errors = _cw_drop_tables(client, database, targets)

    if session_id in _writer_sessions:
        dropped_upper = {str(t).upper() for t in dropped}
        _writer_sessions[session_id]["created_tables"] = [
            t
            for t in resolved["session_allowed"]
            if _cw_normalize_table_name(t).upper() not in dropped_upper
        ]

    return jsonify({
        "dropped": dropped,
        "errors": errors,
        "skipped_non_bot": resolved["skipped_non_bot"],
        "not_found": resolved["not_found"],
        "available_bot_tables": resolved["available"],
    })


# ===========================================================================
# ClickHouse Writer Agent — helper functions
# ===========================================================================

def _cw_get_schema_info(client, database: str, max_tables: int = 30) -> str:
    """Return a compact schema overview (tables + columns + row counts)."""
    try:
        res = client.query(f"SHOW TABLES FROM `{database}`")
        tables = [row[0] for row in res.result_rows]
    except Exception as exc:
        return f"Impossible de lister les tables: {exc}"

    # Approximate row counts from system metadata (much cheaper than COUNT(*) per table).
    row_count_map = {}
    try:
        counts_res = client.query(
            f"SELECT name, total_rows FROM system.tables WHERE database = '{database}'"
        )
        row_count_map = {str(r[0]): r[1] for r in counts_res.result_rows}
    except Exception:
        row_count_map = {}

    parts = []
    for tbl in tables[:max_tables]:
        try:
            desc = client.query(f"DESCRIBE TABLE `{database}`.`{tbl}`")
            cols = ", ".join(f"{r[0]}:{r[1]}" for r in desc.result_rows[:15])
            cnt = row_count_map.get(tbl, "?")
            parts.append(f"TABLE {tbl} ({cnt} rows): {cols}")
        except Exception as exc:
            parts.append(f"TABLE {tbl}: (erreur schema: {exc})")

    if len(tables) > max_tables:
        parts.append(f"... et {len(tables) - max_tables} autres tables.")
    return "\n".join(parts)


def _cw_normalize_table_name(table_name: str) -> str:
    """Normalize table identifier by removing db qualifier and backticks."""
    txt = str(table_name or "").strip()
    if not txt:
        return ""
    if "." in txt:
        txt = txt.split(".")[-1]
    return txt.strip().strip("`").strip()


def _cw_is_bot_table_name(table_name: str) -> bool:
    clean = _cw_normalize_table_name(table_name)
    return bool(clean) and clean.upper().startswith("BOT_")


def _cw_filter_bot_tables(table_names: list[str]) -> tuple[list[str], list[str]]:
    """Split a table list into BOT_* and non-BOT tables (normalized unique names)."""
    allowed, skipped = [], []
    seen_allowed, seen_skipped = set(), set()
    for raw in table_names or []:
        clean = _cw_normalize_table_name(raw)
        if not clean:
            continue
        if _cw_is_bot_table_name(clean):
            if clean not in seen_allowed:
                seen_allowed.add(clean)
                allowed.append(clean)
        else:
            if clean not in seen_skipped:
                seen_skipped.add(clean)
                skipped.append(clean)
    return allowed, skipped


def _cw_list_existing_bot_tables(client, database: str) -> list[str]:
    """List current BOT_* tables in the target database."""
    try:
        res = client.query(f"SHOW TABLES FROM `{database}`")
        raw_tables = [str(row[0]) for row in res.result_rows]
    except Exception:
        raw_tables = []
    allowed, _ = _cw_filter_bot_tables(raw_tables)
    return allowed


def _cw_extract_bot_table_mentions(text: str) -> list[str]:
    """Extract BOT_* table names mentioned in free text."""
    raw = str(text or "")
    mentions = re.findall(r"\bBOT_[A-Za-z0-9_]+\b", raw, flags=re.IGNORECASE)
    allowed, _ = _cw_filter_bot_tables(mentions)
    return allowed


def _cw_resolve_cleanup_targets(
    existing_bot_tables: list[str],
    session_tables: list[str],
    requested_tables: list[str],
) -> dict:
    """
    Resolve which BOT_* tables can be dropped.

    - available BOT tables = existing BOT tables in DB + session BOT tables
    - if requested_tables provided, only those are targeted (case-insensitive match)
    - otherwise all available BOT tables are targeted
    """
    existing_allowed, _ = _cw_filter_bot_tables(existing_bot_tables)
    session_allowed, session_skipped_non_bot = _cw_filter_bot_tables(session_tables)
    requested_allowed, requested_skipped_non_bot = _cw_filter_bot_tables(requested_tables)
    available, _ = _cw_filter_bot_tables(existing_allowed + session_allowed)

    if requested_allowed:
        upper_map = {}
        for name in available:
            upper_map[str(name).upper()] = name
        targets = []
        not_found = []
        seen_targets = set()
        for req in requested_allowed:
            match = upper_map.get(str(req).upper())
            if match:
                match_norm = _cw_normalize_table_name(match)
                if match_norm not in seen_targets:
                    seen_targets.add(match_norm)
                    targets.append(match_norm)
            else:
                not_found.append(_cw_normalize_table_name(req))
    else:
        targets = available
        not_found = []

    return {
        "targets": targets,
        "not_found": not_found,
        "available": available,
        "session_allowed": session_allowed,
        "skipped_non_bot": sorted(set(session_skipped_non_bot + requested_skipped_non_bot)),
    }


def _cw_drop_tables(client, database: str, table_names: list[str]) -> tuple[list[str], list[str]]:
    dropped, errors = [], []
    for table in table_names:
        try:
            client.command(f"DROP TABLE IF EXISTS `{database}`.`{table}`")
            dropped.append(table)
        except Exception as exc:
            errors.append(f"{table}: {str(exc)}")
    return dropped, errors


def _cw_compact_action_log_for_prompt(
    action_log: list[dict],
    *,
    max_items: int = 6,
    max_sql_chars: int = 200,
    max_result_chars: int = 200,
) -> str:
    lines = []
    for entry in action_log[-max_items:]:
        status = "OK" if entry.get("ok") else "FAIL"
        sql_preview = str(entry.get("sql", "")).replace("\n", " ")[:max_sql_chars]
        result_preview = str(entry.get("result_preview", "")).replace("\n", " ")[:max_result_chars]
        lines.append(
            f"Step {entry.get('step_id')} [{status}] {entry.get('description', '')[:120]} | "
            f"SQL: {sql_preview} | Result: {result_preview}"
        )
    return "\n".join(lines) if lines else "(first step)"


def _cw_detect_bot_table(sql: str) -> str | None:
    """Extract the BOT_ table name from a CREATE TABLE statement."""
    m = re.search(r"CREATE\s+TABLE\s+(?:IF\s+NOT\s+EXISTS\s+)?(?:(?:`[^`]+`|\w+)\.)?`?(BOT_\w+)`?",
                  sql, re.IGNORECASE)
    return m.group(1) if m else None


def _cw_extract_table_name(sql: str, pattern: str) -> str | None:
    """Extract a table name (with optional db prefix) from a SQL statement."""
    m = re.match(pattern, sql.strip(), re.IGNORECASE)
    if not m:
        return None
    # The table name may be backtick-quoted; strip them
    return _cw_normalize_table_name(m.group(1))


def _cw_is_sql_safe(sql: str) -> tuple[bool, str]:
    """
    Security guard: verify that a SQL statement cannot destroy or alter
    pre-existing (non-BOT_) tables.

    Rules
    -----
    - SELECT / SHOW / DESCRIBE / EXPLAIN / WITH  → always allowed (read-only)
    - CREATE TABLE                               → allowed only for BOT_* tables
    - INSERT INTO                                → allowed only into BOT_* tables
    - DROP TABLE [IF EXISTS]                     → allowed only for BOT_* tables
    - ALTER TABLE                                → allowed only for BOT_* tables
    - TRUNCATE [TABLE]                           → allowed only for BOT_* tables
    - DELETE FROM                                → allowed only for BOT_* tables
    - DROP DATABASE / DROP SCHEMA                → always blocked
    - RENAME TABLE                               → always blocked

    Returns (is_safe: bool, reason: str).
    """
    sql_stripped = sql.strip()
    sql_upper = sql_stripped.upper()

    # ── Read-only operations ─────────────────────────────────────────────────
    read_prefixes = ("SELECT", "SHOW", "DESCRIBE", "EXPLAIN", "WITH")
    if any(sql_upper.startswith(k) for k in read_prefixes):
        return True, ""

    # ── DROP DATABASE / DROP SCHEMA – never allowed ──────────────────────────
    if re.match(r"DROP\s+(DATABASE|SCHEMA)\b", sql_stripped, re.IGNORECASE):
        return False, "DROP DATABASE/SCHEMA n'est pas autorisé par l'agent Writer."

    # ── All non-table DROP statements are blocked ────────────────────────────
    if re.match(r"DROP\s+\w+\b", sql_stripped, re.IGNORECASE) and not re.match(
        r"DROP\s+TABLE\b", sql_stripped, re.IGNORECASE
    ):
        return False, "Seul DROP TABLE est autorisé, et uniquement sur des tables BOT_*."

    # ── RENAME TABLE – never allowed ─────────────────────────────────────────
    if re.match(r"RENAME\s+TABLE\b", sql_stripped, re.IGNORECASE):
        return False, "RENAME TABLE n'est pas autorisé par l'agent Writer."

    # Helper: return (False, message) if table does not start with BOT_
    def _require_bot(table_name: str, op: str) -> tuple[bool, str] | None:
        clean = _cw_normalize_table_name(table_name)
        if not _cw_is_bot_table_name(clean):
            return (
                False,
                f"[SÉCURITÉ] {op} refusé : la table '{clean or table_name}' n'est pas une table "
                f"temporaire BOT_. L'agent Writer ne peut pas modifier les tables existantes.",
            )
        return None  # OK

    # ── DROP TABLE [IF EXISTS] [db.]table ────────────────────────────────────
    drop_match = re.match(
        r"DROP\s+TABLE\s+(?:IF\s+EXISTS\s+)?(.+)$",
        sql_stripped,
        re.IGNORECASE,
    )
    if drop_match is not None:
        drop_tail = drop_match.group(1).strip()
        # Disallow multi-target drop to prevent bypasses like
        # DROP TABLE BOT_tmp, prod_table.
        if "," in drop_tail:
            return False, "[SÉCURITÉ] DROP TABLE multi-cibles interdit. Supprimez une seule table BOT_* par commande."
        table_token = re.split(r"\s+", drop_tail, maxsplit=1)[0]
        table = _cw_normalize_table_name(table_token)
        err = _require_bot(table, "DROP TABLE")
        if err:
            return err
        return True, ""

    # ── ALTER TABLE [db.]table ───────────────────────────────────────────────
    table = _cw_extract_table_name(
        sql_stripped,
        r"ALTER\s+TABLE\s+(?:(?:`[^`]+`|\w+)\.)?(`?\w+`?)",
    )
    if table is not None:
        err = _require_bot(table, "ALTER TABLE")
        if err:
            return err
        return True, ""

    # ── TRUNCATE [TABLE] [db.]table ──────────────────────────────────────────
    table = _cw_extract_table_name(
        sql_stripped,
        r"TRUNCATE\s+(?:TABLE\s+)?(?:(?:`[^`]+`|\w+)\.)?(`?\w+`?)",
    )
    if table is not None:
        err = _require_bot(table, "TRUNCATE")
        if err:
            return err
        return True, ""

    # ── DELETE FROM [db.]table ───────────────────────────────────────────────
    table = _cw_extract_table_name(
        sql_stripped,
        r"DELETE\s+FROM\s+(?:(?:`[^`]+`|\w+)\.)?(`?\w+`?)",
    )
    if table is not None:
        err = _require_bot(table, "DELETE FROM")
        if err:
            return err
        return True, ""

    # ── CREATE TABLE [IF NOT EXISTS] [db.]table ──────────────────────────────
    table = _cw_extract_table_name(
        sql_stripped,
        r"CREATE\s+TABLE\s+(?:IF\s+NOT\s+EXISTS\s+)?(?:(?:`[^`]+`|\w+)\.)?(`?\w+`?)",
    )
    if table is not None:
        err = _require_bot(table, "CREATE TABLE")
        if err:
            return err
        return True, ""

    # ── INSERT INTO [db.]table ───────────────────────────────────────────────
    table = _cw_extract_table_name(
        sql_stripped,
        r"INSERT\s+INTO\s+(?:(?:`[^`]+`|\w+)\.)?(`?\w+`?)",
    )
    if table is not None:
        err = _require_bot(table, "INSERT INTO")
        if err:
            return err
        return True, ""

    # Block other non-read operations by default for safer writer behavior.
    return False, (
        "Statement non autorisé par l'agent Writer. "
        "Utilisez uniquement SELECT/SHOW/DESCRIBE/EXPLAIN/WITH "
        "ou des écritures strictement cadrées sur des tables BOT_*."
    )


def _cw_plan(user_request: str, database: str, schema_info: str) -> dict:
    """LLM generates a detailed execution plan."""
    system_prompt = (
        "Tu es un expert ClickHouse et architecte de données senior. "
        "Tu analyses des demandes complexes et crées des plans d'exécution précis et efficaces. "
        "Tu penses comme un ingénieur senior qui doit traiter de très grandes tables. "
        "Tu réponds UNIQUEMENT avec du JSON valide, sans markdown ni texte supplémentaire."
    )
    user_content = (
        f"Base de données: {database}\n\n"
        f"Tables disponibles:\n{schema_info}\n\n"
        f"Demande de l'utilisateur: {user_request}\n\n"
        "Crée un plan d'exécution détaillé avec AU MAXIMUM 12 étapes pour réaliser cette demande.\n"
        "Tu peux créer des tables intermédiaires (MUST start with BOT_) pour les calculs complexes.\n"
        "RÈGLE DE SÉCURITÉ ABSOLUE: tu n'as JAMAIS le droit de DROP, ALTER, TRUNCATE, DELETE, "
        "INSERT ou CREATE sur des tables existantes. Seules les tables préfixées BOT_ peuvent être "
        "créées ou modifiées. Tout écart sera bloqué côté serveur.\n"
        "RÈGLE DATES: privilégie des plages temporelles explicites avec BETWEEN. "
        "Évite les transformations coûteuses sur les colonnes de date dans WHERE. "
        "N'impose pas d'année fixe sauf demande explicite de l'utilisateur.\n"
        "RÈGLE DÉCOMPOSITION: Décompose les opérations complexes en sous-requêtes simples enchaînées. "
        "Pour les agrégations multi-niveaux: étape 1 = SELECT simple avec GROUP BY, "
        "étape 2 = INSERT dans table BOT_, étape 3 = vérification. "
        "Ne jamais tenter de faire en une seule requête ce qui nécessite plusieurs passes.\n"
        "RÈGLE FONCTIONS: Évite les fonctions ClickHouse avancées (windowFunnel, retention, argMax, "
        "topK, uniqHLL12, etc.). Préfère COUNT, SUM, AVG, MIN, MAX, groupArray si nécessaire.\n"
        "Sois stratégique: identifie les données nécessaires, les transformations requises, "
        "les vérifications à faire.\n\n"
        "Réponds UNIQUEMENT avec ce JSON:\n"
        "{\n"
        '  "objective": "description claire de l\'objectif",\n'
        '  "approach": "approche haut niveau expliquée",\n'
        '  "estimated_steps": N,\n'
        '  "complexity": "simple|medium|complex|very_complex",\n'
        '  "steps": [\n'
        "    {\n"
        '      "id": 1,\n'
        '      "description": "ce que cette étape accomplit",\n'
        '      "type": "explore|compute|create_table|insert|verify|aggregate|cleanup",\n'
        '      "rationale": "pourquoi cette étape est nécessaire",\n'
        '      "creates_table": null\n'
        "    }\n"
        "  ]\n"
        "}"
    )
    user_content = _truncate_text_to_budget(
        user_content,
        int(_get_effective_context_limit() * 0.56),
    )
    try:
        raw = _call_llm(system_prompt, [{"role": "user", "content": user_content}], temperature=0.1)
        result = _parse_llm_json(raw)
    except Exception:
        result = None
    if not result or "steps" not in result:
        result = {
            "objective": user_request,
            "approach": (
                "Plan de secours déterministe: explorer les données, créer une table BOT_ si nécessaire, "
                "puis vérifier le résultat."
            ),
            "estimated_steps": 3,
            "complexity": "medium",
            "steps": [
                {
                    "id": 1,
                    "description": "Explorer rapidement les tables et colonnes utiles avec une requête simple.",
                    "type": "explore",
                    "rationale": "Valider le périmètre avant écriture.",
                    "creates_table": None,
                },
                {
                    "id": 2,
                    "description": "Créer/peupler une table BOT_ ciblée pour matérialiser le résultat.",
                    "type": "create_table",
                    "rationale": "Isoler le résultat sans toucher aux tables métier existantes.",
                    "creates_table": "BOT_result_tmp",
                },
                {
                    "id": 3,
                    "description": "Vérifier les données produites et préparer la conclusion.",
                    "type": "verify",
                    "rationale": "Confirmer la qualité et l'utilité du résultat.",
                    "creates_table": None,
                },
            ],
        }
    # Enforce max 12 steps
    result["steps"] = result["steps"][:12]
    return result


def _cw_generate_sql(session: dict, step: dict, client, database: str) -> dict:
    """LLM generates the SQL query (or a question) for a given plan step."""
    plan = session["plan"]
    action_log = session["action_log"]
    user_context = session.get("user_context", {})

    prev_summaries = _cw_compact_action_log_for_prompt(
        action_log,
        max_items=5,
        max_sql_chars=170,
        max_result_chars=190,
    )

    try:
        res = client.query(f"SHOW TABLES FROM `{database}`")
        all_tables = [row[0] for row in res.result_rows]
    except Exception:
        all_tables = []
    tables_text = ", ".join(all_tables[:60])
    if len(all_tables) > 60:
        tables_text += f" ... (+{len(all_tables) - 60} autres)"

    system_prompt = (
        "Tu es un expert SQL ClickHouse en train d'exécuter une étape d'un plan d'analyse. "
        "Tu écris du SQL ClickHouse efficace et correct. "
        "Tu es précis et tiens compte des volumes de données. "
        "Tu réponds UNIQUEMENT avec du JSON valide, sans markdown ni texte supplémentaire."
    )
    user_context_text = _truncate_text_to_budget(
        json.dumps(user_context or {}, ensure_ascii=False),
        280,
    )
    user_content = (
        f"Base: {database} | Tables disponibles: {tables_text}\n\n"
        f"Objectif global: {plan.get('objective', '')}\n"
        f"Approche: {plan.get('approach', '')}\n\n"
        f"Étape actuelle {step['id']}/{len(plan.get('steps', []))}: {step['description']}\n"
        f"Type: {step.get('type', 'compute')} | "
        f"Crée une table: {step.get('creates_table', 'non')}\n"
        f"Raison: {step.get('rationale', '')}\n\n"
        f"Résultats des étapes précédentes:\n{prev_summaries}\n\n"
        f"Contexte fourni par l'utilisateur: {user_context_text}\n\n"
        "Génère le SQL pour cette étape.\n"
        "Règles OBLIGATOIRES:\n"
        "- SÉCURITÉ ABSOLUE: tu n'as le droit d'écrire (CREATE, INSERT, DROP, ALTER, TRUNCATE, DELETE) "
        "QUE sur des tables dont le nom commence par BOT_. "
        "Toute tentative d'écrire sur une table existante (non-BOT_) sera bloquée par le système.\n"
        "- Tables temporaires: TOUJOURS préfixées BOT_ (ex: BOT_aggreg_results)\n"
        "- Syntaxe ClickHouse valide uniquement\n"
        "- Pour grandes tables: utilise LIMIT ou SAMPLE pour les explorations\n"
        "- DATES — CRITIQUE: privilégie des plages explicites avec BETWEEN et évite "
        "les transformations coûteuses sur les colonnes de date dans WHERE. "
        "Traduis correctement les périodes relatives demandées (ex: 30 derniers jours, mois en cours) "
        "sans hardcoder d'année fixe. Pour le regroupement périodique, toYYYYMM() ou formatDateTime() "
        "sont autorisés si nécessaire.\n"
        "- CREATE TABLE: utilise ENGINE = MergeTree() ORDER BY tuple() si pas d'ordre naturel. "
        "Crée la table et l'INSERT dans des SQL séparés — ne jamais combiner CREATE + INSERT.\n"
        "- Pour INSERT: INSERT INTO `db`.`BOT_table` SELECT ...\n"
        "- DÉCOMPOSITION: si le calcul est complexe (multi-niveaux, jointures multiples), "
        "génère la PREMIÈRE étape simple uniquement. Les étapes suivantes du plan feront le reste.\n"
        "- FONCTIONS INTERDITES: windowFunnel, retention, argMax, topK, uniqHLL12, uniqExact (utilise COUNT DISTINCT), "
        "runningAccumulate, neighbor, and toRelativeXNum functions.\n"
        "- Si tu dois choisir entre plusieurs options et que l'utilisateur doit décider: "
        "pose une question avec des choix\n\n"
        "Réponds UNIQUEMENT avec ce JSON:\n"
        "{\n"
        '  "sql": "ta requête SQL ici",\n'
        '  "explanation": "ce que cette requête fait et pourquoi",\n'
        '  "creates_table": null,\n'
        '  "needs_user_input": false,\n'
        '  "question": null\n'
        "}\n\n"
        "Si needs_user_input est true, mets sql à null et question à:\n"
        "{\n"
        '  "text": "Question à poser à l\'utilisateur",\n'
        '  "choices": [{"label": "Description option", "value": "valeur"}]\n'
        "}"
    )
    user_content = _truncate_text_to_budget(
        user_content,
        int(_get_effective_context_limit() * 0.58),
    )
    raw = _call_llm(system_prompt, [{"role": "user", "content": user_content}], temperature=0.05)
    result = _parse_llm_json(raw)
    if not result:
        step_id = step["id"]
        step_desc = step["description"]
        return {
            "sql": f"SELECT '{step_id}' AS step_id, '{step_desc}' AS description",
            "explanation": "Fallback SQL (parsing LLM échoué)",
            "creates_table": None,
            "needs_user_input": False,
            "question": None,
        }
    return result


def _cw_replan(session: dict, remaining_credits: int) -> dict:
    """LLM re-evaluates the plan every 3 steps and adjusts if needed."""
    plan = session["plan"]
    action_log = session["action_log"]
    current_idx = session["action_index"]

    results_summary = _cw_compact_action_log_for_prompt(
        action_log,
        max_items=8,
        max_sql_chars=120,
        max_result_chars=160,
    )

    remaining_steps = [s for s in plan.get("steps", []) if s["id"] > current_idx]
    next_id = current_idx + 1

    system_prompt = (
        "Tu es un analyste stratégique réévaluant un plan d'exécution à mi-parcours. "
        "Tu décides si le plan est encore optimal ou s'il faut l'adapter. "
        "Tu réponds UNIQUEMENT avec du JSON valide, sans markdown."
    )
    user_content = (
        f"Objectif: {plan.get('objective', '')}\n"
        f"Approche originale: {plan.get('approach', '')}\n\n"
        f"Crédits restants (actions max): {remaining_credits}\n\n"
        f"Résultats obtenus jusqu'ici:\n{results_summary}\n\n"
        f"Étapes restantes prévues:\n{json.dumps(remaining_steps, ensure_ascii=False, indent=2)}\n\n"
        "Réévalue: Le plan actuel est-il encore optimal compte tenu des résultats?\n"
        "Considère:\n"
        "1. Les résultats sont-ils conformes aux attentes?\n"
        "2. Des étapes ont-elles échoué ou produit des données inattendues?\n"
        "3. Peut-on atteindre l'objectif plus efficacement?\n"
        "4. Faut-il ajouter des étapes basées sur ce qu'on a découvert?\n\n"
        "Réponds UNIQUEMENT avec:\n"
        "{\n"
        '  "should_replan": false,\n'
        '  "reason": "explication de la décision",\n'
        '  "new_remaining_steps": null\n'
        "}\n"
        f"Si should_replan est true, fournis new_remaining_steps: tableau d'au plus "
        f"{remaining_credits} étapes avec id (commençant à {next_id}), description, type, "
        "rationale, creates_table."
    )
    user_content = _truncate_text_to_budget(
        user_content,
        int(_get_effective_context_limit() * 0.55),
    )
    try:
        raw = _call_llm(system_prompt, [{"role": "user", "content": user_content}], temperature=0.1)
        result = _parse_llm_json(raw)
    except Exception:
        result = None
    if not result:
        return {"should_replan": False, "reason": "Pas de réévaluation nécessaire."}
    return result


def _cw_synthesize(session: dict) -> dict:
    """LLM generates a comprehensive synthesis of the entire operation."""
    plan = session["plan"]
    action_log = session["action_log"]
    created_tables, _ = _cw_filter_bot_tables(session.get("created_tables", []))
    replan_log = session.get("replan_log", [])
    replan_summary = []
    for i, r in enumerate(replan_log[-8:]):
        replan_summary.append(
            f"Après étape {i * 3 + 3}: {str(r.get('reason', ''))[:180]} "
            f"(replanifié: {bool(r.get('should_replan', False))})"
        )
    action_summary = _cw_compact_action_log_for_prompt(
        action_log,
        max_items=12,
        max_sql_chars=220,
        max_result_chars=240,
    )

    system_prompt = (
        "Tu es un analyste senior rédigeant un rapport de synthèse complet et perspicace. "
        "Tu expliques des opérations techniques complexes de façon claire, avec des insights métier. "
        "Ta synthèse doit être précieuse à la fois pour les profils techniques et métier. "
        "Tu réponds UNIQUEMENT avec du JSON valide, sans markdown ni texte superflu."
    )
    user_content = (
        f"Objectif: {plan.get('objective', '')}\n"
        f"Approche: {plan.get('approach', '')}\n\n"
        f"Journal d'exécution (compact):\n{action_summary}\n\n"
        f"Réévaluations du plan:\n" + ("\n".join(replan_summary) if replan_summary else "Aucune réévaluation.") + "\n\n"
        f"Tables temporaires créées: {', '.join(created_tables) if created_tables else 'aucune'}\n\n"
        "Génère une synthèse complète et détaillée.\n"
        "Réponds UNIQUEMENT avec:\n"
        "{\n"
        '  "executive_summary": "résumé exécutif 2-3 phrases de ce qui a été accompli",\n'
        '  "key_findings": ["découverte 1", "découverte 2", "..."],\n'
        '  "step_reflections": [\n'
        "    {\n"
        '      "step_id": 1,\n'
        '      "description": "description de l\'étape",\n'
        '      "outcome": "ce qui s\'est passé",\n'
        '      "insight": "ce que ça révèle sur les données",\n'
        '      "status": "success|failed|partial"\n'
        "    }\n"
        "  ],\n"
        '  "data_insights": "observations analytiques approfondies sur les patterns découverts",\n'
        '  "recommendations": ["recommandation 1", "recommandation 2"],\n'
        '  "conclusion": "conclusion finale et prochaines étapes suggérées",\n'
        '  "tables_created": [{"name": "BOT_xxx", "purpose": "contenu", "useful_for": "usages futurs"}]\n'
        "}"
    )
    user_content = _truncate_text_to_budget(
        user_content,
        int(_get_effective_context_limit() * 0.6),
    )
    result = None
    try:
        raw = _call_llm(system_prompt, [{"role": "user", "content": user_content}], temperature=0.3)
        result = _parse_llm_json(raw)
    except Exception as exc:
        if _is_context_overflow_error(str(exc)):
            compact_prompt = _truncate_text_to_budget(
                user_content,
                int(_get_effective_context_limit() * 0.35),
            )
            try:
                raw = _call_llm(
                    system_prompt,
                    [{"role": "user", "content": compact_prompt}],
                    temperature=0.2,
                )
                result = _parse_llm_json(raw)
            except Exception:
                result = None
        else:
            result = None
    if not result:
        successful = [e for e in action_log if e.get("ok")]
        failed = [e for e in action_log if not e.get("ok")]
        return {
            "executive_summary": (
                f"Analyse complétée en {len(action_log)} étape(s), "
                f"avec {len(successful)} succès et {len(failed)} échec(s)."
            ),
            "key_findings": [e["description"] for e in successful[:8]],
            "step_reflections": [],
            "data_insights": (
                "Les résultats confirment une exécution structurée, avec des étapes à creuser "
                "lorsque les résultats intermédiaires étaient vides ou en échec."
            ),
            "recommendations": [
                "Conserver les tables BOT_* utiles pour accélérer les analyses suivantes.",
                "Relancer les étapes en échec avec un périmètre plus ciblé.",
            ],
            "conclusion": "Analyse terminée avec synthèse de secours (mode résilient).",
            "tables_created": [{"name": t, "purpose": "Table intermédiaire", "useful_for": ""}
                                for t in created_tables],
        }
    return result


def _cw_generate_simple_sql(session: dict, step: dict, client, database: str,
                            failed_sql: str, error_msg: str) -> str:
    """Generate a simpler alternative SQL when the original was rejected by ClickHouse.

    Returns the simpler SQL string, or an empty string if generation failed.
    This call does NOT count against the agent's action credit.
    """
    plan = session["plan"]
    action_log = session["action_log"]

    prev_summaries = _cw_compact_action_log_for_prompt(
        action_log,
        max_items=5,
        max_sql_chars=150,
        max_result_chars=170,
    )

    try:
        res = client.query(f"SHOW TABLES FROM `{database}`")
        all_tables = [row[0] for row in res.result_rows]
    except Exception:
        all_tables = []
    tables_text = ", ".join(all_tables[:60])
    if len(all_tables) > 60:
        tables_text += f" ... (+{len(all_tables) - 60} autres)"

    system_prompt = (
        "Tu es un expert SQL ClickHouse. Une requête SQL a été rejetée par ClickHouse avec une erreur technique. "
        "Tu dois générer une version PLUS SIMPLE qui atteint le même objectif analytique. "
        "Tu réponds UNIQUEMENT avec du JSON valide, sans markdown ni texte supplémentaire."
    )
    user_content = (
        f"Base: {database} | Tables disponibles: {tables_text}\n\n"
        f"Objectif global: {plan.get('objective', '')}\n\n"
        f"Étape actuelle {step['id']}: {step['description']}\n\n"
        f"SQL qui a échoué:\n{failed_sql}\n\n"
        f"Erreur ClickHouse:\n{error_msg}\n\n"
        f"Contexte des étapes précédentes:\n{prev_summaries}\n\n"
        "RÈGLES DE SIMPLIFICATION OBLIGATOIRES:\n"
        "- Utilise MOINS de fonctions; remplace les fonctions complexes par COUNT, SUM, AVG, MIN, MAX simples\n"
        "- Évite les window functions, CTEs, sous-requêtes imbriquées — utilise un SELECT … GROUP BY plat\n"
        "- Évite les fonctions ClickHouse avancées (windowFunnel, retention, argMax, topK, uniqHLL12, etc.)\n"
        "- DATES: privilégie BETWEEN avec bornes explicites; évite les transformations coûteuses dans WHERE; pas d'année fixe hardcodée sauf demande explicite\n"
        "- Ajoute LIMIT 100 si absent\n"
        "- Corrige tout problème de syntaxe indiqué par le message d'erreur\n\n"
        "Réponds UNIQUEMENT avec ce JSON:\n"
        "{\n"
        '  "sql": "ta requête SQL simplifiée ici",\n'
        '  "explanation": "ce qui a été simplifié et pourquoi"\n'
        "}"
    )
    user_content = _truncate_text_to_budget(
        user_content,
        int(_get_effective_context_limit() * 0.56),
    )
    try:
        raw = _call_llm(system_prompt, [{"role": "user", "content": user_content}], temperature=0.05)
        result = _parse_llm_json(raw)
    except Exception:
        result = None
    if not result:
        return ""
    return (result.get("sql") or "").strip()


def _cw_rethink_strategy(session: dict, step: dict, client, database: str,
                          failed_sql: str, error_msg: str) -> str:
    """When both original SQL and the simplified retry fail, force the LLM to devise
    a completely different strategy for the step.  This call costs one credit.

    Returns the rethought SQL string, or empty string if generation failed.
    """
    plan = session["plan"]
    action_log = session["action_log"]

    prev_summaries = _cw_compact_action_log_for_prompt(
        action_log,
        max_items=7,
        max_sql_chars=170,
        max_result_chars=170,
    )

    try:
        res = client.query(f"SHOW TABLES FROM `{database}`")
        all_tables = [row[0] for row in res.result_rows]
    except Exception:
        all_tables = []
    tables_text = ", ".join(all_tables[:60])
    if len(all_tables) > 60:
        tables_text += f" ... (+{len(all_tables) - 60} autres)"

    system_prompt = (
        "Tu es un expert ClickHouse devant résoudre un problème persistant. "
        "Deux tentatives de SQL ont échoué pour cette étape. Tu DOIS proposer une NOUVELLE STRATÉGIE "
        "complètement différente — ne répète pas les approches qui ont déjà échoué. "
        "Décompose l'opération en étapes plus simples si nécessaire. "
        "Tu réponds UNIQUEMENT avec du JSON valide, sans markdown."
    )
    user_content = (
        f"Base: {database} | Tables disponibles: {tables_text}\n\n"
        f"Objectif global: {plan.get('objective', '')}\n\n"
        f"Étape {step['id']}: {step['description']}\n"
        f"Type: {step.get('type', 'compute')}\n\n"
        f"SQL original qui a échoué:\n{failed_sql}\n\n"
        f"Erreur: {error_msg}\n\n"
        f"Contexte des étapes précédentes:\n{prev_summaries}\n\n"
        "NOUVELLE STRATÉGIE REQUISE — règles:\n"
        "- Utilise UNIQUEMENT SELECT, COUNT, SUM, MIN, MAX, AVG avec GROUP BY simple\n"
        "- Si CREATE TABLE nécessaire: utilise ENGINE = MergeTree() ORDER BY tuple(), "
        "puis un INSERT séparé dans un second SQL\n"
        "- Évite toute fonction complexe (window functions, CTEs imbriquées, argMax, topK, etc.)\n"
        "- Préfère plusieurs requêtes simples enchaînées plutôt qu'une seule requête complexe\n"
        "- DATES: utilise des bornes explicites cohérentes avec la demande utilisateur (pas d'année hardcodée par défaut)\n"
        "- Ajoute LIMIT 1000 pour les explorations\n"
        "- Si l'objectif ne peut pas être atteint simplement, propose l'approximation la plus proche\n\n"
        "Réponds UNIQUEMENT avec:\n"
        "{\n"
        '  "sql": "ta nouvelle requête SQL ici",\n'
        '  "explanation": "en quoi cette nouvelle approche est différente"\n'
        "}"
    )
    user_content = _truncate_text_to_budget(
        user_content,
        int(_get_effective_context_limit() * 0.56),
    )
    try:
        raw = _call_llm(system_prompt, [{"role": "user", "content": user_content}], temperature=0.1)
        result = _parse_llm_json(raw)
    except Exception:
        result = None
    if not result:
        return ""
    return (result.get("sql") or "").strip()


def _cw_generate_suggestions(user_request: str, schema_info: str, plan: dict) -> list:
    """Generate optional follow-up suggestions for the user based on the plan and schema.

    Returns a list of {label, value} dicts (max 5).
    """
    system_prompt = (
        "Tu es un conseiller data expert. En regardant un plan d'exécution ClickHouse, "
        "tu suggères des analyses complémentaires pertinentes que l'utilisateur pourrait vouloir réaliser. "
        "Tu réponds UNIQUEMENT avec du JSON valide, sans markdown."
    )
    # Extract table names from schema info (first line mentions table names)
    user_content = (
        f"Demande originale: {user_request}\n\n"
        f"Schéma disponible:\n{schema_info[:800]}\n\n"
        f"Plan généré:\n{json.dumps({'objective': plan.get('objective'), 'steps': [s['description'] for s in plan.get('steps', [])]}, ensure_ascii=False)}\n\n"
        "Génère 4 à 5 suggestions de suivi pertinentes que l'utilisateur pourrait vouloir faire après cette analyse. "
        "Pense à: tables connexes à explorer, métriques complémentaires, calculs dérivés, comparaisons temporelles, "
        "vérifications de cohérence, agrégations par dimension, création de tables de synthèse.\n\n"
        "Réponds UNIQUEMENT avec:\n"
        '{"suggestions": [{"label": "Courte description action", "value": "Demande complète à envoyer à l\'agent"}]}'
    )
    user_content = _truncate_text_to_budget(
        user_content,
        int(_get_effective_context_limit() * 0.45),
    )
    try:
        raw = _call_llm(system_prompt, [{"role": "user", "content": user_content}], temperature=0.3)
        result = _parse_llm_json(raw)
    except Exception:
        result = None
    if not result:
        return []
    return (result.get("suggestions") or [])[:5]


def _cw_execute_steps(session: dict, client, database: str, preview_rows: int = 5) -> dict:
    """
    Core execution loop: runs plan steps until done, needs user input, or max actions reached.
    Returns a response dict ready to be jsonified.
    """
    plan = session["plan"]
    steps = plan.get("steps", [])
    max_actions = session.get("max_actions", 12)
    max_technical_retries = max(0, int(session.get("max_technical_retries", 3) or 0))

    while (session["action_index"] < len(steps)
           and session["action_count"] < max_actions):

        step = steps[session["action_index"]]

        # ── Generate SQL via LLM ──────────────────────────────────────────
        try:
            sql_result = _cw_generate_sql(session, step, client, database)
        except Exception as exc:
            sql_result = {
                "sql": "",
                "explanation": f"Erreur LLM lors de la génération SQL: {str(exc)}",
                "creates_table": None,
                "needs_user_input": False,
                "question": None,
            }

        # ── Agent needs user input: pause and return question ─────────────
        if sql_result.get("needs_user_input"):
            session["status"] = "asking_user"
            session["pending_step_index"] = session["action_index"]
            return {
                "content": (
                    f"⏸ L'agent a besoin d'une clarification avant l'étape {step['id']}: "
                    f"{step['description']}"
                ),
                "status": "asking_user",
                "plan": plan,
                "action_log": session["action_log"],
                "action_count": session["action_count"],
                "remaining_credits": max_actions - session["action_count"],
                "question": sql_result["question"],
                "created_tables": _cw_filter_bot_tables(session["created_tables"])[0],
            }

        # ── Execute SQL ───────────────────────────────────────────────────
        sql = (sql_result.get("sql") or "").strip()
        entry = {
            "step_id": step["id"],
            "description": step["description"],
            "sql": sql,
            "ok": False,
            "result_preview": None,
            "rows_affected": None,
            "explanation": sql_result.get("explanation", ""),
        }

        if not sql:
            entry["ok"] = False
            entry["result_preview"] = "SQL vide généré par le LLM."
        else:
            # ── Security guard: block any write on non-BOT_ tables ────────
            is_safe, safety_reason = _cw_is_sql_safe(sql)
            if not is_safe:
                entry["ok"] = False
                entry["result_preview"] = safety_reason
                print(f"[SÉCURITÉ] SQL bloqué — {safety_reason}\nSQL: {sql[:300]}")
            else:
                sql_upper = sql.lstrip().upper()
                is_read = any(sql_upper.startswith(k) for k in
                              ("SELECT", "SHOW", "DESCRIBE", "EXPLAIN", "WITH"))
                exec_main = _execute_sql_guarded(
                    sql,
                    read_only=is_read,
                    max_preview_rows=preview_rows,
                    max_execution_time=20,
                    default_limit=2000 if is_read else 1000,
                    hard_limit_cap=10000,
                    client=client,
                )
                if exec_main["ok"]:
                    entry["ok"] = True
                    entry["sql"] = exec_main.get("normalized_sql", sql)
                    if is_read:
                        entry["result_preview"] = exec_main["preview_rows"]
                        entry["rows_affected"] = exec_main["total_rows"]
                    else:
                        entry["result_preview"] = "Commande exécutée avec succès."
                        entry["rows_affected"] = None
                        bot_table = (sql_result.get("creates_table")
                                     or _cw_detect_bot_table(entry["sql"]))
                        if _cw_is_bot_table_name(bot_table):
                            normalized_bot = _cw_normalize_table_name(bot_table)
                            if normalized_bot and normalized_bot not in session["created_tables"]:
                                session["created_tables"].append(normalized_bot)
                else:
                    original_error = exec_main.get("error", exec_main.get("summary", "Unknown error"))
                    error_class = exec_main.get("error_class", _classify_clickhouse_error(original_error))
                    entry["ok"] = False
                    entry["error_class"] = error_class
                    entry["result_preview"] = f"[{error_class}] {original_error}"

                    # ── Technical retry with simpler SQL (free — no credit consumed) ──
                    technical_retries_used = int(session.get("technical_retries", 0) or 0)
                    alt_sql = ""
                    if technical_retries_used < max_technical_retries:
                        alt_sql = _cw_generate_simple_sql(
                            session,
                            step,
                            client,
                            database,
                            sql,
                            f"[{error_class}] {original_error}",
                        )
                        session["technical_retries"] = technical_retries_used + 1
                    else:
                        entry["result_preview"] += (
                            f" | Retry technique ignoré: budget atteint ({max_technical_retries})."
                        )
                    if alt_sql:
                        alt_is_safe, alt_safety_reason = _cw_is_sql_safe(alt_sql)
                        if alt_is_safe:
                            alt_sql_upper = alt_sql.lstrip().upper()
                            alt_is_read = any(alt_sql_upper.startswith(k) for k in
                                              ("SELECT", "SHOW", "DESCRIBE", "EXPLAIN", "WITH"))
                            alt_exec = _execute_sql_guarded(
                                alt_sql,
                                read_only=alt_is_read,
                                max_preview_rows=preview_rows,
                                max_execution_time=20,
                                default_limit=2000 if alt_is_read else 1000,
                                hard_limit_cap=10000,
                                client=client,
                            )
                            if alt_exec["ok"]:
                                entry["ok"] = True
                                entry["sql"] = alt_exec.get("normalized_sql", alt_sql)
                                entry["error_class"] = ""
                                if alt_is_read:
                                    entry["result_preview"] = alt_exec["preview_rows"]
                                    entry["rows_affected"] = alt_exec["total_rows"]
                                else:
                                    entry["result_preview"] = "Commande exécutée avec succès (alternative simplifiée)."
                                    entry["rows_affected"] = None
                                    bot_table = _cw_detect_bot_table(entry["sql"])
                                    if _cw_is_bot_table_name(bot_table):
                                        normalized_bot = _cw_normalize_table_name(bot_table)
                                        if normalized_bot and normalized_bot not in session["created_tables"]:
                                            session["created_tables"].append(normalized_bot)
                                entry["explanation"] = (
                                    entry.get("explanation", "") + " [alternative simplifiée]"
                                )
                            else:
                                alt_error = alt_exec.get("error", alt_exec.get("summary", "unknown"))
                                entry["result_preview"] = (
                                    f"[{error_class}] {original_error}"
                                    f" | Alternative simplifiée aussi échouée: {alt_error}"
                                )
                                remaining_after_current = max_actions - session["action_count"] - 1
                                if remaining_after_current > 0:
                                    rethink_sql = _cw_rethink_strategy(
                                        session, step, client, database, sql, original_error
                                    )
                                    if rethink_sql:
                                        rethink_is_safe, _ = _cw_is_sql_safe(rethink_sql)
                                        if rethink_is_safe:
                                            # Log the failed original entry first (consume its credit)
                                            entry["explanation"] = (
                                                entry.get("explanation", "")
                                                + " [échec — nouvelle stratégie en cours]"
                                            )
                                            session["action_log"].append(entry)
                                            session["action_index"] += 1
                                            session["action_count"] += 1
                                            # Now try the rethought strategy (uses one more credit)
                                            rethink_entry = {
                                                "step_id": step["id"],
                                                "description": f"[Stratégie alternative] {step['description']}",
                                                "sql": rethink_sql,
                                                "ok": False,
                                                "result_preview": None,
                                                "rows_affected": None,
                                                "explanation": "Nouvelle stratégie après deux échecs consécutifs",
                                            }
                                            rt_upper = rethink_sql.lstrip().upper()
                                            rt_is_read = any(rt_upper.startswith(k) for k in
                                                             ("SELECT", "SHOW", "DESCRIBE", "EXPLAIN", "WITH"))
                                            rt_exec = _execute_sql_guarded(
                                                rethink_sql,
                                                read_only=rt_is_read,
                                                max_preview_rows=preview_rows,
                                                max_execution_time=20,
                                                default_limit=2000 if rt_is_read else 1000,
                                                hard_limit_cap=10000,
                                                client=client,
                                            )
                                            if rt_exec["ok"]:
                                                rethink_entry["ok"] = True
                                                rethink_entry["sql"] = rt_exec.get("normalized_sql", rethink_sql)
                                                if rt_is_read:
                                                    rethink_entry["result_preview"] = rt_exec["preview_rows"]
                                                    rethink_entry["rows_affected"] = rt_exec["total_rows"]
                                                else:
                                                    rethink_entry["result_preview"] = "Commande exécutée (stratégie alternative)."
                                                    bot_t = _cw_detect_bot_table(rethink_entry["sql"])
                                                    if _cw_is_bot_table_name(bot_t):
                                                        normalized_bot = _cw_normalize_table_name(bot_t)
                                                        if normalized_bot and normalized_bot not in session["created_tables"]:
                                                            session["created_tables"].append(normalized_bot)
                                            else:
                                                rt_error = rt_exec.get("error", rt_exec.get("summary", "unknown"))
                                                rethink_entry["result_preview"] = (
                                                    f"Échec stratégie alternative: {rt_error}"
                                                )
                                            session["action_log"].append(rethink_entry)
                                            session["action_index"] += 1
                                            session["action_count"] += 1
                                            # Skip normal end-of-loop append/increment
                                            remaining_credits = max_actions - session["action_count"]
                                            if (session["action_count"] % 3 == 0
                                                    and session["action_index"] < len(steps)
                                                    and remaining_credits > 0):
                                                replan = _cw_replan(session, remaining_credits)
                                                replan["checked_after_step"] = session["action_count"]
                                                session["replan_log"].append(replan)
                                                if replan.get("should_replan") and replan.get("new_remaining_steps"):
                                                    completed_steps = [s for s in steps if s["id"] <= step["id"]]
                                                    new_rem = replan["new_remaining_steps"][:remaining_credits]
                                                    plan["steps"] = completed_steps + new_rem
                                                    plan["replan_note"] = (
                                                        f"Réévalué après étape {session['action_count']}: "
                                                        + replan.get("reason", "")
                                                    )
                                                    session["plan"] = plan
                                                    steps = plan["steps"]
                                            continue
                        else:
                            entry["result_preview"] += f" | Alternative rejetée (sécurité): {alt_safety_reason}"

        # Lightweight quality flag to enable earlier replanning when needed.
        if entry["ok"] and isinstance(entry.get("rows_affected"), int) and entry["rows_affected"] == 0:
            entry["quality_flag"] = "empty_result"
        elif not entry["ok"]:
            entry["quality_flag"] = "failed"
        else:
            entry["quality_flag"] = "ok"

        session["action_log"].append(entry)
        session["action_index"] += 1
        session["action_count"] += 1

        # ── Replan check every 3 actions (or earlier if quality degrades) ──
        remaining_credits = max_actions - session["action_count"]
        recent_flags = [e.get("quality_flag") for e in session["action_log"][-2:]]
        force_early_replan = (
            len(recent_flags) == 2
            and all(flag in {"failed", "empty_result"} for flag in recent_flags)
        )
        if (((session["action_count"] % 3 == 0) or force_early_replan)
                and session["action_index"] < len(steps)
                and remaining_credits > 0):
            replan = _cw_replan(session, remaining_credits)
            replan["checked_after_step"] = session["action_count"]
            if force_early_replan:
                reason = replan.get("reason", "")
                replan["reason"] = (
                    ("Déclenchement anticipé (2 étapes faibles/échouées). " + reason).strip()
                )
            session["replan_log"].append(replan)

            if replan.get("should_replan") and replan.get("new_remaining_steps"):
                completed = [s for s in steps if s["id"] <= step["id"]]
                new_remaining = replan["new_remaining_steps"][:remaining_credits]
                plan["steps"] = completed + new_remaining
                plan["replan_note"] = (
                    f"Réévalué après étape {session['action_count']}: "
                    + replan.get("reason", "")
                )
                session["plan"] = plan
                steps = plan["steps"]

    # ── All steps done (or max reached): synthesize ───────────────────────
    allowed_created_tables, skipped_non_bot = _cw_filter_bot_tables(session.get("created_tables", []))
    existing_bot_tables = _cw_list_existing_bot_tables(client, database)
    cleanup_candidates, _ = _cw_filter_bot_tables(existing_bot_tables + allowed_created_tables)
    session["created_tables"] = allowed_created_tables
    session["cleanup_candidates"] = cleanup_candidates
    synthesis = _cw_synthesize(session)
    session["synthesis"] = synthesis
    session["status"] = "awaiting_cleanup"

    cleanup_question = None
    if cleanup_candidates:
        preview_tables = cleanup_candidates[:12]
        table_preview_txt = ", ".join(preview_tables)
        remaining = max(0, len(cleanup_candidates) - len(preview_tables))
        if remaining > 0:
            table_preview_txt += f", ... (+{remaining})"
        cleanup_question = {
            "text": (
                f"{len(cleanup_candidates)} table(s) BOT_* détectée(s): "
                f"{table_preview_txt}. "
                "Souhaitez-vous les supprimer maintenant ? "
                "Vous pouvez aussi écrire des noms précis (ex: BOT_tmp_a, BOT_tmp_b)."
            ),
            "choices": [
                {"label": "Oui — supprimer toutes les tables BOT_", "value": "delete"},
                {"label": "Non — les conserver pour des analyses futures", "value": "keep"},
            ],
        }

    return {
        "content": (
            f"✅ Analyse terminée en {session['action_count']} action(s) "
            f"({'/' + str(max_actions) + ' max'}). "
            "Voici la synthèse complète ci-dessous."
        ),
        "status": "awaiting_cleanup" if cleanup_candidates else "done",
        "plan": plan,
        "action_log": session["action_log"],
        "action_count": session["action_count"],
        "remaining_credits": max_actions - session["action_count"],
        "technical_retries_used": int(session.get("technical_retries", 0) or 0),
        "max_technical_retries": max_technical_retries,
        "synthesis": synthesis,
        "created_tables": allowed_created_tables,
        "cleanup_candidates": cleanup_candidates,
        "skipped_non_bot_tables": skipped_non_bot,
        "replan_log": session["replan_log"],
        "question": cleanup_question,
    }


# ---------------------------------------------------------------------------
# ClickHouse Writer Agent — main handler
# ---------------------------------------------------------------------------

def _run_clickhouse_writer_agent():
    """
    ClickHouse Writer Agent.

    Autonomous agent that plans and executes up to 12 sequential ClickHouse
    operations (read + write), re-evaluates its plan every 3 steps, asks the
    user clarifying questions when needed, creates BOT_* intermediate tables,
    and delivers a comprehensive reflective synthesis.

    Request body (JSON):
        messages      – conversation history [{role, content}]
        params        – agent parameters (database, max_actions, sample_preview)
        session_id    – (optional) resume an existing session

    Response (JSON):
        content        – human-readable message
        status         – planning|executing|asking_user|awaiting_cleanup|done
        plan           – execution plan
        action_log     – per-step execution details
        action_count   – number of actions executed so far
        remaining_credits – actions remaining
        synthesis      – final synthesis (when done)
        created_tables – list of BOT_* tables created
        question       – optional {text, choices} for user interaction
        replan_log     – history of plan re-evaluations
        session_id     – session identifier to include in next request
    """
    data = request.get_json(silent=True) or {}
    messages = data.get("messages", [])
    params = data.get("params", {})
    session_id = data.get("session_id", "")

    # ── Resolve parameters ────────────────────────────────────────────────
    database = (
        (params.get("database") or "").strip()
        or clickhouse_config.get("database", "default")
    )
    try:
        max_actions = max(1, min(int(params.get("max_actions", 12)), 12))
    except (TypeError, ValueError):
        max_actions = 12
    try:
        preview_rows = max(1, min(int(params.get("sample_preview", 5)), 20))
    except (TypeError, ValueError):
        preview_rows = 5
    try:
        max_technical_retries = max(0, min(int(params.get("max_technical_retries", 3)), 8))
    except (TypeError, ValueError):
        max_technical_retries = 3

    # ── Session management ────────────────────────────────────────────────
    if not session_id or session_id not in _writer_sessions:
        session_id = str(uuid.uuid4())
        session = {
            "status": "new",
            "database": database,
            "max_actions": max_actions,
            "plan": None,
            "action_index": 0,
            "action_count": 0,
            "technical_retries": 0,
            "max_technical_retries": max_technical_retries,
            "action_log": [],
            "created_tables": [],
            "replan_log": [],
            "synthesis": None,
            "user_context": {},
            "pending_step_index": None,
        }
        _writer_sessions[session_id] = session
    else:
        session = _writer_sessions[session_id]
    # Keep runtime tunable from the UI on resumed sessions.
    session["database"] = database
    session["max_actions"] = max_actions
    session["max_technical_retries"] = max_technical_retries

    status = session["status"]
    user_message = messages[-1]["content"].strip() if messages else ""

    # ── ClickHouse connection ─────────────────────────────────────────────
    try:
        client = get_clickhouse_client()
    except Exception as exc:
        return jsonify({
            "error": f"Connexion ClickHouse impossible: {exc}",
            "session_id": session_id,
        }), 500

    db = session["database"]

    # ── State machine ─────────────────────────────────────────────────────
    try:
        if status == "new":
            # Phase 1: explore schema & generate plan
            session["status"] = "executing"
            try:
                schema_info = _cw_get_schema_info(client, db)
            except Exception as exc:
                schema_info = f"Erreur lors de la récupération du schéma: {exc}"

            plan = _cw_plan(user_message, db, schema_info)
            session["plan"] = plan

            # Generate optional follow-up suggestions (non-blocking)
            try:
                optional_suggestions = _cw_generate_suggestions(user_message, schema_info, plan)
            except Exception:
                optional_suggestions = []

            result = _cw_execute_steps(session, client, db, preview_rows)
            # Attach suggestions to the first response
            if optional_suggestions:
                result["optional_suggestions"] = optional_suggestions

        elif status in ("executing",):
            result = _cw_execute_steps(session, client, db, preview_rows)

        elif status == "asking_user":
            # User answered a question: store answer and resume
            step_idx = session.get("pending_step_index", session["action_index"])
            session["user_context"][f"answer_step_{step_idx}"] = user_message
            session["status"] = "executing"
            result = _cw_execute_steps(session, client, db, preview_rows)

        elif status == "awaiting_cleanup":
            requested_bot_tables = _cw_extract_bot_table_mentions(user_message)
            affirmative = bool(requested_bot_tables) or any(
                w in user_message.lower()
                for w in ["oui", "yes", "delete", "drop", "supprimer", "cleanup", "clean", "purge", "ok", "confirme", "1"]
            )
            if affirmative:
                existing_bot_tables = _cw_list_existing_bot_tables(client, db)
                resolved = _cw_resolve_cleanup_targets(
                    existing_bot_tables=existing_bot_tables,
                    session_tables=session.get("created_tables", []),
                    requested_tables=requested_bot_tables,
                )
                dropped, errors = _cw_drop_tables(client, db, resolved["targets"])
                session["status"] = "done"
                dropped_upper = {str(t).upper() for t in dropped}
                session["created_tables"] = [
                    t
                    for t in resolved["session_allowed"]
                    if _cw_normalize_table_name(t).upper() not in dropped_upper
                ]
                remaining_bot_tables = _cw_list_existing_bot_tables(client, db)
                session["cleanup_candidates"] = remaining_bot_tables

                dropped_txt = ", ".join(dropped) if dropped else "aucune"
                error_txt = f" (erreurs: {'; '.join(errors)})" if errors else ""
                skipped_txt = (
                    f" Tables ignorées (hors BOT_*): {', '.join(resolved['skipped_non_bot'])}."
                    if resolved["skipped_non_bot"] else ""
                )
                missing_txt = (
                    f" Tables BOT_* introuvables: {', '.join(resolved['not_found'])}."
                    if resolved["not_found"] else ""
                )
                if not resolved["targets"]:
                    content = (
                        "Aucune table BOT_* correspondante à supprimer."
                        f"{missing_txt}{skipped_txt}"
                    )
                else:
                    content = (
                        f"🗑️ Tables supprimées: {dropped_txt}{error_txt}.{missing_txt}{skipped_txt} "
                        "La session est terminée."
                    )
                result = {
                    "content": content,
                    "status": "done",
                    "cleanup_done": bool(resolved["targets"]),
                    "tables_dropped": dropped,
                    "skipped_non_bot_tables": resolved["skipped_non_bot"],
                    "not_found_tables": resolved["not_found"],
                    "remaining_bot_tables": remaining_bot_tables,
                    "technical_retries_used": int(session.get("technical_retries", 0) or 0),
                    "max_technical_retries": int(session.get("max_technical_retries", 3) or 3),
                    "synthesis": session.get("synthesis"),
                    "plan": session.get("plan"),
                    "action_log": session["action_log"],
                }
            else:
                session["status"] = "done"
                kept_tables, _ = _cw_filter_bot_tables(session.get("created_tables", []))
                session["created_tables"] = kept_tables
                result = {
                    "content": (
                        "✅ Tables BOT_ conservées. Vous pouvez les utiliser pour des analyses futures. "
                        "La session est terminée."
                    ),
                    "status": "done",
                    "cleanup_done": False,
                    "created_tables": kept_tables,
                    "technical_retries_used": int(session.get("technical_retries", 0) or 0),
                    "max_technical_retries": int(session.get("max_technical_retries", 3) or 3),
                    "synthesis": session.get("synthesis"),
                    "plan": session.get("plan"),
                    "action_log": session["action_log"],
                }

        elif status == "done":
            requested_bot_tables = _cw_extract_bot_table_mentions(user_message)
            cleanup_intent = bool(requested_bot_tables) or any(
                w in user_message.lower()
                for w in ["delete", "drop", "supprimer", "cleanup", "clean", "purge"]
            )
            if cleanup_intent:
                existing_bot_tables = _cw_list_existing_bot_tables(client, db)
                resolved = _cw_resolve_cleanup_targets(
                    existing_bot_tables=existing_bot_tables,
                    session_tables=session.get("created_tables", []),
                    requested_tables=requested_bot_tables,
                )
                dropped, errors = _cw_drop_tables(client, db, resolved["targets"])
                dropped_upper = {str(t).upper() for t in dropped}
                session["created_tables"] = [
                    t
                    for t in resolved["session_allowed"]
                    if _cw_normalize_table_name(t).upper() not in dropped_upper
                ]
                session["cleanup_candidates"] = _cw_list_existing_bot_tables(client, db)
                dropped_txt = ", ".join(dropped) if dropped else "aucune"
                error_txt = f" (erreurs: {'; '.join(errors)})" if errors else ""
                missing_txt = (
                    f" Tables BOT_* introuvables: {', '.join(resolved['not_found'])}."
                    if resolved["not_found"] else ""
                )
                result = {
                    "content": (
                        f"🧹 Cleanup BOT_* exécuté. Tables supprimées: {dropped_txt}{error_txt}.{missing_txt}"
                        if resolved["targets"]
                        else f"Aucune table BOT_* correspondante à supprimer.{missing_txt}"
                    ),
                    "status": "done",
                    "cleanup_done": bool(resolved["targets"]),
                    "tables_dropped": dropped,
                    "not_found_tables": resolved["not_found"],
                    "remaining_bot_tables": session.get("cleanup_candidates", []),
                    "technical_retries_used": int(session.get("technical_retries", 0) or 0),
                    "max_technical_retries": int(session.get("max_technical_retries", 3) or 3),
                    "synthesis": session.get("synthesis"),
                    "plan": session.get("plan"),
                    "action_log": session.get("action_log", []),
                }
            else:
                result = {
                    "content": "Session terminée. Démarrez une nouvelle conversation.",
                    "status": "done",
                }

        else:
            result = {
                "content": "Session terminée. Démarrez une nouvelle conversation.",
                "status": "done",
            }

    except Exception as exc:
        # Any unhandled exception (LLM failure, JSON parse error, etc.) must return
        # a proper JSON error so the frontend doesn't show a generic "Erreur de connexion".
        session["status"] = "error"
        _writer_sessions[session_id] = session
        return jsonify({
            "error": f"Erreur interne de l'agent: {str(exc)}",
            "session_id": session_id,
            "status": "error",
        }), 500

    _writer_sessions[session_id] = session
    result["session_id"] = session_id
    return jsonify(result)


def _run_key_identifier_agent():
    """
    Key Identifier Agent.

    Scans all tables and columns in ClickHouse, samples up to 5 distinct non-null
    values per column, then uses the local LLM to identify potential foreign-key
    relationships by field-name similarity and value-type matching.

    Request body (JSON):
        params:
            database    : str  – database name (empty = default)
            sample_size : int  – distinct values to sample per column (1–5, default 5)
            confidence  : str  – "low" | "medium" | "high" (default "medium")

    Response (JSON):
        steps       – per-table analysis log
        suggestions – list of FK candidates with table_a/field_a/table_b/field_b/direction/confidence/reason
        total_fields – total fields analysed
    """
    data = request.get_json(silent=True) or {}
    params = data.get("params", {})

    database = (params.get("database") or "").strip() or clickhouse_config.get("database", "default")
    try:
        sample_size = max(1, min(int(params.get("sample_size", 5)), 5))
    except (TypeError, ValueError):
        sample_size = 5
    confidence_threshold = params.get("confidence", "medium")

    # ── Connect ──────────────────────────────────────────────────────────────
    try:
        client = get_clickhouse_client()
    except Exception as exc:
        return jsonify({"error": f"Impossible de se connecter à ClickHouse : {exc}"}), 500

    # ── List tables ──────────────────────────────────────────────────────────
    try:
        res = client.query(f"SHOW TABLES FROM `{database}`")
        table_list = [row[0] for row in res.result_rows]
    except Exception as exc:
        return jsonify({"error": f"Impossible de lister les tables de '{database}' : {exc}"}), 500

    if not table_list:
        return jsonify({"error": f"Aucune table trouvée dans la base '{database}'."}), 404

    # ── Heuristic: field names that often carry key semantics ─────────────────
    KEY_PATTERNS = (
        "id", "key", "code", "ref", "fk", "pk", "num", "no", "number",
        "uuid", "guid", "hash", "token", "identifier",
    )
    KEY_TYPES = (
        "int", "uint", "int8", "int16", "int32", "int64",
        "uint8", "uint16", "uint32", "uint64",
        "uuid", "fixedstring", "string",
    )

    def _is_key_field(name: str, col_type: str) -> bool:
        n = name.lower()
        t = col_type.lower().split("(")[0]
        name_matches = any(p in n for p in KEY_PATTERNS)
        type_matches = any(t.startswith(k) for k in KEY_TYPES)
        return name_matches or type_matches

    steps = []
    # field_catalog: list of {table, field, type, samples}
    field_catalog = []

    for table_name in table_list:
        try:
            schema_res = client.query(f"DESCRIBE TABLE `{database}`.`{table_name}`")
            columns = [(row[0], row[1]) for row in schema_res.result_rows]
        except Exception as exc:
            steps.append({"table": table_name, "ok": False, "error": str(exc)})
            continue

        # Sample all rows at once to minimise round-trips
        try:
            sample_res = client.query(
                f"SELECT * FROM `{database}`.`{table_name}` LIMIT {sample_size * 3}"
            )
            # column_names is a tuple/list of strings — use directly, not c[0]
            raw_names = sample_res.column_names if hasattr(sample_res, 'column_names') else None
            col_names = list(raw_names) if raw_names is not None else [c[0] for c in columns]
            rows = sample_res.result_rows
        except Exception:
            rows = []
            col_names = [c[0] for c in columns]

        # Build per-column value sets from sampled rows
        col_samples: dict[str, set] = {c[0]: set() for c in columns}
        try:
            for row in rows:
                for i, val in enumerate(row):
                    if i < len(col_names) and val is not None and str(val).strip() != "":
                        col_key = col_names[i]
                        if col_key in col_samples:
                            col_samples[col_key].add(str(val))
        except Exception:
            pass  # Sampling failure is non-fatal

        kept = 0
        for col_name, col_type in columns:
            if not _is_key_field(col_name, col_type):
                continue
            samples = sorted(list(col_samples.get(col_name, set())))[:sample_size]
            field_catalog.append({
                "table": table_name,
                "field": col_name,
                "type": col_type,
                "samples": samples,
            })
            kept += 1

        steps.append({"table": table_name, "ok": True, "fields_kept": kept, "total_fields": len(columns)})

    if len(field_catalog) < 2:
        return jsonify({
            "steps": steps,
            "suggestions": [],
            "total_fields": len(field_catalog),
            "message": "Pas assez de champs candidats trouvés pour établir des relations.",
        })

    # ── Build compact prompt for the LLM ─────────────────────────────────────
    # Format: TABLE.FIELD [TYPE]: val1, val2, ...
    lines = []
    for f in field_catalog:
        sample_str = ", ".join(f["samples"]) if f["samples"] else "(no sample)"
        lines.append(f"{f['table']}.{f['field']} [{f['type']}]: {sample_str}")

    fields_block = "\n".join(lines)

    confidence_instruction = {
        "low": "Include low, medium and high confidence matches.",
        "medium": "Include only medium and high confidence matches.",
        "high": "Include only high confidence matches.",
    }.get(confidence_threshold, "Include only medium and high confidence matches.")

    system_prompt = (
        "You are a database analyst expert in identifying foreign key relationships. "
        "Analyse field names, types and sample values to find FK→PK links between tables. "
        "Reply with valid JSON only — no markdown, no extra text."
    )

    user_content = (
        f"Database: {database}\n\n"
        "Fields (TABLE.FIELD [TYPE]: sample_values):\n"
        f"{fields_block}\n\n"
        f"Task: identify potential foreign-key relationships. {confidence_instruction}\n"
        "For each relationship, determine which field references which (the FK side has values "
        "that are a subset of the PK side).\n\n"
        "Return ONLY a JSON array (max 20 items):\n"
        '[\n'
        '  {\n'
        '    "table_a": "source_table",\n'
        '    "field_a": "fk_field",\n'
        '    "table_b": "target_table",\n'
        '    "field_b": "pk_field",\n'
        '    "direction": "table_a.field_a → table_b.field_b  (field_a references field_b)",\n'
        '    "confidence": "high|medium|low",\n'
        '    "reason": "brief explanation (field name similarity / value overlap / type match)"\n'
        '  }\n'
        ']'
    )

    try:
        llm_raw = _call_llm(
            system_prompt,
            [{"role": "user", "content": user_content}],
            temperature=0.1,
        )
        try:
            suggestions = json.loads(llm_raw)
            if not isinstance(suggestions, list):
                suggestions = []
        except json.JSONDecodeError:
            # Try to extract JSON array from response
            import re as _re
            m = _re.search(r"\[.*\]", llm_raw, _re.DOTALL)
            suggestions = json.loads(m.group(0)) if m else []
    except Exception as exc:
        return jsonify({"error": f"LLM error: {exc}", "steps": steps}), 500

    # Normalise and filter
    clean_suggestions = []
    for s in suggestions:
        if not isinstance(s, dict):
            continue
        if not all(k in s for k in ("table_a", "field_a", "table_b", "field_b")):
            continue
        # Filter by confidence threshold
        conf = s.get("confidence", "medium").lower()
        if confidence_threshold == "high" and conf != "high":
            continue
        if confidence_threshold == "medium" and conf == "low":
            continue
        clean_suggestions.append({
            "table_a": str(s.get("table_a", "")),
            "field_a": str(s.get("field_a", "")),
            "table_b": str(s.get("table_b", "")),
            "field_b": str(s.get("field_b", "")),
            "direction": str(s.get("direction", "")),
            "confidence": conf,
            "reason": str(s.get("reason", "")),
        })

    return jsonify({
        "steps": steps,
        "suggestions": clean_suggestions,
        "total_fields": len(field_catalog),
    })


def _run_data_dictionary_agent():
    """
    Data Dictionary Generator agent.

    Connects to the configured ClickHouse instance, lists/filters tables,
    fetches their schema and a small sample of rows, then calls the local LLM
    to produce a structured, business-friendly data dictionary.

    Request body (JSON):
        messages  – conversation history (list of {role, content})
        params    – agent parameters:
            database    : str   – database name (empty = clickhouse_config default)
            tables      : str   – comma-separated table filter (empty = all tables)
            language    : str   – "fr" | "en"
            sample_rows : int   – rows sampled per table (default 5)

    Response (JSON):
        steps            – per-table processing log
        data_dictionary  – list of documented tables
        tables_processed – count of successfully documented tables
        total_tables     – total tables attempted
    """
    data = request.get_json(silent=True) or {}
    params = data.get("params", {})

    database = (params.get("database") or "").strip() or clickhouse_config.get("database", "default")
    tables_filter = (params.get("tables") or "").strip()
    language = params.get("language", "fr")
    try:
        sample_rows = max(1, min(int(params.get("sample_rows", 5)), 20))
    except (TypeError, ValueError):
        sample_rows = 5

    lang_label = "français" if language == "fr" else "English"

    try:
        client = get_clickhouse_client()
    except Exception as exc:
        return jsonify({"error": f"Impossible de se connecter à ClickHouse : {exc}"}), 500

    # Build table list
    if tables_filter:
        table_list = [t.strip() for t in tables_filter.split(",") if t.strip()]
    else:
        try:
            res = client.query(f"SHOW TABLES FROM `{database}`")
            table_list = [row[0] for row in res.result_rows]
        except Exception as exc:
            return jsonify({"error": f"Impossible de lister les tables de '{database}' : {exc}"}), 500

    if not table_list:
        return jsonify({"error": f"Aucune table trouvée dans la base '{database}'."}), 404

    steps = []
    dict_entries = []

    system_prompt = (
        "You are a senior data engineer specialized in writing clear, precise data dictionaries "
        "for business users. You produce concise, accurate, and helpful documentation. "
        f"All descriptions MUST be written in {lang_label}. "
        "Always reply with valid JSON only — no markdown fences, no extra text."
    )

    for table_name in table_list:
        # ── Fetch schema ──────────────────────────────────────────────────
        try:
            schema_res = client.query(f"DESCRIBE TABLE `{database}`.`{table_name}`")
            # DESCRIBE returns: name, type, default_type, default_expression,
            #                   comment, codec_expression, ttl_expression, ...
            columns = []
            for row in schema_res.result_rows:
                col = {
                    "name": row[0],
                    "type": row[1],
                    "comment": row[4] if len(row) > 4 else "",
                }
                columns.append(col)
        except Exception as exc:
            steps.append({"table": table_name, "ok": False, "error": str(exc)})
            continue

        # ── Sample data ───────────────────────────────────────────────────
        sample_data = []
        try:
            sample_res = client.query(
                f"SELECT * FROM `{database}`.`{table_name}` LIMIT {sample_rows}"
            )
            sample_data = _rows_to_dicts(sample_res)
        except Exception:
            pass  # proceed without sample data

        # ── Call LLM ─────────────────────────────────────────────────────
        user_content = (
            f"Table: {database}.{table_name}\n\n"
            f"Columns:\n{json.dumps(columns, ensure_ascii=False, indent=2)}\n\n"
            + (
                f"Sample rows ({len(sample_data)}):\n"
                f"{json.dumps(sample_data, ensure_ascii=False, default=str, indent=2)}\n\n"
                if sample_data else ""
            )
            + "Generate a complete data dictionary for this table.\n"
            "Return ONLY a JSON object with this exact structure:\n"
            "{\n"
            '  "table_description": "<business description of the table>",\n'
            '  "columns": [\n'
            "    {\n"
            '      "name": "<column name>",\n'
            '      "type": "<clickhouse type>",\n'
            '      "business_description": "<plain-language meaning>",\n'
            '      "format": "<expected format or unit>",\n'
            '      "possible_values": "<enum values, ranges, or free-form description>"\n'
            "    }\n"
            "  ]\n"
            "}"
        )

        try:
            llm_raw = _call_llm(
                system_prompt,
                [{"role": "user", "content": user_content}],
                temperature=0.2,
            )
            try:
                doc = json.loads(llm_raw)
            except json.JSONDecodeError:
                # Fallback: wrap raw text
                doc = {
                    "table_description": llm_raw,
                    "columns": [
                        {"name": c["name"], "type": c["type"],
                         "business_description": c.get("comment", ""),
                         "format": "", "possible_values": ""}
                        for c in columns
                    ],
                }
            doc["table"] = f"{database}.{table_name}"
            dict_entries.append(doc)
            steps.append({"table": table_name, "ok": True, "columns_count": len(columns)})
        except Exception as exc:
            steps.append({"table": table_name, "ok": False, "error": str(exc)})

    tables_processed = sum(1 for s in steps if s.get("ok"))

    return jsonify({
        "steps": steps,
        "data_dictionary": dict_entries,
        "tables_processed": tables_processed,
        "total_tables": len(table_list),
    })


# ===========================================================================
# ETL Agent — helper functions
# ===========================================================================

_ETL_EXTENSIONS = {".csv", ".tsv", ".txt", ".xlsx", ".xls", ".parquet", ".json"}


def _etl_list_files(folder_path: str, recursive: bool = False) -> list[dict]:
    """Return a list of supported data files found in the folder."""
    import pathlib
    root = pathlib.Path(folder_path)
    results = []
    pattern = "**/*" if recursive else "*"
    for p in sorted(root.glob(pattern)):
        if p.is_file() and p.suffix.lower() in _ETL_EXTENSIONS:
            try:
                size_bytes = p.stat().st_size
            except Exception:
                size_bytes = 0
            results.append({
                "path": str(p),
                "name": p.name,
                "relative": str(p.relative_to(root)),
                "extension": p.suffix.lower(),
                "size_bytes": size_bytes,
                "size_human": _etl_human_size(size_bytes),
            })
    return results


def _etl_human_size(size: int) -> str:
    for unit in ("o", "Ko", "Mo", "Go"):
        if size < 1024:
            return f"{size:.0f} {unit}"
        size /= 1024
    return f"{size:.1f} To"


def _etl_parse_file(file_path: str, max_rows: int = 5000) -> tuple[list[str], list[dict], str]:
    """
    Parse a data file and return (column_names, sample_rows_as_dicts, error).
    Supports CSV/TSV/TXT, Excel, Parquet, JSON.
    Returns up to max_rows rows.
    """
    import pathlib
    ext = pathlib.Path(file_path).suffix.lower()
    try:
        import pandas as pd
        if ext in (".csv", ".tsv", ".txt"):
            sep = "\t" if ext in (".tsv", ".txt") else ","
            # Try comma first for txt
            try:
                df = pd.read_csv(file_path, sep=sep, nrows=max_rows, low_memory=False)
                if len(df.columns) == 1 and ext == ".txt":
                    df = pd.read_csv(file_path, sep=",", nrows=max_rows, low_memory=False)
            except Exception:
                df = pd.read_csv(file_path, sep=None, engine="python", nrows=max_rows, low_memory=False)
        elif ext in (".xlsx", ".xls"):
            df = pd.read_excel(file_path, nrows=max_rows)
        elif ext == ".parquet":
            df = pd.read_parquet(file_path)
            if len(df) > max_rows:
                df = df.head(max_rows)
        elif ext == ".json":
            df = pd.read_json(file_path, lines=True) if _etl_is_jsonl(file_path) else pd.read_json(file_path)
            if len(df) > max_rows:
                df = df.head(max_rows)
        else:
            return [], [], f"Extension non supportée: {ext}"

        # Clean column names
        df.columns = [str(c).strip().replace(" ", "_").replace("-", "_") for c in df.columns]
        columns = list(df.columns)
        rows = df.head(5).astype(str).replace("nan", "").to_dict(orient="records")
        return columns, rows, ""
    except Exception as exc:
        return [], [], str(exc)


def _etl_parse_file_full(file_path: str) -> tuple[list[str], list[dict], int, str]:
    """
    Parse a data file fully and return (column_names, all_rows_as_dicts, row_count, error).
    """
    import pathlib
    ext = pathlib.Path(file_path).suffix.lower()
    try:
        import pandas as pd
        if ext in (".csv", ".tsv", ".txt"):
            sep = "\t" if ext in (".tsv", ".txt") else ","
            try:
                df = pd.read_csv(file_path, sep=sep, low_memory=False)
                if len(df.columns) == 1 and ext == ".txt":
                    df = pd.read_csv(file_path, sep=",", low_memory=False)
            except Exception:
                df = pd.read_csv(file_path, sep=None, engine="python", low_memory=False)
        elif ext in (".xlsx", ".xls"):
            df = pd.read_excel(file_path)
        elif ext == ".parquet":
            df = pd.read_parquet(file_path)
        elif ext == ".json":
            df = pd.read_json(file_path, lines=True) if _etl_is_jsonl(file_path) else pd.read_json(file_path)
        else:
            return [], [], 0, f"Extension non supportée: {ext}"

        df.columns = [str(c).strip().replace(" ", "_").replace("-", "_") for c in df.columns]
        # Replace NaN with None for JSON serialization
        df = df.where(df.notna(), None)
        columns = list(df.columns)
        rows = df.to_dict(orient="records")
        return columns, rows, len(rows), ""
    except Exception as exc:
        return [], [], 0, str(exc)


def _etl_is_jsonl(file_path: str) -> bool:
    """Check if a JSON file is in JSONL format (one JSON object per line)."""
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            first_line = f.readline().strip()
            return first_line.startswith("{")
    except Exception:
        return False


def _etl_infer_ch_type(series_dtype, sample_values: list) -> str:
    """Infer a ClickHouse column type from a pandas dtype and sample values."""
    dtype_str = str(series_dtype).lower()
    if "int" in dtype_str:
        return "Int64"
    if "float" in dtype_str:
        return "Float64"
    if "bool" in dtype_str:
        return "UInt8"
    if "datetime" in dtype_str or "timestamp" in dtype_str:
        return "DateTime"
    if "date" in dtype_str:
        return "Date"
    # Try to detect from sample values
    non_null = [v for v in sample_values if v is not None and str(v).strip() not in ("", "nan", "None")]
    if non_null:
        # Integer detection
        try:
            all(int(str(v)) == float(str(v)) for v in non_null[:10])
            return "Int64"
        except (ValueError, TypeError):
            pass
        # Float detection
        try:
            all(float(str(v)) for v in non_null[:10])
            return "Float64"
        except (ValueError, TypeError):
            pass
    return "String"


def _etl_infer_schema_from_file(file_path: str) -> tuple[list[dict], str]:
    """
    Return a list of {name, ch_type, py_dtype} for each column in the file.
    Returns (schema, error).
    """
    import pathlib
    ext = pathlib.Path(file_path).suffix.lower()
    try:
        import pandas as pd
        if ext in (".csv", ".tsv", ".txt"):
            sep = "\t" if ext in (".tsv", ".txt") else ","
            try:
                df = pd.read_csv(file_path, sep=sep, nrows=200, low_memory=False)
            except Exception:
                df = pd.read_csv(file_path, sep=None, engine="python", nrows=200, low_memory=False)
        elif ext in (".xlsx", ".xls"):
            df = pd.read_excel(file_path, nrows=200)
        elif ext == ".parquet":
            df = pd.read_parquet(file_path).head(200)
        elif ext == ".json":
            df = pd.read_json(file_path, lines=True).head(200) if _etl_is_jsonl(file_path) else pd.read_json(file_path).head(200)
        else:
            return [], f"Extension non supportée: {ext}"

        df.columns = [str(c).strip().replace(" ", "_").replace("-", "_") for c in df.columns]
        schema = []
        for col in df.columns:
            sample = df[col].dropna().head(10).tolist()
            ch_type = _etl_infer_ch_type(df[col].dtype, sample)
            schema.append({"name": col, "ch_type": ch_type, "py_dtype": str(df[col].dtype)})
        return schema, ""
    except Exception as exc:
        return [], str(exc)


def _etl_is_bot_etl_table(name: str) -> bool:
    return name.upper().startswith("BOT_ETL_")


def _etl_safe_table_check(table_name: str) -> tuple[bool, str]:
    """Ensure the table name starts with BOT_ETL_."""
    if not _etl_is_bot_etl_table(table_name):
        return False, (
            f"[SÉCURITÉ ETL] Table '{table_name}' refusée. "
            "L'agent ETL ne peut créer/modifier que des tables préfixées BOT_ETL_."
        )
    return True, ""


def _etl_get_db_schema(client, database: str) -> str:
    """Return a compact schema string (existing tables + column info)."""
    try:
        res = client.query(f"SHOW TABLES FROM `{database}`")
        tables = [row[0] for row in res.result_rows]
    except Exception as exc:
        return f"Impossible de lister les tables: {exc}"
    parts = []
    for tbl in tables[:40]:
        try:
            desc = client.query(f"DESCRIBE TABLE `{database}`.`{tbl}`")
            cols = ", ".join(f"{r[0]}:{r[1]}" for r in desc.result_rows[:20])
            parts.append(f"TABLE {tbl}: {cols}")
        except Exception:
            parts.append(f"TABLE {tbl}: (erreur schema)")
    if len(tables) > 40:
        parts.append(f"... et {len(tables) - 40} autres tables.")
    return "\n".join(parts)


def _etl_get_knowledge_context() -> str:
    """Return knowledge base content as context string."""
    global knowledge_base
    if not knowledge_base:
        return ""
    parts = []
    for folder in knowledge_base[:5]:
        title = folder.get("title", "")
        content = folder.get("content", "")[:500]
        if title or content:
            parts.append(f"[KB: {title}]\n{content}")
    return "\n\n".join(parts)


def _etl_plan(session: dict, files_info: list[dict], db_schema: str, kb_context: str) -> dict:
    """LLM generates an ETL plan from files and user request."""
    user_request = session["user_request"]
    max_actions = session["max_actions"]
    database = session["database"]

    files_summary = []
    for f in files_info[:20]:
        cols_str = ", ".join(f.get("columns", [])[:10])
        if len(f.get("columns", [])) > 10:
            cols_str += f" ... (+{len(f['columns'])-10} autres)"
        sample_str = json.dumps(f.get("sample_rows", [])[:2], ensure_ascii=False, default=str)[:300]
        files_summary.append(
            f"Fichier: {f['relative']} ({f['extension']}, {f.get('row_count','?')} lignes, {f.get('size_human','')})\n"
            f"  Colonnes ({len(f.get('columns',[]))}): {cols_str}\n"
            f"  Exemples: {sample_str}"
        )

    system_prompt = (
        "Tu es un expert ETL et architecte de données. "
        "Tu crées des plans d'intégration de données précis et adaptés à la demande. "
        "Tu réponds UNIQUEMENT avec du JSON valide, sans markdown ni texte supplémentaire."
    )
    user_content = (
        f"Base de données cible: {database}\n\n"
        f"Schema ClickHouse existant:\n{db_schema}\n\n"
        f"Fichiers disponibles:\n" + "\n\n".join(files_summary) + "\n\n"
        + (f"Contexte Knowledge Base:\n{kb_context}\n\n" if kb_context else "")
        + f"Demande de l'utilisateur: {user_request}\n\n"
        f"Crée un plan ETL avec AU MAXIMUM {max_actions} étapes.\n\n"
        "RÈGLES OBLIGATOIRES:\n"
        "- Toutes les tables créées DOIVENT commencer par BOT_ETL_ (ex: BOT_ETL_clients)\n"
        "- Tables intermédiaires autorisées (ex: BOT_ETL_tmp_calcul)\n"
        "- Champs calculés préfixés C_* (ex: C_montant_total, C_client_adresse)\n"
        "- L'enrichissement depuis des tables existantes est possible via JOIN\n"
        "- En cas d'ambiguïté (choix de table, clé étrangère, mapping colonne), "
        "  marque l'étape avec needs_user_input=true\n\n"
        "Types d'étapes disponibles:\n"
        "  - browse_files: lister/confirmer les fichiers\n"
        "  - parse_file: lire un fichier et préparer les données\n"
        "  - create_table: créer une table BOT_ETL_*\n"
        "  - insert_data: insérer des données dans une table BOT_ETL_*\n"
        "  - add_calculated_field: ajouter un champ C_* via ALTER + UPDATE ou CREATE TABLE AS SELECT\n"
        "  - enrich_from_table: enrichir depuis une table existante via JOIN\n"
        "  - delete_rows: supprimer des lignes d'une table BOT_ETL_*\n"
        "  - drop_table: supprimer une table BOT_ETL_* (confirmation user requise)\n"
        "  - verify: vérification/contrôle qualité\n"
        "  - ask_user: poser une question à l'utilisateur\n\n"
        "Réponds UNIQUEMENT avec ce JSON:\n"
        "{\n"
        '  "objective": "description de l\'objectif ETL",\n'
        '  "approach": "approche haut niveau",\n'
        '  "tables_to_create": ["BOT_ETL_nom1", "BOT_ETL_nom2"],\n'
        '  "estimated_steps": N,\n'
        '  "steps": [\n'
        "    {\n"
        '      "id": 1,\n'
        '      "type": "create_table|insert_data|add_calculated_field|...",\n'
        '      "description": "description de l\'étape",\n'
        '      "rationale": "pourquoi cette étape",\n'
        '      "target_table": "BOT_ETL_nom ou null",\n'
        '      "source_file": "chemin/fichier ou null",\n'
        '      "needs_user_input": false,\n'
        '      "question": null\n'
        "    }\n"
        "  ]\n"
        "}"
    )
    raw = _call_llm(system_prompt, [{"role": "user", "content": user_content}], temperature=0.1)
    result = _parse_llm_json(raw)
    if not result or "steps" not in result:
        result = {
            "objective": user_request,
            "approach": "Import direct des fichiers vers ClickHouse.",
            "tables_to_create": [],
            "estimated_steps": 1,
            "steps": [{"id": 1, "type": "ask_user", "description": "Clarifier la demande",
                        "rationale": "Plan non généré", "target_table": None,
                        "source_file": None, "needs_user_input": True, "question": None}],
        }
    result["steps"] = result["steps"][:max_actions]
    return result


def _etl_generate_action(session: dict, step: dict, client, database: str) -> dict:
    """LLM generates the concrete action (SQL or question) for a given ETL step."""
    plan = session["plan"]
    action_log = session["action_log"]
    files_info = session.get("files_info", [])

    prev_summaries = []
    for entry in action_log[-4:]:
        status_str = "✓ OK" if entry["ok"] else "✗ ECHEC"
        prev_summaries.append(
            f"Étape {entry['step_id']} ({status_str}): {entry['description']}\n"
            f"  Action: {str(entry.get('action_detail',''))[:200]}\n"
            f"  Résultat: {str(entry.get('result_preview',''))[:200]}"
        )

    try:
        res = client.query(f"SHOW TABLES FROM `{database}`")
        all_tables = [row[0] for row in res.result_rows]
    except Exception:
        all_tables = []

    bot_etl_tables = [t for t in all_tables if t.upper().startswith("BOT_ETL_")]

    files_summary = []
    for f in files_info[:10]:
        cols_str = ", ".join(f.get("columns", [])[:8])
        files_summary.append(f"{f['relative']} → colonnes: {cols_str}")

    system_prompt = (
        "Tu es un expert ETL ClickHouse. Tu génères les actions concrètes pour chaque étape du plan ETL. "
        "Tu réponds UNIQUEMENT avec du JSON valide, sans markdown ni texte supplémentaire."
    )
    user_content = (
        f"Base: {database}\n"
        f"Tables existantes: {', '.join(all_tables[:30])}\n"
        f"Tables BOT_ETL_ créées: {', '.join(bot_etl_tables)}\n\n"
        f"Fichiers disponibles:\n" + "\n".join(files_summary) + "\n\n"
        f"Objectif global: {plan.get('objective','')}\n\n"
        f"Étape actuelle {step['id']}: {step['description']}\n"
        f"Type: {step.get('type','')}\n"
        f"Table cible: {step.get('target_table','')}\n"
        f"Fichier source: {step.get('source_file','')}\n"
        f"Raison: {step.get('rationale','')}\n\n"
        f"Résultats précédents:\n"
        + ("\n".join(prev_summaries) if prev_summaries else "(première étape)")
        + "\n\n"
        "Génère l'action pour cette étape ETL.\n"
        "RÈGLES:\n"
        "- CREATE TABLE → uniquement des tables BOT_ETL_*\n"
        "- INSERT/DELETE/DROP/ALTER → uniquement des tables BOT_ETL_*\n"
        "- Champs calculés → préfixe C_*\n"
        "- Pour create_table: utilise ENGINE = MergeTree() ORDER BY tuple()\n"
        "- Pour insert_data: l'agent va insérer les données via pandas, "
        "  fournis juste le CREATE TABLE SQL si la table n'existe pas encore\n"
        "- Pour add_calculated_field: fournis le SQL ALTER TABLE ADD COLUMN + "
        "  UPDATE ou un CREATE TABLE AS SELECT avec le champ C_* calculé\n"
        "- Pour enrich_from_table: fournis le SQL INSERT INTO BOT_ETL_... SELECT ... FROM ... JOIN ...\n"
        "- Pour verify: fournis un SELECT COUNT/DISTINCT pour vérification\n"
        "- Pour drop_table: needs_user_input doit être true\n"
        "- Si tu as besoin d'info de l'utilisateur: needs_user_input=true avec choices\n\n"
        "Réponds UNIQUEMENT avec ce JSON:\n"
        "{\n"
        '  "sql": "SQL ClickHouse ou null",\n'
        '  "explanation": "ce que fait cette action",\n'
        '  "action_type": "create_table|insert_data|verify|ask_user|...",\n'
        '  "target_table": "BOT_ETL_nom ou null",\n'
        '  "needs_user_input": false,\n'
        '  "question": null\n'
        "}\n"
        "Si needs_user_input=true, mets sql=null et question={\n"
        '  "text": "Question à l\'utilisateur",\n'
        '  "choices": [{"label": "Option A", "value": "a"}, {"label": "Option B", "value": "b"}]\n'
        "}"
    )
    raw = _call_llm(system_prompt, [{"role": "user", "content": user_content}], temperature=0.05)
    result = _parse_llm_json(raw)
    if not result:
        return {
            "sql": None,
            "explanation": "Fallback — parsing LLM échoué",
            "action_type": step.get("type", "verify"),
            "target_table": step.get("target_table"),
            "needs_user_input": False,
            "question": None,
        }
    return result


def _etl_generate_action_with_error(session: dict, step: dict, client, database: str, error: str) -> dict:
    """LLM generates an alternative action when the previous SQL attempt failed."""
    plan = session["plan"]
    action_log = session["action_log"]
    files_info = session.get("files_info", [])

    try:
        res = client.query(f"SHOW TABLES FROM `{database}`")
        all_tables = [row[0] for row in res.result_rows]
    except Exception:
        all_tables = []

    files_summary = []
    for f in files_info[:5]:
        cols_str = ", ".join(f.get("columns", [])[:6])
        files_summary.append(f"{f['relative']} → colonnes: {cols_str}")

    last_error = step.get("_last_error", error)
    last_sql = step.get("_last_sql", "")

    system_prompt = (
        "Tu es un expert ETL ClickHouse. La dernière tentative d'action a échoué. "
        "Tu dois proposer une approche alternative pour accomplir le même objectif. "
        "Tu réponds UNIQUEMENT avec du JSON valide, sans markdown ni texte supplémentaire."
    )
    user_content = (
        f"Base: {database}\n"
        f"Tables existantes: {', '.join(all_tables[:20])}\n\n"
        f"Fichiers: {chr(10).join(files_summary)}\n\n"
        f"Objectif global: {plan.get('objective', '')}\n\n"
        f"Étape {step['id']}: {step['description']}\n"
        f"Type: {step.get('type', '')}\n"
        f"Table cible: {step.get('target_table', '')}\n\n"
        f"SQL qui a échoué:\n{last_sql[:300]}\n\n"
        f"Erreur obtenue:\n{last_error[:300]}\n\n"
        f"Génère une approche alternative pour cette étape ETL.\n"
        "RÈGLES:\n"
        "- CREATE TABLE → uniquement des tables BOT_ETL_*\n"
        "- Utilise ENGINE = MergeTree() ORDER BY tuple()\n"
        "- Adapte les types de colonnes pour éviter l'erreur\n"
        "- Si l'erreur indique une incompatibilité de type, corrige le cast\n\n"
        "Réponds UNIQUEMENT avec ce JSON:\n"
        "{\n"
        '  "sql": "SQL ClickHouse alternatif",\n'
        '  "explanation": "explication de l\'alternative",\n'
        '  "action_type": "type_action",\n'
        '  "target_table": "BOT_ETL_nom ou null",\n'
        '  "needs_user_input": false,\n'
        '  "question": null\n'
        "}\n"
    )
    try:
        raw = _call_llm(system_prompt, [{"role": "user", "content": user_content}], temperature=0.1)
        result = _parse_llm_json(raw)
        if not result:
            return {"sql": None, "explanation": "Alternative LLM parsing failed"}
        return result
    except Exception as exc:
        return {"sql": None, "explanation": f"Alternative generation error: {exc}"}


def _etl_execute_sql(client, database: str, sql: str) -> tuple[bool, str, list]:
    """Execute a SQL statement and return (ok, error, preview_rows)."""
    if not sql or not sql.strip():
        return False, "SQL vide.", []
    # Security: ETL agent can only write to BOT_ETL_* tables
    ok, reason = _etl_sql_safe(sql)
    if not ok:
        return False, reason, []
    try:
        sql_upper = sql.strip().upper()
        if any(sql_upper.startswith(k) for k in ("SELECT", "SHOW", "DESCRIBE", "EXPLAIN", "WITH")):
            res = client.query(sql)
            rows = _rows_to_dicts(res)[:10]
            return True, "", rows
        else:
            client.command(sql)
            return True, "", []
    except Exception as exc:
        return False, str(exc), []


def _etl_sql_safe(sql: str) -> tuple[bool, str]:
    """Security guard for ETL agent: writes only allowed on BOT_ETL_* tables."""
    sql_stripped = sql.strip()
    sql_upper = sql_stripped.upper()

    # Read-only
    if any(sql_upper.startswith(k) for k in ("SELECT", "SHOW", "DESCRIBE", "EXPLAIN", "WITH")):
        return True, ""

    # Never allow DROP DATABASE/SCHEMA
    if re.match(r"DROP\s+(DATABASE|SCHEMA)\b", sql_stripped, re.IGNORECASE):
        return False, "DROP DATABASE/SCHEMA non autorisé par l'agent ETL."

    def _require_bot_etl(table_name: str, op: str) -> tuple[bool, str] | None:
        clean = table_name.strip("`").split(".")[-1].strip("`")
        if not clean.upper().startswith("BOT_ETL_"):
            return (
                False,
                f"[SÉCURITÉ ETL] {op} refusé : '{table_name}' n'est pas une table BOT_ETL_. "
                "L'agent ETL ne modifie que ses propres tables BOT_ETL_*.",
            )
        return None

    for pattern, op in [
        (r"DROP\s+TABLE\s+(?:IF\s+EXISTS\s+)?(?:(?:`[^`]+`|\w+)\.)?(`?\w+`?)", "DROP TABLE"),
        (r"ALTER\s+TABLE\s+(?:(?:`[^`]+`|\w+)\.)?(`?\w+`?)", "ALTER TABLE"),
        (r"TRUNCATE\s+(?:TABLE\s+)?(?:(?:`[^`]+`|\w+)\.)?(`?\w+`?)", "TRUNCATE"),
        (r"DELETE\s+FROM\s+(?:(?:`[^`]+`|\w+)\.)?(`?\w+`?)", "DELETE FROM"),
        (r"CREATE\s+TABLE\s+(?:IF\s+NOT\s+EXISTS\s+)?(?:(?:`[^`]+`|\w+)\.)?(`?\w+`?)", "CREATE TABLE"),
        (r"INSERT\s+INTO\s+(?:(?:`[^`]+`|\w+)\.)?(`?\w+`?)", "INSERT INTO"),
    ]:
        m = re.match(pattern, sql_stripped, re.IGNORECASE)
        if m:
            err = _require_bot_etl(m.group(1), op)
            if err:
                return err
            return True, ""

    return True, ""


def _etl_insert_dataframe_to_ch(client, database: str, table: str, file_path: str) -> tuple[bool, int, str]:
    """
    Read file_path fully, then insert all rows into database.table via clickhouse-connect.
    Returns (ok, rows_inserted, error).
    """
    cols, rows, row_count, err = _etl_parse_file_full(file_path)
    if err:
        return False, 0, err
    if not rows:
        return True, 0, ""
    try:
        import pandas as pd
        df = pd.DataFrame(rows)
        # Get existing table schema from ClickHouse to cast dtypes
        desc = client.query(f"DESCRIBE TABLE `{database}`.`{table}`")
        ch_cols = {r[0]: r[1] for r in desc.result_rows}
        for col in df.columns:
            if col not in ch_cols:
                continue
            ch_type = ch_cols[col].upper()
            try:
                if "INT" in ch_type:
                    df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype("int64")
                elif "FLOAT" in ch_type or "DOUBLE" in ch_type:
                    df[col] = pd.to_numeric(df[col], errors="coerce")
                elif "DATETIME" in ch_type:
                    df[col] = pd.to_datetime(df[col], errors="coerce")
                elif "DATE" in ch_type and "DATETIME" not in ch_type:
                    df[col] = pd.to_datetime(df[col], errors="coerce").dt.date
                else:
                    df[col] = df[col].astype(str).replace("None", "").replace("nan", "")
            except Exception:
                df[col] = df[col].astype(str).replace("None", "").replace("nan", "")
        # Only insert columns that exist in the table
        insert_cols = [c for c in df.columns if c in ch_cols]
        if not insert_cols:
            return False, 0, "Aucune colonne commune entre le fichier et la table cible."
        df_insert = df[insert_cols]
        client.insert_df(f"`{database}`.`{table}`", df_insert)
        return True, len(df_insert), ""
    except Exception as exc:
        return False, 0, str(exc)


def _etl_synthesize(session: dict) -> dict:
    """LLM generates a synthesis of the ETL operation."""
    plan = session.get("plan", {})
    action_log = session.get("action_log", [])
    created_tables = session.get("created_tables", [])

    log_text = []
    for entry in action_log:
        log_text.append(
            f"Étape {entry['step_id']} ({'OK' if entry['ok'] else 'ECHEC'}): "
            f"{entry['description']} → {str(entry.get('result_preview',''))[:150]}"
        )

    system_prompt = (
        "Tu es un expert ETL. Tu génères une synthèse claire et structurée d'une opération ETL. "
        "Tu réponds UNIQUEMENT avec du JSON valide."
    )
    user_content = (
        f"Objectif: {plan.get('objective','')}\n\n"
        f"Log des actions:\n" + "\n".join(log_text) + "\n\n"
        f"Tables créées: {', '.join(created_tables)}\n\n"
        "Génère une synthèse de l'opération ETL.\n"
        "Réponds avec:\n"
        "{\n"
        '  "summary": "résumé exécutif de l\'opération",\n'
        '  "tables_created": [{"name": "BOT_ETL_...", "description": "...", "row_count": N}],\n'
        '  "key_points": ["point clé 1", "point clé 2"],\n'
        '  "warnings": ["avertissement éventuel"],\n'
        '  "next_steps": ["suggestion d\'utilisation 1"]\n'
        "}"
    )
    raw = _call_llm(system_prompt, [{"role": "user", "content": user_content}], temperature=0.2)
    result = _parse_llm_json(raw)
    if not result:
        ok_count = sum(1 for e in action_log if e.get("ok"))
        result = {
            "summary": f"Opération ETL terminée: {ok_count}/{len(action_log)} étapes réussies.",
            "tables_created": [{"name": t, "description": "", "row_count": 0} for t in created_tables],
            "key_points": [],
            "warnings": [],
            "next_steps": [],
        }
    return result


# ---------------------------------------------------------------------------
# ETL Agent — main entry point
# ---------------------------------------------------------------------------

def _run_etl_agent():
    """Main handler for the ETL agent chat endpoint."""
    data = request.get_json(silent=True) or {}
    messages = data.get("messages", [])
    params = data.get("params", {})
    session_id = data.get("session_id", "")

    folder_path = str(params.get("folder_path", "")).strip()
    recursive = str(params.get("recursive", "non")).lower() in ("oui", "true", "1", "yes")
    database = str(params.get("database", "")).strip() or clickhouse_config.get("database", "default")
    max_actions = int(params.get("max_actions", 15))
    max_actions = max(1, min(30, max_actions))

    # Session management
    if not session_id or session_id not in _etl_sessions:
        session_id = str(uuid.uuid4())
        session = {
            "status": "new",
            "folder_path": folder_path,
            "recursive": recursive,
            "database": database,
            "max_actions": max_actions,
            "messages": [],
            "user_request": "",
            "plan": None,
            "action_log": [],
            "action_index": 0,
            "created_tables": [],
            "files_info": [],
            "pending_question": None,
            "pending_step_index": None,
            "user_context": {},
        }
        _etl_sessions[session_id] = session
    else:
        session = _etl_sessions[session_id]
        # Update mutable params from UI
        if folder_path:
            session["folder_path"] = folder_path
        if params.get("database"):
            session["database"] = database
        session["max_actions"] = max_actions
        session["recursive"] = recursive

    user_message = messages[-1]["content"].strip() if messages else ""
    session["messages"] = messages

    status = session["status"]
    result = {}

    try:
        # ── Handle user answer to a pending question ──────────────────────────
        if status == "awaiting_user" and user_message:
            session["user_context"][f"answer_step_{session.get('pending_step_index', '?')}"] = user_message
            # Store for cleanup confirmation
            if session.get("pending_question", {}).get("is_cleanup_confirmation"):
                if user_message.lower() in ("oui", "yes", "1", "confirmer", "confirm", "drop"):
                    tables = session.get("pending_drop_tables", [])
                    try:
                        client = get_clickhouse_client()
                        dropped = []
                        for tbl in tables:
                            try:
                                client.command(f"DROP TABLE IF EXISTS `{session['database']}`.`{tbl}`")
                                dropped.append(tbl)
                                if tbl in session["created_tables"]:
                                    session["created_tables"].remove(tbl)
                            except Exception:
                                pass
                        session["status"] = "executing"
                        session["action_index"] += 1
                        session["action_log"].append({
                            "step_id": session.get("pending_step_index", 0),
                            "description": f"Suppression confirmée: {', '.join(dropped)}",
                            "action_detail": f"DROP TABLE {', '.join(dropped)}",
                            "ok": True,
                            "result_preview": f"{len(dropped)} table(s) supprimée(s).",
                        })
                        result = {
                            "content": f"Tables supprimées: {', '.join(dropped)}.",
                            "status": "executing",
                            "action_log": session["action_log"],
                            "created_tables": session["created_tables"],
                        }
                        # Continue execution
                        session["status"] = "executing"
                    except Exception as exc:
                        result = {"content": f"Erreur suppression: {exc}", "status": "error"}
                else:
                    session["action_index"] += 1
                    session["status"] = "executing"
                    result = {
                        "content": "Suppression annulée. L'agent continue.",
                        "status": "executing",
                        "action_log": session["action_log"],
                    }
                session["pending_question"] = None
                session["pending_drop_tables"] = []
            else:
                session["status"] = "executing"
                session["pending_question"] = None

        # ── New session: browse files first ───────────────────────────────────
        if status == "new":
            if not folder_path:
                session["status"] = "awaiting_user"
                session["pending_question"] = {
                    "text": "Quel est le chemin du dossier contenant les fichiers à importer ?",
                    "choices": [],
                }
                _etl_sessions[session_id] = session
                return jsonify({
                    "content": "Veuillez spécifier le chemin du dossier source dans les paramètres.",
                    "status": "awaiting_user",
                    "question": session["pending_question"],
                    "session_id": session_id,
                })

            if not os.path.isdir(folder_path):
                _etl_sessions[session_id] = session
                return jsonify({
                    "content": f"Dossier introuvable: `{folder_path}`. Vérifiez le chemin dans les paramètres.",
                    "status": "error",
                    "session_id": session_id,
                })

            # List files
            files = _etl_list_files(folder_path, recursive)
            if not files:
                _etl_sessions[session_id] = session
                return jsonify({
                    "content": f"Aucun fichier de données trouvé dans `{folder_path}` (extensions supportées: csv, xlsx, parquet, txt, json).",
                    "status": "error",
                    "session_id": session_id,
                })

            # Infer schema for each file
            files_info = []
            for f in files[:20]:  # limit to 20 files
                cols, sample_rows, parse_err = _etl_parse_file(f["path"])
                schema, schema_err = _etl_infer_schema_from_file(f["path"])
                # Get approximate row count
                try:
                    import pandas as pd
                    import pathlib
                    ext = pathlib.Path(f["path"]).suffix.lower()
                    if ext == ".parquet":
                        import pyarrow.parquet as pq
                        meta = pq.read_metadata(f["path"])
                        row_count = meta.num_rows
                    elif ext in (".csv", ".tsv", ".txt"):
                        row_count = sum(1 for _ in open(f["path"], "r", encoding="utf-8", errors="ignore")) - 1
                    else:
                        row_count = "?"
                except Exception:
                    row_count = "?"

                files_info.append({
                    **f,
                    "columns": cols,
                    "sample_rows": sample_rows,
                    "schema": schema,
                    "parse_error": parse_err or schema_err,
                    "row_count": row_count,
                })

            session["files_info"] = files_info
            session["user_request"] = user_message

            # Display files found and ask for confirmation/instruction
            files_display = []
            for fi in files_info:
                err_note = f" ⚠ {fi['parse_error']}" if fi.get("parse_error") else ""
                files_display.append(
                    f"• {fi['relative']} — {len(fi.get('columns', []))} colonnes, "
                    f"~{fi['row_count']} lignes, {fi['size_human']}{err_note}"
                )

            content = (
                f"J'ai trouvé **{len(files_info)} fichier(s)** dans `{folder_path}`:\n\n"
                + "\n".join(files_display)
                + "\n\nQue souhaitez-vous faire avec ces fichiers ? Décrivez votre besoin "
                "(ex: *importer dans une table clients, enrichir avec la table t_ref sur le champ id*, etc.)"
            )

            session["status"] = "awaiting_request"
            _etl_sessions[session_id] = session
            return jsonify({
                "content": content,
                "status": "awaiting_request",
                "files_found": files_info,
                "session_id": session_id,
            })

        # ── User described what they want — generate plan ─────────────────────
        if status == "awaiting_request":
            session["user_request"] = user_message
            try:
                client = get_clickhouse_client()
                db_schema = _etl_get_db_schema(client, database)
            except Exception as exc:
                db_schema = f"Impossible de se connecter à ClickHouse: {exc}"
                client = None

            kb_context = _etl_get_knowledge_context()
            plan = _etl_plan(session, session["files_info"], db_schema, kb_context)
            session["plan"] = plan
            session["status"] = "plan_ready"
            _etl_sessions[session_id] = session

            steps_display = []
            for s in plan.get("steps", []):
                needs_input = " ❓ (confirmation requise)" if s.get("needs_user_input") else ""
                steps_display.append(f"  {s['id']}. [{s['type']}] {s['description']}{needs_input}")

            content = (
                f"Plan ETL généré — **{len(plan.get('steps', []))} étapes** :\n\n"
                f"**Objectif:** {plan.get('objective','')}\n"
                f"**Approche:** {plan.get('approach','')}\n\n"
                + "\n".join(steps_display)
                + "\n\nRépondez **oui** pour démarrer l'exécution, ou précisez des ajustements."
            )
            return jsonify({
                "content": content,
                "status": "plan_ready",
                "plan": plan,
                "session_id": session_id,
            })

        # ── User confirmed plan — start execution ─────────────────────────────
        if status == "plan_ready":
            if user_message.lower() in ("oui", "yes", "ok", "go", "start", "démarrer", "lancer", "valider", "continuer"):
                session["status"] = "executing"
                session["action_index"] = 0
            else:
                # User wants to adjust — replan
                session["user_request"] = user_message
                try:
                    client = get_clickhouse_client()
                    db_schema = _etl_get_db_schema(client, database)
                except Exception:
                    db_schema = ""
                kb_context = _etl_get_knowledge_context()
                plan = _etl_plan(session, session["files_info"], db_schema, kb_context)
                session["plan"] = plan
                _etl_sessions[session_id] = session
                steps_display = [
                    f"  {s['id']}. [{s['type']}] {s['description']}"
                    for s in plan.get("steps", [])
                ]
                return jsonify({
                    "content": (
                        f"Plan mis à jour — {len(plan.get('steps', []))} étapes :\n"
                        + "\n".join(steps_display)
                        + "\n\nRépondez **oui** pour démarrer."
                    ),
                    "status": "plan_ready",
                    "plan": plan,
                    "session_id": session_id,
                })

        # ── Execution loop ────────────────────────────────────────────────────
        if status in ("executing",):
            try:
                client = get_clickhouse_client()
            except Exception as exc:
                _etl_sessions[session_id] = session
                return jsonify({
                    "content": f"Impossible de se connecter à ClickHouse: {exc}",
                    "status": "error",
                    "session_id": session_id,
                }), 500

            plan = session.get("plan", {})
            steps = plan.get("steps", [])
            action_index = session.get("action_index", 0)

            # Execute remaining steps
            while action_index < len(steps) and len(session["action_log"]) < max_actions:
                step = steps[action_index]
                _log(f"ETL executing step {step['id']}/{len(steps)}: {step['description'][:60]}", source="etl-agent")
                action_index += 1
                session["action_index"] = action_index

                step_type = step.get("type", "")

                # ask_user step
                if step_type == "ask_user" or step.get("needs_user_input"):
                    question = step.get("question") or {
                        "text": step.get("description", "Que souhaitez-vous faire ?"),
                        "choices": [],
                    }
                    session["status"] = "awaiting_user"
                    session["pending_question"] = question
                    session["pending_step_index"] = step["id"]
                    _etl_sessions[session_id] = session
                    return jsonify({
                        "content": f"Question étape {step['id']}: {question.get('text','')}",
                        "status": "awaiting_user",
                        "question": question,
                        "action_log": session["action_log"],
                        "created_tables": session["created_tables"],
                        "session_id": session_id,
                    })

                # Generate action
                action = _etl_generate_action(session, step, client, database)

                if action.get("needs_user_input"):
                    question = action.get("question") or {"text": step["description"], "choices": []}
                    session["status"] = "awaiting_user"
                    session["pending_question"] = question
                    session["pending_step_index"] = step["id"]
                    _etl_sessions[session_id] = session
                    return jsonify({
                        "content": f"Étape {step['id']} — {question.get('text', step['description'])}",
                        "status": "awaiting_user",
                        "question": question,
                        "action_log": session["action_log"],
                        "created_tables": session["created_tables"],
                        "plan": plan,
                        "session_id": session_id,
                    })

                # drop_table requires user confirmation
                if step_type == "drop_table" or (action.get("sql", "") or "").upper().strip().startswith("DROP TABLE"):
                    target = action.get("target_table") or step.get("target_table", "")
                    session["status"] = "awaiting_user"
                    session["pending_question"] = {
                        "text": f"Confirmez-vous la suppression de la table `{target}` ? Cette action est irréversible.",
                        "choices": [
                            {"label": "Oui, supprimer", "value": "oui"},
                            {"label": "Non, annuler", "value": "non"},
                        ],
                        "is_cleanup_confirmation": True,
                    }
                    session["pending_drop_tables"] = [target] if target else []
                    session["pending_step_index"] = step["id"]
                    _etl_sessions[session_id] = session
                    return jsonify({
                        "content": f"Confirmation requise avant suppression de `{target}`.",
                        "status": "awaiting_user",
                        "question": session["pending_question"],
                        "action_log": session["action_log"],
                        "session_id": session_id,
                    })

                # Handle insert_data: use pandas to insert
                if step_type == "insert_data" and action.get("sql", "").upper().strip().startswith("CREATE TABLE"):
                    # Create table first
                    ok, err, rows = _etl_execute_sql(client, database, action["sql"])
                    if ok:
                        tbl_name = action.get("target_table") or step.get("target_table", "")
                        tbl_clean = tbl_name.strip("`").split(".")[-1].strip("`")
                        if _etl_is_bot_etl_table(tbl_clean) and tbl_clean not in session["created_tables"]:
                            session["created_tables"].append(tbl_clean)
                        # Now insert data from file
                        source_file = step.get("source_file") or ""
                        if source_file:
                            ins_ok, ins_count, ins_err = _etl_insert_dataframe_to_ch(
                                client, database, tbl_clean, source_file
                            )
                            entry = {
                                "step_id": step["id"],
                                "description": step["description"],
                                "action_detail": f"CREATE TABLE + INSERT {ins_count} lignes depuis {os.path.basename(source_file)}",
                                "ok": ins_ok,
                                "result_preview": f"{ins_count} lignes insérées." if ins_ok else ins_err,
                                "rows_affected": ins_count,
                            }
                        else:
                            entry = {
                                "step_id": step["id"],
                                "description": step["description"],
                                "action_detail": action["sql"][:150],
                                "ok": True,
                                "result_preview": "Table créée.",
                                "rows_affected": None,
                            }
                    else:
                        entry = {
                            "step_id": step["id"],
                            "description": step["description"],
                            "action_detail": action.get("sql", "")[:150],
                            "ok": False,
                            "result_preview": err,
                            "rows_affected": None,
                        }
                    session["action_log"].append(entry)
                    continue

                # Standard SQL execution — with retry on failure
                sql = action.get("sql", "")
                if sql:
                    ok, err, rows = _etl_execute_sql(client, database, sql)
                    _log(f"ETL step {step['id']} SQL: {'OK' if ok else 'FAIL: ' + err[:80]}", source="etl-agent")

                    # Retry up to 2 times if the step fails and credits remain
                    retry_count = 0
                    while not ok and retry_count < 2 and len(session["action_log"]) < max_actions:
                        retry_count += 1
                        _log(f"ETL step {step['id']} retry {retry_count} after error: {err[:80]}", source="etl-agent", level="warn")
                        # Re-generate action with error context
                        step_with_error = dict(step)
                        step_with_error["_last_error"] = err
                        step_with_error["_last_sql"] = sql
                        # Inject error into session action log for context
                        session["action_log"].append({
                            "step_id": step["id"],
                            "description": f"[retry {retry_count}] {step['description']}",
                            "action_detail": f"FAILED: {sql[:150]}",
                            "ok": False,
                            "result_preview": err[:200],
                            "rows_affected": None,
                        })
                        # Ask LLM for alternative approach
                        alt_action = _etl_generate_action_with_error(session, step_with_error, client, database, err)
                        sql = alt_action.get("sql", "")
                        if not sql:
                            break
                        ok, err, rows = _etl_execute_sql(client, database, sql)
                        _log(f"ETL step {step['id']} retry {retry_count} result: {'OK' if ok else 'FAIL: ' + err[:60]}", source="etl-agent")

                    # Track created tables
                    if ok and sql.strip().upper().startswith("CREATE TABLE"):
                        m = re.search(
                            r"CREATE\s+TABLE\s+(?:IF\s+NOT\s+EXISTS\s+)?(?:(?:`[^`]+`|\w+)\.)?`?(\w+)`?",
                            sql, re.IGNORECASE
                        )
                        if m:
                            tbl = m.group(1)
                            if _etl_is_bot_etl_table(tbl) and tbl not in session["created_tables"]:
                                session["created_tables"].append(tbl)

                    # If it's insert_data and we also have source file, insert via pandas
                    if step_type == "insert_data" and not sql.strip().upper().startswith("CREATE TABLE"):
                        source_file = step.get("source_file") or ""
                        target_table = action.get("target_table") or step.get("target_table", "")
                        tbl_clean = target_table.strip("`").split(".")[-1].strip("`")
                        if source_file and _etl_is_bot_etl_table(tbl_clean):
                            ins_ok, ins_count, ins_err = _etl_insert_dataframe_to_ch(
                                client, database, tbl_clean, source_file
                            )
                            entry = {
                                "step_id": step["id"],
                                "description": step["description"],
                                "action_detail": f"INSERT pandas {ins_count} lignes depuis {os.path.basename(source_file)}",
                                "ok": ins_ok,
                                "result_preview": f"{ins_count} lignes insérées." if ins_ok else ins_err,
                                "rows_affected": ins_count,
                            }
                            session["action_log"].append(entry)
                            continue

                    preview = rows if rows else (f"Erreur: {err}" if not ok else "OK")
                    entry = {
                        "step_id": step["id"],
                        "description": step["description"],
                        "action_detail": sql[:200] if sql else action.get("explanation", "")[:200],
                        "ok": ok,
                        "result_preview": preview,
                        "rows_affected": None,
                    }
                else:
                    # No SQL — informational step
                    entry = {
                        "step_id": step["id"],
                        "description": step["description"],
                        "action_detail": action.get("explanation", "Étape informative")[:200],
                        "ok": True,
                        "result_preview": action.get("explanation", "")[:200],
                        "rows_affected": None,
                    }
                session["action_log"].append(entry)

            # All steps done → synthesize
            session["status"] = "done"
            synthesis = _etl_synthesize(session)
            session["synthesis"] = synthesis
            _etl_sessions[session_id] = session

            tables_created = session.get("created_tables", [])
            ok_count = sum(1 for e in session["action_log"] if e.get("ok"))
            result = {
                "content": (
                    f"ETL terminé — {ok_count}/{len(session['action_log'])} actions réussies. "
                    f"{len(tables_created)} table(s) créée(s): {', '.join(tables_created)}."
                ),
                "status": "done",
                "plan": plan,
                "action_log": session["action_log"],
                "created_tables": tables_created,
                "synthesis": synthesis,
                "session_id": session_id,
            }
            _etl_sessions[session_id] = session
            return jsonify(result)

        # Fallback if in an unexpected state
        _etl_sessions[session_id] = session
        result["content"] = result.get("content", "En attente de votre réponse.")
        result["status"] = session.get("status", "unknown")
        result["session_id"] = session_id
        result["action_log"] = session.get("action_log", [])
        result["created_tables"] = session.get("created_tables", [])
        if session.get("plan"):
            result["plan"] = session["plan"]
        if session.get("pending_question"):
            result["question"] = session["pending_question"]
        return jsonify(result)

    except Exception as exc:
        session["status"] = "error"
        _etl_sessions[session_id] = session
        return jsonify({
            "error": f"Erreur interne agent ETL: {str(exc)}",
            "session_id": session_id,
            "status": "error",
        }), 500


# ---------------------------------------------------------------------------
# Serve built frontend
# ---------------------------------------------------------------------------
@app.route("/", defaults={"path": ""})
@app.route("/<path:path>")
def serve_frontend(path):
    if path.startswith("api/"):
        return jsonify({"error": "Not found"}), 404
    if os.path.exists(DIST_DIR):
        full_path = os.path.join(DIST_DIR, path)
        if path and os.path.isfile(full_path):
            return send_from_directory(DIST_DIR, path)
        return send_from_directory(DIST_DIR, "index.html")
    return (
        "<h2>Frontend not built.</h2><p>Run <code>npm run build</code> first.</p>",
        404,
    )


# ---------------------------------------------------------------------------
if __name__ == "__main__":
    print(f"ClickSense backend running on http://localhost:{PORT}")
    app.run(
        host="0.0.0.0",
        port=PORT,
        debug=os.environ.get("FLASK_DEBUG", "false").lower() == "true",
    )
